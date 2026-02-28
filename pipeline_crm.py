"""
eduit. — Movimentação de Pipeline.

Fase 2: mover negócios para etapas corretas e alterar status (lost/restore)
com base no cruzamento da planilha de matriculados com o CRM.

Regras:
  - Em Curso + Nova Matrícula/Recompra/Retorno → Calouro (in_process)
  - Em Curso + Rematrícula → Veterano (in_process)
  - Em Curso no CRM, ausente na planilha → Sem Rematrícula (in_process)
  - Cancelado/Trancado/Transferido/outro → Perdido (lost)
  - Lost no CRM mas Em Curso na planilha → Restaurar + mover

Endpoints batch (CRM API):
  POST /businesses/actions/move     {ids, destinationStageId}
  POST /businesses/actions/lose     {ids, lossReasonId, justification}
  POST /businesses/actions/restore  {ids}

Uso:
    python pipeline_crm.py --dry-run       # (padrão) resumo
    python pipeline_crm.py --test          # testa 1 de cada ação
    python pipeline_crm.py --execute       # executa tudo em batch
"""

import sys
import io
import os
import csv
import json
import time
import logging
from datetime import date, datetime, timezone, timedelta
from pathlib import Path
from collections import Counter

import warnings
import requests
import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

BRT = timezone(timedelta(hours=-3))

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

load_dotenv(Path(__file__).parent / ".env")

API_BASE = "https://api.g1.datacrazy.io/api/v1"
API_TOKEN = os.getenv("DATACRAZY_API_TOKEN", "")

DB_DSN = dict(
    host=os.getenv("DB_HOST", "localhost"),
    port=os.getenv("DB_PORT", "5432"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASS"),
    dbname=os.getenv("DB_NAME", "dcz_sync"),
)

REPORTS_DIR = Path(__file__).parent / "reports"
LOG_DIR = Path(__file__).parent / "logs"

BIZ_FIELD_IDS = {
    "RGM":           "2ac4e30f-cfd7-435f-b688-fbce27f76c38",
    "Situacao":      "fd08d44b-a4a5-4343-b7a9-37f75e2c1caa",
    "DataMatricula": "bf93a8e9-42c0-4517-8518-6f604746a300",
    "Nivel":         "233fcf6f-0bed-49d7-89a1-d1cd54fb9c12",
}

ROUTE_RATE_LIMIT = 60
BATCH_SIZE = 50

STAGE_NAMES = {
    "calouro":               ["Calouro", "Calouros", "CALOURO"],
    "veterano":              ["Veterano", "Veteranos", "VETERANO"],
    "inadimplente":          ["Inadimplente", "INADIMPLENTE", "Inadimplentes"],
    "sem_remat_adimplente":  ["Sem Rematricula Adimplente", "SEM REMATRICULA ADIMPLENTE", "Sem Rematrícula Adimplente"],
    "sem_remat_inadimplente":["Sem Rematricula Inadimplente", "SEM REMATRICULA INADIMPLENTE", "Sem Rematrícula Inadimplente"],
    "perdido":               ["Perdido", "Perdidos", "PERDIDO"],
}

LOSS_REASON_NAMES = {
    "cancelado":        "Cancelado",
    "trancado":         "Trancado",
    "transferido":      "Transferido",
    "concluinte":       "Concluinte",
    "sem_rematricula":  "Sem Rematrícula",
    "outros":           "Outros",
}

LOSS_JUSTIFICATIONS = {
    "cancelado":        "Matrícula cancelada conforme base acadêmica",
    "trancado":         "Matrícula trancada conforme base acadêmica",
    "transferido":      "Aluno transferido conforme base acadêmica",
    "concluinte":       "Aluno concluiu o curso conforme base acadêmica",
    "sem_rematricula":  "Aluno não realizou rematrícula no ciclo vigente",
    "outros":           "Situação irregular conforme base acadêmica",
}

GRACE_PERIOD_DAYS = 5

CALOURO_TIPOS = {"calouro", "calouro (recompra)", "nova matrícula", "nova matricula", "recompra", "retorno", "regresso (retorno)"}
VETERANO_TIPOS = {"veterano", "rematrícula", "rematricula"}

LOST_SITUACOES = {"cancelado", "trancado", "transferido"}


class _BRTFormatter(logging.Formatter):
    def formatTime(self, record, datefmt=None):
        dt = datetime.fromtimestamp(record.created, tz=BRT)
        return dt.strftime(datefmt or "%H:%M:%S")


logging.basicConfig(level=logging.INFO)
_handler = logging.StreamHandler()
_handler.setFormatter(_BRTFormatter("%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S"))
logging.root.handlers = [_handler]
log = logging.getLogger("pipeline")


# ---------------------------------------------------------------------------
# API Client
# ---------------------------------------------------------------------------

class ApiClient:
    def __init__(self, rate_limit=None):
        self.s = requests.Session()
        self.s.headers["Authorization"] = f"Bearer {API_TOKEN}"
        self.s.headers["Content-Type"] = "application/json"
        rl = rate_limit or ROUTE_RATE_LIMIT
        self._remaining = rl
        self._reset = 0
        self._last_req = 0.0
        self.total_calls = 0
        self.base_delay = 60.0 / rl

    def _throttle(self):
        if self._remaining <= 5 and self._reset > 0:
            wait = self._reset + 1
            log.warning("Rate-limit crítico (%d restantes) — pausando %ds",
                        self._remaining, wait)
            time.sleep(wait)
            return
        elapsed = time.monotonic() - self._last_req
        if elapsed < self.base_delay:
            time.sleep(self.base_delay - elapsed)

    def _read_headers(self, r):
        self._remaining = int(r.headers.get("X-RateLimit-Remaining", self._remaining))
        self._reset = int(r.headers.get("X-RateLimit-Reset", 0))

    def _request(self, method, path, payload=None, retries=4):
        url = f"{API_BASE}{path}"
        for attempt in range(retries):
            self._throttle()
            self._last_req = time.monotonic()
            self.total_calls += 1

            r = self.s.request(method, url, json=payload, timeout=30)

            if r.status_code == 429:
                retry = int(r.headers.get("Retry-After", 30))
                log.warning("429 — Retry-After %ds (tentativa %d/%d)", retry, attempt + 1, retries)
                time.sleep(retry + 1)
                continue

            self._read_headers(r)

            if r.status_code >= 400:
                return {"ok": False, "status": r.status_code, "body": r.text[:500]}

            try:
                body = r.json()
            except Exception:
                body = r.text[:200]
            return {"ok": True, "status": r.status_code, "body": body}

        return {"ok": False, "status": 429, "body": "Falha após tentativas"}

    def get(self, path, params=None):
        url = f"{API_BASE}{path}"
        self._throttle()
        self._last_req = time.monotonic()
        self.total_calls += 1
        r = self.s.get(url, params=params, timeout=30)
        self._read_headers(r)
        if r.status_code >= 400:
            return {"ok": False, "status": r.status_code, "body": r.text[:500]}
        return {"ok": True, "status": r.status_code, "body": r.json()}

    def post(self, path, payload):
        return self._request("POST", path, payload)

    def move_businesses(self, ids, destination_stage_id):
        """POST /businesses/actions/move"""
        return self.post("/businesses/actions/move", {
            "ids": ids,
            "destinationStageId": destination_stage_id,
        })

    def lose_businesses(self, ids, loss_reason_id, justification=""):
        """POST /businesses/actions/lose"""
        payload = {"ids": ids, "lossReasonId": loss_reason_id}
        if justification:
            payload["justification"] = justification
        return self.post("/businesses/actions/lose", payload)

    def restore_businesses(self, ids):
        """POST /businesses/actions/restore"""
        return self.post("/businesses/actions/restore", {"ids": ids})

    def patch(self, path, payload):
        return self._request("PATCH", path, payload)

    def patch_business_tags(self, biz_id, tag_ids):
        """PATCH /businesses/{id} com tagIds."""
        return self.patch(f"/businesses/{biz_id}", {"tagIds": tag_ids})

    def create_tag(self, name):
        """POST /tags"""
        return self.post("/tags", {"name": name})

    def create_loss_reason(self, name):
        """POST /business-loss-reasons"""
        return self.post("/business-loss-reasons", {
            "name": name,
            "requiredJustification": False,
        })


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def resolve_tag_id(conn, api, tag_name):
    """Encontra ou cria uma tag pelo nome. Retorna o ID."""
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, data->>'name' AS nome FROM tags")
    for row in cur.fetchall():
        if (row["nome"] or "").strip().upper() == tag_name.upper():
            cur.close()
            log.info("  Tag '%s' encontrada → %s", tag_name, row["id"][:12] + "…")
            return row["id"]
    cur.close()

    log.info("  Tag '%s' não encontrada — criando...", tag_name)
    r = api.create_tag(tag_name)
    if r["ok"]:
        tid = r["body"].get("id", "")
        log.info("  Tag '%s' criada → %s", tag_name, tid[:12] + "…")
        return tid
    log.error("  Falha ao criar tag '%s': %s", tag_name, r["body"])
    return None


def get_conn():
    return psycopg2.connect(**DB_DSN)


def get_biz_field(biz_data, field_id):
    for f in biz_data.get("additionalFields", []):
        af = f.get("additionalField", {})
        if isinstance(af, dict) and af.get("id") == field_id:
            return f.get("value", "")
        if isinstance(af, str) and af == field_id:
            return f.get("value", "")
    return ""


def lead_name(biz_data):
    lead = biz_data.get("lead")
    if isinstance(lead, dict):
        return lead.get("name", "")
    return ""


def _col_find(col_map, *candidates):
    for c in candidates:
        if c in col_map:
            return c
    for c in candidates:
        cl = c.lower()
        for k in col_map:
            if k and cl in k.lower():
                return k
    return candidates[0]


# ---------------------------------------------------------------------------
# Carregamento
# ---------------------------------------------------------------------------

def load_excel():
    """Carrega planilha → {RGM: {situacao, tipo_matricula, nome}}"""
    log.info("Carregando planilha...")
    xlsx = None
    for f in Path(__file__).parent.iterdir():
        if f.suffix.lower() == ".xlsx" and "matriculados" in f.name.lower():
            xlsx = f
            break
    if not xlsx:
        raise FileNotFoundError("Planilha de matriculados não encontrada")

    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    ws = wb["Export"]
    raw_header = [cell.value for cell in ws[1]]
    raw_col = {h: i for i, h in enumerate(raw_header) if h}

    col_rgm = _col_find(raw_col, "RGM")
    col_sit = _col_find(raw_col, "Situação Matrícula", "Situa")
    col_tipo = _col_find(raw_col, "Tipo Matrícula", "Tipo Matr")
    col_nome = _col_find(raw_col, "Nome")

    if col_rgm not in raw_col:
        raise ValueError(f"Coluna RGM não encontrada. Colunas: {list(raw_col.keys())[:10]}")

    idx_rgm = raw_col[col_rgm]
    idx_sit = raw_col.get(col_sit)
    idx_tipo = raw_col.get(col_tipo)
    idx_nome = raw_col.get(col_nome)

    by_rgm = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rgm = str(row[idx_rgm]).strip() if row[idx_rgm] else ""
        if not rgm:
            continue
        sit = str(row[idx_sit] or "").strip() if idx_sit is not None else ""
        tipo = str(row[idx_tipo] or "").strip() if idx_tipo is not None else ""
        nome = str(row[idx_nome] or "").strip() if idx_nome is not None else ""
        by_rgm[rgm] = {
            "situacao": sit,
            "tipo_matricula": tipo,
            "nome": nome,
        }

    wb.close()
    log.info("  %d RGMs únicos na planilha", len(by_rgm))
    return by_rgm


def load_sem_rematricula_snapshot(conn):
    """Carrega RGMs do snapshot de sem_rematricula separados por status financeiro."""
    adimplentes = set()
    inadimplentes = set()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id FROM xl_snapshots WHERE tipo='sem_rematricula' ORDER BY id DESC LIMIT 1")
        snap = cur.fetchone()
        if not snap:
            cur.close()
            return adimplentes, inadimplentes
        cur.execute("""
            SELECT data->>'rgm_digits' AS rgm, data->>'status_financeiro' AS sf
            FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''
        """, (snap["id"],))
        for r in cur.fetchall():
            if r["sf"] == "inadimplente":
                inadimplentes.add(r["rgm"])
            else:
                adimplentes.add(r["rgm"])
        cur.close()
        log.info("  Snapshot sem_rematricula: %d adimplentes, %d inadimplentes",
                 len(adimplentes), len(inadimplentes))
        return adimplentes, inadimplentes
    except Exception as e:
        log.warning("  Sem snapshot de sem_rematricula: %s", e)
        return adimplentes, inadimplentes


def load_inadimplentes_snapshot(conn):
    """Carrega RGMs do snapshot de inadimplentes (se existir)."""
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id FROM xl_snapshots WHERE tipo='inadimplentes' ORDER BY id DESC LIMIT 1")
        snap = cur.fetchone()
        if not snap:
            cur.close()
            return set()
        cur.execute("SELECT data->>'rgm_digits' AS rgm FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (snap["id"],))
        rgms = {r["rgm"] for r in cur.fetchall()}
        cur.close()
        log.info("  Snapshot inadimplentes: %d RGMs", len(rgms))
        return rgms
    except Exception as e:
        log.warning("  Sem snapshot de inadimplentes: %s", e)
        return set()


def load_concluintes_snapshot(conn):
    """Carrega RGMs do snapshot de concluintes (se existir)."""
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id FROM xl_snapshots WHERE tipo='concluintes' ORDER BY id DESC LIMIT 1")
        snap = cur.fetchone()
        if not snap:
            cur.close()
            return set()
        cur.execute("SELECT data->>'rgm_digits' AS rgm FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (snap["id"],))
        rgms = {r["rgm"] for r in cur.fetchall()}
        cur.close()
        log.info("  Snapshot concluintes: %d RGMs", len(rgms))
        return rgms
    except Exception as e:
        log.warning("  Sem snapshot de concluintes: %s", e)
        return set()


def load_ciclo_cutoffs(conn):
    """Retorna {nivel: dt_inicio} do ciclo vigente por nível.

    Alunos com data_matricula anterior a dt_inicio são elegíveis
    para 'Sem Rematrícula'. Se não houver ciclo configurado para
    um nível, esse nível não aparece no dict (sem trava).
    """
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT DISTINCT ON (nivel) nivel, nome, dt_inicio, dt_fim "
            "FROM ciclos "
            "WHERE dt_inicio <= CURRENT_DATE AND dt_fim >= CURRENT_DATE "
            "ORDER BY nivel, dt_inicio DESC"
        )
        rows = cur.fetchall()
        if not rows:
            cur.execute(
                "SELECT DISTINCT ON (nivel) nivel, nome, dt_inicio, dt_fim "
                "FROM ciclos "
                "ORDER BY nivel, dt_inicio DESC"
            )
            rows = cur.fetchall()
        cur.close()

        cutoffs = {}
        for r in rows:
            cutoffs[r["nivel"]] = r["dt_inicio"]
            log.info("  Ciclo %s (%s): corte em %s", r["nivel"], r["nome"], r["dt_inicio"])

        return cutoffs
    except Exception as e:
        log.warning("  Sem ciclos configurados: %s", e)
        return {}


def load_pipeline_stages(conn):
    """Carrega etapas do pipeline do banco local."""
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT ps.id, ps.pipeline_id,
               ps.data->>'name' AS nome,
               p.data->>'name' AS pipeline_nome
        FROM pipeline_stages ps
        JOIN pipelines p ON p.id = ps.pipeline_id
        ORDER BY p.data->>'name', ps.data->>'order'
    """)
    rows = cur.fetchall()
    cur.close()

    log.info("  %d etapas de pipeline carregadas", len(rows))
    for r in rows:
        log.info("    [%s] %s → %s", r["pipeline_nome"], r["nome"], r["id"][:12] + "…")

    return rows


def resolve_stage_ids(stages):
    """Encontra IDs das etapas necessárias por nome (case-insensitive)."""
    found = {}
    for key, name_variants in STAGE_NAMES.items():
        lower_variants = [n.lower().strip() for n in name_variants]
        for stage in stages:
            sname = (stage["nome"] or "").strip()
            if sname.lower() in lower_variants:
                found[key] = stage["id"]
                log.info("  Etapa '%s' → %s (%s)", key, sname, stage["id"][:12] + "…")
                break

    missing = [k for k in STAGE_NAMES if k not in found]
    if missing:
        log.error("  ETAPAS NÃO ENCONTRADAS: %s", ", ".join(missing))
        log.error("  Crie-as manualmente no CRM antes de executar.")
        return None

    return found


def load_crm_businesses(conn):
    """Carrega todos os negócios com RGM."""
    log.info("Carregando negócios do banco local...")
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, data FROM businesses")
    rows = cur.fetchall()
    cur.close()

    by_rgm = {}
    total_with_rgm = 0
    for row in rows:
        rgm = get_biz_field(row["data"], BIZ_FIELD_IDS["RGM"])
        if rgm and rgm.strip():
            rgm = rgm.strip()
            by_rgm.setdefault(rgm, []).append(row)
            total_with_rgm += 1

    log.info("  %d negócios total | %d com RGM | %d RGMs únicos",
             len(rows), total_with_rgm, len(by_rgm))
    return by_rgm


def ensure_loss_reasons(api):
    """Busca motivos de perda existentes e cria os que faltam."""
    log.info("Buscando motivos de perda...")
    result = api.get("/business-loss-reasons")
    if not result["ok"]:
        log.error("Falha ao buscar motivos: %s", result["body"])
        return None

    body = result["body"]
    existing = body.get("data", body) if isinstance(body, dict) else body
    if isinstance(existing, list):
        by_name = {r["name"]: r["id"] for r in existing}
    else:
        by_name = {}

    log.info("  Motivos existentes: %s", list(by_name.keys()) if by_name else "(nenhum)")

    reason_ids = {}
    for key, name in LOSS_REASON_NAMES.items():
        if name in by_name:
            reason_ids[key] = by_name[name]
            log.info("  '%s' já existe → %s", name, by_name[name][:12] + "…")
        else:
            log.info("  Criando motivo '%s'...", name)
            r = api.create_loss_reason(name)
            if r["ok"]:
                new_id = r["body"].get("id", "")
                reason_ids[key] = new_id
                log.info("  Criado '%s' → %s", name, new_id[:12] + "…")
            else:
                log.error("  Falha ao criar '%s': %s", name, r["body"])
                return None

    return reason_ids


# ---------------------------------------------------------------------------
# Análise
# ---------------------------------------------------------------------------

def _parse_created_at(biz_data):
    """Extrai datetime de createdAt do negócio. Retorna None se inválido."""
    raw = biz_data.get("createdAt", "")
    if not raw:
        return None
    try:
        if "T" in raw:
            return datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return datetime.strptime(raw[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except (ValueError, TypeError):
        return None


def analyze(xl_by_rgm, crm_by_rgm, stage_ids, sem_remat_adim=None,
            sem_remat_inadim=None,
            inadimplentes_rgms=None, concluintes_rgms=None):
    """Cruza planilha com CRM e determina ações necessárias.

    Regras de prioridade:
      1. Won → pular
      2. Na planilha + Em Curso + Inadimplente → etapa INADIMPLENTE
      3. Na planilha + Em Curso → Calouro ou Veterano
      4. Na planilha + situação perdida + Concluinte → Perdido (Concluinte) + TAG
      5. Na planilha + situação perdida → Perdido (por situação)
      6. Não na planilha + Concluinte → Perdido (Concluinte) + TAG
      7. Não na planilha + Sem Rematrícula Adimplente → SEM REMAT ADIMPLENTE
      8. Não na planilha + Sem Rematrícula Inadimplente → SEM REMAT INADIMPLENTE
      9. Não na planilha + criado < 5 dias → pular (grace period)
     10. Não na planilha + nenhum match → Perdido (Sem Rematrícula)

    Returns:
        to_restore, to_move, to_lose, to_tag, stats
    """
    sem_remat_adim = sem_remat_adim or set()
    sem_remat_inadim = sem_remat_inadim or set()
    inadimplentes_rgms = inadimplentes_rgms or set()
    concluintes_rgms = concluintes_rgms or set()

    to_restore = []
    to_move = {sid: [] for sid in stage_ids.values()}
    to_lose = {}
    to_tag = []
    stats = Counter()

    now = datetime.now(timezone.utc)
    grace_cutoff = now - timedelta(days=GRACE_PERIOD_DAYS)

    for rgm, biz_list in crm_by_rgm.items():
        for biz in biz_list:
            biz_id = biz["id"]
            biz_data = biz["data"]
            crm_status = biz_data.get("status", "")
            crm_stage_id = biz_data.get("stageId", "")
            crm_sit = get_biz_field(biz_data, BIZ_FIELD_IDS["Situacao"])
            nome = lead_name(biz_data) or rgm

            if crm_status == "won":
                stats["skip_won"] += 1
                continue

            xl = xl_by_rgm.get(rgm)
            is_concluinte = rgm in concluintes_rgms
            is_inadimplente = rgm in inadimplentes_rgms

            if xl:
                xl_sit = xl["situacao"].lower().strip()
                xl_tipo = xl["tipo_matricula"].lower().strip()
                xl_nome = xl["nome"] or nome

                if xl_sit == "em curso":
                    # Em Curso + Inadimplente → etapa INADIMPLENTE
                    if is_inadimplente and "inadimplente" in stage_ids:
                        target_stage = stage_ids["inadimplente"]
                        stage_label = "Inadimplente"
                    elif xl_tipo in CALOURO_TIPOS:
                        target_stage = stage_ids["calouro"]
                        stage_label = "Calouro"
                    elif xl_tipo in VETERANO_TIPOS:
                        target_stage = stage_ids["veterano"]
                        stage_label = "Veterano"
                    else:
                        target_stage = stage_ids["calouro"]
                        stage_label = f"Calouro (tipo '{xl['tipo_matricula']}')"
                        stats["tipo_desconhecido"] += 1

                    if crm_status == "lost":
                        to_restore.append({
                            "biz_id": biz_id, "rgm": rgm, "nome": xl_nome,
                            "motivo": f"Em Curso na planilha mas lost no CRM",
                        })
                        stats["restore"] += 1

                    if crm_stage_id != target_stage:
                        to_move[target_stage].append({
                            "biz_id": biz_id, "rgm": rgm, "nome": xl_nome,
                            "motivo": stage_label,
                        })
                        stats[f"move_{stage_label.split()[0].lower()}"] += 1
                    else:
                        stats["already_correct"] += 1

                else:
                    # Concluinte tem prioridade sobre o motivo genérico
                    if is_concluinte:
                        reason_key = "concluinte"
                        to_tag.append({"biz_id": biz_id, "rgm": rgm, "nome": xl_nome})
                        stats["tag_concluinte"] += 1
                    elif xl_sit in LOST_SITUACOES:
                        reason_key = xl_sit
                        if reason_key not in LOSS_REASON_NAMES:
                            reason_key = "outros"
                    else:
                        reason_key = "outros"

                    target_stage = stage_ids["perdido"]

                    if crm_stage_id != target_stage:
                        to_move[target_stage].append({
                            "biz_id": biz_id, "rgm": rgm, "nome": xl_nome,
                            "motivo": f"Perdido ({LOSS_REASON_NAMES.get(reason_key, xl['situacao'])})",
                        })
                        stats["move_perdido"] += 1

                    if crm_status != "lost":
                        to_lose.setdefault(reason_key, []).append({
                            "biz_id": biz_id, "rgm": rgm, "nome": xl_nome,
                        })
                        stats[f"lose_{reason_key}"] += 1
                    else:
                        stats["already_lost"] += 1

            else:
                # RGM não está na planilha de matriculados

                # 1) Concluinte → Perdido + TAG
                if is_concluinte:
                    target_stage = stage_ids["perdido"]
                    to_tag.append({"biz_id": biz_id, "rgm": rgm, "nome": nome})
                    stats["tag_concluinte"] += 1

                    if crm_stage_id != target_stage:
                        to_move[target_stage].append({
                            "biz_id": biz_id, "rgm": rgm, "nome": nome,
                            "motivo": "Perdido (Concluinte)",
                        })
                        stats["move_perdido"] += 1

                    if crm_status != "lost":
                        to_lose.setdefault("concluinte", []).append({
                            "biz_id": biz_id, "rgm": rgm, "nome": nome,
                        })
                        stats["lose_concluinte"] += 1
                    else:
                        stats["already_lost"] += 1
                    continue

                # 2) Sem Rematrícula (adimplente ou inadimplente)
                in_adim = rgm in sem_remat_adim
                in_inadim = rgm in sem_remat_inadim
                in_sem_remat = in_adim or in_inadim

                if in_sem_remat:
                    if in_inadim:
                        target_stage = stage_ids["sem_remat_inadimplente"]
                        label = "Sem Rematrícula Inadimplente"
                    else:
                        target_stage = stage_ids["sem_remat_adimplente"]
                        label = "Sem Rematrícula Adimplente"

                    if crm_status == "lost":
                        to_restore.append({
                            "biz_id": biz_id, "rgm": rgm, "nome": nome,
                            "motivo": f"{label} — restaurar de lost",
                        })
                        stats["restore_sem_remat"] += 1

                    if crm_stage_id != target_stage:
                        to_move[target_stage].append({
                            "biz_id": biz_id, "rgm": rgm, "nome": nome,
                            "motivo": label,
                        })
                        stats["move_sem_remat_adim" if in_adim else "move_sem_remat_inadim"] += 1
                    else:
                        stats["already_correct"] += 1
                    continue

                # 3) Grace period — negócio criado nos últimos N dias
                created_at = _parse_created_at(biz_data)
                if created_at and created_at > grace_cutoff:
                    stats["skip_recente"] += 1
                    continue

                # 4) Órfão → Perdido (Sem Rematrícula)
                target_stage = stage_ids["perdido"]
                if crm_stage_id != target_stage:
                    to_move[target_stage].append({
                        "biz_id": biz_id, "rgm": rgm, "nome": nome,
                        "motivo": "Perdido (ausente em todas as planilhas)",
                    })
                    stats["move_perdido_orfao"] += 1

                if crm_status != "lost":
                    to_lose.setdefault("sem_rematricula", []).append({
                        "biz_id": biz_id, "rgm": rgm, "nome": nome,
                    })
                    stats["lose_sem_rematricula"] += 1
                else:
                    stats["already_lost"] += 1

    to_move = {k: v for k, v in to_move.items() if v}

    return to_restore, to_move, to_lose, to_tag, stats


# ---------------------------------------------------------------------------
# Dry-run
# ---------------------------------------------------------------------------

def dry_run_summary(to_restore, to_move, to_lose, to_tag, stats, stage_ids, rate_limit=None):
    stage_id_to_name = {}
    for key, sid in stage_ids.items():
        stage_id_to_name[sid] = key.replace("_", " ").title()

    total_api_calls = 0
    if to_restore:
        total_api_calls += (len(to_restore) + BATCH_SIZE - 1) // BATCH_SIZE
    for items in to_move.values():
        total_api_calls += (len(items) + BATCH_SIZE - 1) // BATCH_SIZE
    for items in to_lose.values():
        total_api_calls += (len(items) + BATCH_SIZE - 1) // BATCH_SIZE
    total_api_calls += len(to_tag)

    print("\n" + "=" * 60)
    print("PIPELINE — DRY-RUN — Resumo")
    print("=" * 60)

    print(f"\n  Restaurar (lost → in_process):  {len(to_restore):,}")
    print(f"\n  Mover para etapas:")
    total_moves = 0
    for sid, items in to_move.items():
        label = stage_id_to_name.get(sid, sid[:12])
        print(f"    {label:25s}: {len(items):,}")
        total_moves += len(items)
    if not to_move:
        print("    (nenhuma movimentação)")

    print(f"\n  Marcar como perdido:")
    total_loses = 0
    for reason_key, items in to_lose.items():
        label = LOSS_REASON_NAMES.get(reason_key, reason_key)
        print(f"    {label:25s}: {len(items):,}")
        total_loses += len(items)
    if not to_lose:
        print("    (nenhum)")

    if to_tag:
        print(f"\n  TAG CONCLUINTE:                 {len(to_tag):,}")

    print(f"\n  Estatísticas:")
    for key, val in sorted(stats.items()):
        print(f"    {key:25s}: {val:,}")

    rl = rate_limit or ROUTE_RATE_LIMIT
    print(f"\n  Total API calls estimadas:     {total_api_calls:,} (batches de {BATCH_SIZE} + tags individuais)")
    est_time = total_api_calls * (60.0 / rl) / 60
    print(f"  Tempo estimado:                {est_time:.1f} min")

    print()
    print("  Para executar: python pipeline_crm.py --execute")
    print("  Para testar:   python pipeline_crm.py --test")
    print("=" * 60)

    REPORTS_DIR.mkdir(exist_ok=True)
    preview = REPORTS_DIR / "pipeline_preview.csv"
    with open(preview, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["acao", "biz_id", "rgm", "nome", "destino_ou_motivo"])
        for item in to_restore:
            w.writerow(["RESTORE", item["biz_id"], item["rgm"], item["nome"], item["motivo"]])
        for sid, items in to_move.items():
            label = stage_id_to_name.get(sid, sid[:12])
            for item in items:
                w.writerow(["MOVE", item["biz_id"], item["rgm"], item["nome"], label])
        for reason_key, items in to_lose.items():
            label = LOSS_REASON_NAMES.get(reason_key, reason_key)
            for item in items:
                w.writerow(["LOSE", item["biz_id"], item["rgm"], item["nome"], label])
        for item in to_tag:
            w.writerow(["TAG", item["biz_id"], item["rgm"], item["nome"], "CONCLUINTE"])
    log.info("  Preview: %s", preview)


# ---------------------------------------------------------------------------
# Execução
# ---------------------------------------------------------------------------

def _batch(items, size):
    for i in range(0, len(items), size):
        yield items[i:i + size]


def test_one(api, to_restore, to_move, to_lose, to_tag, reason_ids, concluinte_tag_id):
    """Testa 1 operação de cada tipo disponível."""
    log.info("=== TESTE: 1 de cada ação ===")
    ok = True

    if to_restore:
        item = to_restore[0]
        log.info("RESTORE: %s (RGM %s, %s)", item["biz_id"][:12], item["rgm"], item["nome"])
        r = api.restore_businesses([item["biz_id"]])
        log.info("  Status: %s | OK: %s", r["status"], r["ok"])
        if not r["ok"]:
            log.error("  FALHOU: %s", r["body"])
            ok = False

    if to_move:
        sid, items = next(iter(to_move.items()))
        item = items[0]
        log.info("MOVE: %s → %s (RGM %s, %s)", item["biz_id"][:12], sid[:12], item["rgm"], item["nome"])
        r = api.move_businesses([item["biz_id"]], sid)
        log.info("  Status: %s | OK: %s", r["status"], r["ok"])
        if not r["ok"]:
            log.error("  FALHOU: %s", r["body"])
            ok = False

    if to_lose:
        reason_key, items = next(iter(to_lose.items()))
        item = items[0]
        rid = reason_ids.get(reason_key, "")
        if rid:
            justification = LOSS_JUSTIFICATIONS.get(reason_key, "Atualização automática via pipeline")
            log.info("LOSE: %s motivo=%s (RGM %s, %s)", item["biz_id"][:12], reason_key, item["rgm"], item["nome"])
            r = api.lose_businesses([item["biz_id"]], rid, justification)
            log.info("  Status: %s | OK: %s", r["status"], r["ok"])
            if not r["ok"]:
                log.error("  FALHOU: %s", r["body"])
                ok = False

    if to_tag and concluinte_tag_id:
        item = to_tag[0]
        log.info("TAG: %s → CONCLUINTE (RGM %s, %s)", item["biz_id"][:12], item["rgm"], item["nome"])
        r = api.patch_business_tags(item["biz_id"], [concluinte_tag_id])
        log.info("  Status: %s | OK: %s", r["status"], r["ok"])
        if not r["ok"]:
            log.error("  FALHOU: %s", r["body"])
            ok = False

    if ok:
        log.info("Teste OK! Execute com: python pipeline_crm.py --execute")
    else:
        log.error("Teste FALHOU em alguma ação. Revise os erros acima.")
    return ok


def execute(api, to_restore, to_move, to_lose, to_tag, reason_ids, concluinte_tag_id):
    """Executa todas as ações em batch: restore → move → lose → tag."""
    LOG_DIR.mkdir(exist_ok=True)
    ts = datetime.now(BRT).strftime("%Y%m%d_%H%M%S")
    log_file = LOG_DIR / f"pipeline_{ts}.csv"

    ok_count = 0
    err_count = 0
    start = time.monotonic()

    with open(log_file, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["timestamp", "acao", "batch_n", "batch_size", "destino_ou_motivo",
                     "status_http", "resultado", "ids_amostra"])

        # 1. Restore
        if to_restore:
            ids = [i["biz_id"] for i in to_restore]
            log.info("=== RESTORE: %d negócios ===", len(ids))
            for bn, batch in enumerate(_batch(ids, BATCH_SIZE), 1):
                log.info("  Restore batch %d (%d IDs)...", bn, len(batch))
                r = api.restore_businesses(batch)
                status = "OK" if r["ok"] else "ERRO"
                w.writerow([
                    datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S"),
                    "RESTORE", bn, len(batch), "",
                    r["status"], status, ";".join(batch[:3]),
                ])
                if r["ok"]:
                    ok_count += len(batch)
                else:
                    body = r.get("body", "")
                    if "not found" in str(body).lower() or r.get("status") == 404:
                        log.warning("  Batch %d: %d negócio(s) não encontrado(s) — provável merge anterior", bn, len(batch))
                    else:
                        err_count += len(batch)
                        log.warning("  ERRO restore batch %d: %s", bn, body[:200])

        # 2. Move
        for sid, items in to_move.items():
            ids = [i["biz_id"] for i in items]
            label = items[0]["motivo"] if items else sid[:12]
            log.info("=== MOVE → %s: %d negócios ===", label, len(ids))
            for bn, batch in enumerate(_batch(ids, BATCH_SIZE), 1):
                log.info("  Move batch %d (%d IDs) → %s...", bn, len(batch), label)
                r = api.move_businesses(batch, sid)
                status = "OK" if r["ok"] else "ERRO"
                w.writerow([
                    datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S"),
                    "MOVE", bn, len(batch), label,
                    r["status"], status, ";".join(batch[:3]),
                ])
                if r["ok"]:
                    ok_count += len(batch)
                else:
                    body = r.get("body", "")
                    if "not found" in str(body).lower() or r.get("status") == 404:
                        log.warning("  Batch %d: %d negócio(s) não encontrado(s) — provável merge anterior", bn, len(batch))
                    else:
                        err_count += len(batch)
                        log.warning("  ERRO move batch %d → %s: %s", bn, label, body[:200])

        # 3. Lose
        for reason_key, items in to_lose.items():
            rid = reason_ids.get(reason_key, "")
            if not rid:
                log.warning("Motivo '%s' sem ID — pulando %d negócios", reason_key, len(items))
                continue

            ids = [i["biz_id"] for i in items]
            label = LOSS_REASON_NAMES.get(reason_key, reason_key)
            justification = LOSS_JUSTIFICATIONS.get(reason_key, "Atualização automática via pipeline")
            log.info("=== LOSE (%s): %d negócios ===", label, len(ids))
            for bn, batch in enumerate(_batch(ids, BATCH_SIZE), 1):
                log.info("  Lose batch %d (%d IDs) motivo=%s...", bn, len(batch), label)
                r = api.lose_businesses(batch, rid, justification)
                status = "OK" if r["ok"] else "ERRO"
                w.writerow([
                    datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S"),
                    "LOSE", bn, len(batch), label,
                    r["status"], status, ";".join(batch[:3]),
                ])
                if r["ok"]:
                    ok_count += len(batch)
                else:
                    body = r.get("body", "")
                    if "not found" in str(body).lower() or r.get("status") == 404:
                        log.warning("  Batch %d: %d negócio(s) não encontrado(s) — provável merge anterior", bn, len(batch))
                    else:
                        err_count += len(batch)
                        log.warning("  ERRO lose batch %d (%s): %s", bn, label, body[:200])

        # 4. Tag CONCLUINTE
        if to_tag and concluinte_tag_id:
            log.info("=== TAG CONCLUINTE: %d negócios ===", len(to_tag))
            for i, item in enumerate(to_tag, 1):
                r = api.patch_business_tags(item["biz_id"], [concluinte_tag_id])
                status = "OK" if r["ok"] else "ERRO"
                w.writerow([
                    datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S"),
                    "TAG", i, 1, "CONCLUINTE",
                    r["status"], status, item["biz_id"],
                ])
                if r["ok"]:
                    ok_count += 1
                else:
                    body = r.get("body", "")
                    if "not found" in str(body).lower() or r.get("status") == 404:
                        log.warning("  TAG %d: negócio não encontrado — provável merge anterior", i)
                    else:
                        err_count += 1
                        log.warning("  ERRO tag %d: %s", i, str(body)[:200])
                if i % 50 == 0:
                    log.info("  TAG progresso: %d/%d", i, len(to_tag))

    elapsed = time.monotonic() - start
    log.info("Concluído em %.1f min. OK: %d | Erros: %d | API calls: %d",
             elapsed / 60, ok_count, err_count, api.total_calls)
    log.info("Log detalhado: %s", log_file)
    return ok_count, err_count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    mode = "--dry-run"
    limit = None

    for arg in sys.argv[1:]:
        if arg in ("--test", "--dry-run", "--execute"):
            mode = arg
    rate_limit = None
    for i, arg in enumerate(sys.argv):
        if arg == "--limit" and i + 1 < len(sys.argv):
            limit = int(sys.argv[i + 1])
        if arg == "--rate" and i + 1 < len(sys.argv):
            rate_limit = int(sys.argv[i + 1])

    log.info("=" * 50)
    log.info("Pipeline CRM — modo: %s", mode.upper())
    if rate_limit:
        log.info("Rate-limit: %d req/min", rate_limit)
    log.info("=" * 50)

    xl_by_rgm = load_excel()

    conn = get_conn()
    try:
        stages = load_pipeline_stages(conn)
        stage_ids = resolve_stage_ids(stages)
        if not stage_ids:
            log.error("Abortado: etapas não encontradas. Crie-as no CRM.")
            return

        crm_by_rgm = load_crm_businesses(conn)
        sem_adim, sem_inadim = load_sem_rematricula_snapshot(conn)
        inad_rgms = load_inadimplentes_snapshot(conn)
        conc_rgms = load_concluintes_snapshot(conn)
    finally:
        conn.close()

    log.info("Analisando ações necessárias...")
    to_restore, to_move, to_lose, to_tag, stats = analyze(
        xl_by_rgm, crm_by_rgm, stage_ids,
        sem_remat_adim=sem_adim, sem_remat_inadim=sem_inadim,
        inadimplentes_rgms=inad_rgms, concluintes_rgms=conc_rgms,
    )
    log.info("  Restore: %d | Move: %d | Lose: %d | Tag: %d",
             len(to_restore),
             sum(len(v) for v in to_move.values()),
             sum(len(v) for v in to_lose.values()),
             len(to_tag))

    if mode == "--dry-run":
        dry_run_summary(to_restore, to_move, to_lose, to_tag, stats, stage_ids, rate_limit=rate_limit)
        return

    api = ApiClient(rate_limit=rate_limit)

    reason_ids = ensure_loss_reasons(api)
    if not reason_ids:
        log.error("Abortado: falha ao configurar motivos de perda.")
        return

    concluinte_tag_id = None
    if to_tag:
        conn2 = get_conn()
        try:
            concluinte_tag_id = resolve_tag_id(conn2, api, "CONCLUINTE")
        finally:
            conn2.close()
        if not concluinte_tag_id:
            log.warning("Tag CONCLUINTE não resolvida — tags não serão aplicadas")

    if mode == "--test":
        test_one(api, to_restore, to_move, to_lose, to_tag, reason_ids, concluinte_tag_id)
        return

    if mode == "--execute":
        log.info("Iniciando execução em massa (batch de %d)...", BATCH_SIZE)
        ok, err = execute(api, to_restore, to_move, to_lose, to_tag, reason_ids, concluinte_tag_id)
        print(f"\nResultado: {ok} OK, {err} erros. API calls: {api.total_calls}")


if __name__ == "__main__":
    main()
