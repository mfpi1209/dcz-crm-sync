"""Match Inadimplentes (TI): cruza inadimplentes x matriculados x leads DataCrazy.

POST /api/match-inadimplentes
    multipart/form-data:
        inadimplentes: .xlsx (RGM_ALUN, NOM_FILI, NOME, DES_CURS, VAL_TITU, ATRASO, PORTADOR, ...)
        matriculados:  .xlsx (RGM, Fone celular, Fone Residencial, Fone Comercial, Email, Curso, Polo, Nome)
        leads:         .csv  (id, name, phone, rawPhone, email, ...)
        filtro:        'pos' | 'grad' | 'ambos'
        export:        'json' (default) | 'xlsx'
    Retorna JSON com os matches (RGM/nome/curso/polo/telefone/email/valor/atraso/portador/lead_id)
    ou stream XLSX quando export=xlsx.
"""
from __future__ import annotations

import io
import csv
import logging
import re
from datetime import date
from typing import Any

from flask import Blueprint, jsonify, request, session, send_file
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)
match_inadimplentes_bp = Blueprint("match_inadimplentes_bp", __name__)

POS_NOM_FILI = "CRUZEIRO DO SUL - PÓS-EAD"
POS_NORM = "cruzeirodosulposead"

# Colunas de saida (ordem estavel para UI e XLSX)
OUTPUT_COLS = [
    ("rgm",        "RGM"),
    ("nome",       "Nome"),
    ("curso",      "Curso"),
    ("polo",       "Polo"),
    ("telefone",   "Telefone matchado"),
    ("email",      "E-mail"),
    ("valor",      "Valor em atraso"),
    ("atraso",     "Dias em atraso"),
    ("portador",   "Portador"),
    ("lead_id",    "Lead ID (DataCrazy)"),
]


# ---------- Helpers ----------

def _s(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_rgm(v: Any) -> str:
    s = _s(v)
    if not s:
        return ""
    # se veio como float ("48690406.0"), corta o .0
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"\D", "", s)
    return s.lstrip("0") or s


def _norm_phone(v: Any) -> str:
    """Retorna o telefone com so digitos, colapsado nos ultimos 11 caracteres.
    (aceita numeros com/sem DDI 55). Retorna "" se < 10 digitos."""
    s = re.sub(r"\D", "", _s(v))
    if len(s) < 10:
        return ""
    # se tem DDI 55 (13 digitos), tira
    if len(s) > 11 and s.startswith("55"):
        s = s[2:]
    # normaliza pra 11 digitos (chave usada nos dois lados)
    return s[-11:]


def _norm_str(s: Any) -> str:
    """Normaliza para comparacao case/acento-insensitive."""
    import unicodedata
    t = unicodedata.normalize("NFKD", _s(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", t.lower())


def _to_float(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = _s(v).replace(".", "").replace(",", ".") if "," in _s(v) else _s(v)
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_int(v: Any) -> int:
    try:
        return int(float(_s(v)))
    except (TypeError, ValueError):
        return 0


def _header_index(header_row: tuple) -> dict:
    """{coluna_normalizada: indice}. Aceita variacoes de acentuacao/case."""
    idx = {}
    for i, h in enumerate(header_row or ()):
        idx[_norm_str(h)] = i
    return idx


def _get(row: tuple, idx: dict, *keys: str) -> Any:
    for k in keys:
        i = idx.get(_norm_str(k))
        if i is not None and i < len(row):
            v = row[i]
            if v not in (None, ""):
                return v
    return None


# ---------- Leitura dos arquivos ----------

def _read_inadimplentes(file_storage, filtro: str) -> list[dict]:
    """Le a planilha de inadimplentes ja aplicando o filtro por NOM_FILI."""
    wb = load_workbook(io.BytesIO(file_storage.read()), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    idx = _header_index(header)

    out: list[dict] = []
    for row in it:
        if row is None:
            continue
        rgm = _norm_rgm(_get(row, idx, "RGM_ALUN", "RGM"))
        if not rgm:
            continue
        nom_fili = _s(_get(row, idx, "NOM_FILI"))
        nf_norm = _norm_str(nom_fili)
        if filtro == "pos" and nf_norm != POS_NORM:
            continue
        if filtro == "grad" and nf_norm == POS_NORM:
            continue
        out.append({
            "rgm":       rgm,
            "nome":      _s(_get(row, idx, "NOME")),
            "curso":     _s(_get(row, idx, "DES_CURS")),
            "nom_fili":  nom_fili,
            "valor":     _to_float(_get(row, idx, "VAL_TITU")),
            "atraso":    _to_int(_get(row, idx, "ATRASO")),
            "portador":  _s(_get(row, idx, "PORTADOR")),
        })
    return out


def _read_matriculados(file_storage) -> dict[str, dict]:
    """Le matriculados por polo e retorna {rgm_normalizado: {telefones[], email, curso, polo, nome}}."""
    wb = load_workbook(io.BytesIO(file_storage.read()), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return {}
    idx = _header_index(header)

    result: dict[str, dict] = {}
    for row in it:
        if row is None:
            continue
        rgm = _norm_rgm(_get(row, idx, "RGM"))
        if not rgm:
            continue
        cel = _norm_phone(_get(row, idx, "Fone celular", "Fone Celular", "Telefone Celular"))
        res = _norm_phone(_get(row, idx, "Fone Residencial", "Fone residencial"))
        com = _norm_phone(_get(row, idx, "Fone Comercial", "Fone comercial"))
        telefones = [t for t in (cel, res, com) if t]
        if not telefones and rgm in result:
            continue
        entry = result.setdefault(rgm, {
            "telefones": [], "email": "", "curso": "", "polo": "", "nome": "",
        })
        for t in telefones:
            if t not in entry["telefones"]:
                entry["telefones"].append(t)
        if not entry["email"]:
            entry["email"] = _s(_get(row, idx, "Email", "E-mail"))
        if not entry["curso"]:
            entry["curso"] = _s(_get(row, idx, "Curso"))
        if not entry["polo"]:
            entry["polo"] = _s(_get(row, idx, "Polo"))
        if not entry["nome"]:
            entry["nome"] = _s(_get(row, idx, "Nome"))
    return result


def _read_leads(file_storage) -> dict[str, dict]:
    """Le o CSV do DataCrazy e retorna {telefone_normalizado: {id, name, email}}."""
    raw = file_storage.read()
    # tenta utf-8-sig; se falhar, cai pra latin-1
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")

    # detecta delimitador
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delim = dialect.delimiter
    except csv.Error:
        delim = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    result: dict[str, dict] = {}
    for r in reader:
        if not r:
            continue
        lead_id = _s(r.get("id"))
        name = _s(r.get("name"))
        email = _s(r.get("email"))
        for k in ("phone", "rawPhone", "rawphone"):
            phone = _norm_phone(r.get(k))
            if not phone:
                continue
            result.setdefault(phone, {"id": lead_id, "name": name, "email": email})
    return result


# ---------- Match ----------

def _match(inad: list[dict], matr: dict[str, dict], leads: dict[str, dict]) -> list[dict]:
    """Cruza os 3 datasets e retorna 1 linha por RGM que bateu com um lead."""
    agregado: dict[str, dict] = {}
    for r in inad:
        rgm = r["rgm"]
        m = matr.get(rgm)
        if not m or not m["telefones"]:
            continue
        # tenta cada telefone contra a base de leads
        tel_match = None
        lead = None
        for t in m["telefones"]:
            if t in leads:
                tel_match = t
                lead = leads[t]
                break
        if not lead:
            continue

        if rgm in agregado:
            g = agregado[rgm]
            g["valor"] += r["valor"]
            if r["atraso"] > g["atraso"]:
                g["atraso"] = r["atraso"]
        else:
            # formata telefone matchado (11 digitos -> (XX) XXXXX-XXXX ou 10 -> (XX) XXXX-XXXX)
            agregado[rgm] = {
                "rgm":      rgm,
                "nome":     r["nome"] or m["nome"],
                "curso":    r["curso"] or m["curso"],
                "polo":     m["polo"],
                "telefone": _fmt_phone(tel_match),
                "email":    lead.get("email") or m["email"],
                "valor":    r["valor"],
                "atraso":   r["atraso"],
                "portador": r["portador"],
                "lead_id":  lead.get("id", ""),
            }
    return list(agregado.values())


def _fmt_phone(digits: str) -> str:
    if not digits:
        return ""
    d = digits[-11:] if len(digits) >= 11 else digits
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    if len(d) == 10:
        return f"({d[:2]}) {d[2:6]}-{d[6:]}"
    return d


# ---------- Export ----------

def _to_xlsx(rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Match"
    ws.append([label for _key, label in OUTPUT_COLS])
    for r in rows:
        ws.append([r.get(key, "") for key, _label in OUTPUT_COLS])
    # Larguras
    widths = [12, 32, 42, 32, 20, 30, 14, 10, 16, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"match_inadimplentes_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ---------- Rota ----------

@match_inadimplentes_bp.route("/api/match-inadimplentes", methods=["POST"])
def api_match_inadimplentes():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "forbidden"}), 403

    filtro = (request.form.get("filtro") or "ambos").lower()
    if filtro not in ("pos", "grad", "ambos"):
        return jsonify({"ok": False, "error": "filtro invalido"}), 400
    export = (request.form.get("export") or "json").lower()

    f_inad = request.files.get("inadimplentes")
    f_matr = request.files.get("matriculados")
    f_leads = request.files.get("leads")
    if not (f_inad and f_matr and f_leads):
        return jsonify({"ok": False, "error": "envie inadimplentes, matriculados e leads"}), 400

    try:
        inad = _read_inadimplentes(f_inad, filtro)
    except Exception as e:
        logger.exception("erro lendo inadimplentes")
        return jsonify({"ok": False, "error": f"inadimplentes: {e}"}), 400
    try:
        matr = _read_matriculados(f_matr)
    except Exception as e:
        logger.exception("erro lendo matriculados")
        return jsonify({"ok": False, "error": f"matriculados: {e}"}), 400
    try:
        leads = _read_leads(f_leads)
    except Exception as e:
        logger.exception("erro lendo leads")
        return jsonify({"ok": False, "error": f"leads: {e}"}), 400

    rows = _match(inad, matr, leads)
    # ordena por atraso desc, depois valor desc
    rows.sort(key=lambda r: (r["atraso"], r["valor"]), reverse=True)

    stats = {
        "inadimplentes": len(inad),
        "matriculados":  len(matr),
        "leads":         len(leads),
        "matches":       len(rows),
    }
    logger.info("match-inadimplentes filtro=%s stats=%s", filtro, stats)

    if export == "xlsx":
        return _to_xlsx(rows)
    return jsonify({"ok": True, "stats": stats, "rows": rows})
