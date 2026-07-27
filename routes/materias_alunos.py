"""Materias Alunos (TI, admin-only): consulta em lote no SIAA + persistencia.

POST /api/materias-alunos/preview  (multipart: matriculados xlsx + data_matricula)
    Le a planilha, filtra por Negocio=GRADUACAO, Tipo Matricula=NOVA MATRICULA
    e Data Matricula==data escolhida. Retorna quais RGMs ja foram consultados
    (via materias_alunos_consultas) e quais estao pendentes.

POST /api/materias-alunos/start    (multipart: matriculados xlsx + data_matricula + intervalo)
    Dispara job em background. Um job global por vez.

GET  /api/materias-alunos/status
    Retorna snapshot do job atual.

POST /api/materias-alunos/cancel
    Sinaliza cancelamento (o loop verifica entre RGMs).
"""
from __future__ import annotations

import io
import logging
import os
import re
import threading
import time
from datetime import date, datetime, timedelta, timezone
from typing import Any
from urllib.parse import urlencode as _urlencode

import requests
from flask import Blueprint, jsonify, request, session
from openpyxl import load_workbook

from siaa.materias_alunos_client import (
    SessaoExpirada, SESSAO_EXPIRADA_MSG, _load_cookie_academico, buscar_materias,
)

logger = logging.getLogger(__name__)
materias_alunos_bp = Blueprint("materias_alunos_bp", __name__)

# ---------- Estado do job (global, 1 execucao por vez) ----------

_job_lock = threading.Lock()
_job: dict | None = None  # {id, status, total, processados, sucesso, sem_materias, erros, ...}


def _new_job(total: int, intervalo: int) -> dict:
    return {
        "id": int(time.time()),
        "status": "running",   # 'running' | 'done' | 'cancelled' | 'session_expired' | 'error'
        "total": total,
        "processados": 0,
        "sucesso": 0,
        "sem_materias": 0,
        "erros": 0,
        "ultimo_rgm": None,
        "ultimo_aluno": None,
        "ultima_msg": None,
        "iniciado_em": datetime.now(timezone.utc).isoformat(),
        "terminado_em": None,
        "intervalo": intervalo,
        "cancel": False,
    }


def _job_snapshot() -> dict | None:
    with _job_lock:
        if _job is None:
            return None
        # copia sem o flag interno de cancel
        snap = {k: v for k, v in _job.items() if k != "cancel"}
    # eta
    if snap["status"] == "running":
        restantes = max(0, snap["total"] - snap["processados"])
        snap["eta_segundos"] = restantes * snap["intervalo"]
    else:
        snap["eta_segundos"] = 0
    return snap


# ---------- Helpers ----------

def _s(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _norm_rgm(v: Any) -> str:
    s = _s(v)
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"\D", "", s)
    return s.lstrip("0") or s


def _norm_str(s: Any) -> str:
    import unicodedata
    t = unicodedata.normalize("NFKD", _s(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", t.lower())


def _parse_data(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _s(v)
    # 2026-02-20 ou 2026-02-20T00:00:00 ou 20/02/2026
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.split("T")[0] if "T" in s else s, fmt).date()
        except ValueError:
            continue
    return None


def _header_index(header_row: tuple) -> dict:
    idx = {}
    for i, h in enumerate(header_row or ()):
        idx[_norm_str(h)] = i
    return idx


def _get_col(row: tuple, idx: dict, *keys: str) -> Any:
    for k in keys:
        i = idx.get(_norm_str(k))
        if i is not None and i < len(row):
            v = row[i]
            if v not in (None, ""):
                return v
    return None


# ---------- Leitura da planilha + filtro ----------

def _ler_rgms(file_bytes: bytes, data_matricula_iso: str) -> tuple[list[dict], str | None]:
    """Retorna (rgms_com_nome, err). Filtra Negocio=GRADUACAO, Tipo=NOVA MATRICULA, Data==data escolhida."""
    alvo = _parse_data(data_matricula_iso)
    if not alvo:
        return [], "data_matricula invalida"

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return [], f"planilha invalida: {e}"

    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return [], "planilha vazia"
    idx = _header_index(header)

    out: list[dict] = []
    seen: set[str] = set()
    for row in it:
        if row is None:
            continue
        negocio = _norm_str(_get_col(row, idx, "Negócio", "Negocio"))
        if "graduacao" not in negocio or "pos" in negocio:
            continue
        tipo = _norm_str(_get_col(row, idx, "Tipo Matrícula", "Tipo Matricula"))
        if tipo != "novamatricula":
            continue
        dt = _parse_data(_get_col(row, idx, "Data Matrícula", "Data Matricula"))
        if dt != alvo:
            continue
        rgm = _norm_rgm(_get_col(row, idx, "RGM", "RGM_ALUN"))
        if not rgm or rgm in seen:
            continue
        seen.add(rgm)
        out.append({"rgm": rgm, "nome": _s(_get_col(row, idx, "Nome"))})
    return out, None


# ---------- Supabase helpers ----------

def _sb_headers(prefer: str = "return=minimal") -> dict[str, str]:
    url = (os.environ.get("SUPABASE_URL", "") or "").rstrip("/")
    key = os.environ.get("SUPABASE_KEY", "") or ""
    if not url or not key:
        raise RuntimeError("SUPABASE_URL e SUPABASE_KEY são obrigatórios")
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Prefer": prefer,
    }


def _sb_base() -> str:
    return (os.environ.get("SUPABASE_URL", "") or "").rstrip("/") + "/rest/v1"


def _sb_upsert(table: str, rows: list[dict], on_conflict: str) -> None:
    if not rows:
        return
    url = f"{_sb_base()}/{table}?{_urlencode({'on_conflict': on_conflict})}"
    r = requests.post(
        url,
        headers=_sb_headers("resolution=merge-duplicates,return=minimal"),
        json=rows,
        timeout=30,
    )
    if r.status_code >= 300:
        raise RuntimeError(f"Supabase upsert {table} {r.status_code}: {(r.text or '')[:500]}")


def _sb_get(table: str, params: dict[str, str]) -> list[dict]:
    r = requests.get(
        f"{_sb_base()}/{table}?{_urlencode(params)}",
        headers=_sb_headers("return=representation"),
        timeout=30,
    )
    if r.status_code >= 300:
        raise RuntimeError(f"Supabase GET {table} {r.status_code}: {(r.text or '')[:500]}")
    data = r.json()
    return data if isinstance(data, list) else [data]


def _ja_consultados(rgms: list[str]) -> set[str]:
    """RGMs que ja aparecem em materias_alunos (Supabase), consultados em lotes."""
    if not rgms:
        return set()
    result: set[str] = set()
    CHUNK = 200
    try:
        for i in range(0, len(rgms), CHUNK):
            batch = rgms[i:i + CHUNK]
            in_list = "(" + ",".join(batch) + ")"
            rows = _sb_get(
                "materias_alunos",
                {"select": "rgm", "rgm": f"in.{in_list}"},
            )
            for row in rows:
                v = row.get("rgm")
                if v:
                    result.add(str(v))
        return result
    except Exception as e:
        logger.warning("ja_consultados: %s", e)
        return set()


# ---------- Persistencia por RGM ----------

def _upsert_materias(rgm: str, aluno: str, materias: list[dict]):
    """Upsert em materias_alunos: 1 linha por aluno com jsonb [{disciplina, data}]."""
    itens: list[dict] = []
    vistos: set[tuple] = set()
    for m in materias or ():
        nome = _s(m.get("disciplina")) or _s(m.get("sigla"))
        data = _s(m.get("data"))
        if not nome:
            continue
        chave = (nome, data)
        if chave in vistos:
            continue
        vistos.add(chave)
        itens.append({"disciplina": nome, "data": data})
    _sb_upsert("materias_alunos", [{
        "rgm": rgm,
        "aluno": aluno,
        "materias": itens,
        "qtd_materias": len(itens),
        "consultado_em": datetime.now(timezone.utc).isoformat(),
    }], on_conflict="rgm")


def _registrar_erro(rgm: str, aluno: str | None, msg: str):
    logger.warning("materias erro rgm=%s aluno=%s: %s", rgm, aluno, (msg or "")[:200])


# ---------- Worker (thread) ----------

def _worker(rgms: list[str], intervalo: int):
    global _job
    try:
        cookie = _load_cookie_academico()
    except SessaoExpirada as e:
        with _job_lock:
            if _job is not None:
                _job["status"] = "session_expired"
                _job["ultima_msg"] = str(e)
                _job["terminado_em"] = datetime.now(timezone.utc).isoformat()
        return

    for i, rgm in enumerate(rgms):
        with _job_lock:
            if _job is None or _job.get("cancel"):
                if _job is not None:
                    _job["status"] = "cancelled"
                    _job["terminado_em"] = datetime.now(timezone.utc).isoformat()
                return

        aluno_nome = None
        try:
            res = buscar_materias(cookie, rgm)
            aluno_nome = res.get("aluno")
            materias = res.get("materias") or []
            _upsert_materias(rgm, aluno_nome, materias)
            with _job_lock:
                if materias:
                    _job["sucesso"] += 1
                else:
                    _job["sem_materias"] += 1
        except SessaoExpirada as e:
            with _job_lock:
                _job["status"] = "session_expired"
                _job["ultima_msg"] = str(e)
                _job["terminado_em"] = datetime.now(timezone.utc).isoformat()
            return
        except Exception as e:
            logger.exception("materias worker rgm=%s falhou", rgm)
            _registrar_erro(rgm, aluno_nome, f"{type(e).__name__}: {e}")
            with _job_lock:
                _job["erros"] += 1
                _job["ultima_msg"] = f"{rgm}: {e}"
        finally:
            with _job_lock:
                _job["processados"] += 1
                _job["ultimo_rgm"] = rgm
                _job["ultimo_aluno"] = aluno_nome

        # sleep intervalo (mas verifica cancel)
        if i < len(rgms) - 1:
            slept = 0.0
            while slept < intervalo:
                with _job_lock:
                    if _job is None or _job.get("cancel"):
                        break
                time.sleep(min(1.0, intervalo - slept))
                slept += 1.0

    with _job_lock:
        if _job is not None and _job["status"] == "running":
            _job["status"] = "done"
            _job["terminado_em"] = datetime.now(timezone.utc).isoformat()


# ---------- Rotas ----------

def _forbidden():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "forbidden"}), 403
    return None


def _read_form_file() -> tuple[bytes | None, str | None, Any]:
    f = request.files.get("matriculados")
    if not f:
        return None, None, (jsonify({"ok": False, "error": "envie o arquivo 'matriculados'"}), 400)
    data_iso = _s(request.form.get("data_matricula"))
    if not data_iso:
        return None, None, (jsonify({"ok": False, "error": "informe data_matricula (YYYY-MM-DD)"}), 400)
    return f.read(), data_iso, None


@materias_alunos_bp.route("/api/materias-alunos/preview", methods=["POST"])
def api_preview():
    fb = _forbidden()
    if fb: return fb
    file_bytes, data_iso, err = _read_form_file()
    if err: return err
    rgms_info, err_msg = _ler_rgms(file_bytes, data_iso)
    if err_msg:
        return jsonify({"ok": False, "error": err_msg}), 400
    all_rgms = [r["rgm"] for r in rgms_info]
    ja = _ja_consultados(all_rgms)
    pendentes = [r for r in all_rgms if r not in ja]
    return jsonify({
        "ok": True,
        "total": len(all_rgms),
        "ja_consultados": len(ja),
        "pendentes": len(pendentes),
        "eta_segundos": len(pendentes) * 15,
        "amostra_pendentes": pendentes[:5],
    })


@materias_alunos_bp.route("/api/materias-alunos/start", methods=["POST"])
def api_start():
    fb = _forbidden()
    if fb: return fb

    global _job
    with _job_lock:
        if _job is not None and _job.get("status") == "running":
            return jsonify({"ok": False, "error": "ja existe um job rodando"}), 409

    file_bytes, data_iso, err = _read_form_file()
    if err: return err

    try:
        intervalo = int(request.form.get("intervalo") or 15)
    except (TypeError, ValueError):
        intervalo = 15
    intervalo = max(5, min(300, intervalo))

    pular = (request.form.get("pular_ja_consultados") or "true").lower() != "false"

    rgms_info, err_msg = _ler_rgms(file_bytes, data_iso)
    if err_msg:
        return jsonify({"ok": False, "error": err_msg}), 400

    todos = [r["rgm"] for r in rgms_info]
    ja = _ja_consultados(todos) if pular else set()
    rgms = [r for r in todos if r not in ja]

    if not rgms:
        return jsonify({"ok": False, "error": "nenhum RGM pendente para processar"}), 400

    # valida cookie antes de subir a thread (falha rapido)
    try:
        _load_cookie_academico()
    except SessaoExpirada as e:
        return jsonify({"ok": False, "error": str(e), "session_expired": True}), 409

    with _job_lock:
        _job = _new_job(len(rgms), intervalo)
        _job["filtros"] = {"data_matricula": data_iso, "pular_ja_consultados": pular}

    t = threading.Thread(target=_worker, args=(rgms, intervalo), daemon=True)
    t.start()

    return jsonify({"ok": True, "job": _job_snapshot()})


@materias_alunos_bp.route("/api/materias-alunos/status", methods=["GET"])
def api_status():
    fb = _forbidden()
    if fb: return fb
    snap = _job_snapshot()
    return jsonify({"ok": True, "job": snap})


@materias_alunos_bp.route("/api/materias-alunos/diag", methods=["GET"])
def api_diag():
    """Diagnostica a leitura do cookie SIAA Academico (admin-only).
    Passar ?rgm=XXXX faz uma chamada real ao SIAA para ver onde falha."""
    fb = _forbidden()
    if fb: return fb
    info = {
        "SIAA_SESSION_URL": (os.environ.get("SIAA_SESSION_URL", "") or "").strip() or None,
        "SIAA_SESSION_TOKEN_len": len((os.environ.get("SIAA_SESSION_TOKEN", "") or "").strip()),
    }
    try:
        cookie = _load_cookie_academico()
        info["cookie_len"] = len(cookie)
        info["cookie_preview"] = cookie[:120]
    except SessaoExpirada as e:
        info["error"] = str(e)
        return jsonify({"ok": False, "info": info}), 502

    rgm = _s(request.args.get("rgm"))
    if not rgm:
        return jsonify({"ok": True, "info": info, "hint": "passe ?rgm=NNNN para testar chamada real ao SIAA"})

    # Chamada instrumentada ao SIAA
    from urllib.parse import quote
    from siaa.materias_alunos_client import (
        _sessao_cookie, _viewstate, _turmas, _parse_grid, _redirecionou,
        CONS_JSF, CONS_HIST, HDR_NAV, HDR_AJAX,
    )
    steps: list[dict] = []
    rgm_num = re.sub(r"\D", "", rgm)
    T = 12  # timeout curto pra caber no timeout do proxy (Traefik/EasyPanel)
    try:
        s = _sessao_cookie(cookie)
        r0 = s.get(CONS_JSF, params={"init": "true"}, headers=HDR_NAV, timeout=T, allow_redirects=False)
        steps.append({
            "step": "GET init",
            "status": r0.status_code,
            "location": r0.headers.get("Location"),
            "content_type": r0.headers.get("Content-Type"),
            "len": len(r0.text or ""),
            "redirected": _redirecionou(r0),
            "body_preview": (r0.text or "")[:400],
        })
        if _redirecionou(r0):
            return jsonify({"ok": False, "info": info, "steps": steps, "diag": "redirect no GET init"}), 502
        vs = _viewstate(r0.text)
        steps[-1]["has_viewstate"] = bool(vs)
        if not vs:
            return jsonify({"ok": False, "info": info, "steps": steps, "diag": "sem ViewState no GET init"}), 502

        body = ("javax.faces.partial.ajax=true&javax.faces.source=formPrincipal%3AbtnBuscar"
                "&javax.faces.partial.execute=%40all&javax.faces.partial.render=formPrincipal"
                "&formPrincipal%3AbtnBuscar=formPrincipal%3AbtnBuscar&formPrincipal=formPrincipal"
                f"&formPrincipal%3Aempresas_focus=&formPrincipal%3Aempresas_input=12"
                f"&formPrincipal%3AfilterAluno={rgm_num}&formPrincipal%3AtabelaListaAlunos_rppDD=10"
                "&formPrincipal%3AtipoEnade_focus=&formPrincipal%3AtipoEnade_input=1"
                f"&javax.faces.ViewState={quote(vs, safe='')}")
        rb = s.post(CONS_JSF, data=body, headers={**HDR_AJAX, "Referer": CONS_JSF + "?init=true"}, timeout=T, allow_redirects=False)
        steps.append({
            "step": "POST buscar",
            "status": rb.status_code,
            "location": rb.headers.get("Location"),
            "content_type": rb.headers.get("Content-Type"),
            "len": len(rb.text or ""),
            "redirected": _redirecionou(rb),
            "body_preview": (rb.text or "")[:400],
        })
        if _redirecionou(rb):
            return jsonify({"ok": False, "info": info, "steps": steps, "diag": "redirect no POST buscar"}), 502

        rh = s.get(CONS_HIST, headers={**HDR_NAV, "Referer": CONS_JSF, "Sec-Fetch-Dest": "iframe"}, timeout=60, allow_redirects=False)
        steps.append({
            "step": "GET historico",
            "status": rh.status_code,
            "location": rh.headers.get("Location"),
            "content_type": rh.headers.get("Content-Type"),
            "len": len(rh.text or ""),
            "redirected": _redirecionou(rh),
            "has_viewstate": bool(_viewstate(rh.text)),
            "body_preview": (rh.text or "")[:400],
        })
        if _redirecionou(rh):
            return jsonify({"ok": False, "info": info, "steps": steps, "diag": "redirect no GET historico"}), 502
        sel, turmas = _turmas(rh.text)
        materias = _parse_grid(rh.text)
        return jsonify({
            "ok": True,
            "info": info,
            "steps": steps,
            "turma_selected": sel,
            "turmas_total": len(turmas),
            "materias_encontradas": len(materias),
            "materias_amostra": [m.get("disciplina") for m in materias[:5]],
        })
    except Exception as e:
        return jsonify({"ok": False, "info": info, "steps": steps, "exception": f"{type(e).__name__}: {e}"}), 502


@materias_alunos_bp.route("/api/materias-alunos/cancel", methods=["POST"])
def api_cancel():
    fb = _forbidden()
    if fb: return fb
    with _job_lock:
        if _job is None or _job.get("status") != "running":
            return jsonify({"ok": False, "error": "nenhum job em execucao"}), 400
        _job["cancel"] = True
    return jsonify({"ok": True})
