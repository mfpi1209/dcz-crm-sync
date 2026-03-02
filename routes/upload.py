"""
eduit. — Blueprint de upload e snapshots XL.
"""

import os
import re
import json
import glob
import shutil
import tempfile
import unicodedata
import zipfile
import traceback
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras
import csv as _csv
import io as _io
from flask import Blueprint, request, jsonify, current_app, Response

from db import get_conn
from helpers import BRT, to_brt, XL_TIPOS, BASE_DIR, LOG_DIR, REPORTS_DIR, _normalize_digits

upload_bp = Blueprint("upload_bp", __name__)

UPLOAD_DIR = BASE_DIR


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_xlsx():
    for f in UPLOAD_DIR.iterdir():
        if f.suffix.lower() == ".xlsx" and "matriculados" in f.name.lower():
            stat = f.stat()
            return {
                "name": f.name,
                "size": stat.st_size,
                "modified": to_brt(datetime.fromtimestamp(stat.st_mtime, tz=BRT)),
            }
    return None


def _compute_snapshot_stats(snap_id, tipo):
    """Calcula e grava métricas agregadas de um snapshot."""
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT data FROM xl_rows WHERE snapshot_id = %s", (snap_id,))
            rows = [r["data"] for r in cur.fetchall()]

        if not rows:
            return

        metrics = {"total": len(rows)}
        by_polo, by_curso = {}, {}
        for r in rows:
            polo = r.get("polo", "N/I") or "N/I"
            by_polo[polo] = by_polo.get(polo, 0) + 1
            curso = r.get("curso", "N/I") or "N/I"
            by_curso[curso] = by_curso.get(curso, 0) + 1

        if tipo == "matriculados":
            by_tipo_aluno = {}
            for r in rows:
                ta = r.get("tipo_matricula", "N/I") or "N/I"
                by_tipo_aluno[ta] = by_tipo_aluno.get(ta, 0) + 1
            metrics["by_tipo_aluno"] = by_tipo_aluno

        elif tipo == "inadimplentes":
            total_valor = 0.0
            max_atraso = 0
            for r in rows:
                try:
                    total_valor += float(r.get("valor_total", "0") or "0")
                except (ValueError, TypeError):
                    pass
                try:
                    a = int(float(r.get("max_atraso", "0") or "0"))
                    if a > max_atraso:
                        max_atraso = a
                except (ValueError, TypeError):
                    pass
            metrics["valor_total"] = round(total_valor, 2)
            metrics["max_atraso"] = max_atraso
            metrics["total_titulos"] = sum(int(r.get("total_titulos", "0") or "0") for r in rows)

        elif tipo == "acesso_ava":
            now = datetime.now()
            ativos_7d, ativos_30d, inativos = 0, 0, 0
            total_interacoes, total_minutos = 0, 0
            for r in rows:
                try:
                    total_interacoes += int(float(r.get("interacoes", "0") or "0"))
                except (ValueError, TypeError):
                    pass
                try:
                    total_minutos += int(float(r.get("minutos", "0") or "0"))
                except (ValueError, TypeError):
                    pass
                ua = r.get("ultimo_acesso", "")
                if ua:
                    try:
                        dt = datetime.strptime(ua[:10], "%Y-%m-%d") if "-" in ua else datetime.strptime(ua[:10], "%d/%m/%Y")
                        delta = (now - dt).days
                        if delta <= 7:
                            ativos_7d += 1
                        if delta <= 30:
                            ativos_30d += 1
                        else:
                            inativos += 1
                    except (ValueError, TypeError):
                        inativos += 1
                else:
                    inativos += 1
            metrics["ativos_7d"] = ativos_7d
            metrics["ativos_30d"] = ativos_30d
            metrics["inativos"] = inativos
            metrics["media_interacoes"] = round(total_interacoes / max(len(rows), 1), 1)
            metrics["media_minutos"] = round(total_minutos / max(len(rows), 1), 1)

        elif tipo == "sem_rematricula":
            adim, inadim = 0, 0
            for r in rows:
                if r.get("status_financeiro") == "adimplente":
                    adim += 1
                else:
                    inadim += 1
            metrics["adimplentes"] = adim
            metrics["inadimplentes"] = inadim

        elif tipo == "concluintes":
            by_inst = {}
            for r in rows:
                inst = r.get("instituicao", "N/I") or "N/I"
                by_inst[inst] = by_inst.get(inst, 0) + 1
            metrics["by_instituicao"] = by_inst

        metrics["by_polo"] = dict(sorted(by_polo.items(), key=lambda x: -x[1])[:20])
        metrics["by_curso"] = dict(sorted(by_curso.items(), key=lambda x: -x[1])[:20])

        with conn.cursor() as cur:
            for metric_name, value in metrics.items():
                if isinstance(value, dict):
                    cur.execute(
                        "INSERT INTO xl_snapshot_stats (snapshot_id, metric, value, detail) "
                        "VALUES (%s, %s, %s, %s::jsonb) "
                        "ON CONFLICT (snapshot_id, metric) DO UPDATE SET value=EXCLUDED.value, detail=EXCLUDED.detail",
                        (snap_id, metric_name, None, json.dumps(value, ensure_ascii=False)),
                    )
                else:
                    cur.execute(
                        "INSERT INTO xl_snapshot_stats (snapshot_id, metric, value, detail) "
                        "VALUES (%s, %s, %s, NULL) "
                        "ON CONFLICT (snapshot_id, metric) DO UPDATE SET value=EXCLUDED.value",
                        (snap_id, metric_name, value),
                    )
        conn.commit()
    except Exception as e:
        current_app.logger.warning("Erro ao computar stats snapshot %s: %s", snap_id, e)
    finally:
        conn.close()


_XL_COLUMN_MAP = {
    "nome": ["Nome", "NOME", "Aluno", "Nome Aluno"],
    "cpf": ["CPF"],
    "rgm": ["RGM", "RGM_ALUN"],
    "curso": ["Curso", "DES_CURS"],
    "polo": ["Polo", "NOME_POL"],
    "serie": ["Série", "Serie"],
    "situacao": ["Situação Matrícula", "Situa", "Situação"],
    "tipo_matricula": ["Tipo Matrícula", "Tipo Matr", "Tipo matricula"],
    "data_mat": ["Data Matrícula", "Data Matr"],
    "email": ["Email", "E-mail"],
    "email_acad": ["Email acadêmico", "Email acad", "Email Acadêmico"],
    "fone_cel": ["Fone celular", "Celular", "Telefone"],
    "fone_res": ["Fone Residencial"],
    "fone_com": ["Fone Comercial"],
    "negocio": ["Negócio", "Neg"],
    "nivel": ["Nível", "Nivel", "Nível de Ensino", "Nível Ensino", "Tipo Curso"],
    "empresa": ["Empresa", "NOM_FILI"],
    "bairro": ["Bairro"],
    "cidade": ["Cidade"],
    "sexo": ["Sexo"],
    "data_nasc": ["Data Nascimento"],
    "ciclo": ["Ciclo"],
    "valor": ["Valor", "Valor Devido", "Saldo"],
    "parcela": ["Parcela", "Parcelas"],
    "vencimento": ["Vencimento", "Data Vencimento"],
    "status_financeiro": ["Status", "Status Financeiro", "Situação Financeira"],
    "data_conclusao": ["Data Conclusão", "Data Formatura", "Conclusão"],
    "periodo": ["Período", "Periodo"],
    "modalidade": ["Modalidade"],
    "instituicao": ["Instituição", "Institui"],
    "ultimo_acesso": ["Último Acesso", "Ultimo Acesso", "Ult Acesso", "Last Access"],
    "interacoes": ["Interações", "Interacoes", "Interações"],
    "minutos": ["Minutos", "Minutes"],
    "total_registros": ["Total Registros", "Total Regis"],
    "id_polo": ["ID_POLO"],
    "cod_inst": ["COD_INST"],
    "tipo_titulo": ["TIPO_TIT"],
    "descricao_titulo": ["DESCRICA"],
    "nr_titulo": ["NR_TITUL"],
    "dt_emissao": ["DTA_EMIS"],
    "dt_vencimento": ["DTA_VCTO"],
    "desconto": ["DESCONTO"],
    "juros": ["JUROS"],
    "valor_titulo": ["VAL_TITU"],
    "dias_atraso": ["ATRASO"],
    "portador": ["PORTADOR"],
    "apto_rematricula": ["Apto Rematricula", "Apta-Rematricula", "Apto-Rematricula"],
}


def _save_xl_snapshot(filepath, filename, tipo="matriculados"):
    """Lê o xlsx e grava um snapshot no banco de dados."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    if not sheet_names:
        wb.close()
        return 0
    ws = wb[sheet_names[0]]
    first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if not first_row:
        wb.close()
        return 0
    header = [cell.value for cell in first_row]
    col_map = {h: i for i, h in enumerate(header) if h}

    def _strip_accents(s):
        return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")

    def _find(names):
        for n in names:
            if n in col_map:
                return col_map[n]
            n_norm = _strip_accents(n.lower())
            for k in col_map:
                if k and (n.lower() in k.lower() or n_norm in _strip_accents(k.lower())):
                    return col_map[k]
        return None

    idx = {}
    for field, aliases in _XL_COLUMN_MAP.items():
        pos = _find(aliases)
        if pos is not None:
            idx[field] = pos

    unmapped = {}
    for h, i in col_map.items():
        if i not in idx.values() and h:
            safe_key = re.sub(r"\W+", "_", h.strip().lower())[:40]
            if safe_key:
                unmapped[safe_key] = i

    def _get(row, col_idx):
        if col_idx is None or col_idx >= len(row):
            return ""
        v = row[col_idx]
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y")
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v).strip()

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
            continue
        entry = {k: _get(row, v) for k, v in idx.items()}
        for k, v in unmapped.items():
            val = _get(row, v)
            if val:
                entry[k] = val
        entry["cpf_digits"] = _normalize_digits(entry.get("cpf", ""))
        entry["rgm_digits"] = _normalize_digits(entry.get("rgm", ""))
        phones = []
        for pk in ("fone_cel", "fone_res", "fone_com"):
            d = _normalize_digits(entry.get(pk, ""))
            if d:
                phones.append(d)
        entry["phones_digits"] = phones
        entries.append(entry)
    wb.close()

    return _persist_snapshot_entries(entries, tipo, filename)


def _persist_snapshot_entries(entries, tipo, filename, nivel=None):
    """Grava uma lista de dicts no banco como snapshot e retorna row count."""
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO xl_snapshots (tipo, filename, row_count, nivel) VALUES (%s, %s, %s, %s) RETURNING id",
                (tipo, filename, len(entries), nivel),
            )
            snap_id = cur.fetchone()[0]

            batch = [(snap_id, json.dumps(e, ensure_ascii=False)) for e in entries]
            psycopg2.extras.execute_batch(
                cur,
                "INSERT INTO xl_rows (snapshot_id, data) VALUES (%s, %s::jsonb)",
                batch,
                page_size=500,
            )
        conn.commit()
    finally:
        conn.close()

    try:
        _compute_snapshot_stats(snap_id, tipo)
    except Exception as e:
        current_app.logger.warning("Erro ao computar stats para snapshot %s: %s", snap_id, e)

    return len(entries)


def _handle_zip_upload(zip_path, tipo):
    """Extrai um .zip e processa conforme o tipo."""
    tmp_dir = Path(tempfile.mkdtemp(prefix=f"eduit_{tipo}_"))
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(str(tmp_dir))

        zip_name = Path(zip_path).name

        if tipo == "inadimplentes":
            entries = _parse_inadimplentes_batch(str(tmp_dir))
            return _persist_snapshot_entries(entries, tipo, zip_name) if entries else 0
        elif tipo == "sem_rematricula":
            entries = _parse_sem_rematricula(str(tmp_dir))
            return _persist_snapshot_entries(entries, tipo, zip_name) if entries else 0
        else:
            xlsx_files = list(tmp_dir.glob("*.xlsx"))
            if xlsx_files:
                return _save_xl_snapshot(str(xlsx_files[0]), zip_name, tipo)
            return 0
    finally:
        shutil.rmtree(str(tmp_dir), ignore_errors=True)


def _parse_inadimplentes_batch(folder_path):
    """Consolida todos os .xlsm de inadimplentes de um diretório em uma lista de entries."""
    import glob as _glob

    HEADER_MAP = {
        "ID_POLO": "id_polo", "NOME_POL": "polo", "RGM_ALUN": "rgm",
        "NOME": "nome", "COD_INST": "cod_inst", "NOM_FILI": "empresa",
        "DES_CURS": "curso", "TIPO_TIT": "tipo_titulo", "NR_TITUL": "nr_titulo",
        "DTA_EMIS": "dt_emissao", "DTA_VCTO": "dt_vencimento",
        "DESCONTO": "desconto", "JUROS": "juros", "VAL_TITU": "valor_titulo",
        "ATRASO": "dias_atraso", "PORTADOR": "portador",
    }
    DESCRICA_COLS = ["descricao_titulo", "portador_nome"]

    files = sorted(set(
        _glob.glob(os.path.join(folder_path, "*.xlsm"))
        + _glob.glob(os.path.join(folder_path, "*.xlsx"))
        + _glob.glob(os.path.join(folder_path, "**", "*.xlsm"), recursive=True)
        + _glob.glob(os.path.join(folder_path, "**", "*.xlsx"), recursive=True)
    ))
    current_app.logger.info("[INADIMPLENTES] Pasta: %s | Arquivos encontrados: %d", folder_path, len(files))
    if not files:
        current_app.logger.warning("[INADIMPLENTES] Nenhum .xlsm/.xlsx encontrado em %s", folder_path)
        return []

    raw_rows = []
    files_ok = 0
    files_no_header = 0
    files_error = 0
    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=False)
            ws = wb[wb.sheetnames[0]]
            header_row = None
            first_cells = []
            for i, row in enumerate(ws.iter_rows(max_col=18, values_only=True), 1):
                if i <= 5 and row and row[0]:
                    first_cells.append(f"L{i}:{str(row[0]).strip()[:30]}")
                if row and row[0] and str(row[0]).strip().upper() == "ID_POLO":
                    header_row = i
                    headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(row)]
                    break
            if not header_row:
                files_no_header += 1
                current_app.logger.warning("[INADIMPLENTES] %s: cabeçalho ID_POLO não encontrado. Primeiras células: %s", fname, first_cells)
                wb.close()
                continue
            for row in ws.iter_rows(min_row=header_row + 1, max_col=18, values_only=True):
                if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                    continue
                entry = {}
                descrica_idx = 0
                for j, val in enumerate(row):
                    if j >= len(headers):
                        break
                    h = headers[j].upper().strip()
                    if h == "DESCRICA":
                        key = DESCRICA_COLS[descrica_idx] if descrica_idx < len(DESCRICA_COLS) else f"descrica_{descrica_idx}"
                        descrica_idx += 1
                    else:
                        key = HEADER_MAP.get(h, h.lower()[:40])
                    v = val
                    if v is None:
                        v = ""
                    elif isinstance(v, float) and key in ("rgm", "id_polo", "cod_inst", "tipo_titulo", "portador", "dias_atraso"):
                        v = str(int(v))
                    elif isinstance(v, datetime):
                        v = v.strftime("%d/%m/%Y")
                    else:
                        v = str(v).strip()
                        if key in ("rgm",) and v.endswith(".0"):
                            v = v[:-2]
                    entry[key] = v
                entry["rgm_digits"] = _normalize_digits(entry.get("rgm", ""))
                for fk in ("valor_titulo", "desconto", "juros"):
                    try:
                        entry[fk] = str(round(float(entry.get(fk, "0") or "0"), 2))
                    except (ValueError, TypeError):
                        pass
                raw_rows.append(entry)
            files_ok += 1
            wb.close()
        except Exception as exc:
            files_error += 1
            current_app.logger.warning("[INADIMPLENTES] Erro ao processar %s: %s", fname, exc)

    current_app.logger.info(
        "[INADIMPLENTES] Resultado: %d arquivos OK, %d sem cabeçalho, %d com erro. Total linhas brutas: %d",
        files_ok, files_no_header, files_error, len(raw_rows),
    )

    aggregated = {}
    for row in raw_rows:
        rgm = row.get("rgm_digits", "")
        if not rgm:
            rgm = row.get("rgm", "unknown")
        if rgm not in aggregated:
            aggregated[rgm] = {
                "nome": row.get("nome", ""),
                "rgm": row.get("rgm", ""),
                "rgm_digits": rgm,
                "polo": row.get("polo", ""),
                "empresa": row.get("empresa", ""),
                "curso": row.get("curso", ""),
                "cpf_digits": "",
                "phones_digits": [],
                "total_titulos": 0,
                "valor_total": 0.0,
                "max_atraso": 0,
                "titulos": [],
            }
        agg = aggregated[rgm]
        agg["total_titulos"] += 1
        try:
            agg["valor_total"] += float(row.get("valor_titulo", "0") or "0")
        except (ValueError, TypeError):
            pass
        try:
            atraso = int(float(row.get("dias_atraso", "0") or "0"))
            if atraso > agg["max_atraso"]:
                agg["max_atraso"] = atraso
        except (ValueError, TypeError):
            pass
        agg["titulos"].append({
            "tipo": row.get("tipo_titulo", ""),
            "descricao": row.get("descricao_titulo", ""),
            "nr": row.get("nr_titulo", ""),
            "dt_emissao": row.get("dt_emissao", ""),
            "dt_vencimento": row.get("dt_vencimento", ""),
            "valor": row.get("valor_titulo", ""),
            "desconto": row.get("desconto", ""),
            "juros": row.get("juros", ""),
            "atraso": row.get("dias_atraso", ""),
            "portador": row.get("portador", ""),
            "portador_nome": row.get("portador_nome", ""),
        })

    entries = []
    for agg in aggregated.values():
        agg["valor_total"] = str(round(agg["valor_total"], 2))
        agg["max_atraso"] = str(agg["max_atraso"])
        agg["total_titulos"] = str(agg["total_titulos"])
        entries.append(agg)
    return entries


def _parse_sem_rematricula(folder_path):
    """Lê adimplentes.xlsx e inadimplentes.xlsx, unifica com flag financeiro."""
    HEADER_NORM = {
        "polo": "polo", "aluno": "nome", "telefone": "fone_cel",
        "e-mail": "email", "e_mail": "email", "rgm": "rgm",
        "serie": "serie", "série": "serie",
        "curso": "curso",
        "apto rematricula": "apto_rematricula",
        "apta-rematricula": "apto_rematricula",
        "apto-rematricula": "apto_rematricula",
    }

    def _find_file(folder, keyword, exclude_keyword=None):
        exact = os.path.join(folder, f"{keyword}s.xlsx")
        if os.path.isfile(exact):
            return exact
        exact2 = os.path.join(folder, f"{keyword}.xlsx")
        if os.path.isfile(exact2):
            return exact2
        for fn in os.listdir(folder):
            fl = fn.lower()
            if not fl.endswith((".xlsx", ".xlsm")):
                continue
            if keyword in fl and (exclude_keyword is None or exclude_keyword not in fl):
                return os.path.join(folder, fn)
        return None

    entries = []
    for keyword, flag, excl in [("adimplente", "adimplente", "inadimplente"), ("inadimplente", "inadimplente", None)]:
        fpath = _find_file(folder_path, keyword, exclude_keyword=excl)
        if not fpath:
            continue
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        if not wb.sheetnames:
            wb.close()
            continue
        ws = wb[wb.sheetnames[0]]
        first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not first_row:
            wb.close()
            continue
        raw_header = [c.value for c in first_row]
        col_idx = {}
        for i, h in enumerate(raw_header):
            if h:
                key = HEADER_NORM.get(h.strip().lower(), h.strip().lower()[:40])
                col_idx[key] = i

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                continue
            entry = {}
            for key, idx in col_idx.items():
                v = row[idx] if idx < len(row) else None
                if v is None:
                    v = ""
                elif isinstance(v, datetime):
                    v = v.strftime("%d/%m/%Y")
                elif isinstance(v, float) and v == int(v):
                    v = str(int(v))
                else:
                    v = str(v).strip()
                entry[key] = v
            entry["status_financeiro"] = flag
            entry["rgm_digits"] = _normalize_digits(entry.get("rgm", ""))
            entry["cpf_digits"] = _normalize_digits(entry.get("cpf", ""))
            phones = []
            d = _normalize_digits(entry.get("fone_cel", ""))
            if d:
                phones.append(d)
            entry["phones_digits"] = phones
            entries.append(entry)
        wb.close()
    return entries


# ---------------------------------------------------------------------------
# Rotas — Snapshots
# ---------------------------------------------------------------------------

@upload_bp.route("/api/xl-snapshots")
def api_xl_snapshots():
    tipo = request.args.get("tipo", "").strip().lower()
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if tipo:
                cur.execute("""
                    SELECT id, tipo, filename, row_count, uploaded_at
                    FROM xl_snapshots WHERE tipo = %s ORDER BY id DESC LIMIT 20
                """, (tipo,))
            else:
                cur.execute("""
                    SELECT id, tipo, filename, row_count, uploaded_at
                    FROM xl_snapshots ORDER BY id DESC LIMIT 20
                """)
            snaps = cur.fetchall()
        for s in snaps:
            s["uploaded_at"] = to_brt(s["uploaded_at"])
        return jsonify({"snapshots": snaps, "tipos": XL_TIPOS})
    except Exception as e:
        return jsonify({"snapshots": [], "tipos": XL_TIPOS, "error": str(e)}), 500
    finally:
        conn.close()


@upload_bp.route("/api/snapshots/compare")
def api_snapshots_compare():
    """Compara dois snapshots do mesmo tipo."""
    tipo = request.args.get("tipo", "").strip().lower()
    periodo = request.args.get("periodo", "")
    snap_a = request.args.get("snap_a", "")
    snap_b = request.args.get("snap_b", "")

    if not tipo and not snap_a:
        return jsonify({"error": "Informe tipo ou snap_a"}), 400

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if snap_a and snap_b:
                cur.execute("SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots WHERE id IN (%s,%s) ORDER BY id DESC", (snap_a, snap_b))
                snaps = cur.fetchall()
                if len(snaps) < 2:
                    return jsonify({"error": "Snapshots não encontrados"}), 404
                sa, sb = snaps[0], snaps[1]
            else:
                cur.execute("SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots WHERE tipo=%s ORDER BY id DESC LIMIT 1", (tipo,))
                sa = cur.fetchone()
                if not sa:
                    return jsonify({"error": f"Nenhum snapshot para tipo '{tipo}'"}), 404

                period_map = {"6m": 180, "1y": 365, "2y": 730, "3m": 90}
                days = period_map.get(periodo, 180)
                target_date = sa["uploaded_at"] - timedelta(days=days)
                cur.execute(
                    "SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots "
                    "WHERE tipo=%s AND uploaded_at <= %s ORDER BY uploaded_at DESC LIMIT 1",
                    (tipo, target_date),
                )
                sb = cur.fetchone()
                if not sb:
                    cur.execute(
                        "SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots "
                        "WHERE tipo=%s AND id < %s ORDER BY id ASC LIMIT 1",
                        (tipo, sa["id"]),
                    )
                    sb = cur.fetchone()

            def _get_stats(sid):
                cur.execute("SELECT metric, value, detail FROM xl_snapshot_stats WHERE snapshot_id=%s", (sid,))
                stats = {}
                for r in cur.fetchall():
                    stats[r["metric"]] = r["detail"] if r["detail"] is not None else (float(r["value"]) if r["value"] is not None else None)
                return stats

            stats_a = _get_stats(sa["id"])
            stats_b = _get_stats(sb["id"]) if sb else {}

            cur.execute("SELECT data->>'rgm_digits' AS rgm FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (sa["id"],))
            rgms_a = {r["rgm"] for r in cur.fetchall()}
            rgms_b = set()
            if sb:
                cur.execute("SELECT data->>'rgm_digits' AS rgm FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (sb["id"],))
                rgms_b = {r["rgm"] for r in cur.fetchall()}

            for s in [sa, sb]:
                if s:
                    s["uploaded_at"] = to_brt(s["uploaded_at"])

        return jsonify({
            "snap_a": sa,
            "snap_b": sb,
            "stats_a": stats_a,
            "stats_b": stats_b,
            "novos": len(rgms_a - rgms_b),
            "removidos": len(rgms_b - rgms_a),
            "mantidos": len(rgms_a & rgms_b),
            "delta_total": (sa["row_count"] - sb["row_count"]) if sb else 0,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@upload_bp.route("/api/snapshots/timeline")
def api_snapshots_timeline():
    """Série temporal de uma métrica para um tipo de snapshot."""
    tipo = request.args.get("tipo", "").strip().lower()
    metric = request.args.get("metric", "total")
    months = int(request.args.get("months", 24))

    if not tipo:
        return jsonify({"error": "Informe tipo"}), 400

    conn = get_conn()
    try:
        cutoff = datetime.now(BRT) - timedelta(days=months * 30)
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT s.id, s.uploaded_at, s.row_count, st.value, st.detail "
                "FROM xl_snapshots s "
                "LEFT JOIN xl_snapshot_stats st ON st.snapshot_id = s.id AND st.metric = %s "
                "WHERE s.tipo = %s AND s.uploaded_at >= %s "
                "ORDER BY s.uploaded_at",
                (metric, tipo, cutoff),
            )
            rows = cur.fetchall()

        points = []
        for r in rows:
            val = None
            if r["detail"] is not None:
                val = r["detail"]
            elif r["value"] is not None:
                val = float(r["value"])
            elif metric == "total":
                val = r["row_count"]
            points.append({
                "date": to_brt(r["uploaded_at"]),
                "snapshot_id": r["id"],
                "value": val,
            })

        return jsonify({"tipo": tipo, "metric": metric, "points": points})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


_POS_RE_UPLOAD = re.compile(r'p[oó]s', re.IGNORECASE)
_POS_CURSO_RE_UPLOAD = re.compile(r'(mba|especializa[cç][aã]o|p[oó]s.gradua|lato.sensu|stricto)', re.IGNORECASE)


def _classify_nivel_row(data):
    """Classify a row's nivel from its data dict (Python-side counterpart of SQL CASE)."""
    raw_nivel = data.get("nivel", "")
    if raw_nivel and _POS_RE_UPLOAD.search(raw_nivel):
        return "Pós-Graduação"
    raw_negocio = data.get("negocio", "")
    if _POS_RE_UPLOAD.search(raw_negocio or ""):
        return "Pós-Graduação"
    raw_curso = data.get("curso", "")
    if _POS_CURSO_RE_UPLOAD.search(raw_curso or ""):
        return "Pós-Graduação"
    return "Graduação"


@upload_bp.route("/api/snapshots/crossref")
def api_snapshots_crossref():
    """Cruzamento entre dois tipos de snapshot por RGM, com filtro opcional por nivel."""
    tipo_a = request.args.get("tipo_a", "").strip().lower()
    tipo_b = request.args.get("tipo_b", "").strip().lower()
    nivel = request.args.get("nivel", "").strip()

    if not tipo_a or not tipo_b:
        return jsonify({"error": "Informe tipo_a e tipo_b"}), 400

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            def _latest_rgms(tipo, filter_nivel=None):
                if tipo == 'inadimplentes':
                    if filter_nivel:
                        cur.execute("""
                            SELECT id FROM xl_snapshots
                            WHERE tipo = 'inadimplentes' AND nivel = %s
                            ORDER BY id DESC LIMIT 1
                        """, (filter_nivel,))
                        snap = cur.fetchone()
                        if not snap:
                            return set(), None
                        cur.execute(
                            "SELECT data->>'rgm_digits' AS rgm FROM xl_rows "
                            "WHERE snapshot_id = %s AND data->>'rgm_digits' != ''",
                            (snap["id"],)
                        )
                        return {r["rgm"] for r in cur.fetchall()}, snap["id"]
                    else:
                        cur.execute("""
                            SELECT DISTINCT ON (COALESCE(nivel, ''))
                                   id FROM xl_snapshots
                            WHERE tipo = 'inadimplentes'
                            ORDER BY COALESCE(nivel, ''), id DESC
                        """)
                        snap_ids = [r["id"] for r in cur.fetchall()]
                        if not snap_ids:
                            return set(), None
                        placeholders = ','.join(['%s'] * len(snap_ids))
                        cur.execute(
                            f"SELECT data->>'rgm_digits' AS rgm FROM xl_rows "
                            f"WHERE snapshot_id IN ({placeholders}) AND data->>'rgm_digits' != ''",
                            snap_ids
                        )
                        return {r["rgm"] for r in cur.fetchall()}, snap_ids[0]

                cur.execute("SELECT id FROM xl_snapshots WHERE tipo=%s ORDER BY id DESC LIMIT 1", (tipo,))
                snap = cur.fetchone()
                if not snap:
                    return set(), None

                if filter_nivel and tipo == 'matriculados':
                    cur.execute("SELECT data FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (snap["id"],))
                    rgms = set()
                    for r in cur.fetchall():
                        row_nivel = _classify_nivel_row(r["data"])
                        if row_nivel == filter_nivel:
                            rgm = r["data"].get("rgm_digits", "")
                            if rgm:
                                rgms.add(rgm)
                    return rgms, snap["id"]

                cur.execute("SELECT data->>'rgm_digits' AS rgm FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (snap["id"],))
                return {r["rgm"] for r in cur.fetchall()}, snap["id"]

            rgms_a, sid_a = _latest_rgms(tipo_a, filter_nivel=nivel if nivel else None)
            rgms_b, sid_b = _latest_rgms(tipo_b, filter_nivel=nivel if nivel else None)

        return jsonify({
            "tipo_a": tipo_a, "tipo_b": tipo_b,
            "total_a": len(rgms_a), "total_b": len(rgms_b),
            "em_ambos": len(rgms_a & rgms_b),
            "apenas_a": len(rgms_a - rgms_b),
            "apenas_b": len(rgms_b - rgms_a),
            "snap_a": sid_a, "snap_b": sid_b,
            "nivel": nivel or "Consolidado",
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@upload_bp.route("/api/inadimplencia/historico")
def api_inadimplencia_historico():
    """Retorna séries temporais de inadimplentes agrupados por nivel, tipo_aluno e turma."""
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    conn = get_conn()
    try:
        date_clause = ""
        params = []
        if date_from:
            date_clause += " AND s.uploaded_at >= %s::date"
            params.append(date_from)
        if date_to:
            date_clause += " AND s.uploaded_at < (%s::date + interval '1 day')"
            params.append(date_to)

        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"""
                SELECT s.id, s.uploaded_at, s.row_count, s.nivel AS snap_nivel
                FROM xl_snapshots s
                WHERE s.tipo = 'inadimplentes'{date_clause}
                ORDER BY s.uploaded_at
            """, params)
            snapshots = cur.fetchall()

            if not snapshots:
                return jsonify({"snapshots": [], "series": [], "message": "Nenhum snapshot de inadimplentes encontrado."})

            cur.execute(f"""
                SELECT s.id AS snap_id, s.uploaded_at, s.nivel AS snap_nivel,
                       r.data
                FROM xl_snapshots s
                JOIN xl_rows r ON r.snapshot_id = s.id
                WHERE s.tipo = 'inadimplentes'{date_clause}
                ORDER BY s.uploaded_at
            """, params)
            all_rows = cur.fetchall()

            cur.execute("SELECT id FROM xl_snapshots WHERE tipo='matriculados' ORDER BY id DESC LIMIT 1")
            snap_mat = cur.fetchone()
            mat_nivel_map = {}
            mat_tipo_map = {}
            mat_turma_map = {}
            if snap_mat:
                cur.execute("SELECT data FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (snap_mat["id"],))
                for r in cur.fetchall():
                    d = r["data"]
                    rgm = d.get("rgm_digits", "")
                    if rgm:
                        mat_nivel_map[rgm] = _classify_nivel_row(d)
                        tipo_raw = (d.get("tipo_matricula", "") or "").strip()
                        mat_tipo_map[rgm] = tipo_raw if tipo_raw else "N/I"
                        turma_raw = (d.get("serie", "") or d.get("ciclo", "") or "").strip()
                        mat_turma_map[rgm] = turma_raw if turma_raw else "N/I"

        snapshot_data = {}
        for row in all_rows:
            sid = row["snap_id"]
            uploaded = row["uploaded_at"]
            snap_nivel = row.get("snap_nivel") or None
            d = row["data"]
            rgm = d.get("rgm_digits", "")
            if not rgm:
                continue

            if sid not in snapshot_data:
                snapshot_data[sid] = {
                    "date": to_brt(uploaded),
                    "snap_nivel": snap_nivel,
                    "by_nivel": {},
                    "by_tipo": {},
                    "by_turma": {},
                    "total": 0,
                }
            sd = snapshot_data[sid]
            sd["total"] += 1

            nivel = snap_nivel or mat_nivel_map.get(rgm, "Graduação")
            sd["by_nivel"][nivel] = sd["by_nivel"].get(nivel, 0) + 1

            tipo = mat_tipo_map.get(rgm, "N/I")
            sd["by_tipo"][tipo] = sd["by_tipo"].get(tipo, 0) + 1

            turma = mat_turma_map.get(rgm, "N/I")
            sd["by_turma"][turma] = sd["by_turma"].get(turma, 0) + 1

        series = []
        for sid in sorted(snapshot_data.keys()):
            sd = snapshot_data[sid]
            series.append({
                "snapshot_id": sid,
                "date": sd["date"],
                "snap_nivel": sd["snap_nivel"],
                "total": sd["total"],
                "by_nivel": sd["by_nivel"],
                "by_tipo": sd["by_tipo"],
                "by_turma": dict(sorted(sd["by_turma"].items(), key=lambda x: -x[1])[:20]),
            })

        return jsonify({
            "snapshots_count": len(series),
            "series": series,
            "has_history": len(series) >= 2,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@upload_bp.route("/api/snapshots/crossref/export")
def api_snapshots_crossref_export():
    """Exporta CSV dos alunos de um subset do cruzamento entre dois tipos."""
    tipo_a = request.args.get("tipo_a", "").strip().lower()
    tipo_b = request.args.get("tipo_b", "").strip().lower()
    subset = request.args.get("subset", "em_ambos").strip()
    nivel = request.args.get("nivel", "").strip()

    if not tipo_a or not tipo_b:
        return jsonify({"error": "Informe tipo_a e tipo_b"}), 400
    if subset not in ("em_ambos", "apenas_a", "apenas_b"):
        return jsonify({"error": "subset deve ser em_ambos, apenas_a ou apenas_b"}), 400

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            def _get_snap_data(tipo, filter_nivel=None):
                if tipo == 'inadimplentes':
                    if filter_nivel:
                        cur.execute("SELECT id FROM xl_snapshots WHERE tipo='inadimplentes' AND nivel=%s ORDER BY id DESC LIMIT 1", (filter_nivel,))
                    else:
                        cur.execute("SELECT id FROM xl_snapshots WHERE tipo='inadimplentes' ORDER BY id DESC LIMIT 1")
                    snap = cur.fetchone()
                    if not snap:
                        return {}
                    cur.execute("SELECT data FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (snap["id"],))
                    return {r["data"]["rgm_digits"]: r["data"] for r in cur.fetchall()}

                cur.execute("SELECT id FROM xl_snapshots WHERE tipo=%s ORDER BY id DESC LIMIT 1", (tipo,))
                snap = cur.fetchone()
                if not snap:
                    return {}
                cur.execute("SELECT data FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (snap["id"],))
                all_data = {r["data"]["rgm_digits"]: r["data"] for r in cur.fetchall()}
                if filter_nivel and tipo == 'matriculados':
                    return {k: v for k, v in all_data.items() if _classify_nivel_row(v) == filter_nivel}
                return all_data

            data_a = _get_snap_data(tipo_a, filter_nivel=nivel if nivel else None)
            data_b = _get_snap_data(tipo_b, filter_nivel=nivel if nivel else None)

            rgms_a = set(data_a.keys())
            rgms_b = set(data_b.keys())

            if subset == "em_ambos":
                target_rgms = rgms_a & rgms_b
                source = data_a
            elif subset == "apenas_a":
                target_rgms = rgms_a - rgms_b
                source = data_a
            else:
                target_rgms = rgms_b - rgms_a
                source = data_b

        buf = _io.StringIO()
        writer = _csv.writer(buf, delimiter=';')
        writer.writerow(["Nome", "RGM", "CPF", "Curso", "Polo", "Email", "Nivel", "Tipo"])
        for rgm in sorted(target_rgms):
            d = source.get(rgm, {})
            row_nivel = _classify_nivel_row(d)
            writer.writerow([
                d.get("nome", ""), d.get("rgm", ""), d.get("cpf", ""),
                d.get("curso", ""), d.get("polo", ""),
                d.get("email", ""), row_nivel,
                d.get("tipo_matricula", d.get("tipo_titulo", "")),
            ])

        output = buf.getvalue()
        fname = f"crossref_{tipo_a}_{tipo_b}_{subset}"
        if nivel:
            fname += f"_{nivel.replace(' ','_').lower()}"
        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={fname}.csv"},
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@upload_bp.route("/api/upload/info")
def api_upload_info():
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            result = {}
            for t in XL_TIPOS:
                cur.execute("""
                    SELECT id, tipo, filename, row_count, uploaded_at
                    FROM xl_snapshots WHERE tipo = %s ORDER BY id DESC LIMIT 1
                """, (t,))
                snap = cur.fetchone()
                if snap:
                    snap["uploaded_at"] = to_brt(snap["uploaded_at"])
                result[t] = snap
        file_disk = _find_xlsx()
        return jsonify({"file": file_disk, "snapshots": result})
    except Exception as e:
        return jsonify({"file": _find_xlsx(), "snapshots": {}, "error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Upload
# ---------------------------------------------------------------------------

@upload_bp.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Nenhum arquivo selecionado."}), 400

    fname_lower = f.filename.lower()
    allowed_ext = (".xlsx", ".xlsm", ".zip")
    if not any(fname_lower.endswith(ext) for ext in allowed_ext):
        return jsonify({"error": "Aceitos: .xlsx, .xlsm ou .zip"}), 400

    tipo = request.form.get("tipo", "matriculados").strip().lower()
    if tipo not in XL_TIPOS:
        return jsonify({"error": f"Tipo inválido. Use: {', '.join(XL_TIPOS)}"}), 400

    if tipo == "matriculados":
        for old in UPLOAD_DIR.iterdir():
            if old.suffix.lower() == ".xlsx" and "matriculados" in old.name.lower():
                old.unlink()

    safe_name = f.filename
    if tipo == "matriculados" and "matriculados" not in safe_name.lower():
        safe_name = "Relação de matriculados por polo.xlsx"

    dest = UPLOAD_DIR / safe_name
    f.save(str(dest))

    try:
        if fname_lower.endswith(".zip"):
            snap_count = _handle_zip_upload(str(dest), tipo)
        elif tipo == "inadimplentes" and fname_lower.endswith(".xlsm"):
            tmp_dir = UPLOAD_DIR / f"_tmp_{tipo}"
            tmp_dir.mkdir(exist_ok=True)
            shutil.copy2(str(dest), str(tmp_dir / safe_name))
            entries = _parse_inadimplentes_batch(str(tmp_dir))
            snap_count = _persist_snapshot_entries(entries, tipo, safe_name) if entries else 0
        elif tipo == "sem_rematricula" and fname_lower.endswith((".xlsx", ".xlsm")):
            staging = UPLOAD_DIR / "_staging_sem_rematricula"
            staging.mkdir(exist_ok=True)
            subtipo = request.form.get("subtipo", "").strip().lower()
            if subtipo in ("adimplente", "inadimplente"):
                canonical = f"{subtipo}s.xlsx"
            elif "inadimplente" in safe_name.lower():
                canonical = "inadimplentes.xlsx"
            elif "adimplente" in safe_name.lower():
                canonical = "adimplentes.xlsx"
            else:
                canonical = safe_name
            shutil.copy2(str(dest), str(staging / canonical))
            staged_files = [p.name for p in staging.iterdir() if p.suffix.lower() in (".xlsx", ".xlsm")]
            current_app.logger.info("sem_rematricula staging: salvou '%s' como '%s'. Arquivos: %s", safe_name, canonical, staged_files)
            entries = _parse_sem_rematricula(str(staging))
            if entries:
                snap_count = _persist_snapshot_entries(entries, tipo, safe_name)
                current_app.logger.info("sem_rematricula snapshot criado: %d linhas", snap_count)
            else:
                snap_count = 0
                current_app.logger.info("sem_rematricula: '%s' salvo, aguardando o outro arquivo", canonical)
        else:
            snap_count = _save_xl_snapshot(str(dest), safe_name, tipo)
    except Exception as e:
        current_app.logger.warning("Erro ao gravar snapshot (%s): %s", tipo, e)
        current_app.logger.warning("Traceback: %s", traceback.format_exc())
        snap_count = -1

    stat = dest.stat()
    return jsonify({
        "ok": True,
        "tipo": tipo,
        "file": {
            "name": dest.name,
            "size": stat.st_size,
            "modified": to_brt(datetime.fromtimestamp(stat.st_mtime, tz=BRT)),
        },
        "snapshot_rows": snap_count,
    })


@upload_bp.route("/api/upload-batch", methods=["POST"])
def api_upload_batch():
    """Recebe múltiplos arquivos .xlsm/.xlsx para inadimplentes e processa em lote."""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400

    tipo = request.form.get("tipo", "").strip().lower()
    if tipo not in ("inadimplentes",):
        return jsonify({"error": "Upload em lote só suportado para inadimplentes."}), 400

    nivel = request.form.get("nivel", "").strip() or None

    allowed_ext = (".xlsx", ".xlsm")
    tmp_dir = UPLOAD_DIR / f"_tmp_{tipo}"
    if tmp_dir.exists():
        shutil.rmtree(str(tmp_dir), ignore_errors=True)
    tmp_dir.mkdir(exist_ok=True)

    saved = []
    for f in files:
        if not f.filename:
            continue
        fname_lower = f.filename.lower()
        if not any(fname_lower.endswith(ext) for ext in allowed_ext):
            continue
        dest = tmp_dir / f.filename
        f.save(str(dest))
        saved.append(f.filename)

    if not saved:
        return jsonify({"error": "Nenhum arquivo .xlsx/.xlsm válido encontrado."}), 400

    current_app.logger.info("[UPLOAD-BATCH] %d arquivos salvos em %s: %s", len(saved), tmp_dir, saved[:5])

    try:
        entries = _parse_inadimplentes_batch(str(tmp_dir))
        snap_count = _persist_snapshot_entries(entries, tipo, f"{len(saved)} arquivos", nivel=nivel) if entries else 0
    except Exception as e:
        current_app.logger.warning("Erro upload-batch (%s): %s", tipo, e)
        current_app.logger.warning("Traceback: %s", traceback.format_exc())
        return jsonify({"error": f"Erro ao processar: {e}"}), 500
    finally:
        shutil.rmtree(str(tmp_dir), ignore_errors=True)

    if snap_count == 0 and saved:
        return jsonify({
            "ok": True,
            "tipo": tipo,
            "files_count": len(saved),
            "snapshot_rows": 0,
            "warning": "Arquivos recebidos mas nenhum dado extraído. Verifique se os arquivos contêm a linha de cabeçalho com ID_POLO na primeira coluna.",
        })

    return jsonify({
        "ok": True,
        "tipo": tipo,
        "files_count": len(saved),
        "snapshot_rows": snap_count,
    })


@upload_bp.route("/api/upload-folder", methods=["POST"])
def api_upload_folder():
    """Processa pastas já presentes no servidor (deploy/scp)."""
    body = request.json or {}
    tipo = body.get("tipo", "").strip().lower()
    folder = body.get("path", "").strip()

    if tipo not in XL_TIPOS:
        return jsonify({"error": f"Tipo inválido. Use: {', '.join(XL_TIPOS)}"}), 400

    if not folder:
        default_folders = {
            "inadimplentes": str(BASE_DIR / "Inadimplentes"),
            "sem_rematricula": str(BASE_DIR / "Sem_Rematricula"),
            "concluintes": str(BASE_DIR / "Concluíntes"),
            "acesso_ava": str(BASE_DIR / "Acesso_AVA"),
        }
        folder = default_folders.get(tipo, "")

    if not folder or not os.path.isdir(folder):
        return jsonify({"error": f"Pasta não encontrada: {folder}"}), 404

    try:
        folder_name = os.path.basename(folder)
        if tipo == "inadimplentes":
            entries = _parse_inadimplentes_batch(folder)
            count = _persist_snapshot_entries(entries, tipo, f"{folder_name} (servidor)") if entries else 0
        elif tipo == "sem_rematricula":
            entries = _parse_sem_rematricula(folder)
            count = _persist_snapshot_entries(entries, tipo, f"{folder_name} (servidor)") if entries else 0
        else:
            xlsx_files = sorted(Path(folder).glob("*.xlsx"))
            if not xlsx_files:
                return jsonify({"error": "Nenhum arquivo .xlsx encontrado na pasta."}), 404
            count = _save_xl_snapshot(str(xlsx_files[0]), xlsx_files[0].name, tipo)

        return jsonify({"ok": True, "tipo": tipo, "snapshot_rows": count, "folder": folder})
    except Exception as e:
        current_app.logger.exception("Erro ao processar pasta %s", folder)
        return jsonify({"error": str(e)}), 500
