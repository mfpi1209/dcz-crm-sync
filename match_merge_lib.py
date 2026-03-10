"""
Match & Merge SIAA — Engine.

Pipeline completo:
1. Lê arquivos .xlsm/.xlsx (candidatos inscritos / matriculados por polo)
2. Merge em consolidado por tipo
3. Normalização (curso, polo, CPF, situação, modalidade, grau, preço)
4. Cruzamento inscritos × matriculados por CPF + similaridade de curso
5. Match SIAA × Kommo por CPF/telefone + curso
6. Geração de ações (APROVADO, MATRICULADO, NOVO)
7. Execução de updates no Kommo via API v4
"""

import io
import os
import re
import json
import time
import logging
import zipfile
import threading
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from difflib import SequenceMatcher
from pathlib import Path

import requests
import psycopg2
from psycopg2.extras import execute_values, RealDictCursor
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

log = logging.getLogger("match_merge")

BRT = timezone(timedelta(hours=-3))

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = BASE_DIR / "uploads" / "match_merge"

DB_DSN = dict(
    host=os.getenv("DB_HOST", "localhost"),
    port=os.getenv("DB_PORT", "5432"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASS"),
    dbname=os.getenv("DB_NAME", "dcz_sync"),
)

KOMMO_DB_DSN = dict(
    host=os.getenv("KOMMO_PG_HOST", os.getenv("DB_HOST", "localhost")),
    port=os.getenv("KOMMO_PG_PORT", os.getenv("DB_PORT", "5432")),
    user=os.getenv("KOMMO_PG_USER", os.getenv("DB_USER")),
    password=os.getenv("KOMMO_PG_PASS", os.getenv("DB_PASS")),
    dbname=os.getenv("KOMMO_PG_DB", "kommo_sync"),
)

KOMMO_BASE_URL = os.getenv("KOMMO_BASE_URL", "https://eduitbr.kommo.com")
KOMMO_TOKEN = os.getenv("KOMMO_TOKEN", "")

DATACRAZY_API_BASE = "https://api.g1.datacrazy.io/api/v1"
DATACRAZY_API_TOKEN = os.getenv("DATACRAZY_API_TOKEN", "")

PREPOSICOES = {"de", "da", "do", "dos", "das", "e"}

EMPRESAS_PERMITIDAS = {"1", "7", "12"}


def get_conn():
    return psycopg2.connect(**DB_DSN)


def get_kommo_conn():
    return psycopg2.connect(**KOMMO_DB_DSN)


# ════════════════════════════════════════════════════════════════
#  REFERENCE DATA
# ════════════════════════════════════════════════════════════════

_precos = None
_procvs = None


def _load_json(name):
    with open(DATA_DIR / name, encoding="utf-8") as f:
        return json.load(f)


def get_precos():
    global _precos
    if _precos is None:
        data = _load_json("precos.json")
        _precos = {item["chave"].upper(): item for item in data}
    return _precos


def get_procvs():
    global _procvs
    if _procvs is None:
        _procvs = _load_json("procvs.json")
    return _procvs


# ════════════════════════════════════════════════════════════════
#  EXCEL READING  (from merge_excel.py)
# ════════════════════════════════════════════════════════════════

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _col_index(col_str):
    result = 0
    for c in col_str.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result - 1


def _parse_cell_ref(ref):
    m = re.match(r"([A-Z]+)(\d+)", ref)
    return (m.group(1), int(m.group(2))) if m else (None, None)


def read_xlsm_xml(filepath, max_col=35):
    """Parse .xlsm/.xlsx via internal XML (handles inlineStr)."""
    rows_data = {}
    try:
        with zipfile.ZipFile(filepath) as z:
            sheet_path = None
            for name in z.namelist():
                if "worksheets/sheet1.xml" in name.lower():
                    sheet_path = name
                    break
            if not sheet_path:
                return []
            with z.open(sheet_path) as f:
                tree = ET.parse(f)

        root = tree.getroot()
        for row_el in root.iter(f"{{{NS}}}row"):
            row_num = int(row_el.get("r", 0))
            for cell_el in row_el.iter(f"{{{NS}}}c"):
                ref = cell_el.get("r", "")
                col_str, _ = _parse_cell_ref(ref)
                if not col_str:
                    continue
                col_idx = _col_index(col_str)
                if col_idx >= max_col:
                    continue

                value = None
                cell_type = cell_el.get("t", "")

                if cell_type == "inlineStr":
                    is_el = cell_el.find(f"{{{NS}}}is")
                    if is_el is not None:
                        t_el = is_el.find(f"{{{NS}}}t")
                        if t_el is not None and t_el.text:
                            value = t_el.text
                elif cell_type == "s":
                    v_el = cell_el.find(f"{{{NS}}}v")
                    if v_el is not None:
                        value = v_el.text
                else:
                    v_el = cell_el.find(f"{{{NS}}}v")
                    if v_el is not None:
                        value = v_el.text

                if value is not None:
                    if row_num not in rows_data:
                        rows_data[row_num] = {}
                    rows_data[row_num][col_idx] = value
    except Exception as e:
        log.error("Erro ao parsear XML de %s: %s", filepath, e)
        return []

    if not rows_data:
        return []

    max_col_found = max(max(cols.keys()) for cols in rows_data.values()) + 1
    result = []
    for row_num in sorted(rows_data.keys()):
        row = [rows_data[row_num].get(c) for c in range(max_col_found)]
        if any(v is not None for v in row):
            result.append(row)
    return result


def read_excel_file(filepath):
    """Read .xlsx (openpyxl) or .xlsm (XML fallback)."""
    rows = []
    if str(filepath).lower().endswith(".xlsx"):
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if any(v is not None for v in row):
                    rows.append(list(row))
            wb.close()
            if len(rows) > 1:
                return rows
        except Exception:
            pass

    rows = read_xlsm_xml(str(filepath))
    filtered = []
    for row in rows:
        first_val = str(row[0] or "").strip()
        if not filtered:
            if first_val == "" or first_val.upper().startswith("RELA"):
                continue
        filtered.append(row)
    return filtered


# ════════════════════════════════════════════════════════════════
#  MERGE
# ════════════════════════════════════════════════════════════════

def merge_uploaded_files(file_paths, tipo_label=""):
    """Merge multiple Excel files into a single list of rows (header + data)."""
    all_rows = []
    header_written = False
    total = 0

    for i, fpath in enumerate(file_paths):
        fname = os.path.basename(fpath)
        log.info("  [%d/%d] %s", i + 1, len(file_paths), fname)
        try:
            rows = read_excel_file(fpath)
            if not rows:
                log.warning("    -> 0 linhas")
                continue
            for idx, row in enumerate(rows):
                if idx == 0:
                    if not header_written:
                        all_rows.append(row + ["Arquivo_Origem"])
                        header_written = True
                    continue
                all_rows.append(row + [fname])
                total += 1
            log.info("    -> %d linhas", total)
        except Exception as e:
            log.error("  Erro ao processar %s: %s", fname, e)

    log.info("Merge %s: %d linhas de %d arquivos", tipo_label, total, len(file_paths))
    return all_rows


# ════════════════════════════════════════════════════════════════
#  NORMALIZATION  (from cruzamento.py)
# ════════════════════════════════════════════════════════════════

def limpar_nome(val):
    if not val or str(val).strip() in ("", "----"):
        return None
    palavras = str(val).strip().split()
    return " ".join(
        p.lower() if p.lower() in PREPOSICOES else p.capitalize()
        for p in palavras
    )


def limpar_cpf(val):
    if not val or str(val).strip() in ("", "----"):
        return None
    try:
        num = int(float(str(val)))
        return str(num).zfill(11)
    except (ValueError, TypeError):
        return re.sub(r"\D", "", str(val)).zfill(11)


def limpar_telefone(ddd, fone):
    def to_digits(v):
        if not v or str(v).strip() in ("", "----"):
            return ""
        try:
            return str(int(float(str(v))))
        except (ValueError, TypeError):
            return re.sub(r"\D", "", str(v))
    d, f = to_digits(ddd), to_digits(fone)
    return (d + f) if (d or f) else None


def limpar_cep(val):
    if not val or str(val).strip() in ("", "----"):
        return None
    try:
        num = int(float(str(val)))
        cep = str(num).zfill(8)
        return f"{cep[:5]}-{cep[5:]}"
    except (ValueError, TypeError):
        return re.sub(r"\D", "", str(val))


def limpar_data(val):
    if not val or str(val).strip() in ("", "----"):
        return None
    try:
        return datetime.strptime(str(val).strip(), "%d/%m/%Y").date()
    except (ValueError, TypeError):
        return None


def limpar_valor(val):
    if not val:
        return None
    s = str(val).strip()
    return None if s in ("----", "") else s


def limpar_rgm(val):
    if not val or str(val).strip() in ("", "----"):
        return None
    try:
        return str(int(float(str(val))))
    except (ValueError, TypeError):
        return re.sub(r"\D", "", str(val)) or None


def detectar_grau(curso_raw):
    if not curso_raw:
        return None
    s = str(curso_raw).upper()
    if "CST" in s or "TECNÓLOGO" in s or "TECNOLOGO" in s:
        return "Tecnólogo"
    if "BACHAREL" in s:
        return "Bacharelado"
    if "LICENCIA" in s:
        return "Licenciatura"
    return None


def detectar_modalidade(curso_raw):
    if not curso_raw:
        return "EAD Digital"
    s = str(curso_raw)
    if "4.0i" in s or "aulas ao vivo" in s.lower():
        return "EaD com aulas ao vivo"
    if "4.0" in s or "semipresencial" in s.lower():
        return "EaD Semipresencial"
    return "EAD Digital"


def limpar_curso(curso_raw):
    if not curso_raw:
        return None
    s = str(curso_raw).strip()
    if s in ("", "----"):
        return None
    s = re.sub(r"^\d+\s*-\s*", "", s)
    s = re.sub(r"^CST\s+EM\s+", "", s, flags=re.IGNORECASE)
    s = re.sub(r"^Cst\s+[Ee]m\s+", "", s)
    s = re.sub(r"\s*\([^)]*\)", "", s)
    s = re.sub(r"\s+\d+\.\d+\w*.*$", "", s)
    s = " ".join(s.split()).strip()
    if not s:
        return None
    return " ".join(
        p.lower() if p.lower() in PREPOSICOES else p.capitalize()
        for p in s.split()
    )


def construir_chave_preco(curso_raw):
    if not curso_raw:
        return None
    s = str(curso_raw).strip()
    procvs = get_procvs()
    for correcao in procvs.get("cursos", []):
        if s.upper() == correcao["de"].upper():
            s = correcao["para"]
            break
    modalidade = detectar_modalidade(curso_raw)
    grau = detectar_grau(curso_raw)
    nome = re.sub(r"^\d+\s*-\s*", "", s)
    nome = re.sub(r"^CST\s+EM\s+", "", nome, flags=re.IGNORECASE)
    nome = re.sub(r"^Cst\s+[Ee]m\s+", "", nome)
    nome = re.sub(r"\s*\([^)]*\)", "", nome)
    nome = re.sub(r"\s+\d+\.\d+\w*.*$", "", nome)
    nome = " ".join(nome.split()).strip().upper()
    if not nome:
        return None
    grau_str = grau.upper() if grau else ""
    return f"{nome} ({modalidade}){grau_str}"


def normalizar_polo_procvs(polo_raw):
    if not polo_raw:
        return None, None
    procvs = get_procvs()
    polo_upper = str(polo_raw).strip().upper()
    for entry in procvs.get("polos", []):
        if entry["de"].upper() == polo_upper:
            return entry["para"], entry["marca"]
    return _normalizar_polo_fallback(polo_raw), None


POLO_MAP_FALLBACK = {
    "mituzi": "Taboão da Serra", "santos dumont": "Taboão da Serra - Centro",
    "ouro verde": "Campinas", "moinho velho": "Freguesia do Ó",
    "vila ema": "Sapopemba", "progredior": "Morumbi",
    "indianópolis": "Ibirapuera", "indianopolis": "Ibirapuera",
    "ibirapuera": "Ibirapuera", "campinas": "Campinas",
    "capivari": "Capivari", "itapira": "Itapira",
    "vila mariana": "Vila Mariana", "barra funda": "Barra Funda",
    "freguesia": "Freguesia do Ó", "morumbi": "Morumbi",
    "santana": "Santana", "sapopemba": "Sapopemba",
    "vila prudente": "Vila Prudente", "taboão": "Taboão da Serra",
    "taboao": "Taboão da Serra",
}

POLO_PRIORITY = [
    "mituzi", "santos dumont", "ouro verde", "moinho velho",
    "vila ema", "progredior", "indianópolis", "indianopolis",
]


def _normalizar_polo_fallback(val):
    if not val:
        return None
    lower = val.lower()
    for term in POLO_PRIORITY:
        if term in lower:
            return POLO_MAP_FALLBACK[term]
    for keyword, polo in POLO_MAP_FALLBACK.items():
        if keyword in lower:
            return polo
    return limpar_nome(val)


def calcular_situacao_final(situacao, data_pagamento):
    if not situacao:
        return situacao
    sit = str(situacao).strip()
    if sit in ("Matriculado", "Reprovado", "0", "Situação", ""):
        return sit
    if data_pagamento and str(data_pagamento).strip() not in ("", "----"):
        return f"{sit} - Boleto Pago"
    return sit


def normalizar_situacao_matriculado(situacao_raw):
    if not situacao_raw:
        return None
    s = str(situacao_raw).strip()
    if s.startswith("0") or "em curso" in s.lower():
        return "Matriculado"
    if s.startswith("5") or "cancelado" in s.lower():
        return "Cancelado"
    if "trancado" in s.lower() or "transferido" in s.lower():
        return "Transferido"
    return s


# ════════════════════════════════════════════════════════════════
#  HEADER-BASED COLUMN MAPPING
# ════════════════════════════════════════════════════════════════

def _build_col_map(header_row):
    """Build a case-insensitive header-name -> index map."""
    col_map = {}
    for idx, val in enumerate(header_row):
        if val is not None:
            col_map[str(val).strip().lower()] = idx
    return col_map


def _get_col(col_map, *aliases):
    """Find column index by trying multiple name aliases (case-insensitive)."""
    for alias in aliases:
        key = alias.strip().lower()
        if key in col_map:
            return col_map[key]
    return None


def _cell(row, idx):
    """Safely get cell value by index."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def limpar_data_flex(val):
    """Parse date from datetime objects or various string formats."""
    if val is None:
        return None
    from datetime import date as _date
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, _date):
        return val
    s = str(val).strip()
    if s in ("", "----", "None"):
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def limpar_telefone_unico(val):
    """Clean a single phone number value (no DDD+number split)."""
    if not val or str(val).strip() in ("", "----"):
        return None
    try:
        return str(int(float(str(val))))
    except (ValueError, TypeError):
        digits = re.sub(r"\D", "", str(val))
        return digits if digits else None


_STATUS_INSCRITO_MAP = {
    "APROVADOS": "Aprovado", "APROVADO": "Aprovado",
    "MATRICULADO": "Matriculado", "MATRICULADOS": "Matriculado",
    "REPROVADO": "Reprovado", "REPROVADOS": "Reprovado",
    "INSCRITO": "Inscrito", "INSCRITOS": "Inscrito",
    "CANCELADO": "Cancelado", "CANCELADOS": "Cancelado",
    "DESISTENTE": "Desistente", "DESISTENTES": "Desistente",
}


def _normalizar_status_inscrito(val):
    """Normalize inscrito status values to consistent casing."""
    if not val:
        return None
    s = str(val).strip()
    return _STATUS_INSCRITO_MAP.get(s.upper(), s)


def _empresa_permitida(val):
    """Check if the Empresa value prefix is in the allowed set."""
    if not val:
        return False
    s = str(val).strip()
    prefix = s.split(" - ")[0].strip() if " - " in s else s
    return prefix in EMPRESAS_PERMITIDAS


# ════════════════════════════════════════════════════════════════
#  NORMALIZE ROWS  (header-aware)
# ════════════════════════════════════════════════════════════════

def normalizar_inscritos(rows, header=None, tipo="grad"):
    """Normalise candidatos inscritos rows using header-based column mapping."""
    if not header:
        log.error("normalizar_inscritos: header obrigatório para mapeamento de colunas")
        return []

    cm = _build_col_map(header)

    i_nome      = _get_col(cm, "nome", "name")
    i_sexo      = _get_col(cm, "sexo", "sex")
    i_cpf       = _get_col(cm, "cpf")
    i_rg        = _get_col(cm, "rg")
    i_email     = _get_col(cm, "email", "e-mail")
    i_celular   = _get_col(cm, "celular", "fone celular", "telefone")
    i_inscricao = _get_col(cm, "inscrição", "inscricao", "código inscrição")
    i_modalidade = _get_col(cm, "modalidade")
    i_curso     = _get_col(cm, "curso")
    i_instituicao = _get_col(cm, "instituição", "instituicao")
    i_polo      = _get_col(cm, "polo")
    i_status    = _get_col(cm, "status", "situação", "situacao")
    i_data_inscr = _get_col(cm, "data inscrição", "data inscricao")
    i_data_aprov = _get_col(cm, "data aprovação", "data aprovacao")
    i_ciclo     = _get_col(cm, "ciclo vestibular", "ciclo")
    i_regional  = _get_col(cm, "regional")
    i_empresa   = _get_col(cm, "empresa")
    i_arquivo   = _get_col(cm, "arquivo_origem")

    log.info("Col-map inscritos: nome=%s cpf=%s curso=%s polo=%s status=%s empresa=%s",
             i_nome, i_cpf, i_curso, i_polo, i_status, i_empresa)

    dados = []
    skipped = 0
    for row in rows:
        if i_empresa is not None and not _empresa_permitida(_cell(row, i_empresa)):
            skipped += 1
            continue
        curso_raw = _cell(row, i_curso)
        polo_raw  = _cell(row, i_polo)
        polo_curto, marca = normalizar_polo_procvs(polo_raw)
        if not marca:
            inst = _cell(row, i_instituicao)
            if inst:
                marca = str(inst).strip()

        chave = construir_chave_preco(curso_raw)
        info_preco = get_precos().get(chave.upper(), {}) if chave else {}

        status_raw  = limpar_valor(_cell(row, i_status))
        status_norm = _normalizar_status_inscrito(status_raw)

        modalidade_val = limpar_valor(_cell(row, i_modalidade))
        if not modalidade_val or modalidade_val.upper() in ("SEM INFORMAÇÃO", "SEM INFORMACAO"):
            modalidade_val = detectar_modalidade(curso_raw)

        dados.append((
            tipo,                                            # tipo
            status_norm,                                     # status
            None,                                            # dt_pag_insc
            limpar_valor(_cell(row, i_inscricao)),           # inscricao
            limpar_nome(_cell(row, i_nome)),                 # nome
            limpar_valor(_cell(row, i_sexo)),                # sexo
            limpar_cpf(_cell(row, i_cpf)),                   # cpf
            limpar_valor(_cell(row, i_rg)),                  # rg
            limpar_valor(curso_raw),                         # curso_raw
            limpar_curso(curso_raw),                         # curso_limpo
            detectar_grau(curso_raw),                        # grau_curso
            modalidade_val,                                  # modalidade
            limpar_valor(polo_raw),                          # polo_raw
            polo_curto,                                      # polo_normalizado
            marca,                                           # marca_instituicao
            limpar_data_flex(_cell(row, i_data_inscr)),      # data_inscr
            limpar_data_flex(_cell(row, i_data_aprov)),      # data_prova
            limpar_telefone_unico(_cell(row, i_celular)),    # telefone
            None,                                            # telefone_res
            None,                                            # telefone_com
            limpar_valor(_cell(row, i_email)),               # email
            None,                                            # cep
            None,                                            # endereco
            None,                                            # bairro
            None,                                            # cidade
            limpar_valor(_cell(row, i_regional)),            # estado
            None,                                            # data_pagamento
            None,                                            # data_matricula
            status_raw,                                      # situacao_raw
            status_norm,                                     # situacao_final
            None,                                            # observacao
            None,                                            # captador
            limpar_valor(_cell(row, i_ciclo)),               # trimestre_ingresso
            chave,                                           # chave_preco
            info_preco.get("preco"),                         # preco_balcao
            info_preco.get("area"),                          # area_curso
            info_preco.get("semestres"),                     # semestres
            limpar_valor(_cell(row, i_arquivo)),             # arquivo_origem
        ))
    if skipped:
        log.info("Inscritos filtrados por Empresa: %d excluídos, %d mantidos", skipped, len(dados))
    return dados


def normalizar_matriculados(rows, header=None, tipo="grad"):
    """Normalise matriculados rows using header-based column mapping."""
    if not header:
        log.error("normalizar_matriculados: header obrigatório para mapeamento de colunas")
        return []

    cm = _build_col_map(header)

    i_ciclo     = _get_col(cm, "ciclo")
    i_nome      = _get_col(cm, "nome", "name")
    i_cpf       = _get_col(cm, "cpf")
    i_rgm       = _get_col(cm, "rgm")
    i_rg        = _get_col(cm, "rg")
    i_sexo      = _get_col(cm, "sexo", "sex")
    i_curso     = _get_col(cm, "curso")
    i_instituicao = _get_col(cm, "instituição", "instituicao")
    i_empresa   = _get_col(cm, "empresa")
    i_polo      = _get_col(cm, "polo")
    i_negocio   = _get_col(cm, "negócio", "negocio")
    i_serie     = _get_col(cm, "série", "serie")
    i_data_nasc = _get_col(cm, "data nascimento")
    i_tipo_mat  = _get_col(cm, "tipo matrícula", "tipo matricula")
    i_data_mat  = _get_col(cm, "data matrícula", "data matricula")
    i_situacao  = _get_col(cm, "situação matrícula", "situacao matricula",
                           "situação", "situacao")
    i_fone_res  = _get_col(cm, "fone residencial")
    i_fone_com  = _get_col(cm, "fone comercial")
    i_fone_cel  = _get_col(cm, "fone celular", "celular")
    i_email     = _get_col(cm, "email", "e-mail")
    i_email_ad  = _get_col(cm, "email acadêmico", "email academico")
    i_endereco  = _get_col(cm, "endereço", "endereco")
    i_bairro    = _get_col(cm, "bairro")
    i_cidade    = _get_col(cm, "cidade")
    i_arquivo   = _get_col(cm, "arquivo_origem")

    log.info("Col-map matriculados: nome=%s cpf=%s curso=%s polo=%s sit=%s tipo_mat=%s empresa=%s",
             i_nome, i_cpf, i_curso, i_polo, i_situacao, i_tipo_mat, i_empresa)

    dados = []
    skipped = 0
    for row in rows:
        if i_empresa is not None and not _empresa_permitida(_cell(row, i_empresa)):
            skipped += 1
            continue
        curso_raw    = _cell(row, i_curso)
        situacao_raw = limpar_valor(_cell(row, i_situacao))
        situacao_norm = normalizar_situacao_matriculado(situacao_raw)

        dn = limpar_data_flex(_cell(row, i_data_nasc))
        dm = limpar_data_flex(_cell(row, i_data_mat))

        dados.append((
            tipo,                                            # tipo
            limpar_nome(_cell(row, i_nome)),                 # nome
            limpar_cpf(_cell(row, i_cpf)),                   # cpf
            limpar_rgm(_cell(row, i_rgm)),                   # rgm
            limpar_valor(_cell(row, i_rg)),                  # rg
            limpar_valor(_cell(row, i_sexo)),                # sexo
            str(dn) if dn else None,                         # data_nasc
            limpar_valor(_cell(row, i_instituicao)),         # polo_captador
            limpar_valor(_cell(row, i_negocio)),             # tipo_polo
            limpar_valor(_cell(row, i_polo)),                # polo_aulas
            limpar_valor(curso_raw),                         # curso_raw
            limpar_curso(curso_raw),                         # curso_limpo
            None,                                            # prouni
            limpar_valor(_cell(row, i_serie)),               # serie
            str(dm) if dm else None,                         # data_matricula
            limpar_valor(_cell(row, i_ciclo)),               # ano_tri_ingresso
            limpar_valor(_cell(row, i_tipo_mat)),            # tipo_matricula
            situacao_raw,                                    # situacao_raw
            situacao_norm,                                   # situacao
            limpar_telefone_unico(_cell(row, i_fone_res)),   # fone_res
            limpar_telefone_unico(_cell(row, i_fone_com)),   # fone_com
            limpar_telefone_unico(_cell(row, i_fone_cel)),   # fone_cel
            limpar_valor(_cell(row, i_email)),               # email
            limpar_valor(_cell(row, i_email_ad)),            # email_ad
            limpar_valor(_cell(row, i_endereco)),            # endereco
            limpar_valor(_cell(row, i_bairro)),              # bairro
            limpar_valor(_cell(row, i_cidade)),              # cidade
            limpar_valor(_cell(row, i_arquivo)),             # arquivo_origem
        ))
    if skipped:
        log.info("Matriculados filtrados por Empresa: %d excluídos, %d mantidos", skipped, len(dados))
    return dados


# ════════════════════════════════════════════════════════════════
#  DATABASE — tables, upload, cruzamento
# ════════════════════════════════════════════════════════════════

CREATE_INSCRITOS_SQL = """
CREATE TABLE IF NOT EXISTS mm_inscritos (
    id SERIAL PRIMARY KEY,
    tipo TEXT, status TEXT, dt_pag_insc TEXT, inscricao TEXT,
    nome TEXT, sexo TEXT, cpf TEXT, rg TEXT,
    curso_raw TEXT, curso_limpo TEXT, grau_curso TEXT, modalidade TEXT,
    polo_raw TEXT, polo_normalizado TEXT, marca_instituicao TEXT,
    data_inscr DATE, data_prova DATE,
    telefone TEXT, telefone_res TEXT, telefone_com TEXT,
    email TEXT, cep TEXT, endereco TEXT, bairro TEXT, cidade TEXT, estado TEXT,
    data_pagamento TEXT, data_matricula TEXT,
    situacao_raw TEXT, situacao_final TEXT,
    observacao TEXT, captador TEXT, trimestre_ingresso TEXT,
    chave_preco TEXT, preco_balcao TEXT, area_curso TEXT, semestres TEXT,
    arquivo_origem TEXT, uploaded_at TIMESTAMP DEFAULT NOW()
);
"""

CREATE_MATRICULADOS_SQL = """
CREATE TABLE IF NOT EXISTS mm_matriculados (
    id SERIAL PRIMARY KEY,
    tipo TEXT, nome TEXT, cpf TEXT, rgm TEXT, rg TEXT, sexo TEXT, data_nasc TEXT,
    polo_captador TEXT, tipo_polo TEXT, polo_aulas TEXT,
    curso_raw TEXT, curso_limpo TEXT,
    prouni TEXT, serie TEXT, data_matricula TEXT, ano_tri_ingresso TEXT,
    tipo_matricula TEXT, situacao_raw TEXT, situacao TEXT,
    fone_res TEXT, fone_com TEXT, fone_cel TEXT, email TEXT, email_ad TEXT,
    endereco TEXT, bairro TEXT, cidade TEXT,
    arquivo_origem TEXT, uploaded_at TIMESTAMP DEFAULT NOW()
);
"""

CREATE_CRUZADO_SQL = """
CREATE TABLE IF NOT EXISTS mm_cruzado (
    id SERIAL PRIMARY KEY,
    inscrito_id INTEGER, matriculado_id INTEGER,
    match_tipo TEXT, match_score REAL,
    nome TEXT, cpf TEXT,
    curso_inscrito TEXT, curso_matriculado TEXT, rgm TEXT,
    situacao_inscrito TEXT, situacao_matriculado TEXT,
    data_matricula TEXT, tipo TEXT,
    uploaded_at TIMESTAMP DEFAULT NOW()
);
"""

INDEX_SQL = """
CREATE INDEX IF NOT EXISTS idx_mm_insc_cpf ON mm_inscritos(cpf);
CREATE INDEX IF NOT EXISTS idx_mm_insc_tipo ON mm_inscritos(tipo);
CREATE INDEX IF NOT EXISTS idx_mm_mat_cpf ON mm_matriculados(cpf);
CREATE INDEX IF NOT EXISTS idx_mm_mat_tipo ON mm_matriculados(tipo);
CREATE INDEX IF NOT EXISTS idx_mm_cruz_cpf ON mm_cruzado(cpf);
"""


def db_prepare():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(CREATE_INSCRITOS_SQL)
    cur.execute(CREATE_MATRICULADOS_SQL)
    cur.execute(CREATE_CRUZADO_SQL)
    cur.execute(INDEX_SQL)
    cur.execute("TRUNCATE TABLE mm_cruzado RESTART IDENTITY CASCADE")
    cur.execute("TRUNCATE TABLE mm_inscritos RESTART IDENTITY CASCADE")
    cur.execute("TRUNCATE TABLE mm_matriculados RESTART IDENTITY CASCADE")
    conn.commit()
    cur.close()
    conn.close()
    log.info("Tabelas mm_* criadas e limpas.")


def db_upload_inscritos(dados):
    conn = get_conn()
    cur = conn.cursor()
    sql = """INSERT INTO mm_inscritos (
        tipo, status, dt_pag_insc, inscricao, nome, sexo, cpf, rg,
        curso_raw, curso_limpo, grau_curso, modalidade,
        polo_raw, polo_normalizado, marca_instituicao,
        data_inscr, data_prova, telefone, telefone_res, telefone_com,
        email, cep, endereco, bairro, cidade, estado,
        data_pagamento, data_matricula, situacao_raw, situacao_final,
        observacao, captador, trimestre_ingresso,
        chave_preco, preco_balcao, area_curso, semestres, arquivo_origem
    ) VALUES %s"""
    execute_values(cur, sql, dados, page_size=500)
    conn.commit()
    cur.close()
    conn.close()
    log.info("Upload inscritos: %d registros.", len(dados))
    return len(dados)


def db_upload_matriculados(dados):
    conn = get_conn()
    cur = conn.cursor()
    sql = """INSERT INTO mm_matriculados (
        tipo, nome, cpf, rgm, rg, sexo, data_nasc,
        polo_captador, tipo_polo, polo_aulas,
        curso_raw, curso_limpo, prouni, serie,
        data_matricula, ano_tri_ingresso, tipo_matricula,
        situacao_raw, situacao, fone_res, fone_com, fone_cel,
        email, email_ad, endereco, bairro, cidade, arquivo_origem
    ) VALUES %s"""
    execute_values(cur, sql, dados, page_size=500)
    conn.commit()
    cur.close()
    conn.close()
    log.info("Upload matriculados: %d registros.", len(dados))
    return len(dados)


# ════════════════════════════════════════════════════════════════
#  CRUZAMENTO inscritos × matriculados
# ════════════════════════════════════════════════════════════════

def _similaridade_curso(a, b):
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def cruzar():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, cpf, curso_limpo, nome, situacao_final, tipo
        FROM mm_inscritos
        WHERE UPPER(COALESCE(situacao_final,'')) IN ('MATRICULADO')
          AND cpf IS NOT NULL
    """)
    inscritos = cur.fetchall()
    log.info("Inscritos com situação Matriculado: %d", len(inscritos))

    cur.execute("""
        SELECT id, cpf, curso_limpo, rgm, situacao, data_matricula, nome
        FROM mm_matriculados
        WHERE cpf IS NOT NULL
          AND UPPER(COALESCE(tipo_matricula,'')) IN (
              'INGRESSANTE', 'NOVA MATRICULA', 'NOVA MATRÍCULA',
              'RETORNO', 'RECOMPRA'
          )
    """)
    mat_rows = cur.fetchall()
    log.info("Matriculados ingressantes: %d", len(mat_rows))

    mat_by_cpf = {}
    for row in mat_rows:
        mat_by_cpf.setdefault(row[1], []).append(row)

    cruzados = []
    matched = no_match = 0

    for insc in inscritos:
        insc_id, cpf, curso_limpo, nome, sit_final, tipo = insc
        candidatos = mat_by_cpf.get(cpf, [])

        if not candidatos:
            cruzados.append((insc_id, None, "sem_match", 0.0,
                             nome, cpf, curso_limpo, None, None,
                             sit_final, None, None, tipo))
            no_match += 1
            continue

        if len(candidatos) == 1:
            m = candidatos[0]
            score = _similaridade_curso(curso_limpo, m[2])
            cruzados.append((insc_id, m[0], "cpf_unico", score,
                             nome, cpf, curso_limpo, m[2], m[3],
                             sit_final, m[4], m[5], tipo))
            matched += 1
            continue

        best_score, best_mat = 0.0, None
        for m in candidatos:
            score = _similaridade_curso(curso_limpo, m[2])
            if score > best_score:
                best_score, best_mat = score, m

        if best_mat and best_score >= 0.4:
            cruzados.append((insc_id, best_mat[0], "cpf_curso", best_score,
                             nome, cpf, curso_limpo, best_mat[2], best_mat[3],
                             sit_final, best_mat[4], best_mat[5], tipo))
            matched += 1
        else:
            cruzados.append((insc_id, None, "sem_match_curso", best_score,
                             nome, cpf, curso_limpo, None, None,
                             sit_final, None, None, tipo))
            no_match += 1

    log.info("Cruzamento: %d matched, %d sem match", matched, no_match)

    if cruzados:
        cur.execute("TRUNCATE TABLE mm_cruzado RESTART IDENTITY")
        sql = """INSERT INTO mm_cruzado (
            inscrito_id, matriculado_id, match_tipo, match_score,
            nome, cpf, curso_inscrito, curso_matriculado, rgm,
            situacao_inscrito, situacao_matriculado, data_matricula, tipo
        ) VALUES %s"""
        execute_values(cur, sql, cruzados, page_size=500)
        conn.commit()
        log.info("Upload cruzado: %d registros.", len(cruzados))

    cur.close()
    conn.close()
    return {"matched": matched, "no_match": no_match, "total": len(cruzados)}


# ════════════════════════════════════════════════════════════════
#  MATCH SIAA × KOMMO
# ════════════════════════════════════════════════════════════════

_MM_INSCRITOS_COLS_FOR_MATCH = [
    "id", "nome", "cpf", "telefone", "inscricao", "curso_raw", "curso_limpo",
    "situacao_raw", "situacao_final", "polo_normalizado", "email",
    "data_inscr", "marca_instituicao", "modalidade", "grau_curso",
]

COMPARE_SQL = """
WITH leads_ativos AS (
    SELECT id AS lead_id FROM leads WHERE status_id NOT IN (142, 143)
),
kommo_cpf AS (
    SELECT lcf.lead_id,
           LPAD(regexp_replace(lcf.values_json->0->>'value', '[^0-9]', '', 'g'), 11, '0') AS cpf
    FROM lead_custom_field_values lcf
    JOIN leads_ativos la ON la.lead_id = lcf.lead_id
    WHERE lcf.field_name = 'CPF'
      AND length(regexp_replace(lcf.values_json->0->>'value', '[^0-9]', '', 'g')) >= 11
),
kommo_telefone AS (
    SELECT DISTINCT lcf.lead_id,
           RIGHT(regexp_replace(lcf.values_json->0->>'value', '[^0-9]', '', 'g'), 11) AS telefone
    FROM lead_custom_field_values lcf
    JOIN leads_ativos la ON la.lead_id = lcf.lead_id
    WHERE lcf.field_name IN ('Telefone Comercial', 'Telefone Inscricao')
      AND length(regexp_replace(lcf.values_json->0->>'value', '[^0-9]', '', 'g')) >= 10
),
kommo_situacao AS (
    SELECT lcf.lead_id,
           lcf.values_json->0->>'value' AS situacao_kommo
    FROM lead_custom_field_values lcf
    WHERE lcf.field_name = 'Situação'
      AND lcf.values_json->0->>'value' IS NOT NULL
),
match_cpf AS (
    SELECT DISTINCT ON (s.id)
        s.id AS siaa_id, kc.lead_id, 'cpf' AS match_tipo
    FROM _tmp_mm_inscritos s
    JOIN kommo_cpf kc ON s.cpf IS NOT NULL AND kc.cpf = s.cpf
    ORDER BY s.id, kc.lead_id
),
match_telefone AS (
    SELECT DISTINCT ON (s.id)
        s.id AS siaa_id, kt.lead_id, 'telefone' AS match_tipo
    FROM _tmp_mm_inscritos s
    JOIN kommo_telefone kt ON s.telefone IS NOT NULL
        AND RIGHT(s.telefone, 11) = kt.telefone
    WHERE s.id NOT IN (SELECT siaa_id FROM match_cpf)
    ORDER BY s.id, kt.lead_id
),
all_matches AS (
    SELECT * FROM match_cpf
    UNION ALL SELECT * FROM match_telefone
)
SELECT
    s.id AS siaa_id, s.nome, s.cpf, s.telefone, s.inscricao,
    s.curso_raw, s.curso_limpo, s.situacao_final AS siaa_situacao,
    s.polo_normalizado, s.email,
    s.data_inscr, s.marca_instituicao, s.modalidade, s.grau_curso,
    m.lead_id AS lead_id_match, m.match_tipo,
    ks.situacao_kommo,
    l.name AS lead_name,
    CASE WHEN m.lead_id IS NOT NULL THEN TRUE ELSE FALSE END AS tem_match
FROM _tmp_mm_inscritos s
LEFT JOIN all_matches m ON m.siaa_id = s.id
LEFT JOIN kommo_situacao ks ON ks.lead_id = m.lead_id
LEFT JOIN leads l ON l.id = m.lead_id
ORDER BY s.id;
"""

_TMP_TABLE_DDL = """
CREATE TEMP TABLE _tmp_mm_inscritos (
    id              INTEGER PRIMARY KEY,
    nome            TEXT,
    cpf             TEXT,
    telefone        TEXT,
    inscricao       TEXT,
    curso_raw       TEXT,
    curso_limpo     TEXT,
    situacao_raw    TEXT,
    situacao_final  TEXT,
    polo_normalizado TEXT,
    email           TEXT,
    data_inscr      TEXT,
    marca_instituicao TEXT,
    modalidade      TEXT,
    grau_curso      TEXT
) ON COMMIT DROP;
"""


COMPARE_SQL_MATRICULADOS = """
WITH kommo_rgm AS (
    SELECT lcf.lead_id,
           regexp_replace(lcf.values_json->0->>'value', '[^0-9]', '', 'g') AS rgm
    FROM lead_custom_field_values lcf
    JOIN leads l ON l.id = lcf.lead_id AND l.status_id NOT IN (142, 143)
    WHERE lcf.field_name = 'RGM'
      AND lcf.values_json->0->>'value' IS NOT NULL
      AND lcf.values_json->0->>'value' != ''
),
kommo_situacao AS (
    SELECT lcf.lead_id,
           lcf.values_json->0->>'value' AS situacao_kommo
    FROM lead_custom_field_values lcf
    WHERE lcf.field_name = 'Situação'
      AND lcf.values_json->0->>'value' IS NOT NULL
),
match_rgm AS (
    SELECT DISTINCT ON (s.id)
        s.id AS mat_id, kr.lead_id, 'rgm' AS match_tipo
    FROM _tmp_mm_matriculados s
    JOIN kommo_rgm kr ON s.rgm IS NOT NULL AND kr.rgm = s.rgm
    ORDER BY s.id, kr.lead_id
)
SELECT
    s.id AS mat_id, s.nome, s.cpf, s.rgm,
    s.curso_raw, s.curso_limpo,
    s.situacao AS mat_situacao,
    s.polo_aulas, s.data_matricula, s.tipo_matricula,
    m.lead_id AS lead_id_match, m.match_tipo,
    ks.situacao_kommo,
    l.name AS lead_name, l.status_id,
    CASE WHEN m.lead_id IS NOT NULL THEN TRUE ELSE FALSE END AS tem_match
FROM _tmp_mm_matriculados s
LEFT JOIN match_rgm m ON m.mat_id = s.id
LEFT JOIN kommo_situacao ks ON ks.lead_id = m.lead_id
LEFT JOIN leads l ON l.id = m.lead_id
ORDER BY s.id;
"""

_TMP_TABLE_MATRICULADOS_DDL = """
CREATE TEMP TABLE _tmp_mm_matriculados (
    id              INTEGER PRIMARY KEY,
    nome            TEXT,
    cpf             TEXT,
    rgm             TEXT,
    curso_raw       TEXT,
    curso_limpo     TEXT,
    situacao_raw    TEXT,
    situacao        TEXT,
    polo_aulas      TEXT,
    data_matricula  TEXT,
    tipo_matricula  TEXT
) ON COMMIT DROP;
"""

_MM_MATRICULADOS_COLS_FOR_MATCH = [
    "id", "nome", "cpf", "rgm", "curso_raw", "curso_limpo",
    "situacao_raw", "situacao", "polo_aulas", "data_matricula", "tipo_matricula",
]


def match_kommo():
    """Compare mm_inscritos (dcz_sync) with Kommo leads (kommo_sync)."""
    dcz = get_conn()
    dcz_cur = dcz.cursor()
    cols_sql = ", ".join(_MM_INSCRITOS_COLS_FOR_MATCH)
    dcz_cur.execute(f"SELECT {cols_sql} FROM mm_inscritos")
    inscritos_rows = dcz_cur.fetchall()
    dcz_cur.close()
    dcz.close()

    if not inscritos_rows:
        return {"total": 0, "com_match": 0, "sem_match": 0,
                "tipos": {}, "divergencias": {}, "detalhes": []}

    kommo = get_kommo_conn()
    kommo.autocommit = False
    kcur = kommo.cursor()
    kcur.execute(_TMP_TABLE_DDL)

    placeholders = ", ".join(["%s"] * len(_MM_INSCRITOS_COLS_FOR_MATCH))
    insert_sql = f"INSERT INTO _tmp_mm_inscritos ({cols_sql}) VALUES ({placeholders})"
    execute_values(
        kcur,
        f"INSERT INTO _tmp_mm_inscritos ({cols_sql}) VALUES %s",
        inscritos_rows,
    )

    kcur.execute(COMPARE_SQL)
    cols = [desc[0] for desc in kcur.description]
    rows = kcur.fetchall()
    kommo.rollback()
    kcur.close()
    kommo.close()

    total = len(rows)
    com_match = sum(1 for r in rows if r[cols.index("tem_match")])
    sem_match = total - com_match

    tipos = {}
    for r in rows:
        tipo = r[cols.index("match_tipo")]
        if tipo:
            tipos[tipo] = tipos.get(tipo, 0) + 1

    divergencias = {"atualizar_matriculado": 0, "atualizar_aprovado": 0,
                    "ok": 0, "sem_situacao_kommo": 0}
    for r in rows:
        if not r[cols.index("tem_match")]:
            continue
        siaa_sit = r[cols.index("siaa_situacao")]
        kommo_sit = r[cols.index("situacao_kommo")]
        if siaa_sit == "Matriculado" and kommo_sit != "Matriculado":
            divergencias["atualizar_matriculado"] += 1
        elif siaa_sit == "Aprovado" and kommo_sit not in ("Aprovado", "Matriculado"):
            divergencias["atualizar_aprovado"] += 1
        elif kommo_sit is None:
            divergencias["sem_situacao_kommo"] += 1
        else:
            divergencias["ok"] += 1

    log.info("Match Inscritos x Kommo: total=%d, match=%d, sem=%d", total, com_match, sem_match)
    log.info("  Tipos: %s", tipos)
    log.info("  Divergencias: %s", divergencias)

    detalhes = [dict(zip(cols, r)) for r in rows]

    return {
        "total": total, "com_match": com_match, "sem_match": sem_match,
        "tipos": tipos, "divergencias": divergencias, "detalhes": detalhes,
    }


def match_matriculados_kommo():
    """Compare mm_matriculados (dcz_sync) with Kommo leads by RGM (kommo_sync)."""
    dcz = get_conn()
    dcz_cur = dcz.cursor()
    cols_sql = ", ".join(_MM_MATRICULADOS_COLS_FOR_MATCH)
    dcz_cur.execute(f"""
        SELECT {cols_sql} FROM mm_matriculados
        WHERE rgm IS NOT NULL AND rgm != ''
          AND UPPER(COALESCE(situacao,'')) = 'MATRICULADO'
    """)
    mat_rows = dcz_cur.fetchall()
    dcz_cur.close()
    dcz.close()

    if not mat_rows:
        return {"total": 0, "com_match": 0, "sem_match": 0,
                "tipos": {}, "detalhes": []}

    kommo = get_kommo_conn()
    kommo.autocommit = False
    kcur = kommo.cursor()
    kcur.execute(_TMP_TABLE_MATRICULADOS_DDL)

    execute_values(
        kcur,
        f"INSERT INTO _tmp_mm_matriculados ({cols_sql}) VALUES %s",
        mat_rows,
    )

    kcur.execute(COMPARE_SQL_MATRICULADOS)
    cols = [desc[0] for desc in kcur.description]
    rows = kcur.fetchall()
    kommo.rollback()
    kcur.close()
    kommo.close()

    total = len(rows)
    com_match = sum(1 for r in rows if r[cols.index("tem_match")])
    sem_match = total - com_match

    tipos = {}
    for r in rows:
        tipo = r[cols.index("match_tipo")]
        if tipo:
            tipos[tipo] = tipos.get(tipo, 0) + 1

    log.info("Match Matriculados x Kommo (RGM): total=%d, match=%d, sem=%d",
             total, com_match, sem_match)

    detalhes = [dict(zip(cols, r)) for r in rows]

    return {
        "total": total, "com_match": com_match, "sem_match": sem_match,
        "tipos": tipos, "detalhes": detalhes,
    }


# ════════════════════════════════════════════════════════════════
#  ACTION GENERATION
# ════════════════════════════════════════════════════════════════

def gerar_acoes(inscritos_match, matriculados_match=None):
    """Generate actions from inscritos + matriculados match results.

    Action types:
      NOVO       - inscrito sem match no CRM -> criar lead
      ATUALIZAR  - inscrito com match, lead precisa dados da inscricao
      MATRICULADO - matriculado com match por RGM -> mover para ganho
    """
    acoes = []

    for row in inscritos_match.get("detalhes", []):
        lead_id = row.get("lead_id_match")
        siaa_sit = row.get("siaa_situacao")
        kommo_sit = row.get("situacao_kommo")

        base = {
            "siaa_id": row["siaa_id"],
            "nome": row["nome"],
            "cpf": row["cpf"],
            "curso_siaa": row.get("curso_limpo"),
            "polo": row.get("polo_normalizado"),
            "situacao_siaa": siaa_sit,
            "situacao_kommo": kommo_sit,
            "match_tipo": row.get("match_tipo"),
            "telefone": row.get("telefone"),
            "email": row.get("email"),
            "marca": row.get("marca_instituicao"),
            "inscricao": row.get("inscricao"),
            "modalidade": row.get("modalidade"),
            "grau": row.get("grau_curso"),
            "data_inscr": str(row.get("data_inscr") or ""),
        }

        if lead_id:
            needs_update = (
                kommo_sit is None
                or kommo_sit == ""
                or kommo_sit != siaa_sit
            )
            if needs_update:
                acoes.append({**base, "acao": "ATUALIZAR", "lead_id": lead_id})
        else:
            acoes.append({**base, "acao": "NOVO", "lead_id": None})

    if matriculados_match:
        for row in matriculados_match.get("detalhes", []):
            lead_id = row.get("lead_id_match")
            if not lead_id:
                continue
            kommo_sit = row.get("situacao_kommo")
            if kommo_sit == "Matriculado":
                continue
            acoes.append({
                "acao": "MATRICULADO",
                "lead_id": lead_id,
                "siaa_id": row.get("mat_id"),
                "nome": row.get("nome"),
                "cpf": row.get("cpf"),
                "rgm": row.get("rgm"),
                "curso_siaa": row.get("curso_limpo"),
                "polo": row.get("polo_aulas"),
                "situacao_siaa": row.get("mat_situacao"),
                "situacao_kommo": kommo_sit,
                "match_tipo": row.get("match_tipo"),
                "data_matricula": str(row.get("data_matricula") or ""),
                "tipo_matricula": row.get("tipo_matricula"),
            })

    n_novo = sum(1 for a in acoes if a["acao"] == "NOVO")
    n_atualizar = sum(1 for a in acoes if a["acao"] == "ATUALIZAR")
    n_matriculado = sum(1 for a in acoes if a["acao"] == "MATRICULADO")
    log.info("Acoes: %d total (NOVO=%d, ATUALIZAR=%d, MATRICULADO=%d)",
             len(acoes), n_novo, n_atualizar, n_matriculado)
    return acoes


# ════════════════════════════════════════════════════════════════
#  KOMMO API CLIENT (for updates)
# ════════════════════════════════════════════════════════════════

class KommoApiClient:
    """Minimal Kommo API v4 client for PATCH operations."""

    def __init__(self, base_url=None, token=None, rate_per_sec=5):
        self.base_url = (base_url or KOMMO_BASE_URL).rstrip("/")
        self.token = token or KOMMO_TOKEN
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json",
        })
        self._min_interval = 1.0 / rate_per_sec
        self._last_req = 0.0
        self.total_calls = 0
        self.errors = 0

    def _throttle(self):
        elapsed = time.monotonic() - self._last_req
        if elapsed < self._min_interval:
            time.sleep(self._min_interval - elapsed)

    def _request(self, method, path, payload=None):
        url = f"{self.base_url}{path}"
        for attempt in range(4):
            self._throttle()
            self._last_req = time.monotonic()
            self.total_calls += 1
            try:
                r = self.session.request(method, url, json=payload, timeout=30)
            except (requests.exceptions.ReadTimeout,
                    requests.exceptions.ConnectionError) as exc:
                wait = min(5 * (2 ** attempt), 60)
                log.warning("Timeout/conexão (tentativa %d/4): %s", attempt + 1, str(exc)[:120])
                time.sleep(wait)
                continue

            if r.status_code == 429:
                retry = int(r.headers.get("Retry-After", 5))
                log.warning("429 — Retry-After %ds (tentativa %d/4)", retry, attempt + 1)
                time.sleep(retry + 1)
                continue

            if r.status_code >= 400:
                self.errors += 1
                return {"ok": False, "status": r.status_code, "body": r.text[:500]}

            body = r.json() if r.text.strip() else {}
            return {"ok": True, "status": r.status_code, "body": body}

        self.errors += 1
        return {"ok": False, "status": 0, "body": "Falha após 4 tentativas"}

    def patch_lead(self, lead_id, payload):
        return self._request("PATCH", f"/api/v4/leads/{lead_id}", payload)

    def create_lead(self, payload):
        """POST /api/v4/leads — create one or more leads."""
        return self._request("POST", "/api/v4/leads", payload)

    def get_custom_field_ids(self, field_names):
        """Lookup numeric field_id for given field names from the kommo_sync database."""
        conn = get_kommo_conn()
        cur = conn.cursor()
        placeholders = ",".join(["%s"] * len(field_names))
        cur.execute(f"""
            SELECT DISTINCT field_name, field_id
            FROM lead_custom_field_values
            WHERE field_name IN ({placeholders})
        """, field_names)
        result = {row[0]: row[1] for row in cur.fetchall()}
        cur.close()
        conn.close()
        return result

    def get_pipeline_stages(self):
        """Load pipeline stages from kommo_sync database."""
        conn = get_kommo_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, pipeline_id, name
            FROM pipeline_statuses
            ORDER BY pipeline_id, sort NULLS LAST
        """)
        stages = cur.fetchall()
        cur.close()
        conn.close()
        return stages


# ════════════════════════════════════════════════════════════════
#  EXECUTE ACTIONS
# ════════════════════════════════════════════════════════════════

_FIELD_NAMES_FOR_UPDATE = [
    "Situação", "Curso_SIAA", "Modalidade_SIAA", "Grau_SIAA",
    "Polo", "Marca", "CPF", "Inscrição", "Telefone Inscricao",
]

_FIELD_NAMES_FOR_MATRICULA = [
    "Situação", "RGM", "Matrícula",
]


def _find_stage_id(stages, name_fragment):
    """Find a pipeline stage by name fragment (case-insensitive)."""
    frag = name_fragment.lower()
    for s in stages:
        if frag in (s.get("name") or s.get("nome") or "").lower():
            return s["id"], s["pipeline_id"]
    return None, None


def _build_custom_fields(field_ids, acao, fields_map):
    """Build custom_fields_values array for Kommo API payload."""
    cf = []
    for field_name, acao_key in fields_map.items():
        fid = field_ids.get(field_name)
        val = acao.get(acao_key)
        if fid and val:
            cf.append({"field_id": fid, "values": [{"value": str(val)}]})
    return cf


def executar_acoes(acoes, limit=None, log_callback=None):
    """Execute Kommo updates/creates for the given actions."""
    api = KommoApiClient()

    all_fields = list(set(_FIELD_NAMES_FOR_UPDATE + _FIELD_NAMES_FOR_MATRICULA))
    field_ids = api.get_custom_field_ids(all_fields)
    stages = api.get_pipeline_stages()

    aprovado_stage, aprovado_pipe = _find_stage_id(stages, "aprovad")
    matriculado_stage, matriculado_pipe = _find_stage_id(stages, "venda ganha")
    if not matriculado_stage:
        matriculado_stage, matriculado_pipe = _find_stage_id(stages, "matriculado")

    log.info("Stages: Aprovado=%s (pipe=%s), Matriculado=%s (pipe=%s)",
             aprovado_stage, aprovado_pipe, matriculado_stage, matriculado_pipe)
    log.info("Fields: %s", field_ids)

    to_process = acoes[:limit] if limit else acoes
    results = {"ok": 0, "erro": 0, "skip": 0, "novo_ok": 0}

    update_fields_map = {
        "Situação": "situacao_siaa",
        "Curso_SIAA": "curso_siaa",
        "Modalidade_SIAA": "modalidade",
        "Grau_SIAA": "grau",
        "Polo": "polo",
        "Marca": "marca",
        "CPF": "cpf",
        "Inscrição": "inscricao",
        "Telefone Inscricao": "telefone",
    }

    for i, acao in enumerate(to_process):
        tipo = acao["acao"]
        lead_id = acao.get("lead_id")

        if tipo == "ATUALIZAR" and lead_id:
            cf = _build_custom_fields(field_ids, acao, update_fields_map)
            payload = {}
            if cf:
                payload["custom_fields_values"] = cf
            if aprovado_stage and acao.get("situacao_siaa") == "Aprovado":
                payload["pipeline_id"] = aprovado_pipe
                payload["status_id"] = aprovado_stage
            if not payload:
                results["skip"] += 1
                continue
            resp = api.patch_lead(lead_id, payload)

        elif tipo == "MATRICULADO" and lead_id:
            cf = _build_custom_fields(field_ids, acao, {
                "Situação": "situacao_siaa",
                "RGM": "rgm",
            })
            payload = {}
            if cf:
                payload["custom_fields_values"] = cf
            else:
                payload["custom_fields_values"] = [{
                    "field_id": field_ids.get("Situação"),
                    "values": [{"value": "Matriculado"}],
                }]
            if matriculado_stage:
                payload["pipeline_id"] = matriculado_pipe
                payload["status_id"] = matriculado_stage
            resp = api.patch_lead(lead_id, payload)

        elif tipo == "NOVO":
            cf = _build_custom_fields(field_ids, acao, update_fields_map)
            nome = acao.get("nome") or "Lead SIAA"
            lead_payload = [{"name": nome}]
            if aprovado_stage:
                lead_payload[0]["pipeline_id"] = aprovado_pipe
                lead_payload[0]["status_id"] = aprovado_stage
            if cf:
                lead_payload[0]["custom_fields_values"] = cf
            resp = api.create_lead(lead_payload)
            if resp["ok"]:
                results["novo_ok"] += 1

        else:
            results["skip"] += 1
            continue

        if resp["ok"]:
            results["ok"] += 1
            msg = f"[{i+1}/{len(to_process)}] OK {tipo} lead={lead_id or 'NOVO'} {acao.get('nome','')}"
        else:
            results["erro"] += 1
            msg = f"[{i+1}/{len(to_process)}] ERRO {tipo} lead={lead_id or 'NOVO'}: {resp.get('body','')[:100]}"

        log.info(msg)
        if log_callback:
            log_callback(msg)

    log.info("Execucao: ok=%d, erro=%d, skip=%d, novos=%d, API calls=%d",
             results["ok"], results["erro"], results["skip"],
             results["novo_ok"], api.total_calls)
    return results


# ════════════════════════════════════════════════════════════════
#  FULL PIPELINE
# ════════════════════════════════════════════════════════════════

def run_pipeline(candidatos_files, matriculados_files, nivel="grad", log_callback=None):
    """
    Full pipeline: merge -> normalize -> db upload -> cruzamento -> match Kommo.
    Returns result dict with stats and actions preview.
    """
    def _log(msg):
        log.info(msg)
        if log_callback:
            log_callback(msg)

    start = datetime.now(BRT)
    _log(f"Pipeline iniciado: {start.strftime('%H:%M:%S')}")
    _log(f"Nível: {nivel} | Candidatos: {len(candidatos_files)} | Matriculados: {len(matriculados_files)}")

    # 1. Merge
    _log(">>> ETAPA 1: MERGE")
    insc_rows = merge_uploaded_files(candidatos_files, f"inscritos-{nivel}")
    mat_rows = merge_uploaded_files(matriculados_files, f"matriculados-{nivel}")

    insc_header = insc_rows[0] if insc_rows else []
    mat_header = mat_rows[0] if mat_rows else []
    insc_data = insc_rows[1:] if len(insc_rows) > 1 else []
    mat_data = mat_rows[1:] if len(mat_rows) > 1 else []

    _log(f"  Inscritos: {len(insc_data)} linhas (cols: {len(insc_header)})")
    _log(f"  Matriculados: {len(mat_data)} linhas (cols: {len(mat_header)})")

    if not insc_data and not mat_data:
        _log("ERRO: Nenhum dado para processar.")
        return {"error": "Nenhum dado encontrado nos arquivos."}

    # 2. Normalize
    _log(">>> ETAPA 2: NORMALIZAÇÃO")
    tipo = nivel  # "grad" or "pos"
    inscritos_norm = normalizar_inscritos(insc_data, header=insc_header, tipo=tipo) if insc_data else []
    matriculados_norm = normalizar_matriculados(mat_data, header=mat_header, tipo=tipo) if mat_data else []
    _log(f"  Inscritos normalizados: {len(inscritos_norm)}")
    _log(f"  Matriculados normalizados: {len(matriculados_norm)}")

    # 3. DB upload
    _log(">>> ETAPA 3: UPLOAD BANCO")
    db_prepare()
    n_insc = db_upload_inscritos(inscritos_norm) if inscritos_norm else 0
    n_mat = db_upload_matriculados(matriculados_norm) if matriculados_norm else 0
    _log(f"  Banco: {n_insc} inscritos + {n_mat} matriculados")

    # 4. Cruzamento inscritos x matriculados
    _log(">>> ETAPA 4: CRUZAMENTO INSCRITOS x MATRICULADOS")
    cruz_result = cruzar()
    _log(f"  Cruzados: {cruz_result['total']} (match={cruz_result['matched']}, sem={cruz_result['no_match']})")

    # 5. Match Inscritos x Kommo (CPF + Telefone)
    _log(">>> ETAPA 5: MATCH INSCRITOS x KOMMO (CPF + Telefone)")
    inscritos_match = match_kommo()
    _log(f"  Total: {inscritos_match['total']} | Match: {inscritos_match['com_match']} | Sem: {inscritos_match['sem_match']}")
    _log(f"  Tipos: {inscritos_match['tipos']}")

    # 6. Match Matriculados x Kommo (RGM)
    _log(">>> ETAPA 6: MATCH MATRICULADOS x KOMMO (RGM)")
    matriculados_match = match_matriculados_kommo()
    _log(f"  Total: {matriculados_match['total']} | Match: {matriculados_match['com_match']} | Sem: {matriculados_match['sem_match']}")

    # 7. Generate actions
    _log(">>> ETAPA 7: GERAR ACOES")
    acoes = gerar_acoes(inscritos_match, matriculados_match)

    n_novo = sum(1 for a in acoes if a["acao"] == "NOVO")
    n_atualizar = sum(1 for a in acoes if a["acao"] == "ATUALIZAR")
    n_matriculado = sum(1 for a in acoes if a["acao"] == "MATRICULADO")

    elapsed = (datetime.now(BRT) - start).total_seconds()
    _log(f"Pipeline concluido em {elapsed:.0f}s.")
    _log(f"  NOVO={n_novo} | ATUALIZAR={n_atualizar} | MATRICULADO={n_matriculado} | Total={len(acoes)}")

    return {
        "inscritos": n_insc,
        "matriculados": n_mat,
        "cruzamento": cruz_result,
        "match": {
            "total": inscritos_match["total"],
            "com_match": inscritos_match["com_match"],
            "sem_match": inscritos_match["sem_match"],
            "tipos": inscritos_match["tipos"],
            "divergencias": inscritos_match.get("divergencias", {}),
        },
        "match_matriculados": {
            "total": matriculados_match["total"],
            "com_match": matriculados_match["com_match"],
            "sem_match": matriculados_match["sem_match"],
        },
        "acoes": acoes,
        "elapsed": elapsed,
    }
