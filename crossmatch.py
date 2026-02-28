"""
eduit. — Cruzamento Planilha Matriculados x CRM.

Modo DRY-RUN por padrão: apenas gera relatórios em c:\\DCz\\reports\\
Nenhuma alteração é feita no CRM sem aprovação explícita.

Uso:
    python crossmatch.py                  # gera todos os relatórios
    python crossmatch.py --resumo         # apenas resumo rápido
"""

import sys
import io
import os
import csv
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from collections import Counter

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

load_dotenv(Path(__file__).parent / ".env")

REPORTS_DIR = Path(__file__).parent / "reports"

def _find_xlsx():
    d = Path(__file__).parent
    for f in d.iterdir():
        if f.suffix.lower() == ".xlsx" and "matriculados" in f.name.lower():
            return f
    raise FileNotFoundError("Planilha de matriculados não encontrada em " + str(d))

EXCEL_PATH = _find_xlsx()

DB_DSN = dict(
    host=os.getenv("DB_HOST", "localhost"),
    port=os.getenv("DB_PORT", "5432"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASS"),
    dbname=os.getenv("DB_NAME", "dcz_sync"),
)

# IDs dos campos personalizados dos negócios
FIELD_IDS = {
    "RGM":              "2ac4e30f-cfd7-435f-b688-fbce27f76c38",
    "Curso":            "4bddb764-658b-48bc-9d70-6e94ad420132",
    "Polo":             "0ec9d8dc-d547-4482-b9ad-d4a3e6ec1b54",
    "Serie":            "b921a702-8e51-4b6c-b4d8-cdea931ea51d",
    "Situacao":         "fd08d44b-a4a5-4343-b7a9-37f75e2c1caa",
    "Data Matricula":   "bf93a8e9-42c0-4517-8518-6f604746a300",
    "Modalidade":       "9c8fc723-d9f7-4074-a0bc-ca4b96d36739",
    "Bairro":           "f7cf5892-573f-45b8-9425-6dafab92cc2c",
    "Cidade":           "7a4407e4-7345-4f7e-8a24-4f51d4a10cf8",
    "Email AD":         "731bd2fd-7cfa-49af-ab24-2e55e0374798",
    "Senha Provisoria": "cccb3046-1906-4465-901d-329ef2fe08dc",
    "Tipo Aluno":       "4230e4db-970b-4444-abaf-c3135a03b79c",
    "Sexo":             "d77aa3c7-cd39-46d2-9c12-7be48b86eb2f",
    "Turma":            "8815a8de-f755-4597-b6f4-8da6d289b6eb",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("crossmatch")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_conn():
    return psycopg2.connect(**DB_DSN)


def _col_find(col_map, *candidates):
    """Busca coluna por nome exato ou parcial (para lidar com encoding)."""
    for c in candidates:
        if c in col_map:
            return c
    for c in candidates:
        cl = c.lower()
        for k in col_map:
            if k and cl in k.lower():
                return k
    return candidates[0]


def clean_cpf(cpf):
    if not cpf:
        return ""
    return str(cpf).replace(".", "").replace("-", "").replace(" ", "").strip()


def clean_phone(phone):
    if not phone:
        return ""
    digits = "".join(c for c in str(phone) if c.isdigit())
    if digits.startswith("55") and len(digits) >= 12:
        digits = digits[2:]
    if len(digits) >= 10:
        return digits[-10:] if len(digits) == 10 else digits[-11:]
    return ""


def normalize_name(name):
    if not name:
        return ""
    import unicodedata
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    return " ".join(name.upper().split())


def get_biz_field(biz_data, field_id):
    """Extrai o valor de um campo personalizado do negócio."""
    for f in biz_data.get("additionalFields", []):
        af = f.get("additionalField", {})
        if isinstance(af, dict) and af.get("id") == field_id:
            return f.get("value", "")
        if isinstance(af, str) and af == field_id:
            return f.get("value", "")
    return ""


def write_csv(filename, rows, headers):
    path = REPORTS_DIR / filename
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(headers)
        for r in rows:
            w.writerow(r)
    log.info("  Relatório: %s (%d linhas)", filename, len(rows))
    return path


# ---------------------------------------------------------------------------
# Carregamento
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    "Ciclo":              ["Ciclo"],
    "Nome":               ["Nome"],
    "CPF":                ["CPF"],
    "RGM":                ["RGM"],
    "Sexo":               ["Sexo"],
    "Curso":              ["Curso"],
    "Instituicao":        ["Instituição", "Institui"],
    "Empresa":            ["Empresa"],
    "Polo":               ["Polo"],
    "Negocio":            ["Negócio", "Neg"],
    "Serie":              ["Série", "Serie", "rie"],
    "DataNascimento":     ["Data Nascimento"],
    "TipoMatricula":      ["Tipo Matrícula", "Tipo Matr"],
    "DataMatricula":      ["Data Matrícula", "Data Matr"],
    "SituacaoMatricula":  ["Situação Matrícula", "Situa"],
    "FoneResidencial":    ["Fone Residencial"],
    "FoneComercial":      ["Fone Comercial"],
    "FoneCelular":        ["Fone celular"],
    "Email":              ["Email"],
    "EmailAcademico":     ["Email acadêmico", "Email acad"],
    "Endereco":           ["Endereço", "Endere"],
    "Bairro":             ["Bairro"],
    "Cidade":             ["Cidade"],
}


def load_excel():
    log.info("Carregando planilha: %s", EXCEL_PATH.name)
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb["Export"]
    raw_header = [cell.value for cell in ws[1]]
    raw_col = {h: i for i, h in enumerate(raw_header) if h}

    col = {}
    for norm_name, aliases in COLUMN_ALIASES.items():
        found = _col_find(raw_col, *aliases)
        if found in raw_col:
            col[norm_name] = raw_col[found]
        else:
            log.warning("  Coluna não encontrada: %s", norm_name)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[1] is None:
            continue
        rows.append(row)
    wb.close()
    log.info("  %d registros carregados, %d colunas mapeadas", len(rows), len(col))
    return rows, col


def load_crm_businesses(conn):
    log.info("Carregando negócios do CRM com RGM...")
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT b.id, b.data
        FROM businesses b
        WHERE EXISTS (
            SELECT 1 FROM jsonb_array_elements(b.data->'additionalFields') elem
            WHERE elem->'additionalField'->>'id' = %s
              AND elem->>'value' IS NOT NULL AND elem->>'value' != ''
        )
    """, (FIELD_IDS["RGM"],))
    result = cur.fetchall()
    cur.close()

    by_rgm = {}
    for row in result:
        rgm = get_biz_field(row["data"], FIELD_IDS["RGM"])
        if rgm:
            by_rgm.setdefault(rgm, []).append(row)
    log.info("  %d negócios com RGM, %d RGMs únicos", len(result), len(by_rgm))
    return by_rgm


def load_crm_leads(conn):
    log.info("Carregando todos os leads do CRM...")
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT id,
               data->>'name' AS nome,
               REPLACE(REPLACE(COALESCE(data->>'taxId',''), '.', ''), '-', '') AS cpf,
               data->>'rawPhone' AS telefone,
               data->>'email' AS email
        FROM leads
    """)
    result = cur.fetchall()
    cur.close()

    by_cpf = {}
    by_phone = {}
    by_name = {}

    for r in result:
        cpf = r["cpf"].strip() if r["cpf"] else ""
        if cpf:
            by_cpf.setdefault(cpf, []).append(r)

        phone = clean_phone(r["telefone"])
        if phone:
            by_phone.setdefault(phone, []).append(r)

        nome = normalize_name(r["nome"])
        if nome:
            by_name.setdefault(nome, []).append(r)

    log.info("  %d leads total | %d com CPF | %d com telefone | %d nomes únicos",
             len(result), len(by_cpf), len(by_phone), len(by_name))
    return by_cpf, by_phone, by_name


# ---------------------------------------------------------------------------
# AÇÃO 5: Identificar RGMs duplicados errados
# ---------------------------------------------------------------------------

def report_rgm_duplicados(crm_by_rgm):
    log.info("=== AÇÃO 5: RGMs duplicados em leads diferentes ===")
    rows_csv = []

    for rgm, bizzes in crm_by_rgm.items():
        lead_ids = set()
        for b in bizzes:
            lid = b["data"].get("leadId", "")
            lead_ids.add(lid)

        if len(lead_ids) > 1:
            for b in bizzes:
                lead_nome = ""
                lead_obj = b["data"].get("lead")
                if isinstance(lead_obj, dict):
                    lead_nome = lead_obj.get("name", "")
                rows_csv.append([
                    rgm,
                    b["id"],
                    b["data"].get("leadId", ""),
                    lead_nome,
                    b["data"].get("status", ""),
                    get_biz_field(b["data"], FIELD_IDS["Curso"]),
                    len(lead_ids),
                ])

    write_csv(
        "acao5_rgm_duplicados.csv", rows_csv,
        ["RGM", "negocio_id", "lead_id", "lead_nome", "status", "curso", "total_leads_diferentes"]
    )

    rgm_unicos = set(r[0] for r in rows_csv)
    log.info("  %d RGMs em %d negócios de leads diferentes", len(rgm_unicos), len(rows_csv))
    return rgm_unicos


# ---------------------------------------------------------------------------
# AÇÃO 1: Match por RGM — preparar atualizações
# ---------------------------------------------------------------------------

def report_match_rgm(xl_rows, col, crm_by_rgm):
    log.info("=== AÇÃO 1: Match por RGM ===")
    matched = []
    status_cruzado = []

    for r in xl_rows:
        rgm = str(r[col["RGM"]]).strip() if r[col["RGM"]] else ""
        if not rgm or rgm not in crm_by_rgm:
            continue

        xl_sit = r[col["SituacaoMatricula"]] or ""
        xl_curso = r[col["Curso"]] or ""
        xl_polo = r[col["Polo"]] or ""
        xl_serie = str(r[col["Serie"]] or "")
        xl_negocio = r[col["Negocio"]] or ""
        xl_bairro = r[col["Bairro"]] or ""
        xl_cidade = r[col["Cidade"]] or ""
        xl_tipo = r[col["TipoMatricula"]] or ""
        xl_data_mat = ""
        if r[col["DataMatricula"]]:
            dm = r[col["DataMatricula"]]
            xl_data_mat = dm.strftime("%Y-%m-%d") if hasattr(dm, "strftime") else str(dm)

        for biz in crm_by_rgm[rgm]:
            crm_status = biz["data"].get("status", "")
            crm_curso = get_biz_field(biz["data"], FIELD_IDS["Curso"])
            crm_polo = get_biz_field(biz["data"], FIELD_IDS["Polo"])
            crm_serie = get_biz_field(biz["data"], FIELD_IDS["Serie"])
            crm_sit = get_biz_field(biz["data"], FIELD_IDS["Situacao"])
            crm_bairro = get_biz_field(biz["data"], FIELD_IDS["Bairro"])
            crm_cidade = get_biz_field(biz["data"], FIELD_IDS["Cidade"])

            mudancas = []
            if xl_curso and xl_curso != crm_curso:
                mudancas.append(f"Curso: '{crm_curso}' → '{xl_curso}'")
            if xl_polo and xl_polo != crm_polo:
                mudancas.append(f"Polo: '{crm_polo}' → '{xl_polo}'")
            if xl_serie and xl_serie != crm_serie:
                mudancas.append(f"Serie: '{crm_serie}' → '{xl_serie}'")
            if xl_sit and xl_sit != crm_sit:
                mudancas.append(f"Situacao: '{crm_sit}' → '{xl_sit}'")
            if xl_bairro and xl_bairro != crm_bairro:
                mudancas.append(f"Bairro: '{crm_bairro}' → '{xl_bairro}'")
            if xl_cidade and xl_cidade != crm_cidade:
                mudancas.append(f"Cidade: '{crm_cidade}' → '{xl_cidade}'")

            lead_nome = ""
            lead_obj = biz["data"].get("lead")
            if isinstance(lead_obj, dict):
                lead_nome = lead_obj.get("name", "")

            alerta = ""
            if xl_sit == "CANCELADO" and crm_status == "in_process":
                alerta = "CANCELADO na planilha, ABERTO no CRM"
            elif xl_sit == "EM CURSO" and crm_status == "lost":
                alerta = "EM CURSO na planilha, PERDIDO no CRM"
            elif xl_sit == "TRANCADO" and crm_status == "in_process":
                alerta = "TRANCADO na planilha, ABERTO no CRM"

            matched.append([
                rgm, biz["id"], biz["data"].get("leadId", ""), lead_nome,
                crm_status, xl_sit, xl_curso, xl_polo, xl_serie,
                xl_data_mat, xl_negocio, xl_tipo,
                " | ".join(mudancas) if mudancas else "(sem mudanças)",
                alerta,
            ])

            if alerta:
                status_cruzado.append([
                    rgm, biz["id"], lead_nome, crm_status, xl_sit, alerta,
                    xl_curso, xl_polo,
                ])

    write_csv(
        "acao1_match_rgm.csv", matched,
        ["RGM", "negocio_id", "lead_id", "lead_nome",
         "crm_status", "planilha_situacao", "planilha_curso", "planilha_polo",
         "planilha_serie", "planilha_data_mat", "planilha_nivel", "planilha_tipo",
         "mudancas_detectadas", "ALERTA_STATUS"]
    )

    write_csv(
        "acao1_alertas_status.csv", status_cruzado,
        ["RGM", "negocio_id", "lead_nome", "crm_status", "planilha_situacao",
         "alerta", "curso", "polo"]
    )

    com_mudanca = sum(1 for r in matched if r[12] != "(sem mudanças)")
    log.info("  %d matches, %d com mudanças, %d alertas de status",
             len(matched), com_mudanca, len(status_cruzado))
    return set(str(r[col["RGM"]]).strip() for r in xl_rows if r[col["RGM"]] and str(r[col["RGM"]]).strip() in crm_by_rgm)


# ---------------------------------------------------------------------------
# AÇÃO 2: Match cascata — CPF → Telefone → Nome (para quem não tem RGM)
# ---------------------------------------------------------------------------

def _xl_phone(r, col):
    raw = str(r[col["FoneCelular"]] or "")
    return clean_phone(raw)


def _xl_name(r, col):
    return normalize_name(r[col["Nome"]] or "")


def report_match_sem_rgm(xl_rows, col, crm_by_rgm, crm_by_cpf, crm_by_phone, crm_by_name):
    log.info("=== AÇÃO 2: Match cascata (CPF → Telefone → Nome) ===")

    cpf_rows = []
    phone_rows = []
    name_rows = []
    sem_match = []

    for r in xl_rows:
        rgm = str(r[col["RGM"]]).strip() if r[col["RGM"]] else ""
        if rgm in crm_by_rgm:
            continue

        cpf = clean_cpf(r[col["CPF"]])
        phone = _xl_phone(r, col)
        nome = _xl_name(r, col)

        xl_data_mat = ""
        if r[col["DataMatricula"]]:
            dm = r[col["DataMatricula"]]
            xl_data_mat = dm.strftime("%Y-%m-%d") if hasattr(dm, "strftime") else str(dm)

        base_info = [
            rgm, cpf,
            r[col["Nome"]] or "",
            r[col["Curso"]] or "",
            r[col["Polo"]] or "",
            r[col["SituacaoMatricula"]] or "",
            r[col["TipoMatricula"]] or "",
            xl_data_mat,
        ]

        # Camada 1: CPF
        if cpf and cpf in crm_by_cpf:
            for lead in crm_by_cpf[cpf]:
                cpf_rows.append(base_info + [
                    "CPF", lead["id"], lead["nome"],
                    lead.get("telefone", ""), lead.get("email", ""),
                ])
            continue

        # Camada 2: Telefone
        if phone and phone in crm_by_phone:
            for lead in crm_by_phone[phone]:
                phone_rows.append(base_info + [
                    "TELEFONE", lead["id"], lead["nome"],
                    lead.get("telefone", ""), lead.get("email", ""),
                ])
            continue

        # Camada 3: Nome normalizado
        if nome and nome in crm_by_name:
            for lead in crm_by_name[nome]:
                name_rows.append(base_info + [
                    "NOME", lead["id"], lead["nome"],
                    lead.get("telefone", ""), lead.get("email", ""),
                ])
            continue

        # Sem match
        sem_match.append([
            rgm, cpf,
            r[col["Nome"]] or "",
            r[col["Sexo"]] or "",
            r[col["Curso"]] or "",
            r[col["Polo"]] or "",
            r[col["Negocio"]] or "",
            str(r[col["Serie"]] or ""),
            r[col["SituacaoMatricula"]] or "",
            r[col["TipoMatricula"]] or "",
            xl_data_mat,
            r[col["FoneCelular"]] or "",
            r[col["Email"]] or "",
            r[col["EmailAcademico"]] or "",
            r[col["Endereco"]] or "",
            r[col["Bairro"]] or "",
            r[col["Cidade"]] or "",
            r[col["Instituicao"]] or "",
            r[col["Empresa"]] or "",
        ])

    match_headers = [
        "RGM", "CPF", "planilha_nome", "planilha_curso", "planilha_polo",
        "planilha_situacao", "planilha_tipo", "planilha_data_mat",
        "match_tipo", "crm_lead_id", "crm_lead_nome",
        "crm_telefone", "crm_email",
    ]

    all_matched = cpf_rows + phone_rows + name_rows
    write_csv("acao2_match_todos.csv", all_matched, match_headers)
    write_csv("acao2a_match_cpf.csv", cpf_rows, match_headers)
    write_csv("acao2b_match_telefone.csv", phone_rows, match_headers)
    write_csv("acao2c_match_nome.csv", name_rows, match_headers)

    write_csv(
        "acao3_criar_novos.csv", sem_match,
        ["RGM", "CPF", "nome", "sexo", "curso", "polo", "nivel",
         "serie", "situacao", "tipo_matricula", "data_matricula",
         "telefone", "email", "email_acad", "endereco", "bairro", "cidade",
         "instituicao", "empresa"]
    )

    log.info("  Match CPF:      %d", len(cpf_rows))
    log.info("  Match Telefone: %d", len(phone_rows))
    log.info("  Match Nome:     %d", len(name_rows))
    log.info("  Sem match:      %d", len(sem_match))

    situacoes = Counter(r[8] for r in sem_match)
    if situacoes:
        log.info("  Sem match por situação: %s", dict(situacoes.most_common()))

    return len(cpf_rows), len(phone_rows), len(name_rows), len(sem_match)


# ---------------------------------------------------------------------------
# AÇÃO 4: Órfãos (RGM no CRM sem planilha)
# ---------------------------------------------------------------------------

def report_orfaos(xl_rows, col, crm_by_rgm):
    log.info("=== AÇÃO 4: Órfãos (RGM no CRM, não na planilha) ===")

    xl_rgms = set()
    for r in xl_rows:
        rgm = str(r[col["RGM"]]).strip() if r[col["RGM"]] else ""
        if rgm:
            xl_rgms.add(rgm)

    rows_csv = []
    status_count = Counter()

    for rgm, bizzes in crm_by_rgm.items():
        if rgm in xl_rgms:
            continue
        for biz in bizzes:
            status = biz["data"].get("status", "")
            status_count[status] += 1
            lead_nome = ""
            lead_obj = biz["data"].get("lead")
            if isinstance(lead_obj, dict):
                lead_nome = lead_obj.get("name", "")

            rows_csv.append([
                rgm, biz["id"], biz["data"].get("leadId", ""), lead_nome,
                status,
                get_biz_field(biz["data"], FIELD_IDS["Curso"]),
                get_biz_field(biz["data"], FIELD_IDS["Polo"]),
                get_biz_field(biz["data"], FIELD_IDS["Situacao"]),
            ])

    write_csv(
        "acao4_orfaos_crm.csv", rows_csv,
        ["RGM", "negocio_id", "lead_id", "lead_nome", "status",
         "curso", "polo", "situacao_campo"]
    )

    log.info("  %d negócios órfãos", len(rows_csv))
    for s, c in status_count.most_common():
        log.info("    %s: %d", s, c)


# ---------------------------------------------------------------------------
# AÇÃO 6: Leads duplicados por CPF
# ---------------------------------------------------------------------------

def report_leads_duplicados(conn):
    log.info("=== AÇÃO 6: Leads duplicados por CPF ===")
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("""
        WITH cpf_groups AS (
            SELECT
                REPLACE(REPLACE(data->>'taxId', '.', ''), '-', '') AS cpf,
                id,
                data->>'name' AS nome,
                data->>'rawPhone' AS telefone,
                data->>'email' AS email,
                data->>'createdAt' AS criado_em
            FROM leads
            WHERE data->>'taxId' IS NOT NULL AND data->>'taxId' != ''
        )
        SELECT cpf, id, nome, telefone, email, criado_em
        FROM cpf_groups
        WHERE cpf IN (
            SELECT cpf FROM cpf_groups GROUP BY cpf HAVING COUNT(*) > 1
        )
        ORDER BY cpf, criado_em
    """)
    result = cur.fetchall()
    cur.close()

    rows_csv = []
    cpf_groups = {}
    for r in result:
        cpf = r["cpf"].strip() if r["cpf"] else ""
        cpf_groups.setdefault(cpf, []).append(r)

    for cpf, leads in cpf_groups.items():
        for i, lead in enumerate(leads):
            rows_csv.append([
                cpf,
                len(leads),
                lead["id"],
                lead["nome"],
                lead["telefone"],
                lead["email"],
                lead["criado_em"],
                "MANTER" if i == 0 else "DUPLICADO",
            ])

    write_csv(
        "acao6_leads_duplicados_cpf.csv", rows_csv,
        ["CPF", "total_duplicados", "lead_id", "nome", "telefone", "email",
         "criado_em", "sugestao"]
    )
    log.info("  %d CPFs duplicados, %d leads envolvidos", len(cpf_groups), len(rows_csv))


# ---------------------------------------------------------------------------
# RESUMO
# ---------------------------------------------------------------------------

def print_resumo(xl_rows, col, crm_by_rgm, match_counts):
    cpf_count, phone_count, name_count, no_match_count = match_counts

    xl_rgms = set()
    for r in xl_rows:
        rgm = str(r[col["RGM"]]).strip() if r[col["RGM"]] else ""
        if rgm:
            xl_rgms.add(rgm)

    matched_rgm = xl_rgms & set(crm_by_rgm.keys())
    crm_only_rgm = set(crm_by_rgm.keys()) - xl_rgms

    print("\n" + "=" * 60)
    print("RESUMO DO CRUZAMENTO")
    print("=" * 60)
    print(f"  Planilha:                       {len(xl_rows):,} registros")
    print(f"  CRM negócios com RGM:           {sum(len(v) for v in crm_by_rgm.values()):,}")
    print()
    print(f"  ✓ Match por RGM (Ação 1):       {len(matched_rgm):,}")
    print(f"  ✓ Match por CPF (Ação 2a):      {cpf_count:,}")
    print(f"  ✓ Match por Telefone (Ação 2b): {phone_count:,}")
    print(f"  ✓ Match por Nome (Ação 2c):     {name_count:,}")
    print(f"  ✗ Sem match - criar (Ação 3):   {no_match_count:,}")
    print(f"  ⚠ Órfãos no CRM (Ação 4):      {len(crm_only_rgm):,}")
    print()
    print(f"  Relatórios gerados em: {REPORTS_DIR}")
    print("=" * 60)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    REPORTS_DIR.mkdir(exist_ok=True)

    log.info("=" * 50)
    log.info("Cruzamento Planilha x CRM — MODO DRY-RUN")
    log.info("Nenhuma alteração será feita no CRM")
    log.info("=" * 50)

    xl_rows, col = load_excel()

    conn = get_conn()
    try:
        crm_by_rgm = load_crm_businesses(conn)
        crm_by_cpf, crm_by_phone, crm_by_name = load_crm_leads(conn)

        match_counts = (0, 0, 0, 0)
        if "--resumo" not in sys.argv:
            rgm_dups = report_rgm_duplicados(crm_by_rgm)
            report_match_rgm(xl_rows, col, crm_by_rgm)
            match_counts = report_match_sem_rgm(
                xl_rows, col, crm_by_rgm,
                crm_by_cpf, crm_by_phone, crm_by_name,
            )
            report_orfaos(xl_rows, col, crm_by_rgm)
            report_leads_duplicados(conn)

        print_resumo(xl_rows, col, crm_by_rgm, match_counts)

    finally:
        conn.close()

    log.info("Concluído. Revise os relatórios em %s", REPORTS_DIR)


if __name__ == "__main__":
    main()
