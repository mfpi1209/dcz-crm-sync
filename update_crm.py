"""
eduit. — Atualização em massa a partir da planilha de matriculados.

Atualiza campos existentes E cria leads/negócios para registros sem match.

Uso:
    python update_crm.py --test                  # Testa 1 update real
    python update_crm.py --dry-run               # Mostra resumo (padrão)
    python update_crm.py --execute               # Atualiza + cria tudo
    python update_crm.py --execute --no-create   # Só atualiza, sem criar
    python update_crm.py --execute --limit 50    # Primeiros 50 updates (sem criar)
    python update_crm.py --execute --rate 180    # Rate limit customizado
"""

import sys
import io
import os
import csv
import json
import time
import hashlib
import logging
import unicodedata
from datetime import datetime, timezone, timedelta

BRT = timezone(timedelta(hours=-3))
from pathlib import Path
from collections import Counter

import warnings
import requests
import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

load_dotenv(Path(__file__).parent / ".env")

API_BASE = "https://api.g1.datacrazy.io/api/v1"
API_TOKEN = os.getenv("DATACRAZY_API_TOKEN", "")

DB_DSN = dict(
    host=os.getenv("DB_HOST", "localhost"),
    port=os.getenv("DB_PORT", "5432"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASS"),
    dbname=os.getenv("DB_NAME", "dcz_sync"),
)

REPORTS_DIR = Path(__file__).parent / "reports"
LOG_DIR = Path(__file__).parent / "logs"

BIZ_FIELD_IDS = {
    "RGM":              "2ac4e30f-cfd7-435f-b688-fbce27f76c38",
    "Curso":            "4bddb764-658b-48bc-9d70-6e94ad420132",
    "Polo":             "0ec9d8dc-d547-4482-b9ad-d4a3e6ec1b54",
    "Serie":            "b921a702-8e51-4b6c-b4d8-cdea931ea51d",
    "Situacao":         "fd08d44b-a4a5-4343-b7a9-37f75e2c1caa",
    "DataMatricula":    "bf93a8e9-42c0-4517-8518-6f604746a300",
    "Modalidade":       "9c8fc723-d9f7-4074-a0bc-ca4b96d36739",
    "Bairro":           "f7cf5892-573f-45b8-9425-6dafab92cc2c",
    "Cidade":           "7a4407e4-7345-4f7e-8a24-4f51d4a10cf8",
    "EmailAD":          "731bd2fd-7cfa-49af-ab24-2e55e0374798",
    "SenhaProvisoria":  "cccb3046-1906-4465-901d-329ef2fe08dc",
    "TipoAluno":        "4230e4db-970b-4444-abaf-c3135a03b79c",
    "Turma":            "8815a8de-f755-4597-b6f4-8da6d289b6eb",
    "Ciclo":            "b9dce12b-30b7-4a0f-a764-298031f5b84e",
    "Nivel":            "233fcf6f-0bed-49d7-89a1-d1cd54fb9c12",
}

LEAD_FIELD_IDS = {
    "Sexo":             "802d0e93-53e9-4c3e-b593-14dafd890bf8",
}

FIELD_IDS = {**BIZ_FIELD_IDS, **LEAD_FIELD_IDS}

API_RATE_LIMIT = 240            # requests/min allowed by the API
DEFAULT_TARGET_RATE = 120       # requests/min default target (50% margin)
CRITICAL_REMAINING = 20        # below this → pause until window resets
SKIP_ADDRESS = True             # pular atualizações de endereço (rua/bairro/cidade)

class _BRTFormatter(logging.Formatter):
    def formatTime(self, record, datefmt=None):
        dt = datetime.fromtimestamp(record.created, tz=BRT)
        return dt.strftime(datefmt or "%H:%M:%S")

logging.basicConfig(level=logging.INFO)
_handler = logging.StreamHandler()
_handler.setFormatter(_BRTFormatter("%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S"))
logging.root.handlers = [_handler]
log = logging.getLogger("update_crm")

# ---------------------------------------------------------------------------
# API Client
# ---------------------------------------------------------------------------

class ApiClient:
    def __init__(self, target_rate=None):
        self.s = requests.Session()
        self.s.headers["Authorization"] = f"Bearer {API_TOKEN}"
        self.s.headers["Content-Type"] = "application/json"
        self._remaining = API_RATE_LIMIT
        self._reset = 0
        self._last_req = 0.0
        self.total_calls = 0
        self._window_start = time.monotonic()
        self._window_calls = 0
        self.target_rate = max(1, min(target_rate or DEFAULT_TARGET_RATE, API_RATE_LIMIT))
        self.base_delay = 60.0 / self.target_rate
        log.info("Rate-limit configurado: %d req/min (delay base %.2fs)",
                 self.target_rate, self.base_delay)

    def _throttle(self):
        now = time.monotonic()

        if now - self._window_start >= 60:
            self._window_start = now
            self._window_calls = 0

        if self._remaining <= CRITICAL_REMAINING and self._reset > 0:
            wait = self._reset + 1
            log.warning("Rate-limit crítico (%d restantes) — pausando %ds",
                        self._remaining, wait)
            time.sleep(wait)
            self._window_start = time.monotonic()
            self._window_calls = 0
            return

        ratio = self._remaining / API_RATE_LIMIT
        if ratio > 0.5:
            delay = self.base_delay
        elif ratio > 0.25:
            delay = self.base_delay * 1.5
        else:
            delay = self.base_delay * 3.0

        if self._window_calls >= self.target_rate:
            remaining_window = 60 - (now - self._window_start)
            if remaining_window > 0:
                log.info("Limite interno atingido (%d calls) — pausando %.1fs",
                         self._window_calls, remaining_window)
                time.sleep(remaining_window + 0.5)
                self._window_start = time.monotonic()
                self._window_calls = 0
                return

        elapsed = now - self._last_req
        if elapsed < delay:
            time.sleep(delay - elapsed)

    def _read_headers(self, r):
        self._remaining = int(r.headers.get("X-RateLimit-Remaining", self._remaining))
        self._reset = int(r.headers.get("X-RateLimit-Reset", 0))

    def _request(self, method, url, payload=None):
        for attempt in range(4):
            self._throttle()
            self._last_req = time.monotonic()
            self.total_calls += 1
            self._window_calls += 1

            try:
                r = self.s.request(method, url, json=payload, timeout=30)
            except (requests.exceptions.ReadTimeout,
                    requests.exceptions.ConnectionError) as exc:
                wait = min(5 * (2 ** attempt), 60)
                log.warning("Timeout/conexão (tentativa %d/4) — retry em %ds: %s",
                            attempt + 1, wait, str(exc)[:120])
                time.sleep(wait)
                continue

            if r.status_code == 429:
                retry = int(r.headers.get("Retry-After", 30))
                log.warning("429 — Retry-After %ds (tentativa %d/4)", retry, attempt + 1)
                time.sleep(retry + 1)
                continue

            self._read_headers(r)

            if r.status_code >= 400:
                return {"ok": False, "status": r.status_code, "body": r.text[:500]}

            body = r.json() if r.text.strip() else {}
            return {"ok": True, "status": r.status_code, "body": body}

        return {"ok": False, "status": 429, "body": "Falha após 4 tentativas (timeout/429)"}

    def patch(self, path, payload):
        return self._request("PATCH", f"{API_BASE}{path}", payload)

    def put(self, path, payload):
        return self._request("PUT", f"{API_BASE}{path}", payload)

    def put_biz_field(self, biz_id, field_id, value):
        """PUT /crm/crm/additional-fields/business/{bizId}/{fieldId}"""
        path = f"/crm/crm/additional-fields/business/{biz_id}/{field_id}"
        return self.put(path, {"value": str(value)})

    def put_lead_field(self, lead_id, field_id, value):
        """PUT /crm/crm/additional-fields/lead/{leadId}/{fieldId}"""
        path = f"/crm/crm/additional-fields/lead/{lead_id}/{field_id}"
        return self.put(path, {"value": str(value)})

    def post(self, path, payload):
        return self._request("POST", f"{API_BASE}{path}", payload)

    def get(self, path, params=None):
        url = f"{API_BASE}{path}"
        self._throttle()
        self._last_req = time.monotonic()
        self.total_calls += 1
        r = self.s.get(url, params=params, timeout=30)
        self._read_headers(r)
        if r.status_code >= 400:
            return {"ok": False, "status": r.status_code, "body": r.text[:500]}
        return {"ok": True, "status": r.status_code, "body": r.json()}

    def create_lead(self, payload):
        return self._request("POST", f"{API_BASE}/leads", payload)

    def create_business(self, lead_id, stage_id):
        return self.post("/businesses", {"leadId": lead_id, "stageId": stage_id})

    def move_businesses(self, ids, destination_stage_id):
        return self.post("/businesses/actions/move", {"ids": ids, "destinationStageId": destination_stage_id})

    def search_leads(self, query):
        from urllib.parse import quote
        return self._request("GET", f"{API_BASE}/leads?search={quote(str(query))}&take=5")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_cpf(cpf):
    if not cpf:
        return ""
    return str(cpf).replace(".", "").replace("-", "").replace(" ", "").strip()


def format_cpf(cpf):
    """Only digits, zero-padded to 11 chars."""
    c = clean_cpf(cpf)
    if not c:
        return ""
    return c.zfill(11)


def clean_phone(phone):
    if not phone:
        return ""
    digits = "".join(c for c in str(phone) if c.isdigit())
    if digits.startswith("55") and len(digits) >= 12:
        digits = digits[2:]
    if len(digits) >= 10:
        return digits[-10:] if len(digits) == 10 else digits[-11:]
    return ""


def normalize_name(name):
    if not name:
        return ""
    import unicodedata
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    return " ".join(name.upper().split())


# ---------------------------------------------------------------------------
# Normalização de dados para o CRM
# ---------------------------------------------------------------------------

KEEP_UPPER = {"cst", "ead", "ti", "rh", "ii", "iii", "iv"}

def title_case(text):
    """Primeira maiúscula em cada palavra, preservando siglas conhecidas."""
    if not text:
        return ""
    text = str(text).strip()
    LOWERCASE_WORDS = {"de", "do", "da", "dos", "das", "e", "em", "o", "a"}
    words = text.lower().split()
    result = []
    for i, w in enumerate(words):
        if w in KEEP_UPPER:
            result.append(w.upper())
        elif i == 0 or w not in LOWERCASE_WORDS:
            result.append(w.capitalize())
        else:
            result.append(w)
    return " ".join(result)


SEXO_MAP = {
    "m": "Masculino",
    "f": "Feminino",
    "masculino": "Masculino",
    "feminino": "Feminino",
}

def normalize_sexo(sexo):
    if not sexo:
        return ""
    return SEXO_MAP.get(str(sexo).strip().lower(), title_case(sexo))


POLO_MAP = {
    "barra funda": "Barra Funda",
    "sapopemba": "Sapopemba",
    "vila prudente": "Vila Prudente",
    "morumbi": "Morumbi",
    "santana": "Santana",
    "vila mariana": "Vila Mariana",
    "ibirapuera": "Ibirapuera",
    "freguesia do o": "Freguesia do Ó",
    "freguesia do ó": "Freguesia do Ó",
    "taboao da serra - centro": "Taboão da Serra - Centro",
    "taboão da serra - centro": "Taboão da Serra - Centro",
    "taboao da serra centro": "Taboão da Serra - Centro",
    "taboão da serra centro": "Taboão da Serra - Centro",
    "taboao da serra - mituzi": "Taboão da Serra - Mituzi",
    "taboão da serra - mituzi": "Taboão da Serra - Mituzi",
    "taboao da serra mituzi": "Taboão da Serra - Mituzi",
    "taboão da serra mituzi": "Taboão da Serra - Mituzi",
    "campinas": "Campinas",
    "capivari": "Capivari",
    "itapira": "Itapira",
}


def normalize_polo(polo):
    if not polo:
        return ""
    import unicodedata
    key = unicodedata.normalize("NFKD", str(polo).strip())
    key = "".join(c for c in key if not unicodedata.combining(c)).lower()

    if key in POLO_MAP:
        return POLO_MAP[key]

    # Match parcial para nomes longos tipo "3146 - Polo Taboão da Serra_centro..."
    if "taboao" in key or "taboa" in key:
        if "mituzi" in key or "mituizi" in key or "jardim" in key:
            return "Taboão da Serra - Mituzi"
        if "centro" in key or "santos dumont" in key or "parque" in key:
            return "Taboão da Serra - Centro"
        return "Taboão da Serra - Centro"

    for canonical_key, canonical_val in POLO_MAP.items():
        if canonical_key in key:
            return canonical_val

    return title_case(str(polo).strip())


SITUACAO_MAP = {
    "em curso": "Em Curso",
    "trancado": "Trancado",
    "cancelado": "Cancelado",
    "transferido": "Transferido",
}


def normalize_situacao(sit):
    if not sit:
        return ""
    key = str(sit).strip().lower()
    return SITUACAO_MAP.get(key, title_case(sit))


TIPO_ALUNO_MAP = {
    "nova matricula": "Calouro",
    "nova matrícula": "Calouro",
    "recompra": "Calouro (Recompra)",
    "retorno": "Regresso (Retorno)",
    "rematricula": "Veterano",
    "rematrícula": "Veterano",
}


def normalize_tipo_aluno(tipo):
    if not tipo:
        return ""
    key = str(tipo).strip().lower()
    return TIPO_ALUNO_MAP.get(key, "")


NIVEL_MAP = {
    "graduação": "Graduação",
    "graduacao": "Graduação",
    "pós-graduação": "Pós-Graduação",
    "pos-graduacao": "Pós-Graduação",
    "pós graduação": "Pós-Graduação",
    "pos graduacao": "Pós-Graduação",
    "posgraduacao": "Pós-Graduação",
}


def normalize_nivel(nivel):
    if not nivel:
        return ""
    key = _strip_accents(str(nivel).strip()).lower()
    return NIVEL_MAP.get(key, "")


def _strip_accents(text):
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def generate_senha(nome, rgm, cpf):
    """Senha padrão: Primeiras 3 letras do nome (sem acento, capitalizado) + @ + 3 primeiros dígitos RGM + 4 primeiros dígitos CPF.
    Ex: Fernanda, RGM 36925847, CPF 12345678912 → Fer@3691234"""
    if not nome or not rgm or not cpf:
        return ""
    first_name = nome.strip().split()[0] if nome.strip() else ""
    if not first_name or len(first_name) < 2:
        return ""
    prefix = _strip_accents(first_name[:3]).capitalize()
    rgm_digits = "".join(c for c in str(rgm) if c.isdigit())[:3]
    cpf_digits = "".join(c for c in str(cpf) if c.isdigit())[:4]
    if len(rgm_digits) < 3 or len(cpf_digits) < 4:
        return ""
    return f"{prefix}@{rgm_digits}{cpf_digits}"


def get_additional_field(data, field_id):
    """Retorna o valor de um additionalField pelo ID (funciona para leads e negócios)."""
    for f in data.get("additionalFields", []):
        af = f.get("additionalField", {})
        if isinstance(af, dict) and af.get("id") == field_id:
            return f.get("value", "")
        if isinstance(af, str) and af == field_id:
            return f.get("value", "")
    return ""


def get_biz_field(biz_data, field_id):
    return get_additional_field(biz_data, field_id)


def get_biz_field_value_id(biz_data, field_id):
    """Retorna o ID do registro de valor do campo (para saber se é update ou create)."""
    for f in biz_data.get("additionalFields", []):
        af = f.get("additionalField", {})
        if isinstance(af, dict) and af.get("id") == field_id:
            return f.get("id", "")
    return ""


CALOURO_TIPOS = {"calouro", "calouro (recompra)", "nova matrícula", "nova matricula", "recompra", "retorno", "regresso (retorno)"}
VETERANO_TIPOS = {"veterano", "rematrícula", "rematricula"}

STAGE_NAMES = {
    "calouro":   ["Calouro", "Calouros", "CALOURO"],
    "veterano":  ["Veterano", "Veteranos", "VETERANO"],
}


def get_conn():
    return psycopg2.connect(**DB_DSN)


def load_pipeline_stages(conn):
    """Carrega etapas do pipeline do banco local."""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT ps.id, ps.data->>'name' AS nome
            FROM pipeline_stages ps
            ORDER BY ps.data->>'order'
        """)
        return cur.fetchall()


def resolve_stage_ids(stages):
    """Encontra IDs das etapas Calouro e Veterano."""
    found = {}
    for key, name_variants in STAGE_NAMES.items():
        lower_variants = [n.lower().strip() for n in name_variants]
        for stage in stages:
            sname = (stage["nome"] or "").strip()
            if sname.lower() in lower_variants:
                found[key] = stage["id"]
                break
    return found


def _load_turmas(conn):
    """Carrega configuração de turmas do banco."""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("SELECT nivel, nome, dt_inicio, dt_fim FROM turmas ORDER BY dt_inicio")
        rows = cur.fetchall()
    log.info("  %d turmas configuradas no banco", len(rows))
    return rows


def _load_ciclos(conn):
    """Carrega configuração de ciclos do banco."""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("SELECT nivel, nome, dt_inicio, dt_fim FROM ciclos ORDER BY dt_inicio")
        rows = cur.fetchall()
    log.info("  %d ciclos configurados no banco", len(rows))
    return rows


def _classify_date_range(data_matricula, nivel, ranges):
    """Retorna o nome do range (turma ou ciclo) ou '' se nenhum casar."""
    if not data_matricula or not nivel:
        return ""
    from datetime import date as dt_date
    try:
        parts = data_matricula.split("-")
        dt = dt_date(int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, IndexError):
        return ""
    for t in ranges:
        if t["nivel"] == nivel and t["dt_inicio"] <= dt <= t["dt_fim"]:
            return t["nome"]
    return ""


def _classify_turma(data_matricula, nivel, turmas):
    return _classify_date_range(data_matricula, nivel, turmas)


def _classify_ciclo(data_matricula, nivel, ciclos):
    return _classify_date_range(data_matricula, nivel, ciclos)


def _data_hash(data):
    return hashlib.md5(json.dumps(data, sort_keys=True, default=str).encode()).hexdigest()


def update_local_biz_field(conn, biz_id, field_id, new_value):
    """Atualiza o valor de um campo adicional no JSONB local do negócio."""
    with conn.cursor() as cur:
        cur.execute("SELECT data FROM businesses WHERE id = %s", (biz_id,))
        row = cur.fetchone()
        if not row:
            return
        data = row[0]
        found = False
        for f in data.get("additionalFields", []):
            af = f.get("additionalField", {})
            fid = af.get("id") if isinstance(af, dict) else af
            if fid == field_id:
                f["value"] = str(new_value)
                found = True
                break
        if not found:
            fname = field_id_to_name(field_id)
            data.setdefault("additionalFields", []).append({
                "additionalField": {"id": field_id, "name": fname},
                "value": str(new_value),
            })
        jdata = json.dumps(data)
        cur.execute(
            "UPDATE businesses SET data = %s::jsonb, data_hash = %s WHERE id = %s",
            (jdata, _data_hash(data), biz_id),
        )
        conn.commit()


def update_local_lead_field(conn, lead_id, field_id, field_name, new_value):
    """Atualiza o valor de um campo adicional no JSONB local do lead."""
    with conn.cursor() as cur:
        cur.execute("SELECT data FROM leads WHERE id = %s", (lead_id,))
        row = cur.fetchone()
        if not row:
            return
        data = row[0]
        found = False
        for f in data.get("additionalFields", []):
            af = f.get("additionalField", {})
            fid = af.get("id") if isinstance(af, dict) else af
            if fid == field_id:
                f["value"] = str(new_value)
                found = True
                break
        if not found:
            data.setdefault("additionalFields", []).append({
                "additionalField": {"id": field_id, "name": field_name},
                "value": str(new_value),
            })
        jdata = json.dumps(data)
        cur.execute(
            "UPDATE leads SET data = %s::jsonb, data_hash = %s WHERE id = %s",
            (jdata, _data_hash(data), lead_id),
        )
        conn.commit()


def update_local_lead(conn, lead_id, updates):
    """Atualiza campos do lead no JSONB local."""
    if not updates:
        return
    with conn.cursor() as cur:
        cur.execute("SELECT data FROM leads WHERE id = %s", (lead_id,))
        row = cur.fetchone()
        if not row:
            return
        data = row[0]
        for k, v in updates.items():
            if k == "address":
                if "address" not in data or not isinstance(data["address"], dict):
                    data["address"] = {}
                data["address"].update(v)
            else:
                data[k] = v
        jdata = json.dumps(data)
        cur.execute(
            "UPDATE leads SET data = %s::jsonb, data_hash = %s WHERE id = %s",
            (jdata, _data_hash(data), lead_id),
        )
        conn.commit()


# ---------------------------------------------------------------------------
# Carregamento
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    "Ciclo": ["Ciclo"], "Nome": ["Nome"], "CPF": ["CPF"], "RGM": ["RGM"],
    "Sexo": ["Sexo"], "Curso": ["Curso"],
    "Instituicao": ["Instituição", "Institui"], "Empresa": ["Empresa"],
    "Polo": ["Polo"], "Negocio": ["Negócio", "Neg"],
    "Serie": ["Série", "Serie", "rie"],
    "DataNascimento": ["Data Nascimento"],
    "TipoMatricula": ["Tipo Matrícula", "Tipo Matr"],
    "DataMatricula": ["Data Matrícula", "Data Matr"],
    "SituacaoMatricula": ["Situação Matrícula", "Situa"],
    "FoneResidencial": ["Fone Residencial"], "FoneComercial": ["Fone Comercial"],
    "FoneCelular": ["Fone celular"], "Email": ["Email"],
    "EmailAcademico": ["Email acadêmico", "Email acad"],
    "Endereco": ["Endereço", "Endere"],
    "Bairro": ["Bairro"], "Cidade": ["Cidade"],
}


def _col_find(col_map, *candidates):
    for c in candidates:
        if c in col_map:
            return c
    for c in candidates:
        cl = c.lower()
        for k in col_map:
            if k and cl in k.lower():
                return k
    return candidates[0]


def load_excel():
    log.info("Carregando planilha...")
    xlsx = None
    for f in Path(__file__).parent.iterdir():
        if f.suffix.lower() == ".xlsx" and "matriculados" in f.name.lower():
            xlsx = f
            break
    if not xlsx:
        raise FileNotFoundError("Planilha de matriculados não encontrada")

    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    ws = wb["Export"]
    raw_header = [cell.value for cell in ws[1]]
    raw_col = {h: i for i, h in enumerate(raw_header) if h}

    col = {}
    missing = []
    for norm, aliases in COLUMN_ALIASES.items():
        found = _col_find(raw_col, *aliases)
        if found in raw_col:
            col[norm] = raw_col[found]
        else:
            missing.append(norm)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[1] is None:
            continue
        rows.append(row)
    wb.close()
    log.info("  %d registros, %d colunas mapeadas", len(rows), len(col))
    log.info("  Colunas encontradas: %s", ", ".join(sorted(col.keys())))
    if missing:
        log.warning("  Colunas NÃO encontradas na planilha: %s", ", ".join(sorted(missing)))
    return rows, col


def load_crm_data(conn):
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    log.info("Carregando todos os negócios (pode levar ~2 min)...")
    cur.execute("SELECT b.id, b.data FROM businesses b")
    all_biz = cur.fetchall()

    by_rgm = {}
    biz_by_lead = {}
    for row in all_biz:
        lead_id = row["data"].get("leadId", "")
        if lead_id:
            biz_by_lead.setdefault(lead_id, []).append(row)
        rgm = get_biz_field(row["data"], FIELD_IDS["RGM"])
        if rgm:
            by_rgm.setdefault(rgm, []).append(row)
    log.info("  %d negócios | %d RGMs | %d leads com negócio",
             len(all_biz), len(by_rgm), len(biz_by_lead))

    log.info("Carregando todos os leads (pode levar ~1 min)...")
    cur.execute("""
        SELECT id,
               data->>'name' AS nome,
               REPLACE(REPLACE(COALESCE(data->>'taxId',''), '.', ''), '-', '') AS cpf,
               data->>'rawPhone' AS telefone,
               data->>'email' AS email,
               data
        FROM leads
    """)
    leads = cur.fetchall()

    by_cpf, by_phone, by_name = {}, {}, {}
    leads_by_id = {}
    for r in leads:
        leads_by_id[r["id"]] = r
        cpf = r["cpf"].strip() if r["cpf"] else ""
        if cpf:
            by_cpf.setdefault(cpf, []).append(r)
        phone = clean_phone(r["telefone"])
        if phone:
            by_phone.setdefault(phone, []).append(r)
        nome = normalize_name(r["nome"])
        if nome:
            by_name.setdefault(nome, []).append(r)

    cur.close()
    log.info("  %d leads | CPF: %d | Phone: %d | Name: %d",
             len(leads), len(by_cpf), len(by_phone), len(by_name))

    return by_rgm, by_cpf, by_phone, by_name, leads_by_id, biz_by_lead


# ---------------------------------------------------------------------------
# Preparação das atualizações
# ---------------------------------------------------------------------------

_debug_diff = False
_diff_details = []

def _find_lead(cpf, phone, nome, email, crm_by_cpf, crm_by_phone, crm_by_name, leads_by_id):
    """Stage 1: Find the lead by CPF → Phone → Email → Name.
    Returns (lead_match_type, lead_id) or (None, None).
    For weak matches (phone/name), cross-checks CPF if available."""
    if cpf and cpf in crm_by_cpf:
        return "CPF", crm_by_cpf[cpf][0]["id"]

    if phone and phone in crm_by_phone:
        candidate = crm_by_phone[phone][0]
        cand_cpf = clean_cpf(candidate.get("cpf", ""))
        if cpf and cand_cpf and cpf != cand_cpf:
            pass
        else:
            return "TELEFONE", candidate["id"]

    if nome and nome in crm_by_name:
        candidate = crm_by_name[nome][0]
        cand_cpf = clean_cpf(candidate.get("cpf", ""))
        if cpf and cand_cpf and cpf != cand_cpf:
            pass
        else:
            return "NOME", candidate["id"]

    return None, None


def _find_biz_for_rgm(rgm, biz_list):
    """Stage 2: Within a lead's businesses, find the one with this RGM.
    Returns the matching business dict, or None."""
    if not rgm:
        return None
    for biz in biz_list:
        biz_rgm = get_biz_field(biz["data"], FIELD_IDS["RGM"])
        if biz_rgm and biz_rgm.strip() == rgm:
            return biz
    return None


def _find_empty_biz(biz_list):
    """Find a business with no RGM (candidate for RGM assignment)."""
    for biz in biz_list:
        biz_rgm = get_biz_field(biz["data"], FIELD_IDS["RGM"])
        if not biz_rgm or not biz_rgm.strip():
            return biz
    return None


def _normalize_date(val):
    """Normalize date to YYYY-MM-DD for comparison."""
    if not val:
        return ""
    s = str(val).strip().split("T")[0].split(" ")[0]
    if "/" in s:
        parts = s.split("/")
        if len(parts) == 3:
            if len(parts[0]) == 4:  # YYYY/MM/DD
                return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            else:  # DD/MM/YYYY
                return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    return s


def _compare_field(field_name, crm_val, xl_val):
    """Return True if values are equivalent (no update needed)."""
    if crm_val == xl_val:
        return True
    if not crm_val and not xl_val:
        return True
    if field_name == "DataMatricula":
        return _normalize_date(crm_val) == _normalize_date(xl_val)
    if field_name in ("Curso", "Polo", "Bairro", "Cidade", "Nivel"):
        return crm_val.strip().lower() == xl_val.strip().lower()
    return False


def prepare_updates(xl_rows, col, crm_by_rgm, crm_by_cpf, crm_by_phone, crm_by_name, leads_by_id, biz_by_lead, turmas=None, ciclos=None):
    """Two-stage matching:
    Stage 1 — Find the LEAD (CPF → Phone → Email → Name)
    Stage 2 — Find the BUSINESS within that lead (by RGM)
    Falls back to direct RGM lookup if lead match fails.
    Returns (updates, unmatched) — unmatched contains records to create."""
    global _debug_diff, _diff_details
    _debug_diff = True
    updates = []
    unmatched = []
    _skipped_nivel = []
    _format_samples = {}  # field_name → [(crm_val, xl_val)] first 3 diffs per field

    # Deduplica RGMs — mantém última ocorrência na planilha
    seen_rgm = {}
    dupes = {}
    for idx, r in enumerate(xl_rows):
        rgm_val = str(r[col["RGM"]]).strip() if r[col["RGM"]] else ""
        if not rgm_val:
            continue
        if rgm_val in seen_rgm:
            dupes.setdefault(rgm_val, [seen_rgm[rgm_val]]).append(idx)
        seen_rgm[rgm_val] = idx

    skip_indices = set()
    for rgm_val, indices in dupes.items():
        for i in indices[:-1]:
            skip_indices.add(i)

    if dupes:
        log.info("  RGMs duplicados na planilha: %d (mantendo última ocorrência)", len(dupes))
        REPORTS_DIR.mkdir(exist_ok=True)
        dup_path = REPORTS_DIR / "duplicados_planilha.csv"
        with open(dup_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["RGM", "Ocorrencias", "Linhas"])
            for rgm_val, indices in sorted(dupes.items()):
                w.writerow([rgm_val, len(indices) + 1, ", ".join(str(i + 2) for i in indices)])
        log.info("  Relatório: %s", dup_path)

    for idx, r in enumerate(xl_rows):
        if idx in skip_indices:
            continue

        rgm = str(r[col["RGM"]]).strip() if r[col["RGM"]] else ""
        cpf = clean_cpf(r[col["CPF"]])
        phone = clean_phone(str(r[col["FoneCelular"]] or ""))
        nome = normalize_name(r[col["Nome"]] or "")
        email = (r[col["Email"]] or "").strip().lower() if "Email" in col else ""

        serie_raw = r[col["Serie"]] if "Serie" in col else None
        try:
            serie_str = str(int(float(serie_raw))) if serie_raw else ""
        except (ValueError, TypeError):
            serie_str = str(serie_raw).strip() if serie_raw else ""

        def _parse_date(key):
            val = r[col[key]] if key in col else None
            if val and hasattr(val, "strftime"):
                return val.strftime("%Y-%m-%d")
            elif val:
                return str(val).strip().split("T")[0].split(" ")[0]
            return ""

        dm_str = _parse_date("DataMatricula")
        dn_str = _parse_date("DataNascimento")

        def _col_val(key, default=""):
            return r[col[key]] if key in col and col[key] < len(r) else default

        negocio_raw = str(_col_val("Negocio") or "").strip()
        nivel = normalize_nivel(negocio_raw)

        if not nivel:
            _skipped_nivel.append({
                "nome": title_case(_col_val("Nome") or ""),
                "rgm": rgm,
                "cpf": cpf,
                "negocio": negocio_raw,
            })
            continue

        xl_data = {
            "rgm": rgm,
            "cpf": cpf,
            "nome": title_case(_col_val("Nome") or ""),
            "curso": title_case(_col_val("Curso") or ""),
            "polo": normalize_polo(_col_val("Polo") or ""),
            "serie": serie_str,
            "situacao": normalize_situacao(_col_val("SituacaoMatricula") or ""),
            "tipo": normalize_tipo_aluno(_col_val("TipoMatricula") or ""),
            "endereco": title_case(_col_val("Endereco") or ""),
            "bairro": title_case(_col_val("Bairro") or ""),
            "cidade": title_case(_col_val("Cidade") or ""),
            "sexo": normalize_sexo(_col_val("Sexo") or ""),
            "email": email,
            "email_acad": (_col_val("EmailAcademico") or "").strip().lower(),
            "phone_raw": str(_col_val("FoneCelular") or ""),
            "data_matricula": dm_str,
            "data_nasc": dn_str,
            "nivel": nivel,
        }

        # ── Stage 1: Find the LEAD ──
        lead_match, matched_lead_id = _find_lead(
            cpf, phone, nome, email,
            crm_by_cpf, crm_by_phone, crm_by_name, leads_by_id,
        )

        match_type = None
        target_biz = None

        no_match_reason = None
        if matched_lead_id:
            lead_bizs = biz_by_lead.get(matched_lead_id, [])

            # ── Stage 2: Find the BUSINESS by RGM within this lead ──
            target_biz = _find_biz_for_rgm(rgm, lead_bizs)

            if target_biz:
                match_type = f"{lead_match}+RGM"
            else:
                target_biz = _find_empty_biz(lead_bizs)
                if target_biz:
                    match_type = f"{lead_match}+ASSIGN"
                elif lead_bizs:
                    no_match_reason = "LEAD_SEM_BIZ_RGM"
                else:
                    no_match_reason = "LEAD_SEM_BIZ"
        else:
            if rgm and rgm in crm_by_rgm:
                biz_matches = crm_by_rgm[rgm]
                if len(biz_matches) == 1:
                    target_biz = biz_matches[0]
                    matched_lead_id = target_biz["data"].get("leadId", "")
                    match_type = "RGM_DIRETO"
                else:
                    no_match_reason = "RGM_MULTIPLO"
            else:
                no_match_reason = "SEM_MATCH"

        if no_match_reason:
            unmatched.append({
                "xl_data": xl_data,
                "reason": no_match_reason,
                "lead_id": matched_lead_id,
            })
            continue

        if not target_biz or not match_type:
            continue

        # ── Prepare lead updates ──
        lead_updates = {}
        if matched_lead_id and matched_lead_id in leads_by_id:
            lead = leads_by_id[matched_lead_id]

            if cpf:
                formatted_cpf = format_cpf(cpf)
                crm_cpf = clean_cpf(lead["cpf"] or "")
                if formatted_cpf != crm_cpf:
                    lead_updates["taxId"] = formatted_cpf

            crm_email = (lead["email"] or "").strip().lower()
            if xl_data["email"] and xl_data["email"] != crm_email:
                lead_updates["email"] = xl_data["email"]

            crm_company = (lead["data"].get("company") or "").strip()
            if crm_company:
                lead_updates["company"] = ""

            if not SKIP_ADDRESS:
                crm_addr = lead["data"].get("address") or {}
                addr = {}
                crm_street = (crm_addr.get("address") or crm_addr.get("street") or "").strip()
                if xl_data["endereco"] and xl_data["endereco"] != crm_street:
                    addr["address"] = xl_data["endereco"]
                if xl_data["bairro"] and xl_data["bairro"] != (crm_addr.get("block") or "").strip():
                    addr["block"] = xl_data["bairro"]
                if xl_data["cidade"] and xl_data["cidade"] != (crm_addr.get("city") or "").strip():
                    addr["city"] = xl_data["cidade"]
                if addr:
                    lead_updates["address"] = addr

            if xl_data["data_nasc"]:
                crm_bday = (lead["data"].get("birthDate") or "").strip()
                if _normalize_date(xl_data["data_nasc"]) != _normalize_date(crm_bday):
                    lead_updates["birthDate"] = xl_data["data_nasc"] + "T03:00:00.000Z"

        # ── Prepare lead additional field updates ──
        lead_field_updates = {}
        if matched_lead_id and xl_data["sexo"] and matched_lead_id in leads_by_id:
            fid = LEAD_FIELD_IDS.get("Sexo", "")
            if fid:
                current_sexo = get_additional_field(leads_by_id[matched_lead_id]["data"], fid)
                if current_sexo != xl_data["sexo"]:
                    lead_field_updates["Sexo"] = (fid, xl_data["sexo"])

        # ── Prepare business field updates (single target business) ──
        biz_updates = []
        fields_to_update = {}
        senha = generate_senha(xl_data["nome"], rgm, cpf)
        turma_nome = _classify_turma(xl_data["data_matricula"], xl_data["nivel"], turmas or [])
        ciclo_nome = _classify_ciclo(xl_data["data_matricula"], xl_data["nivel"], ciclos or [])

        mapping = {
            "Curso": xl_data["curso"],
            "Polo": xl_data["polo"],
            "Serie": xl_data["serie"],
            "Situacao": xl_data["situacao"],
            "DataMatricula": xl_data["data_matricula"],
            "TipoAluno": xl_data["tipo"],
            "EmailAD": xl_data["email_acad"],
            "SenhaProvisoria": senha,
            "Nivel": xl_data["nivel"],
            "Turma": turma_nome,
            "Ciclo": ciclo_nome,
        }
        if rgm:
            mapping["RGM"] = rgm

        for field_name, new_val in mapping.items():
            if not new_val:
                continue
            fid = FIELD_IDS.get(field_name, "")
            if not fid:
                continue
            current = str(get_biz_field(target_biz["data"], fid) or "").strip()
            new_clean = str(new_val).strip()
            if _compare_field(field_name, current, new_clean):
                continue
            fields_to_update[fid] = new_clean
            if _debug_diff:
                _diff_details.append(f"  {field_name}: '{current}' → '{new_clean}'")
            samples = _format_samples.setdefault(field_name, [])
            if len(samples) < 3:
                samples.append((current, new_clean))

        if fields_to_update:
            biz_updates.append({
                "biz_id": target_biz["id"],
                "fields": fields_to_update,
            })

        if lead_updates or lead_field_updates or biz_updates:
            updates.append({
                "match_type": match_type,
                "xl_nome": xl_data["nome"],
                "xl_rgm": rgm,
                "lead_id": matched_lead_id,
                "lead_updates": lead_updates,
                "lead_field_updates": lead_field_updates,
                "biz_updates": biz_updates,
                "_diff": list(_diff_details),
            })
            _diff_details.clear()

            if len(updates) <= 5:
                log.info("  DIFF %s (RGM %s):", xl_data["nome"], rgm)
                for d in updates[-1]["_diff"]:
                    log.info("    %s", d)
                if lead_updates:
                    log.info("    lead: %s", lead_updates)
        else:
            _diff_details.clear()

    if _format_samples:
        log.info("  Amostras de diferenças por campo:")
        for fname, samples in sorted(_format_samples.items()):
            log.info("    %s:", fname)
            for crm_v, xl_v in samples:
                log.info("      CRM: '%s'  →  Planilha: '%s'", crm_v, xl_v)

    if _skipped_nivel:
        REPORTS_DIR.mkdir(exist_ok=True)
        skip_path = REPORTS_DIR / "ignorados_nivel.csv"
        with open(skip_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Nome", "RGM", "CPF", "Negocio"])
            for s in _skipped_nivel:
                w.writerow([s["nome"], s["rgm"], s["cpf"], s["negocio"]])
        log.info("Ignorados por nível (não Graduação/Pós): %d → %s", len(_skipped_nivel), skip_path)

    if unmatched:
        REPORTS_DIR.mkdir(exist_ok=True)
        sm_path = REPORTS_DIR / "sem_match.csv"
        with open(sm_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Nome", "RGM", "CPF", "Telefone", "Email", "Curso", "Polo",
                         "Situacao", "TipoAluno", "Nivel", "Motivo"])
            for u in unmatched:
                d = u["xl_data"]
                w.writerow([d["nome"], d["rgm"], d["cpf"], d["phone_raw"], d["email"],
                            d["curso"], d["polo"], d["situacao"], d["tipo"], d["nivel"],
                            u["reason"]])
        log.info("Sem match (a criar): %d → %s", len(unmatched), sm_path)

    return updates, unmatched


# ---------------------------------------------------------------------------
# Execução
# ---------------------------------------------------------------------------

def field_id_to_name(fid):
    for name, fid2 in FIELD_IDS.items():
        if fid == fid2:
            return name
    return fid[:8]


def test_one_update(api, updates):
    """Testa 1 atualização via PUT campo a campo."""
    log.info("=== MODO TESTE: verificando 1 atualização ===")

    upd = None
    for u in updates:
        if u["biz_updates"]:
            upd = u
            break

    if not upd:
        log.error("Nenhum registro com atualização de negócio para testar.")
        return False

    biz = upd["biz_updates"][0]
    log.info("Negócio: %s", biz["biz_id"])
    log.info("  Nome: %s | RGM: %s | Match: %s", upd["xl_nome"], upd["xl_rgm"], upd["match_type"])

    first_fid = list(biz["fields"].keys())[0]
    first_val = biz["fields"][first_fid]
    fname = field_id_to_name(first_fid)

    log.info("  Testando campo: %s = %s", fname, first_val)
    log.info("  PUT /crm/crm/additional-fields/business/%s/%s", biz["biz_id"], first_fid)

    result = api.put_biz_field(biz["biz_id"], first_fid, first_val)
    log.info("  Status: %s", result["status"])

    if not result["ok"]:
        log.error("  FALHOU: %s", result["body"])
        return False

    body = result["body"]
    returned_value = body.get("value", "")
    log.info("  Retorno API: value=%s", returned_value)

    time.sleep(1.5)
    verify = api.get(f"/businesses/{biz['biz_id']}")
    if verify["ok"]:
        actual = get_biz_field(verify["body"], first_fid)
        log.info("  Verificação GET: %s = '%s'", fname, actual)
        if actual == str(first_val):
            log.info("  CONFIRMADO — campo atualizado com sucesso!")
            return True
        else:
            log.warning("  Campo retornou valor diferente: '%s' vs '%s'", actual, first_val)
            return True

    return True


def execute_updates(api, updates, limit=None):
    """Executa as atualizações — PUT campo a campo para negócios."""
    LOG_DIR.mkdir(exist_ok=True)
    ts = datetime.now(BRT).strftime("%Y%m%d_%H%M%S")
    log_file = LOG_DIR / f"update_{ts}.csv"

    conn = get_conn()

    with open(log_file, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["timestamp", "tipo", "match", "nome", "rgm", "id",
                     "campo", "valor", "status", "resultado"])

        total = len(updates)
        if limit:
            updates = updates[:limit]
            log.info("Limitado a %d de %d atualizações", limit, total)

        ok_count = 0
        err_count = 0
        start = time.monotonic()

        for i, upd in enumerate(updates, 1):
            n_fields = sum(len(b["fields"]) for b in upd["biz_updates"])
            lead_tag = "lead+" if upd["lead_updates"] else ""
            biz_field_names = ", ".join(
                field_id_to_name(fid)
                for b in upd["biz_updates"]
                for fid in b["fields"]
            )
            detail = biz_field_names if biz_field_names else ""
            lead_parts = list(upd["lead_updates"].keys()) + list(upd.get("lead_field_updates", {}).keys())
            if lead_parts:
                lead_keys = ",".join(lead_parts)
                detail = f"lead({lead_keys})+{detail}" if detail else f"lead({lead_keys})"
            log.info("[%d/%d] %s | RGM %s | %s | %s",
                     i, len(updates), upd["xl_nome"],
                     upd["xl_rgm"] or "—", detail,
                     upd["match_type"])

            if upd["lead_updates"] and upd["lead_id"]:
                payload = upd["lead_updates"]
                result = api.patch(f"/leads/{upd['lead_id']}", payload)
                status = "OK" if result["ok"] else "ERRO"
                w.writerow([
                    datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S"), "LEAD", upd["match_type"],
                    upd["xl_nome"], upd["xl_rgm"], upd["lead_id"],
                    ",".join(payload.keys()), "", result["status"], status,
                ])
                if result["ok"]:
                    ok_count += 1
                    try:
                        update_local_lead(conn, upd["lead_id"], payload)
                    except Exception as e:
                        log.warning("  CHECKPOINT FALHOU lead %s: %s", upd["lead_id"], e)
                else:
                    err_count += 1
                    log.warning("  ERRO lead %s: %s", upd["lead_id"], result["body"][:200])

            for fname, (fid, val) in upd.get("lead_field_updates", {}).items():
                result = api.put_lead_field(upd["lead_id"], fid, val)
                status = "OK" if result["ok"] else "ERRO"
                w.writerow([
                    datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S"), "LEAD_FIELD", upd["match_type"],
                    upd["xl_nome"], upd["xl_rgm"], upd["lead_id"],
                    fname, val, result["status"], status,
                ])
                if result["ok"]:
                    ok_count += 1
                    try:
                        update_local_lead_field(conn, upd["lead_id"], fid, fname, val)
                    except Exception as e:
                        log.warning("  CHECKPOINT FALHOU lead_field %s: %s", fname, e)
                else:
                    err_count += 1
                    log.warning("  ERRO lead field %s: %s", fname, result["body"][:200])

            for biz in upd["biz_updates"]:
                for fid, val in biz["fields"].items():
                    fname = field_id_to_name(fid)
                    result = api.put_biz_field(biz["biz_id"], fid, val)
                    status = "OK" if result["ok"] else "ERRO"
                    w.writerow([
                        datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S"), "BIZ_FIELD", upd["match_type"],
                        upd["xl_nome"], upd["xl_rgm"], biz["biz_id"],
                        fname, val, result["status"], status,
                    ])
                    if result["ok"]:
                        ok_count += 1
                        try:
                            update_local_biz_field(conn, biz["biz_id"], fid, val)
                        except Exception as e:
                            log.warning("  CHECKPOINT FALHOU biz %s campo %s: %s",
                                        biz["biz_id"], fname, e)
                    else:
                        err_count += 1
                        log.warning("  ERRO biz %s campo %s: %s",
                                    biz["biz_id"], fname, result["body"][:200])

            if i % 50 == 0 or i == len(updates):
                elapsed = time.monotonic() - start
                rate = api.total_calls / elapsed * 60 if elapsed > 0 else 0
                remaining = (len(updates) - i) * (elapsed / i) if i > 0 else 0
                log.info("--- %d/%d (%.0f%%) | OK: %d | Erros: %d | ~%.0f min restantes ---",
                         i, len(updates), i/len(updates)*100,
                         ok_count, err_count, remaining/60)

    conn.close()
    log.info("Concluído. OK: %d | Erros: %d | API calls: %d", ok_count, err_count, api.total_calls)
    log.info("Log detalhado: %s", log_file)
    return ok_count, err_count


def dry_run_summary(updates, unmatched=None, stage_ids=None):
    """Mostra resumo do que seria atualizado e criado."""
    unmatched = unmatched or []
    match_types = Counter(u["match_type"] for u in updates)
    lead_updates_count = sum(1 for u in updates if u["lead_updates"])
    lead_field_calls = sum(len(u.get("lead_field_updates", {})) for u in updates)
    biz_updates_count = sum(len(u["biz_updates"]) for u in updates)
    biz_field_calls = sum(
        len(biz["fields"]) for u in updates for biz in u["biz_updates"]
    )
    total_update_calls = lead_updates_count + lead_field_calls + biz_field_calls

    n_biz_fields_per_create = len(BIZ_FIELD_IDS)
    create_api_calls = len(unmatched) * (1 + 1 + n_biz_fields_per_create)
    total_api_calls = total_update_calls + create_api_calls

    base_delay = 60.0 / DEFAULT_TARGET_RATE
    estimated_minutes = total_api_calls * base_delay / 60

    print("\n" + "=" * 60)
    print("DRY-RUN — Resumo das atualizações pendentes")
    print("=" * 60)
    print(f"  Registros com match:           {len(updates):,}")
    for mt, c in match_types.most_common():
        print(f"    {mt:20s}: {c:,}")
    print()
    print(f"  Leads a atualizar (PATCH):     {lead_updates_count:,}")
    print(f"  Lead fields (PUT):             {lead_field_calls:,}")
    print(f"  Negócios a atualizar:          {biz_updates_count:,}")
    print(f"  Biz fields (PUT):              {biz_field_calls:,}")
    print(f"  Subtotal update calls:         {total_update_calls:,}")
    print()

    lead_fields = Counter()
    for u in updates:
        for k in u["lead_updates"]:
            lead_fields[k] += 1
        for k in u.get("lead_field_updates", {}):
            lead_fields[k] += 1
    if lead_fields:
        print("  Campos de lead atualizados:")
        for k, c in lead_fields.most_common():
            print(f"    {k:20s}: {c:,}")

    biz_fields = Counter()
    for u in updates:
        for biz in u["biz_updates"]:
            for fid in biz["fields"]:
                for name, fid2 in FIELD_IDS.items():
                    if fid == fid2:
                        biz_fields[name] += 1
                        break
    if biz_fields:
        print("  Campos de negócio atualizados:")
        for k, c in biz_fields.most_common():
            print(f"    {k:20s}: {c:,}")

    if unmatched:
        print()
        print("-" * 60)
        print(f"  CRIAÇÕES PENDENTES:            {len(unmatched):,}")
        reasons = Counter(u["reason"] for u in unmatched)
        for reason, c in reasons.most_common():
            print(f"    {reason:24s}: {c:,}")

        tipo_counter = Counter()
        for u in unmatched:
            tipo = u["xl_data"].get("tipo", "?") or "?"
            tipo_counter[tipo] += 1
        print("  Por tipo de aluno:")
        for tipo, c in tipo_counter.most_common():
            print(f"    {tipo:24s}: {c:,}")

        have_calouro = "calouro" in (stage_ids or {})
        have_veterano = "veterano" in (stage_ids or {})
        if not have_calouro:
            print("  ⚠ Etapa 'Calouro' não encontrada — criações NÃO executarão")
        if not have_veterano:
            print("  ⚠ Etapa 'Veterano' não encontrada — criações NÃO executarão")

        print(f"  API calls p/ criações:         ~{create_api_calls:,}")
        print(f"    (POST lead + POST biz + ~{n_biz_fields_per_create} PUT fields cada)")

    print()
    print(f"  Total de API calls:            {total_api_calls:,}")
    print(f"  Tempo estimado (~{DEFAULT_TARGET_RATE} req/min):  {estimated_minutes:.0f} min ({estimated_minutes/60:.1f}h)")
    print()
    print("  Para executar: python update_crm.py --execute")
    print("  Para testar 1: python update_crm.py --test")
    print("=" * 60)

    REPORTS_DIR.mkdir(exist_ok=True)

    preview_rows = []
    for u in updates[:200]:
        lead_chg = json.dumps(u["lead_updates"], ensure_ascii=False) if u["lead_updates"] else ""
        biz_chg = "; ".join(
            f"{biz['biz_id']}: " + ", ".join(
                f"{k}={v}" for k, v in biz["fields"].items()
            ) for biz in u["biz_updates"]
        )
        preview_rows.append([
            u["match_type"], u["xl_rgm"], u["xl_nome"],
            u["lead_id"] or "", lead_chg, biz_chg,
        ])
    preview_path = REPORTS_DIR / "update_preview.csv"
    with open(preview_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["match_tipo", "rgm", "nome", "lead_id", "lead_mudancas", "biz_mudancas"])
        w.writerows(preview_rows)
    log.info("  Preview salvo: %s (primeiros 200)", preview_path)

    if unmatched:
        create_path = REPORTS_DIR / "create_preview.csv"
        with open(create_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Nome", "RGM", "CPF", "Telefone", "Email", "Curso", "Polo",
                         "Situacao", "TipoAluno", "Nivel", "Motivo"])
            for u in unmatched:
                d = u["xl_data"]
                w.writerow([d["nome"], d["rgm"], d["cpf"], d["phone_raw"], d["email"],
                            d["curso"], d["polo"], d["situacao"], d["tipo"], d["nivel"],
                            u["reason"]])
        log.info("  Preview criações: %s (%d registros)", create_path, len(unmatched))


# ---------------------------------------------------------------------------
# Criação de leads + negócios (registros sem match)
# ---------------------------------------------------------------------------

def _build_lead_payload(xl_data):
    """Constrói payload para POST /leads a partir de dados da planilha."""
    payload = {}
    if xl_data.get("nome"):
        payload["name"] = xl_data["nome"]
    cpf = xl_data.get("cpf", "")
    if cpf:
        payload["taxId"] = format_cpf(cpf)
    phone = xl_data.get("phone_raw", "")
    if phone:
        digits = "".join(c for c in str(phone) if c.isdigit())
        if len(digits) >= 10:
            payload["phone"] = digits
    email = xl_data.get("email", "")
    if email:
        payload["email"] = email
    bairro = xl_data.get("bairro", "")
    cidade = xl_data.get("cidade", "")
    endereco = xl_data.get("endereco", "")
    addr = {}
    if endereco:
        addr["address"] = endereco
    if bairro:
        addr["block"] = bairro
    if cidade:
        addr["city"] = cidade
    if addr:
        payload["address"] = addr
    if not payload.get("name"):
        payload["name"] = f"RGM {xl_data.get('rgm', '?')}"
    payload["source"] = "Planilha Matriculados"
    return payload


def _build_biz_field_list(xl_data, turmas=None, ciclos=None):
    """Retorna lista de (field_id, value) para popular os campos do negócio."""
    fields = []
    mapping = {
        "RGM": xl_data.get("rgm", ""),
        "Curso": xl_data.get("curso", ""),
        "Polo": xl_data.get("polo", ""),
        "Serie": xl_data.get("serie", ""),
        "Situacao": xl_data.get("situacao", ""),
        "DataMatricula": xl_data.get("data_matricula", ""),
        "EmailAD": xl_data.get("email_acad", ""),
        "TipoAluno": xl_data.get("tipo", ""),
        "Nivel": xl_data.get("nivel", ""),
    }

    senha = generate_senha(xl_data.get("nome", ""), xl_data.get("rgm", ""), xl_data.get("cpf", ""))
    if senha:
        mapping["SenhaProvisoria"] = senha

    turma_nome = _classify_turma(xl_data.get("data_matricula", ""), xl_data.get("nivel", ""), turmas or [])
    ciclo_nome = _classify_ciclo(xl_data.get("data_matricula", ""), xl_data.get("nivel", ""), ciclos or [])
    if turma_nome:
        mapping["Turma"] = turma_nome
    if ciclo_nome:
        mapping["Ciclo"] = ciclo_nome

    for field_name, value in mapping.items():
        if not value:
            continue
        fid = BIZ_FIELD_IDS.get(field_name, "")
        if fid:
            fields.append((fid, str(value).strip()))
    return fields


def _find_lead_by_contact_local(conn, cpf=None, phone=None, email=None):
    """Busca lead no banco local por CPF, telefone ou e-mail."""
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if cpf:
        cpf_clean = clean_cpf(cpf)
        if cpf_clean:
            cur.execute(
                "SELECT id FROM leads WHERE REPLACE(REPLACE(COALESCE(data->>'taxId',''), '.', ''), '-', '') = %s LIMIT 1",
                (cpf_clean,),
            )
            row = cur.fetchone()
            if row:
                cur.close()
                return row["id"]
    if phone:
        digits = "".join(c for c in str(phone) if c.isdigit())
        if len(digits) >= 10:
            cur.execute(
                "SELECT id FROM leads WHERE regexp_replace(data->>'phone', '[^0-9]', '', 'g') LIKE %s LIMIT 1",
                (f"%{digits[-10:]}",),
            )
            row = cur.fetchone()
            if row:
                cur.close()
                return row["id"]
    if email:
        cur.execute(
            "SELECT id FROM leads WHERE lower(data->>'email') = lower(%s) LIMIT 1",
            (email.strip(),),
        )
        row = cur.fetchone()
        if row:
            cur.close()
            return row["id"]
    cur.close()
    return None


def _resolve_target_stage(tipo_aluno, stage_ids):
    """Determina a etapa (Calouro/Veterano) com base no tipo de aluno."""
    tipo_lower = (tipo_aluno or "").strip().lower()
    if tipo_lower in CALOURO_TIPOS:
        return stage_ids.get("calouro")
    if tipo_lower in VETERANO_TIPOS:
        return stage_ids.get("veterano")
    return stage_ids.get("calouro")


def execute_creations(api, unmatched, stage_ids, turmas=None, ciclos=None):
    """Cria leads e negócios para registros sem match no CRM."""
    if not stage_ids:
        log.error("Nenhuma etapa encontrada (Calouro/Veterano). Abortando criações.")
        return 0, 0

    LOG_DIR.mkdir(exist_ok=True)
    ts = datetime.now(BRT).strftime("%Y%m%d_%H%M%S")
    log_file = LOG_DIR / f"create_{ts}.csv"

    conn = get_conn()
    ok_count = 0
    err_count = 0
    start = time.monotonic()

    with open(log_file, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["timestamp", "tipo", "nome", "rgm", "id", "status", "detalhe"])

        for i, item in enumerate(unmatched, 1):
            xl = item["xl_data"]
            rgm = xl.get("rgm", "")
            nome = xl.get("nome", "?")
            reason = item["reason"]

            target_stage_id = _resolve_target_stage(xl.get("tipo", ""), stage_ids)
            if not target_stage_id:
                log.warning("  [%d/%d] %s RGM %s — sem etapa destino, pulando",
                            i, len(unmatched), nome, rgm)
                err_count += 1
                continue

            lead_id = item.get("lead_id")

            if reason in ("LEAD_SEM_BIZ_RGM", "LEAD_SEM_BIZ"):
                log.info("  [%d/%d] %s | RGM %s | Lead existe → criando negócio | %s",
                         i, len(unmatched), nome, rgm, reason)
            else:
                log.info("  [%d/%d] %s | RGM %s | Criando lead + negócio | %s",
                         i, len(unmatched), nome, rgm, reason)

                lead_payload = _build_lead_payload(xl)
                r_lead = api.create_lead(lead_payload)

                if r_lead["ok"]:
                    body = r_lead["body"]
                    lead_id = body.get("id") if isinstance(body, dict) else None
                    if lead_id:
                        log.info("    Lead criado: %s", lead_id[:12])
                        w.writerow([datetime.now(BRT).strftime("%H:%M:%S"), "LEAD_OK",
                                    nome, rgm, lead_id, "OK", ""])

                if not lead_id:
                    err_body = str(r_lead.get("body", ""))
                    if "lead-with-same-contact-exists" in err_body or "same-contact" in err_body.lower():
                        log.info("    Lead já existe — buscando...")
                        lead_id = _find_lead_by_contact_local(
                            conn, cpf=xl.get("cpf"), phone=xl.get("phone_raw"), email=xl.get("email")
                        )
                        if not lead_id:
                            search_term = xl.get("email") or xl.get("phone_raw", "")
                            if search_term:
                                r_search = api.search_leads(search_term)
                                if r_search["ok"] and isinstance(r_search["body"], dict):
                                    items = r_search["body"].get("data", [])
                                    if items:
                                        lead_id = items[0].get("id")
                        if lead_id:
                            log.info("    Lead existente: %s", lead_id[:12])
                        else:
                            log.warning("    Lead não encontrado para RGM %s — pulando", rgm)
                            err_count += 1
                            w.writerow([datetime.now(BRT).strftime("%H:%M:%S"), "LEAD_SKIP",
                                        nome, rgm, "", "SKIP", "lead existe mas não localizado"])
                            continue
                    else:
                        log.warning("    ERRO criar lead RGM %s: %s", rgm, err_body[:200])
                        err_count += 1
                        w.writerow([datetime.now(BRT).strftime("%H:%M:%S"), "LEAD_ERRO",
                                    nome, rgm, "", "ERRO", err_body[:100]])
                        continue

            r_biz = api.create_business(lead_id, target_stage_id)
            if not r_biz["ok"]:
                log.warning("    ERRO criar negócio RGM %s: %s", rgm, str(r_biz["body"])[:200])
                err_count += 1
                w.writerow([datetime.now(BRT).strftime("%H:%M:%S"), "BIZ_ERRO",
                            nome, rgm, lead_id or "", "ERRO", str(r_biz["body"])[:100]])
                continue

            biz_body = r_biz["body"]
            biz_id = biz_body.get("id") if isinstance(biz_body, dict) else None
            if not biz_id:
                log.warning("    Resposta sem ID de negócio para RGM %s", rgm)
                err_count += 1
                continue

            log.info("    Negócio criado: %s → %s",
                     biz_id[:12], xl.get("tipo", "?"))

            biz_fields = _build_biz_field_list(xl, turmas=turmas, ciclos=ciclos)
            field_ok = 0
            for fid, value in biz_fields:
                rf = api.put_biz_field(biz_id, fid, value)
                if rf["ok"]:
                    field_ok += 1
                else:
                    log.warning("    ERRO campo %s RGM %s: %s",
                                field_id_to_name(fid), rgm, str(rf["body"])[:100])

            lead_field_list = []
            sexo = xl.get("sexo", "")
            if sexo:
                sexo_fid = LEAD_FIELD_IDS.get("Sexo")
                if sexo_fid:
                    lead_field_list.append((sexo_fid, sexo))
            for lfid, lval in lead_field_list:
                api.put_lead_field(lead_id, lfid, lval)

            ok_count += 1
            w.writerow([datetime.now(BRT).strftime("%H:%M:%S"), "CREATE_OK",
                        nome, rgm, biz_id, "OK", f"fields={field_ok}/{len(biz_fields)}"])

            if i % 25 == 0 or i == len(unmatched):
                elapsed = time.monotonic() - start
                rate = api.total_calls / elapsed * 60 if elapsed > 0 else 0
                remaining = (len(unmatched) - i) * (elapsed / i) if i > 0 else 0
                log.info("--- Criações %d/%d (%.0f%%) | OK: %d | Erros: %d | ~%.0f min restantes ---",
                         i, len(unmatched), i / len(unmatched) * 100,
                         ok_count, err_count, remaining / 60)

    conn.close()
    log.info("Criações concluídas. OK: %d | Erros: %d | Log: %s", ok_count, err_count, log_file)
    return ok_count, err_count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global SKIP_ADDRESS
    mode = "--dry-run"
    limit = None
    rate = None
    skip_create = False

    for arg in sys.argv[1:]:
        if arg in ("--test", "--dry-run", "--execute"):
            mode = arg
        if arg == "--with-address":
            SKIP_ADDRESS = False
        if arg == "--no-create":
            skip_create = True

    for i, arg in enumerate(sys.argv):
        if arg == "--limit" and i + 1 < len(sys.argv):
            limit = int(sys.argv[i + 1])
        if arg == "--rate" and i + 1 < len(sys.argv):
            rate = int(sys.argv[i + 1])

    log.info("=" * 50)
    log.info("Atualização CRM — modo: %s%s", mode.upper(),
             "" if SKIP_ADDRESS else " (com address)")
    log.info("=" * 50)

    xl_rows, col = load_excel()

    conn = get_conn()
    try:
        crm_by_rgm, crm_by_cpf, crm_by_phone, crm_by_name, leads_by_id, biz_by_lead = load_crm_data(conn)
        turmas = _load_turmas(conn)
        ciclos = _load_ciclos(conn)
    finally:
        conn.close()

    conn = get_conn()
    stages = load_pipeline_stages(conn)
    stage_ids = resolve_stage_ids(stages)
    conn.close()

    log.info("Preparando atualizações...")
    updates, unmatched = prepare_updates(
        xl_rows, col, crm_by_rgm,
        crm_by_cpf, crm_by_phone, crm_by_name, leads_by_id, biz_by_lead,
        turmas=turmas, ciclos=ciclos,
    )
    log.info("  %d registros com atualizações pendentes", len(updates))
    if unmatched:
        log.info("  %d registros sem match (a criar)", len(unmatched))

    if mode == "--dry-run":
        dry_run_summary(updates, unmatched, stage_ids)
        return

    api = ApiClient(target_rate=rate)

    if mode == "--test":
        result = test_one_update(api, updates)
        if result:
            log.info("Teste OK! Formato aceito: %s", result)
            log.info("Execute com: python update_crm.py --execute")
        else:
            log.error("Teste FALHOU. Verifique os logs acima.")
        return

    if mode == "--execute":
        log.info("Iniciando atualização em massa...")
        ok, err = execute_updates(api, updates, limit)
        print(f"\nResultado: {ok} OK, {err} erros de {len(updates)} registros.")

        if unmatched and not limit and not skip_create:
            print()
            log.info("=" * 50)
            log.info("CRIAÇÃO — %d registros sem match no CRM", len(unmatched))
            log.info("=" * 50)
            c_ok, c_err = execute_creations(api, unmatched, stage_ids, turmas, ciclos)
            print(f"\nCriações: {c_ok} OK, {c_err} erros de {len(unmatched)} registros.")
        elif unmatched and skip_create:
            log.info("Criações puladas (--no-create). %d registros sem match → sem_match.csv", len(unmatched))


if __name__ == "__main__":
    main()
