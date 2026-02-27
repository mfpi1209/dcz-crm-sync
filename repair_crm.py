"""
DataCrazy CRM — Reparo de RGMs incorretos e relatório de leads afetados.

O update_crm.py antigo espalhava o RGM de um aluno para negócios de leads
que matchavam por telefone/nome (match fraco). Este script:
  1. Carrega a planilha de matriculados para obter o CPF correto de cada RGM
  2. Encontra negócios com RGM atribuído ao lead errado (CPF não bate)
  3. Limpa o RGM desses negócios via API
  4. Gera relatório de leads potencialmente renomeados

Uso:
    python repair_crm.py --dry-run       # Analisa sem alterar nada (padrão)
    python repair_crm.py --execute       # Executa a limpeza via API
"""

import sys
import io
import os
import csv
import json
import time
import logging
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import requests
import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

load_dotenv(Path(__file__).parent / ".env")

API_BASE = "https://api.g1.datacrazy.io/api/v1"
API_TOKEN = os.getenv("DATACRAZY_API_TOKEN", "")

DB_DSN = dict(
    host=os.getenv("DB_HOST", "localhost"),
    port=os.getenv("DB_PORT", "5432"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASS"),
    dbname=os.getenv("DB_NAME", "dcz_sync"),
)

REPORTS_DIR = Path(__file__).parent / "reports"
LOG_DIR = Path(__file__).parent / "logs"

FIELD_RGM = "2ac4e30f-cfd7-435f-b688-fbce27f76c38"

MIN_REQUEST_DELAY = 1.05
RATE_LIMIT_BUFFER = 5

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("repair_crm")


# ---------------------------------------------------------------------------
# API Client (same as update_crm.py)
# ---------------------------------------------------------------------------

class ApiClient:
    def __init__(self):
        self.s = requests.Session()
        self.s.headers["Authorization"] = f"Bearer {API_TOKEN}"
        self.s.headers["Content-Type"] = "application/json"
        self._remaining = 60
        self._reset = 0
        self._last_req = 0.0
        self.total_calls = 0

    def _throttle(self):
        elapsed = time.monotonic() - self._last_req
        if elapsed < MIN_REQUEST_DELAY:
            time.sleep(MIN_REQUEST_DELAY - elapsed)
        if self._remaining <= RATE_LIMIT_BUFFER and self._reset > 0:
            wait = self._reset + 1
            log.warning("Rate-limit próximo (%d restantes) — pausando %ds", self._remaining, wait)
            time.sleep(wait)

    def _read_headers(self, r):
        self._remaining = int(r.headers.get("X-RateLimit-Remaining", self._remaining))
        self._reset = int(r.headers.get("X-RateLimit-Reset", 0))

    def put(self, path, payload):
        for attempt in range(4):
            self._throttle()
            self._last_req = time.monotonic()
            self.total_calls += 1
            r = self.s.put(f"{API_BASE}{path}", json=payload, timeout=30)
            if r.status_code == 429:
                retry = int(r.headers.get("Retry-After", 30))
                log.warning("429 — Retry-After %ds (tentativa %d/4)", retry, attempt + 1)
                time.sleep(retry + 1)
                continue
            self._read_headers(r)
            if r.status_code >= 400:
                return {"ok": False, "status": r.status_code, "body": r.text[:500]}
            return {"ok": True, "status": r.status_code, "body": r.json()}
        return {"ok": False, "status": 429, "body": "Falha após 4 tentativas"}

    def clear_biz_field(self, biz_id, field_id):
        path = f"/crm/crm/additional-fields/business/{biz_id}/{field_id}"
        return self.put(path, {"value": ""})


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_cpf(cpf):
    if not cpf:
        return ""
    return str(cpf).replace(".", "").replace("-", "").replace(" ", "").strip()


def get_conn():
    return psycopg2.connect(**DB_DSN)


def get_biz_field(biz_data, field_id):
    for f in biz_data.get("additionalFields", []):
        af = f.get("additionalField", {})
        if isinstance(af, dict) and af.get("id") == field_id:
            return f.get("value", "")
    return ""


def update_local_biz_field(conn, biz_id, field_id, new_value):
    with conn.cursor() as cur:
        cur.execute("SELECT data FROM businesses WHERE id = %s", (biz_id,))
        row = cur.fetchone()
        if not row:
            return
        data = row[0]
        for f in data.get("additionalFields", []):
            af = f.get("additionalField", {})
            fid = af.get("id") if isinstance(af, dict) else af
            if fid == field_id:
                f["value"] = str(new_value)
                break
        cur.execute(
            "UPDATE businesses SET data = %s::jsonb WHERE id = %s",
            (json.dumps(data), biz_id),
        )
        conn.commit()


# ---------------------------------------------------------------------------
# Load Excel — get RGM→CPF mapping
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    "Nome": ["Nome"], "CPF": ["CPF"], "RGM": ["RGM"],
}


def _col_find(col_map, *candidates):
    for c in candidates:
        if c in col_map:
            return c
    for c in candidates:
        cl = c.lower()
        for k in col_map:
            if k and cl in k.lower():
                return k
    return candidates[0]


def load_excel_rgm_cpf():
    """Returns dict mapping RGM → set of clean CPFs from the Excel."""
    log.info("Carregando planilha...")
    xlsx = None
    for f in Path(__file__).parent.iterdir():
        if f.suffix.lower() == ".xlsx" and "matriculados" in f.name.lower():
            xlsx = f
            break
    if not xlsx:
        raise FileNotFoundError("Planilha de matriculados não encontrada")

    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    ws = wb["Export"]
    raw_header = [cell.value for cell in ws[1]]
    raw_col = {h: i for i, h in enumerate(raw_header) if h}

    col = {}
    for norm, aliases in COLUMN_ALIASES.items():
        found = _col_find(raw_col, *aliases)
        if found in raw_col:
            col[norm] = raw_col[found]

    rgm_to_cpfs = defaultdict(set)
    rgm_to_nome = {}
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[1] is None:
            continue
        rgm = str(row[col["RGM"]]).strip() if row[col.get("RGM", -1)] else ""
        cpf = clean_cpf(row[col["CPF"]]) if "CPF" in col and row[col["CPF"]] else ""
        nome = str(row[col["Nome"]] or "").strip() if "Nome" in col else ""
        if rgm and cpf:
            rgm_to_cpfs[rgm].add(cpf)
            if rgm not in rgm_to_nome:
                rgm_to_nome[rgm] = nome
        count += 1

    wb.close()
    log.info("  %d linhas, %d RGMs com CPF", count, len(rgm_to_cpfs))
    return rgm_to_cpfs, rgm_to_nome


# ---------------------------------------------------------------------------
# Find damaged businesses
# ---------------------------------------------------------------------------

def find_damaged_businesses(conn, rgm_to_cpfs, rgm_to_nome):
    """
    For each business with an RGM, check if the linked lead's CPF matches
    any CPF associated with that RGM in the Excel.
    Returns list of dicts describing businesses to repair.
    """
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    log.info("Carregando todos os negócios com RGM...")
    cur.execute("""
        SELECT b.id AS biz_id,
               b.data->>'leadId' AS lead_id,
               b.data->>'code' AS biz_code,
               l.data->>'name' AS lead_nome,
               REPLACE(REPLACE(COALESCE(l.data->>'taxId',''), '.', ''), '-', '') AS lead_cpf,
               l.data->>'rawPhone' AS lead_tel,
               elem->>'value' AS rgm
        FROM businesses b
        LEFT JOIN leads l ON l.id = b.data->>'leadId'
        CROSS JOIN LATERAL jsonb_array_elements(b.data->'additionalFields') elem
        WHERE elem->'additionalField'->>'id' = %s
          AND elem->>'value' IS NOT NULL
          AND elem->>'value' != ''
    """, (FIELD_RGM,))

    all_biz_with_rgm = cur.fetchall()
    cur.close()

    log.info("  %d negócios com RGM no sistema", len(all_biz_with_rgm))

    # Group by RGM
    by_rgm = defaultdict(list)
    for row in all_biz_with_rgm:
        by_rgm[row["rgm"]].append(row)

    damaged = []
    correct = []
    unknown = []

    for rgm, biz_list in by_rgm.items():
        if len(biz_list) <= 1:
            continue

        excel_cpfs = rgm_to_cpfs.get(rgm)
        if not excel_cpfs:
            unknown.extend(biz_list)
            continue

        for biz in biz_list:
            lead_cpf = biz["lead_cpf"].strip() if biz["lead_cpf"] else ""
            if lead_cpf and lead_cpf in excel_cpfs:
                correct.append(biz)
            elif lead_cpf and lead_cpf not in excel_cpfs:
                damaged.append({
                    "biz_id": biz["biz_id"],
                    "biz_code": biz["biz_code"],
                    "lead_id": biz["lead_id"],
                    "lead_nome": biz["lead_nome"],
                    "lead_cpf": lead_cpf,
                    "lead_tel": biz["lead_tel"],
                    "rgm": rgm,
                    "aluno_nome": rgm_to_nome.get(rgm, "?"),
                    "aluno_cpfs": ",".join(excel_cpfs),
                })
            else:
                unknown.append(biz)

    log.info("  Corretos (CPF bate): %d", len(correct))
    log.info("  DANIFICADOS (CPF não bate): %d", len(damaged))
    log.info("  Incertos (sem CPF no lead): %d", len(unknown))

    return damaged


# ---------------------------------------------------------------------------
# Generate reports
# ---------------------------------------------------------------------------

def generate_reports(damaged):
    REPORTS_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Report 1: Businesses to repair
    biz_report = REPORTS_DIR / f"repair_businesses_{ts}.csv"
    with open(biz_report, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["biz_id", "biz_code", "rgm_incorreto", "lead_id",
                     "lead_nome", "lead_cpf", "aluno_nome", "aluno_cpfs"])
        for d in damaged:
            w.writerow([
                d["biz_id"], d["biz_code"], d["rgm"],
                d["lead_id"], d["lead_nome"], d["lead_cpf"],
                d["aluno_nome"], d["aluno_cpfs"],
            ])
    log.info("Relatório de negócios: %s (%d registros)", biz_report, len(damaged))

    # Report 2: Leads potentially renamed (unique leads from damaged list)
    seen_leads = {}
    for d in damaged:
        lid = d["lead_id"]
        if lid not in seen_leads:
            seen_leads[lid] = d

    leads_report = REPORTS_DIR / f"repair_leads_renamed_{ts}.csv"
    with open(leads_report, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["lead_id", "lead_nome_atual", "lead_cpf", "lead_tel",
                     "rgm_incorreto", "aluno_nome_escrito"])
        for lid, d in seen_leads.items():
            w.writerow([
                d["lead_id"], d["lead_nome"], d["lead_cpf"], d["lead_tel"],
                d["rgm"], d["aluno_nome"],
            ])
    log.info("Relatório de leads afetados: %s (%d leads)", leads_report, len(seen_leads))

    return biz_report, leads_report


# ---------------------------------------------------------------------------
# Execute repair
# ---------------------------------------------------------------------------

def execute_repair(api, damaged, conn):
    LOG_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = LOG_DIR / f"repair_{ts}.csv"

    ok_count = 0
    err_count = 0
    start = time.monotonic()

    with open(log_file, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["timestamp", "biz_id", "biz_code", "rgm_removido",
                     "lead_id", "lead_nome", "status", "resultado"])

        for i, d in enumerate(damaged, 1):
            log.info("[%d/%d] Limpando RGM %s do negócio %s (lead: %s)",
                     i, len(damaged), d["rgm"], d["biz_code"], d["lead_nome"])

            result = api.clear_biz_field(d["biz_id"], FIELD_RGM)
            status = "OK" if result["ok"] else "ERRO"

            w.writerow([
                datetime.now().isoformat(), d["biz_id"], d["biz_code"],
                d["rgm"], d["lead_id"], d["lead_nome"],
                result["status"], status,
            ])

            if result["ok"]:
                ok_count += 1
                try:
                    update_local_biz_field(conn, d["biz_id"], FIELD_RGM, "")
                except Exception:
                    pass
            else:
                err_count += 1
                log.warning("  ERRO: %s", result["body"][:200])

            if i % 50 == 0 or i == len(damaged):
                elapsed = time.monotonic() - start
                remaining = (len(damaged) - i) * (elapsed / i) if i > 0 else 0
                log.info("--- %d/%d (%.0f%%) | OK: %d | Erros: %d | ~%.0f min restantes ---",
                         i, len(damaged), i / len(damaged) * 100,
                         ok_count, err_count, remaining / 60)

    log.info("Reparo concluído. OK: %d | Erros: %d | API calls: %d",
             ok_count, err_count, api.total_calls)
    log.info("Log detalhado: %s", log_file)
    return ok_count, err_count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    mode = "--dry-run"
    for arg in sys.argv[1:]:
        if arg in ("--dry-run", "--execute"):
            mode = arg

    log.info("=" * 50)
    log.info("Reparo CRM — modo: %s", mode.upper())
    log.info("=" * 50)

    rgm_to_cpfs, rgm_to_nome = load_excel_rgm_cpf()

    conn = get_conn()
    try:
        damaged = find_damaged_businesses(conn, rgm_to_cpfs, rgm_to_nome)
    finally:
        conn.close()

    if not damaged:
        log.info("Nenhum negócio danificado encontrado!")
        return

    biz_report, leads_report = generate_reports(damaged)

    estimated_min = len(damaged) * MIN_REQUEST_DELAY / 60

    print(f"\n{'=' * 60}")
    print(f"REPARO — Resumo")
    print(f"{'=' * 60}")
    print(f"  Negócios com RGM incorreto:  {len(damaged)}")
    print(f"  Leads únicos afetados:       {len(set(d['lead_id'] for d in damaged))}")
    print(f"  API calls necessárias:       {len(damaged)}")
    print(f"  Tempo estimado (~1 req/s):   {estimated_min:.0f} min")
    print(f"\n  Relatório negócios: {biz_report}")
    print(f"  Relatório leads:    {leads_report}")
    print(f"{'=' * 60}")

    if mode == "--dry-run":
        print("\n  Modo DRY-RUN. Para executar: python repair_crm.py --execute\n")
        return

    if mode == "--execute":
        log.info("Iniciando reparo...")
        conn = get_conn()
        try:
            api = ApiClient()
            ok, err = execute_repair(api, damaged, conn)
            print(f"\nResultado: {ok} OK, {err} erros de {len(damaged)} negócios.")
        finally:
            conn.close()


if __name__ == "__main__":
    main()
