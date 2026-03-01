"""
eduit. — Gestão Acadêmica (Flask).

Uso:
    python app.py
    Acesse http://localhost:5001
"""

import os
import sys
import json
import hashlib
import subprocess
import threading
import time
import re
from datetime import datetime, timezone, timedelta

BRT = timezone(timedelta(hours=-3))
from pathlib import Path
from collections import deque

from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, session, send_file,
)
import psycopg2
import psycopg2.extras
import requests as _requests
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")


def to_brt(dt):
    """Convert a datetime to BRT (UTC-3) string."""
    if dt is None:
        return None
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(BRT).strftime("%d/%m/%Y %H:%M:%S")
    return str(dt)


app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dcz-sync-default-key-change-me")

DB_DSN = dict(
    host=os.getenv("DB_HOST", "localhost"),
    port=os.getenv("DB_PORT", "5432"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASS"),
    dbname=os.getenv("DB_NAME", "dcz_sync"),
)

# ---------------------------------------------------------------------------
# Autenticação por sessão (banco de dados)
# ---------------------------------------------------------------------------

ALL_PAGES = [
    "dashboard", "search", "sync", "update", "pipeline",
    "logs", "distribuicao", "intelligence", "feedback", "config", "schedule",
]

APP_USER_FALLBACK = os.getenv("APP_USER", "admin")
APP_PASS_FALLBACK = os.getenv("APP_PASS", "")


def _hash_pw(password):
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def _db_auth(username, password):
    """Authenticate against app_users table. Returns dict or None."""
    try:
        conn = psycopg2.connect(**DB_DSN)
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, username, pw_hash, role FROM app_users WHERE username = %s",
                        (username,))
            row = cur.fetchone()
        conn.close()
        if row and row["pw_hash"] == _hash_pw(password):
            return dict(row)
    except Exception:
        pass
    return None


def _get_user_permissions(user_id):
    """Returns list of page slugs the user can access."""
    try:
        conn = psycopg2.connect(**DB_DSN)
        with conn.cursor() as cur:
            cur.execute("SELECT page FROM user_permissions WHERE user_id = %s", (user_id,))
            pages = [r[0] for r in cur.fetchall()]
        conn.close()
        return pages
    except Exception:
        return []


@app.before_request
def require_auth():
    if request.path in ("/login",):
        return
    if request.path.startswith("/static/"):
        return
    if not session.get("authenticated"):
        if request.path.startswith("/api/"):
            return jsonify({"error": "Não autenticado"}), 401
        return redirect(url_for("login"))
    if "role" not in session:
        uid = session.get("user_id")
        if uid and uid != 0:
            try:
                conn = get_conn()
                with conn.cursor() as cur:
                    cur.execute("SELECT role FROM app_users WHERE id = %s", (uid,))
                    row = cur.fetchone()
                conn.close()
                if row:
                    session["role"] = row[0]
                    return
            except Exception:
                pass
        session.clear()
        if request.path.startswith("/api/"):
            return jsonify({"error": "Sessão expirada, faça login novamente"}), 401
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        user = request.form.get("username", "")
        pwd = request.form.get("password", "")
        db_user = _db_auth(user, pwd)
        if db_user:
            session["authenticated"] = True
            session["user_id"] = db_user["id"]
            session["username"] = db_user["username"]
            session["role"] = db_user["role"]
            return redirect(url_for("index"))
        if APP_PASS_FALLBACK and user == APP_USER_FALLBACK and pwd == APP_PASS_FALLBACK:
            session["authenticated"] = True
            session["user_id"] = 0
            session["username"] = APP_USER_FALLBACK
            session["role"] = "admin"
            return redirect(url_for("index"))
        error = "Usuário ou senha incorretos."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/api/me")
def api_me():
    """Returns current user info + permissions for sidebar rendering."""
    uid = session.get("user_id", 0)
    role = session.get("role", "admin")
    if role == "admin":
        pages = list(ALL_PAGES)
    else:
        pages = _get_user_permissions(uid)
    return jsonify({
        "user_id": uid,
        "username": session.get("username", ""),
        "role": role,
        "pages": pages,
    })

BASE_DIR = Path(__file__).parent
SYNC_SCRIPT = str(BASE_DIR / "sync.py")
UPDATE_SCRIPT = str(BASE_DIR / "update_crm.py")
SANITIZE_SCRIPT = str(BASE_DIR / "sanitize_crm.py")
PIPELINE_SCRIPT = str(BASE_DIR / "pipeline_crm.py")
ENRICH_SCRIPT = str(BASE_DIR / "enrich_crosslead.py")
MERGE_SCRIPT = str(BASE_DIR / "merge_leads.py")
INADIMPLENTES_SCRIPT = str(BASE_DIR / "update_inadimplentes.py")
CONCLUINTES_SCRIPT = str(BASE_DIR / "update_concluintes.py")
LOG_DIR = BASE_DIR / "logs"
REPORTS_DIR = BASE_DIR / "reports"

MAX_LOG_LINES = 2000

# ---------------------------------------------------------------------------
# Estado global
# ---------------------------------------------------------------------------

_sync_running = False
_sync_proc = None
_sync_logs: deque = deque(maxlen=MAX_LOG_LINES)

_update_running = False
_update_proc = None
_update_logs: deque = deque(maxlen=MAX_LOG_LINES)

_sanitize_running = False
_sanitize_proc = None
_sanitize_logs: deque = deque(maxlen=MAX_LOG_LINES)

_pipeline_running = False
_pipeline_proc = None
_pipeline_logs: deque = deque(maxlen=MAX_LOG_LINES)

_enrich_running = False
_enrich_proc = None
_enrich_logs: deque = deque(maxlen=MAX_LOG_LINES)

_merge_running = False
_merge_proc = None
_merge_logs: deque = deque(maxlen=MAX_LOG_LINES)

_inadimplentes_running = False
_inadimplentes_proc = None
_inadimplentes_logs: deque = deque(maxlen=MAX_LOG_LINES)

_concluintes_running = False
_concluintes_proc = None
_concluintes_logs: deque = deque(maxlen=MAX_LOG_LINES)


def _add_sync_log(line: str):
    _sync_logs.append(line.rstrip())


def _add_update_log(line: str):
    _update_logs.append(line.rstrip())


def _add_sanitize_log(line: str):
    _sanitize_logs.append(line.rstrip())


def _add_pipeline_log(line: str):
    _pipeline_logs.append(line.rstrip())


def _add_enrich_log(line: str):
    _enrich_logs.append(line.rstrip())


def _add_merge_log(line: str):
    _merge_logs.append(line.rstrip())


def _add_inadimplentes_log(line: str):
    _inadimplentes_logs.append(line.rstrip())


def _add_concluintes_log(line: str):
    _concluintes_logs.append(line.rstrip())

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_conn():
    return psycopg2.connect(**DB_DSN)


FIELD_RGM = "2ac4e30f-cfd7-435f-b688-fbce27f76c38"

SEARCH_QUERY = """
SELECT
    l.id                                  AS lead_id,
    l.data->>'name'                       AS lead_nome,
    l.data->>'phone'                      AS lead_telefone,
    l.data->>'rawPhone'                   AS lead_telefone_raw,
    l.data->>'email'                      AS lead_email,
    l.data->>'source'                     AS lead_origem,
    l.data->>'taxId'                      AS lead_cpf,
    l.data->'address'->>'city'            AS lead_cidade,
    l.data->'address'->>'state'           AS lead_estado,
    l.data->>'createdAt'                  AS lead_criado_em,

    b.id                                  AS negocio_id,
    b.data->>'code'                       AS negocio_codigo,
    b.data->>'status'                     AS negocio_status,
    b.data->>'total'                      AS negocio_valor,
    b.data->>'createdAt'                  AS negocio_criado_em,
    b.data->>'lastMovedAt'                AS negocio_movido_em,

    p.data->>'name'                       AS pipeline_nome,
    ps.data->>'name'                      AS etapa_nome,
    ps.data->>'color'                     AS etapa_cor,

    b.data->'attendant'->>'name'          AS atendente,

    biz_cf.campos                         AS campos_negocio,
    lead_cf.campos                        AS campos_lead

FROM businesses b
LEFT JOIN leads l            ON l.id  = b.data->>'leadId'
LEFT JOIN pipeline_stages ps ON ps.id = b.data->>'stageId'
LEFT JOIN pipelines p        ON p.id  = ps.pipeline_id
LEFT JOIN LATERAL (
    SELECT jsonb_object_agg(
        elem->'additionalField'->>'name',
        COALESCE(elem->>'value', '')
    ) AS campos
    FROM jsonb_array_elements(b.data->'additionalFields') elem
    WHERE elem->'additionalField'->>'name' IS NOT NULL
) biz_cf ON true
LEFT JOIN LATERAL (
    SELECT jsonb_object_agg(
        elem->'additionalField'->>'name',
        COALESCE(elem->>'value', '')
    ) AS campos
    FROM jsonb_array_elements(l.data->'additionalFields') elem
    WHERE elem->'additionalField'->>'name' IS NOT NULL
) lead_cf ON true
WHERE (
    (%(cpf)s != '' AND REPLACE(REPLACE(l.data->>'taxId', '.', ''), '-', '') LIKE '%%' || REPLACE(REPLACE(%(cpf)s, '.', ''), '-', '') || '%%')
    OR (%(rgm)s != '' AND EXISTS (
        SELECT 1 FROM jsonb_array_elements(b.data->'additionalFields') e
        WHERE e->'additionalField'->>'id' = '2ac4e30f-cfd7-435f-b688-fbce27f76c38'
          AND e->>'value' LIKE '%%' || %(rgm)s || '%%'
    ))
    OR (%(telefone)s != '' AND (
        l.data->>'rawPhone' LIKE '%%' || %(telefone)s || '%%'
        OR REPLACE(REPLACE(REPLACE(REPLACE(l.data->>'phone', ' ', ''), '(', ''), ')', ''), '-', '') LIKE '%%' || %(telefone)s || '%%'
    ))
)
ORDER BY b.data->>'lastMovedAt' DESC NULLS LAST
LIMIT 50;
"""

RECENT_BIZ_UPDATES_QUERY = """
SELECT
    'negocio' AS tipo,
    b.id,
    b.data->'lead'->>'name' AS nome_lead,
    b.data->>'status' AS status,
    p.data->>'name' AS pipeline,
    ps.data->>'name' AS etapa,
    b.synced_at
FROM businesses b
LEFT JOIN pipeline_stages ps ON ps.id = b.data->>'stageId'
LEFT JOIN pipelines p ON p.id = ps.pipeline_id
WHERE b.synced_at = (SELECT MAX(synced_at) FROM businesses)
ORDER BY b.synced_at DESC
LIMIT 10;
"""

SYNC_STATE_QUERY = """
SELECT entity_type, last_sync_at, last_full_sync_at, run_count
FROM sync_state ORDER BY entity_type;
"""

# ---------------------------------------------------------------------------
# Rotas — Páginas
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    resp = app.make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp


# ---------------------------------------------------------------------------
# Rotas — Dashboard
# ---------------------------------------------------------------------------

@app.route("/api/dashboard")
def api_dashboard():
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT COUNT(*) AS total FROM leads")
            total_leads = cur.fetchone()["total"]

            cur.execute("SELECT COUNT(*) AS total FROM businesses")
            total_biz = cur.fetchone()["total"]

            cur.execute(SYNC_STATE_QUERY)
            states = []
            for r in cur.fetchall():
                row = dict(r)
                for k, v in row.items():
                    if isinstance(v, datetime):
                        row[k] = to_brt(v)
                states.append(row)

            cur.execute(RECENT_BIZ_UPDATES_QUERY)
            recent = []
            for r in cur.fetchall():
                row = dict(r)
                for k, v in row.items():
                    if isinstance(v, datetime):
                        row[k] = to_brt(v)
                recent.append(row)

            cur.execute("SELECT COUNT(*) AS total FROM pipelines")
            total_pipelines = cur.fetchone()["total"]

            # Schedules
            try:
                cur.execute("SELECT * FROM schedules ORDER BY created_at")
                schedules = [dict(r) for r in cur.fetchall()]
                for s in schedules:
                    for k, v in s.items():
                        if isinstance(v, datetime):
                            s[k] = to_brt(v)
            except Exception:
                schedules = []

        return jsonify({
            "total_leads": total_leads,
            "total_businesses": total_biz,
            "total_pipelines": total_pipelines,
            "sync_states": states,
            "recent_updates": recent,
            "schedules": schedules,
            "sync_running": _sync_running,
            "update_running": _update_running,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Dashboard: Métricas de Alunos
# ---------------------------------------------------------------------------

TIPO_ALUNO_FIELD = "4230e4db-970b-4444-abaf-c3135a03b79c"
DATA_MATRICULA_FIELD = "bf93a8e9-42c0-4517-8518-6f604746a300"
SITUACAO_FIELD = "fd08d44b-a4a5-4343-b7a9-37f75e2c1caa"
NIVEL_FIELD = "233fcf6f-0bed-49d7-89a1-d1cd54fb9c12"
POLO_FIELD = "0ec9d8dc-d547-4482-b9ad-d4a3e6ec1b54"
TURMA_FIELD = "8815a8de-f755-4597-b6f4-8da6d289b6eb"

_STUDENT_METRICS_QUERY = """
WITH biz_fields AS (
    SELECT
        b.id,
        MAX(CASE WHEN af->>'additionalFieldId' = %(tipo_id)s OR af->'additionalField'->>'id' = %(tipo_id)s
                 THEN af->>'value' END) AS tipo_aluno,
        MAX(CASE WHEN af->>'additionalFieldId' = %(dt_id)s   OR af->'additionalField'->>'id' = %(dt_id)s
                 THEN af->>'value' END) AS data_matricula,
        MAX(CASE WHEN af->>'additionalFieldId' = %(sit_id)s  OR af->'additionalField'->>'id' = %(sit_id)s
                 THEN af->>'value' END) AS situacao,
        MAX(CASE WHEN af->>'additionalFieldId' = %(niv_id)s  OR af->'additionalField'->>'id' = %(niv_id)s
                 THEN af->>'value' END) AS nivel,
        MAX(CASE WHEN af->>'additionalFieldId' = %(polo_id)s OR af->'additionalField'->>'id' = %(polo_id)s
                 THEN af->>'value' END) AS polo,
        MAX(CASE WHEN af->>'additionalFieldId' = %(turma_id)s OR af->'additionalField'->>'id' = %(turma_id)s
                 THEN af->>'value' END) AS turma
    FROM businesses b,
         jsonb_array_elements(b.data->'additionalFields') af
    GROUP BY b.id
)
SELECT
    COALESCE(bf.tipo_aluno, 'Não informado') AS tipo,
    bf.situacao,
    bf.nivel,
    bf.polo,
    bf.turma,
    c.nome AS ciclo,
    COUNT(*) AS total
FROM biz_fields bf
LEFT JOIN LATERAL (
    SELECT ci.nome FROM ciclos ci
    WHERE ci.nivel = bf.nivel
      AND bf.data_matricula IS NOT NULL
      AND bf.data_matricula ~ '^\\d{4}-\\d{2}-\\d{2}'
      AND bf.data_matricula::date BETWEEN ci.dt_inicio AND ci.dt_fim
    LIMIT 1
) c ON TRUE
WHERE (%(dt_from)s IS NULL OR bf.data_matricula >= %(dt_from)s)
  AND (%(dt_to)s   IS NULL OR bf.data_matricula <= %(dt_to)s)
  AND (%(f_nivel)s IS NULL OR bf.nivel = %(f_nivel)s)
  AND (%(f_sit)s   IS NULL OR bf.situacao = %(f_sit)s)
GROUP BY bf.tipo_aluno, bf.situacao, bf.nivel, bf.polo, bf.turma, c.nome
ORDER BY total DESC
"""


@app.route("/api/dashboard/students")
def api_dashboard_students():
    dt_from = request.args.get("from", "")
    dt_to = request.args.get("to", "")
    f_nivel = request.args.get("nivel", "")
    f_sit = request.args.get("situacao", "")
    f_ciclo = request.args.get("ciclo", "")
    conn = get_conn()
    try:
        if f_ciclo:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT MIN(dt_inicio) AS dt_start, MAX(dt_fim) AS dt_end "
                    "FROM ciclos WHERE nome = %s", (f_ciclo,)
                )
                crow = cur.fetchone()
                if crow and crow["dt_start"]:
                    dt_from = str(crow["dt_start"])
                    dt_to = str(crow["dt_end"])

        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(_STUDENT_METRICS_QUERY, {
                "tipo_id": TIPO_ALUNO_FIELD,
                "dt_id": DATA_MATRICULA_FIELD,
                "sit_id": SITUACAO_FIELD,
                "niv_id": NIVEL_FIELD,
                "polo_id": POLO_FIELD,
                "turma_id": TURMA_FIELD,
                "dt_from": dt_from or None,
                "dt_to": dt_to or None,
                "f_nivel": f_nivel or None,
                "f_sit": f_sit or None,
            })
            rows = cur.fetchall()

        tipo_map = {
            "Calouro": "novos",
            "Regresso (Retorno)": "regresso",
            "Calouro (Recompra)": "recompra",
            "Veterano": "rematricula",
        }

        totals = {"novos": 0, "regresso": 0, "recompra": 0, "rematricula": 0, "outros": 0}
        by_situacao = {}
        by_nivel = {}
        by_polo = {}
        by_turma = {}
        by_ciclo = {}
        by_tipo_detail = {}

        for r in rows:
            tipo = r["tipo"] or "Não informado"
            cat = tipo_map.get(tipo, "outros")
            totals[cat] += r["total"]

            sit = r["situacao"] or "N/I"
            by_situacao[sit] = by_situacao.get(sit, 0) + r["total"]

            niv = r["nivel"] or "N/I"
            by_nivel[niv] = by_nivel.get(niv, 0) + r["total"]

            polo = r["polo"] or "N/I"
            by_polo[polo] = by_polo.get(polo, 0) + r["total"]

            turma = r["turma"] or "N/I"
            by_turma[turma] = by_turma.get(turma, 0) + r["total"]

            ciclo = r["ciclo"] or "N/I"
            by_ciclo[ciclo] = by_ciclo.get(ciclo, 0) + r["total"]

            if cat not in by_tipo_detail:
                by_tipo_detail[cat] = {"by_situacao": {}, "by_nivel": {}, "by_polo": {}}
            td = by_tipo_detail[cat]
            td["by_situacao"][sit] = td["by_situacao"].get(sit, 0) + r["total"]
            td["by_nivel"][niv] = td["by_nivel"].get(niv, 0) + r["total"]
            td["by_polo"][polo] = td["by_polo"].get(polo, 0) + r["total"]

        for cat in by_tipo_detail:
            td = by_tipo_detail[cat]
            td["by_situacao"] = dict(sorted(td["by_situacao"].items(), key=lambda x: -x[1]))
            td["by_nivel"] = dict(sorted(td["by_nivel"].items(), key=lambda x: -x[1]))
            td["by_polo"] = dict(sorted(td["by_polo"].items(), key=lambda x: -x[1])[:8])

        return jsonify({
            "totals": totals,
            "by_tipo_detail": by_tipo_detail,
            "by_situacao": dict(sorted(by_situacao.items(), key=lambda x: -x[1])),
            "by_nivel": dict(sorted(by_nivel.items(), key=lambda x: -x[1])),
            "by_polo": dict(sorted(by_polo.items(), key=lambda x: -x[1])),
            "by_turma": dict(sorted(by_turma.items(), key=lambda x: -x[1])),
            "by_ciclo": dict(sorted(by_ciclo.items(), key=lambda x: -x[1])),
            "grand_total": sum(totals.values()),
            "filter": {"from": dt_from, "to": dt_to},
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Dashboard Timeline (gráficos de linha com drill-down)
# ---------------------------------------------------------------------------

_TIMELINE_QUERY = """
WITH biz_fields AS (
    SELECT
        b.id,
        MAX(CASE WHEN af->>'additionalFieldId' = %(tipo_id)s OR af->'additionalField'->>'id' = %(tipo_id)s
                 THEN af->>'value' END) AS tipo_aluno,
        MAX(CASE WHEN af->>'additionalFieldId' = %(dt_id)s   OR af->'additionalField'->>'id' = %(dt_id)s
                 THEN af->>'value' END) AS data_matricula,
        MAX(CASE WHEN af->>'additionalFieldId' = %(niv_id)s  OR af->'additionalField'->>'id' = %(niv_id)s
                 THEN af->>'value' END) AS nivel
    FROM businesses b,
         jsonb_array_elements(b.data->'additionalFields') af
    GROUP BY b.id
)
SELECT
    CASE WHEN %(granularity)s = 'month'
         THEN TO_CHAR(bf.data_matricula::date, 'YYYY-MM')
         ELSE TO_CHAR(bf.data_matricula::date, 'YYYY-MM-DD')
    END AS period,
    COALESCE(bf.tipo_aluno, 'Não informado') AS tipo,
    COUNT(*) AS total
FROM biz_fields bf
WHERE bf.data_matricula IS NOT NULL
  AND bf.data_matricula ~ '^\\d{4}-\\d{2}-\\d{2}'
  AND bf.data_matricula::date BETWEEN %(range_start)s AND %(range_end)s
  AND (%(f_nivel)s IS NULL OR bf.nivel = %(f_nivel)s)
GROUP BY period, bf.tipo_aluno
ORDER BY period, total DESC
"""


@app.route("/api/dashboard/timeline")
def api_dashboard_timeline():
    """Retorna dados de timeline agrupados por mês ou dia, para gráficos de linha."""
    from dateutil.relativedelta import relativedelta

    granularity = request.args.get("granularity", "month")
    f_nivel = request.args.get("nivel") or None
    dt_from = request.args.get("from", "")
    dt_to = request.args.get("to", "")

    today = datetime.now().date()

    if dt_from:
        range_start = datetime.strptime(dt_from, "%Y-%m-%d").date()
    else:
        range_start = today - relativedelta(months=6)

    if dt_to:
        range_end = datetime.strptime(dt_to, "%Y-%m-%d").date()
    else:
        range_end = today

    tipo_map = {
        "Calouro": "novos",
        "Regresso (Retorno)": "regresso",
        "Calouro (Recompra)": "recompra",
        "Veterano": "rematricula",
    }

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(_TIMELINE_QUERY, {
                "tipo_id": TIPO_ALUNO_FIELD,
                "dt_id": DATA_MATRICULA_FIELD,
                "niv_id": NIVEL_FIELD,
                "granularity": granularity,
                "range_start": range_start,
                "range_end": range_end,
                "f_nivel": f_nivel,
            })
            rows = cur.fetchall()

        series = {}
        all_periods = set()
        for r in rows:
            p = r["period"]
            all_periods.add(p)
            cat = tipo_map.get(r["tipo"] or "", "outros")
            if cat not in series:
                series[cat] = {}
            series[cat][p] = series[cat].get(p, 0) + r["total"]

        periods = sorted(all_periods)

        result = {
            "periods": periods,
            "series": {},
            "granularity": granularity,
            "range": {"from": str(range_start), "to": str(range_end)},
        }
        for cat in ["novos", "rematricula", "regresso", "recompra"]:
            if cat in series:
                result["series"][cat] = [series[cat].get(p, 0) for p in periods]

        total_series = [0] * len(periods)
        for cat, vals in result["series"].items():
            for i, v in enumerate(vals):
                total_series[i] += v
        result["series"]["total"] = total_series

        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Dashboard Ciclos (Master Panel)
# ---------------------------------------------------------------------------

_BIZ_FIELDS_CTE = """
WITH biz_fields AS (
    SELECT
        b.id,
        MAX(CASE WHEN af->>'additionalFieldId' = %(tipo_id)s OR af->'additionalField'->>'id' = %(tipo_id)s
                 THEN af->>'value' END) AS tipo_aluno,
        MAX(CASE WHEN af->>'additionalFieldId' = %(dt_id)s   OR af->'additionalField'->>'id' = %(dt_id)s
                 THEN af->>'value' END) AS data_matricula,
        MAX(CASE WHEN af->>'additionalFieldId' = %(sit_id)s  OR af->'additionalField'->>'id' = %(sit_id)s
                 THEN af->>'value' END) AS situacao,
        MAX(CASE WHEN af->>'additionalFieldId' = %(niv_id)s  OR af->'additionalField'->>'id' = %(niv_id)s
                 THEN af->>'value' END) AS nivel,
        MAX(CASE WHEN af->>'additionalFieldId' = %(polo_id)s OR af->'additionalField'->>'id' = %(polo_id)s
                 THEN af->>'value' END) AS polo
    FROM businesses b,
         jsonb_array_elements(b.data->'additionalFields') af
    GROUP BY b.id
)
"""

_CICLO_COMPARE_QUERY = _BIZ_FIELDS_CTE + """
SELECT
    c.nome AS ciclo, c.nivel AS ciclo_nivel,
    COALESCE(bf.tipo_aluno, 'Não informado') AS tipo,
    bf.situacao, bf.nivel, bf.polo, COUNT(*) AS total
FROM biz_fields bf
INNER JOIN ciclos c ON c.nivel = bf.nivel
    AND bf.data_matricula IS NOT NULL
    AND bf.data_matricula ~ '^\\d{4}-\\d{2}-\\d{2}'
    AND bf.data_matricula::date BETWEEN c.dt_inicio AND c.dt_fim
GROUP BY c.nome, c.nivel, bf.tipo_aluno, bf.situacao, bf.nivel, bf.polo
ORDER BY c.nome, total DESC
"""

_DATE_RANGE_QUERY = _BIZ_FIELDS_CTE + """
SELECT
    COALESCE(bf.tipo_aluno, 'Não informado') AS tipo,
    bf.situacao, bf.nivel, bf.polo, COUNT(*) AS total
FROM biz_fields bf
WHERE bf.data_matricula IS NOT NULL
  AND bf.data_matricula ~ '^\\d{4}-\\d{2}-\\d{2}'
  AND bf.data_matricula::date BETWEEN %(range_start)s AND %(range_end)s
  AND (%(f_nivel)s IS NULL OR bf.nivel = %(f_nivel)s)
GROUP BY bf.tipo_aluno, bf.situacao, bf.nivel, bf.polo
ORDER BY total DESC
"""


def _aggregate_rows(rows, tipo_map):
    result = {
        "totals": {"novos": 0, "regresso": 0, "recompra": 0, "rematricula": 0, "outros": 0},
        "by_situacao": {}, "by_polo": {}, "grand_total": 0,
    }
    for r in rows:
        tipo = r["tipo"] or "Não informado"
        cat = tipo_map.get(tipo, "outros")
        result["totals"][cat] += r["total"]
        result["grand_total"] += r["total"]
        sit = r["situacao"] or "N/I"
        result["by_situacao"][sit] = result["by_situacao"].get(sit, 0) + r["total"]
        polo = r["polo"] or "N/I"
        result["by_polo"][polo] = result["by_polo"].get(polo, 0) + r["total"]
    result["by_situacao"] = dict(sorted(result["by_situacao"].items(), key=lambda x: -x[1]))
    result["by_polo"] = dict(sorted(result["by_polo"].items(), key=lambda x: -x[1]))
    return result


@app.route("/api/dashboard/ciclos")
def api_dashboard_ciclos():
    """Retorna métricas por ciclo + comparações temporais (YTD vs ano anterior, vs 6 meses)."""
    from dateutil.relativedelta import relativedelta
    import traceback

    f_nivel = request.args.get("nivel") or None

    conn = get_conn()
    try:
        today = datetime.now().date()
        field_params = {
            "tipo_id": TIPO_ALUNO_FIELD, "dt_id": DATA_MATRICULA_FIELD,
            "sit_id": SITUACAO_FIELD, "niv_id": NIVEL_FIELD, "polo_id": POLO_FIELD,
        }
        tipo_map = {
            "Calouro": "novos", "Regresso (Retorno)": "regresso",
            "Calouro (Recompra)": "recompra", "Veterano": "rematricula",
        }

        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if f_nivel:
                cur.execute("SELECT nome, nivel, dt_inicio, dt_fim FROM ciclos WHERE nivel = %s ORDER BY dt_inicio", (f_nivel,))
            else:
                cur.execute("SELECT nome, nivel, dt_inicio, dt_fim FROM ciclos ORDER BY dt_inicio")
            ciclos_config = cur.fetchall()

            cur.execute("""
                SELECT DISTINCT bf.nivel, COUNT(*) AS total
                FROM (
                    SELECT MAX(CASE WHEN af->>'additionalFieldId' = %(niv_id)s
                                    OR af->'additionalField'->>'id' = %(niv_id)s
                                THEN af->>'value' END) AS nivel
                    FROM businesses b, jsonb_array_elements(b.data->'additionalFields') af
                    GROUP BY b.id
                ) bf
                WHERE bf.nivel IS NOT NULL
                GROUP BY bf.nivel ORDER BY total DESC
            """, {"niv_id": NIVEL_FIELD})
            distinct_nivels = {r["nivel"]: r["total"] for r in cur.fetchall()}

            cur.execute(_CICLO_COMPARE_QUERY, field_params)
            cycle_rows = cur.fetchall()
            if f_nivel:
                cycle_rows = [r for r in cycle_rows if r.get("ciclo_nivel") == f_nivel]

            ytd_start = today.replace(month=1, day=1)
            cur.execute(_DATE_RANGE_QUERY, {
                **field_params, "range_start": ytd_start, "range_end": today, "f_nivel": f_nivel,
            })
            ytd_current = cur.fetchall()

            ytd_prev_start = ytd_start.replace(year=today.year - 1)
            ytd_prev_end = today.replace(year=today.year - 1)
            cur.execute(_DATE_RANGE_QUERY, {
                **field_params, "range_start": ytd_prev_start, "range_end": ytd_prev_end, "f_nivel": f_nivel,
            })
            ytd_previous = cur.fetchall()

            m6_start = today - relativedelta(months=6)
            cur.execute(_DATE_RANGE_QUERY, {
                **field_params, "range_start": m6_start, "range_end": today, "f_nivel": f_nivel,
            })
            m6_current = cur.fetchall()

            m6_prev_start = today - relativedelta(months=12)
            m6_prev_end = today - relativedelta(months=6)
            cur.execute(_DATE_RANGE_QUERY, {
                **field_params, "range_start": m6_prev_start, "range_end": m6_prev_end, "f_nivel": f_nivel,
            })
            m6_previous = cur.fetchall()

        # Aggregate cycle data
        ciclos = {}
        for r in cycle_rows:
            cn = r["ciclo"]
            if cn not in ciclos:
                ciclos[cn] = {"nome": cn, "nivel": r["ciclo_nivel"],
                              "totals": {"novos": 0, "regresso": 0, "recompra": 0, "rematricula": 0, "outros": 0},
                              "by_situacao": {}, "by_polo": {}, "grand_total": 0}
            c = ciclos[cn]
            cat = tipo_map.get(r["tipo"] or "", "outros")
            c["totals"][cat] += r["total"]
            c["grand_total"] += r["total"]
            sit = r["situacao"] or "N/I"
            c["by_situacao"][sit] = c["by_situacao"].get(sit, 0) + r["total"]
            polo = r["polo"] or "N/I"
            c["by_polo"][polo] = c["by_polo"].get(polo, 0) + r["total"]
        for cn in ciclos:
            ciclos[cn]["by_situacao"] = dict(sorted(ciclos[cn]["by_situacao"].items(), key=lambda x: -x[1]))
            ciclos[cn]["by_polo"] = dict(sorted(ciclos[cn]["by_polo"].items(), key=lambda x: -x[1]))

        config_list = []
        for cc in ciclos_config:
            row = dict(cc)
            for k, v in row.items():
                if hasattr(v, "isoformat"):
                    row[k] = v.isoformat()
            config_list.append(row)

        return jsonify({
            "ciclos": sorted(ciclos.values(), key=lambda x: x["nome"], reverse=True),
            "config": config_list,
            "distinct_nivels": distinct_nivels,
            "comparisons": {
                "ytd": {
                    "label": f"YTD {today.year}",
                    "period": f"{ytd_start.isoformat()} → {today.isoformat()}",
                    "current": _aggregate_rows(ytd_current, tipo_map),
                },
                "ytd_prev": {
                    "label": f"YTD {today.year - 1}",
                    "period": f"{ytd_prev_start.isoformat()} → {ytd_prev_end.isoformat()}",
                    "current": _aggregate_rows(ytd_previous, tipo_map),
                },
                "m6": {
                    "label": f"Últimos 6 meses",
                    "period": f"{m6_start.isoformat()} → {today.isoformat()}",
                    "current": _aggregate_rows(m6_current, tipo_map),
                },
                "m6_prev": {
                    "label": f"6 meses anteriores",
                    "period": f"{m6_prev_start.isoformat()} → {m6_prev_end.isoformat()}",
                    "current": _aggregate_rows(m6_previous, tipo_map),
                },
            },
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Turmas
# ---------------------------------------------------------------------------

GRAD_MONTHS = [2, 3, 4, 5, 8, 9, 10, 11]
POS_MONTHS = list(range(1, 13))
MONTH_NAMES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}


def _turma_defaults(nivel, ano):
    """Gera ranges padrão de turmas para um nível/ano."""
    import calendar
    months = GRAD_MONTHS if nivel == "Graduação" else POS_MONTHS
    rows = []
    for m in months:
        last_day = calendar.monthrange(ano, m)[1]
        rows.append({
            "nivel": nivel,
            "nome": f"{MONTH_NAMES[m]} {ano}",
            "dt_inicio": f"{ano}-{m:02d}-01",
            "dt_fim": f"{ano}-{m:02d}-{last_day:02d}",
            "ano": ano,
        })
    return rows


@app.route("/api/turmas")
def api_turmas_list():
    nivel = request.args.get("nivel", "")
    ano = request.args.get("ano", "")
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            q = "SELECT * FROM turmas WHERE 1=1"
            params = []
            if nivel:
                q += " AND nivel = %s"
                params.append(nivel)
            if ano:
                q += " AND ano = %s"
                params.append(int(ano))
            q += " ORDER BY ano, dt_inicio"
            cur.execute(q, params)
            rows = []
            for r in cur.fetchall():
                row = dict(r)
                for k, v in row.items():
                    if hasattr(v, "isoformat"):
                        row[k] = v.isoformat() if v else None
                rows.append(row)
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/turmas", methods=["POST"])
def api_turmas_create():
    body = request.json or {}
    required = ("nivel", "nome", "dt_inicio", "dt_fim", "ano")
    if not all(body.get(k) for k in required):
        return jsonify({"error": "Campos obrigatórios: " + ", ".join(required)}), 400
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO turmas (nivel, nome, dt_inicio, dt_fim, ano) "
                "VALUES (%s, %s, %s, %s, %s) "
                "ON CONFLICT (nivel, nome) DO UPDATE SET dt_inicio=EXCLUDED.dt_inicio, dt_fim=EXCLUDED.dt_fim, ano=EXCLUDED.ano "
                "RETURNING id",
                (body["nivel"], body["nome"], body["dt_inicio"], body["dt_fim"], int(body["ano"])),
            )
            new_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/turmas/<int:tid>", methods=["PUT"])
def api_turmas_update(tid):
    body = request.json or {}
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE turmas SET nivel=COALESCE(%s,nivel), nome=COALESCE(%s,nome), "
                "dt_inicio=COALESCE(%s,dt_inicio), dt_fim=COALESCE(%s,dt_fim), ano=COALESCE(%s,ano) "
                "WHERE id=%s",
                (body.get("nivel"), body.get("nome"), body.get("dt_inicio"), body.get("dt_fim"),
                 int(body["ano"]) if body.get("ano") else None, tid),
            )
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/turmas/<int:tid>", methods=["DELETE"])
def api_turmas_delete(tid):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM turmas WHERE id=%s", (tid,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/turmas/seed", methods=["POST"])
def api_turmas_seed():
    body = request.json or {}
    ano = int(body.get("ano", datetime.now().year))
    conn = get_conn()
    try:
        created = 0
        with conn.cursor() as cur:
            for nivel in ("Graduação", "Pós-Graduação"):
                for t in _turma_defaults(nivel, ano):
                    cur.execute(
                        "INSERT INTO turmas (nivel, nome, dt_inicio, dt_fim, ano) "
                        "VALUES (%s, %s, %s, %s, %s) "
                        "ON CONFLICT (nivel, nome) DO NOTHING",
                        (t["nivel"], t["nome"], t["dt_inicio"], t["dt_fim"], t["ano"]),
                    )
                    created += cur.rowcount
        conn.commit()
        return jsonify({"ok": True, "created": created, "ano": ano})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Ciclos
# ---------------------------------------------------------------------------

@app.route("/api/ciclos")
def api_ciclos_list():
    nivel = request.args.get("nivel", "")
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            q = "SELECT * FROM ciclos WHERE 1=1"
            params = []
            if nivel:
                q += " AND nivel = %s"
                params.append(nivel)
            q += " ORDER BY dt_inicio"
            cur.execute(q, params)
            rows = []
            for r in cur.fetchall():
                row = dict(r)
                for k, v in row.items():
                    if hasattr(v, "isoformat"):
                        row[k] = v.isoformat() if v else None
                rows.append(row)
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/ciclos", methods=["POST"])
def api_ciclos_create():
    body = request.json or {}
    required = ("nivel", "nome", "dt_inicio", "dt_fim")
    if not all(body.get(k) for k in required):
        return jsonify({"error": "Campos obrigatórios: " + ", ".join(required)}), 400
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO ciclos (nivel, nome, dt_inicio, dt_fim) "
                "VALUES (%s, %s, %s, %s) "
                "ON CONFLICT (nivel, nome) DO UPDATE SET dt_inicio=EXCLUDED.dt_inicio, dt_fim=EXCLUDED.dt_fim "
                "RETURNING id",
                (body["nivel"], body["nome"], body["dt_inicio"], body["dt_fim"]),
            )
            new_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/ciclos/<int:cid>", methods=["PUT"])
def api_ciclos_update(cid):
    body = request.json or {}
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE ciclos SET nivel=COALESCE(%s,nivel), nome=COALESCE(%s,nome), "
                "dt_inicio=COALESCE(%s,dt_inicio), dt_fim=COALESCE(%s,dt_fim) "
                "WHERE id=%s",
                (body.get("nivel"), body.get("nome"), body.get("dt_inicio"), body.get("dt_fim"), cid),
            )
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/ciclos/<int:cid>", methods=["DELETE"])
def api_ciclos_delete(cid):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM ciclos WHERE id=%s", (cid,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/ciclos/seed", methods=["POST"])
def api_ciclos_seed():
    """Gera ciclos padrão: Graduação semestral, Pós-Graduação semestral."""
    body = request.json or {}
    ano = int(body.get("ano", datetime.now().year))
    conn = get_conn()
    try:
        created = 0
        defaults = [
            ("Graduação", f"{ano}.1", f"{ano-1}-11-16", f"{ano}-05-15"),
            ("Graduação", f"{ano}.2", f"{ano}-05-16", f"{ano}-11-15"),
            ("Pós-Graduação", f"{ano}.1", f"{ano-1}-11-16", f"{ano}-05-15"),
            ("Pós-Graduação", f"{ano}.2", f"{ano}-05-16", f"{ano}-11-15"),
        ]
        with conn.cursor() as cur:
            for nivel, nome, dt_ini, dt_fim in defaults:
                cur.execute(
                    "INSERT INTO ciclos (nivel, nome, dt_inicio, dt_fim) "
                    "VALUES (%s, %s, %s, %s) "
                    "ON CONFLICT (nivel, nome) DO NOTHING",
                    (nivel, nome, dt_ini, dt_fim),
                )
                created += cur.rowcount
        conn.commit()
        return jsonify({"ok": True, "created": created, "ano": ano})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Diagnóstico de address
# ---------------------------------------------------------------------------

@app.route("/api/debug/address")
def api_debug_address():
    """Compara address no banco local vs API direta para diagnosticar sync."""
    import requests as req
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, data->>'name' AS nome, data->'address' AS local_address
                FROM leads
                WHERE data->'address' IS NOT NULL
                   AND data->'address' != 'null'::jsonb
                   AND data->'address' != '{}'::jsonb
                LIMIT 3
            """)
            with_addr = cur.fetchall()

            cur.execute("""
                SELECT id, data->>'name' AS nome, data->'address' AS local_address
                FROM leads
                WHERE data->'address' IS NULL
                   OR data->'address' = 'null'::jsonb
                   OR data->'address' = '{}'::jsonb
                LIMIT 3
            """)
            without_addr = cur.fetchall()

            cur.execute("""
                SELECT
                    COUNT(*) FILTER (WHERE data->'address' IS NOT NULL
                        AND data->'address' != 'null'::jsonb
                        AND data->'address' != '{}'::jsonb) AS com_address,
                    COUNT(*) FILTER (WHERE data->'address' IS NULL
                        OR data->'address' = 'null'::jsonb
                        OR data->'address' = '{}'::jsonb) AS sem_address,
                    COUNT(*) AS total
                FROM leads
            """)
            stats = dict(cur.fetchone())

        token = os.getenv("DATACRAZY_API_TOKEN", "")
        headers = {"Authorization": f"Bearer {token}"}
        api_base = "https://api.g1.datacrazy.io/api/v1"

        api_samples = []
        sample_ids = [r["id"] for r in (with_addr[:1] + without_addr[:1])]
        for lid in sample_ids:
            try:
                r = req.get(f"{api_base}/leads/{lid}", headers=headers, timeout=15)
                if r.ok:
                    d = r.json()
                    api_samples.append({
                        "id": lid,
                        "nome": d.get("name"),
                        "api_address": d.get("address"),
                    })
            except Exception as e:
                api_samples.append({"id": lid, "error": str(e)})

        list_sample = []
        try:
            r = req.get(f"{api_base}/leads", headers=headers, params={
                "take": 2,
                "complete[additionalFields]": "true",
            }, timeout=15)
            if r.ok:
                for lead in r.json().get("data", [])[:2]:
                    list_sample.append({
                        "id": lead.get("id"),
                        "nome": lead.get("name"),
                        "list_address": lead.get("address"),
                        "has_address_key": "address" in lead,
                    })
        except Exception as e:
            list_sample = [{"error": str(e)}]

        return jsonify({
            "stats": stats,
            "local_with_addr_samples": [dict(r) for r in with_addr],
            "local_without_addr_samples": [dict(r) for r in without_addr],
            "api_individual_samples": api_samples,
            "api_list_samples": list_sample,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Busca
# ---------------------------------------------------------------------------

@app.route("/api/search")
def api_search():
    cpf = request.args.get("cpf", "").strip()
    rgm = request.args.get("rgm", "").strip()
    telefone = request.args.get("telefone", "").strip()

    if not cpf and not rgm and not telefone:
        return jsonify({"results": [], "error": "Informe pelo menos um critério de busca."})

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(SEARCH_QUERY, {"cpf": cpf, "rgm": rgm, "telefone": telefone})
            rows = cur.fetchall()
            results = []
            for r in rows:
                row = dict(r)
                for k, v in row.items():
                    if isinstance(v, datetime):
                        row[k] = to_brt(v)
                results.append(row)
        return jsonify({"results": results})
    except Exception as e:
        return jsonify({"results": [], "error": str(e)}), 500
    finally:
        conn.close()


_xl_cache = {"data": None, "mtime": 0}

def _normalize_digits(s):
    """Remove tudo exceto dígitos, tratando floats do Excel (46901353.0 → 46901353)."""
    if not s:
        return ""
    if isinstance(s, float) and s == int(s):
        s = int(s)
    raw = str(s).strip()
    if re.match(r"^\d+\.0+$", raw):
        raw = raw.split(".")[0]
    return re.sub(r"\D", "", raw)


@app.route("/api/search-xl")
def api_search_xl():
    cpf = _normalize_digits(request.args.get("cpf", ""))
    rgm = _normalize_digits(request.args.get("rgm", ""))
    telefone = _normalize_digits(request.args.get("telefone", ""))
    snapshot_id = request.args.get("snapshot_id", "")
    tipo = request.args.get("tipo", "").strip().lower()

    if not cpf and not rgm and not telefone:
        return jsonify({"results": [], "snapshot": None})

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if snapshot_id:
                cur.execute("SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots WHERE id = %s", (snapshot_id,))
            elif tipo:
                cur.execute("SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots WHERE tipo = %s ORDER BY id DESC LIMIT 1", (tipo,))
            else:
                cur.execute("SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots ORDER BY id DESC LIMIT 1")
            snap = cur.fetchone()
            if not snap:
                return jsonify({"results": [], "snapshot": None})

            snap_info = {
                "id": snap["id"],
                "tipo": snap["tipo"],
                "filename": snap["filename"],
                "row_count": snap["row_count"],
                "uploaded_at": to_brt(snap["uploaded_at"]),
            }
            sid = snap["id"]

            conditions = []
            params_list = [sid]

            if cpf:
                conditions.append("data->>'cpf_digits' LIKE '%%' || %s || '%%'")
                params_list.append(cpf)
            if rgm:
                conditions.append("data->>'rgm' LIKE '%%' || %s || '%%'")
                params_list.append(rgm)
            if telefone:
                conditions.append("""(
                    EXISTS (SELECT 1 FROM jsonb_array_elements_text(data->'phones_digits') ph WHERE ph LIKE '%%' || %s || '%%')
                )""")
                params_list.append(telefone)

            where = " OR ".join(conditions)
            cur.execute(
                f"SELECT data FROM xl_rows WHERE snapshot_id = %s AND ({where}) LIMIT 20",
                params_list,
            )
            rows = cur.fetchall()

        results = []
        for r in rows:
            d = r["data"]
            results.append({k: v for k, v in d.items()
                            if k not in ("cpf_digits", "rgm_digits", "phones_digits")})

        return jsonify({"results": results, "snapshot": snap_info})
    except Exception as e:
        return jsonify({"results": [], "snapshot": None, "error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/xl-snapshots")
def api_xl_snapshots():
    tipo = request.args.get("tipo", "").strip().lower()
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if tipo:
                cur.execute("""
                    SELECT id, tipo, filename, row_count, uploaded_at
                    FROM xl_snapshots WHERE tipo = %s ORDER BY id DESC LIMIT 20
                """, (tipo,))
            else:
                cur.execute("""
                    SELECT id, tipo, filename, row_count, uploaded_at
                    FROM xl_snapshots ORDER BY id DESC LIMIT 20
                """)
            snaps = cur.fetchall()
        for s in snaps:
            s["uploaded_at"] = to_brt(s["uploaded_at"])
        return jsonify({"snapshots": snaps, "tipos": XL_TIPOS})
    except Exception as e:
        return jsonify({"snapshots": [], "tipos": XL_TIPOS, "error": str(e)}), 500
    finally:
        conn.close()


def _compute_snapshot_stats(snap_id, tipo):
    """Calcula e grava métricas agregadas de um snapshot."""
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT data FROM xl_rows WHERE snapshot_id = %s", (snap_id,))
            rows = [r["data"] for r in cur.fetchall()]

        if not rows:
            return

        metrics = {"total": len(rows)}
        by_polo, by_curso = {}, {}
        for r in rows:
            polo = r.get("polo", "N/I") or "N/I"
            by_polo[polo] = by_polo.get(polo, 0) + 1
            curso = r.get("curso", "N/I") or "N/I"
            by_curso[curso] = by_curso.get(curso, 0) + 1

        if tipo == "matriculados":
            by_tipo_aluno = {}
            for r in rows:
                ta = r.get("tipo_matricula", "N/I") or "N/I"
                by_tipo_aluno[ta] = by_tipo_aluno.get(ta, 0) + 1
            metrics["by_tipo_aluno"] = by_tipo_aluno

        elif tipo == "inadimplentes":
            total_valor = 0.0
            max_atraso = 0
            for r in rows:
                try:
                    total_valor += float(r.get("valor_total", "0") or "0")
                except (ValueError, TypeError):
                    pass
                try:
                    a = int(float(r.get("max_atraso", "0") or "0"))
                    if a > max_atraso:
                        max_atraso = a
                except (ValueError, TypeError):
                    pass
            metrics["valor_total"] = round(total_valor, 2)
            metrics["max_atraso"] = max_atraso
            metrics["total_titulos"] = sum(int(r.get("total_titulos", "0") or "0") for r in rows)

        elif tipo == "acesso_ava":
            now = datetime.now()
            ativos_7d, ativos_30d, inativos = 0, 0, 0
            total_interacoes, total_minutos = 0, 0
            for r in rows:
                try:
                    total_interacoes += int(float(r.get("interacoes", "0") or "0"))
                except (ValueError, TypeError):
                    pass
                try:
                    total_minutos += int(float(r.get("minutos", "0") or "0"))
                except (ValueError, TypeError):
                    pass
                ua = r.get("ultimo_acesso", "")
                if ua:
                    try:
                        dt = datetime.strptime(ua[:10], "%Y-%m-%d") if "-" in ua else datetime.strptime(ua[:10], "%d/%m/%Y")
                        delta = (now - dt).days
                        if delta <= 7:
                            ativos_7d += 1
                        if delta <= 30:
                            ativos_30d += 1
                        else:
                            inativos += 1
                    except (ValueError, TypeError):
                        inativos += 1
                else:
                    inativos += 1
            metrics["ativos_7d"] = ativos_7d
            metrics["ativos_30d"] = ativos_30d
            metrics["inativos"] = inativos
            metrics["media_interacoes"] = round(total_interacoes / max(len(rows), 1), 1)
            metrics["media_minutos"] = round(total_minutos / max(len(rows), 1), 1)

        elif tipo == "sem_rematricula":
            adim, inadim = 0, 0
            for r in rows:
                if r.get("status_financeiro") == "adimplente":
                    adim += 1
                else:
                    inadim += 1
            metrics["adimplentes"] = adim
            metrics["inadimplentes"] = inadim

        elif tipo == "concluintes":
            by_inst = {}
            for r in rows:
                inst = r.get("instituicao", "N/I") or "N/I"
                by_inst[inst] = by_inst.get(inst, 0) + 1
            metrics["by_instituicao"] = by_inst

        metrics["by_polo"] = dict(sorted(by_polo.items(), key=lambda x: -x[1])[:20])
        metrics["by_curso"] = dict(sorted(by_curso.items(), key=lambda x: -x[1])[:20])

        with conn.cursor() as cur:
            for metric_name, value in metrics.items():
                if isinstance(value, dict):
                    cur.execute(
                        "INSERT INTO xl_snapshot_stats (snapshot_id, metric, value, detail) "
                        "VALUES (%s, %s, %s, %s::jsonb) "
                        "ON CONFLICT (snapshot_id, metric) DO UPDATE SET value=EXCLUDED.value, detail=EXCLUDED.detail",
                        (snap_id, metric_name, None, json.dumps(value, ensure_ascii=False)),
                    )
                else:
                    cur.execute(
                        "INSERT INTO xl_snapshot_stats (snapshot_id, metric, value, detail) "
                        "VALUES (%s, %s, %s, NULL) "
                        "ON CONFLICT (snapshot_id, metric) DO UPDATE SET value=EXCLUDED.value",
                        (snap_id, metric_name, value),
                    )
        conn.commit()
    except Exception as e:
        app.logger.warning("Erro ao computar stats snapshot %s: %s", snap_id, e)
    finally:
        conn.close()


@app.route("/api/snapshots/compare")
def api_snapshots_compare():
    """Compara dois snapshots do mesmo tipo."""
    tipo = request.args.get("tipo", "").strip().lower()
    periodo = request.args.get("periodo", "")
    snap_a = request.args.get("snap_a", "")
    snap_b = request.args.get("snap_b", "")

    if not tipo and not snap_a:
        return jsonify({"error": "Informe tipo ou snap_a"}), 400

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if snap_a and snap_b:
                cur.execute("SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots WHERE id IN (%s,%s) ORDER BY id DESC", (snap_a, snap_b))
                snaps = cur.fetchall()
                if len(snaps) < 2:
                    return jsonify({"error": "Snapshots não encontrados"}), 404
                sa, sb = snaps[0], snaps[1]
            else:
                cur.execute("SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots WHERE tipo=%s ORDER BY id DESC LIMIT 1", (tipo,))
                sa = cur.fetchone()
                if not sa:
                    return jsonify({"error": f"Nenhum snapshot para tipo '{tipo}'"}), 404

                period_map = {"6m": 180, "1y": 365, "2y": 730, "3m": 90}
                days = period_map.get(periodo, 180)
                target_date = sa["uploaded_at"] - timedelta(days=days)
                cur.execute(
                    "SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots "
                    "WHERE tipo=%s AND uploaded_at <= %s ORDER BY uploaded_at DESC LIMIT 1",
                    (tipo, target_date),
                )
                sb = cur.fetchone()
                if not sb:
                    cur.execute(
                        "SELECT id, tipo, filename, row_count, uploaded_at FROM xl_snapshots "
                        "WHERE tipo=%s AND id < %s ORDER BY id ASC LIMIT 1",
                        (tipo, sa["id"]),
                    )
                    sb = cur.fetchone()

            def _get_stats(sid):
                cur.execute("SELECT metric, value, detail FROM xl_snapshot_stats WHERE snapshot_id=%s", (sid,))
                stats = {}
                for r in cur.fetchall():
                    stats[r["metric"]] = r["detail"] if r["detail"] is not None else (float(r["value"]) if r["value"] is not None else None)
                return stats

            stats_a = _get_stats(sa["id"])
            stats_b = _get_stats(sb["id"]) if sb else {}

            cur.execute("SELECT data->>'rgm_digits' AS rgm FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (sa["id"],))
            rgms_a = {r["rgm"] for r in cur.fetchall()}
            rgms_b = set()
            if sb:
                cur.execute("SELECT data->>'rgm_digits' AS rgm FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (sb["id"],))
                rgms_b = {r["rgm"] for r in cur.fetchall()}

            for s in [sa, sb]:
                if s:
                    s["uploaded_at"] = to_brt(s["uploaded_at"])

        return jsonify({
            "snap_a": sa,
            "snap_b": sb,
            "stats_a": stats_a,
            "stats_b": stats_b,
            "novos": len(rgms_a - rgms_b),
            "removidos": len(rgms_b - rgms_a),
            "mantidos": len(rgms_a & rgms_b),
            "delta_total": (sa["row_count"] - sb["row_count"]) if sb else 0,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/snapshots/timeline")
def api_snapshots_timeline():
    """Série temporal de uma métrica para um tipo de snapshot."""
    tipo = request.args.get("tipo", "").strip().lower()
    metric = request.args.get("metric", "total")
    months = int(request.args.get("months", 24))

    if not tipo:
        return jsonify({"error": "Informe tipo"}), 400

    conn = get_conn()
    try:
        cutoff = datetime.now(BRT) - timedelta(days=months * 30)
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT s.id, s.uploaded_at, s.row_count, st.value, st.detail "
                "FROM xl_snapshots s "
                "LEFT JOIN xl_snapshot_stats st ON st.snapshot_id = s.id AND st.metric = %s "
                "WHERE s.tipo = %s AND s.uploaded_at >= %s "
                "ORDER BY s.uploaded_at",
                (metric, tipo, cutoff),
            )
            rows = cur.fetchall()

        points = []
        for r in rows:
            val = None
            if r["detail"] is not None:
                val = r["detail"]
            elif r["value"] is not None:
                val = float(r["value"])
            elif metric == "total":
                val = r["row_count"]
            points.append({
                "date": to_brt(r["uploaded_at"]),
                "snapshot_id": r["id"],
                "value": val,
            })

        return jsonify({"tipo": tipo, "metric": metric, "points": points})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/snapshots/crossref")
def api_snapshots_crossref():
    """Cruzamento entre dois tipos de snapshot por RGM."""
    tipo_a = request.args.get("tipo_a", "").strip().lower()
    tipo_b = request.args.get("tipo_b", "").strip().lower()

    if not tipo_a or not tipo_b:
        return jsonify({"error": "Informe tipo_a e tipo_b"}), 400

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            def _latest_rgms(tipo):
                cur.execute("SELECT id FROM xl_snapshots WHERE tipo=%s ORDER BY id DESC LIMIT 1", (tipo,))
                snap = cur.fetchone()
                if not snap:
                    return set(), None
                cur.execute("SELECT data->>'rgm_digits' AS rgm FROM xl_rows WHERE snapshot_id=%s AND data->>'rgm_digits' != ''", (snap["id"],))
                return {r["rgm"] for r in cur.fetchall()}, snap["id"]

            rgms_a, sid_a = _latest_rgms(tipo_a)
            rgms_b, sid_b = _latest_rgms(tipo_b)

        return jsonify({
            "tipo_a": tipo_a, "tipo_b": tipo_b,
            "total_a": len(rgms_a), "total_b": len(rgms_b),
            "em_ambos": len(rgms_a & rgms_b),
            "apenas_a": len(rgms_a - rgms_b),
            "apenas_b": len(rgms_b - rgms_a),
            "snap_a": sid_a, "snap_b": sid_b,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/upload/info")
def api_upload_info():
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            result = {}
            for t in XL_TIPOS:
                cur.execute("""
                    SELECT id, tipo, filename, row_count, uploaded_at
                    FROM xl_snapshots WHERE tipo = %s ORDER BY id DESC LIMIT 1
                """, (t,))
                snap = cur.fetchone()
                if snap:
                    snap["uploaded_at"] = to_brt(snap["uploaded_at"])
                result[t] = snap
        file_disk = _find_xlsx()
        return jsonify({"file": file_disk, "snapshots": result})
    except Exception as e:
        return jsonify({"file": _find_xlsx(), "snapshots": {}, "error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/sync-state")
def api_sync_state():
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(SYNC_STATE_QUERY)
            states = []
            for r in cur.fetchall():
                row = dict(r)
                for k, v in row.items():
                    if isinstance(v, datetime):
                        row[k] = to_brt(v)
                states.append(row)

            cur.execute(RECENT_BIZ_UPDATES_QUERY)
            recent = []
            for r in cur.fetchall():
                row = dict(r)
                for k, v in row.items():
                    if isinstance(v, datetime):
                        row[k] = to_brt(v)
                recent.append(row)

        return jsonify({"states": states, "recent_updates": recent})
    except Exception as e:
        return jsonify({"states": [], "recent_updates": [], "error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Sync
# ---------------------------------------------------------------------------

@app.route("/api/sync/<mode>", methods=["POST"])
def api_sync(mode):
    global _sync_running

    if mode not in ("delta", "full"):
        return jsonify({"error": "Modo inválido. Use 'delta' ou 'full'."}), 400

    if _sync_running:
        return jsonify({"error": "Sincronização já em andamento."}), 409

    _sync_running = True
    _sync_logs.clear()

    def run():
        global _sync_running, _sync_proc
        try:
            cmd = [sys.executable, SYNC_SCRIPT]
            if mode == "full":
                cmd.append("--full")

            _add_sync_log(f"[INÍCIO] Sincronização {mode.upper()} iniciada")

            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                cwd=str(BASE_DIR),
                env=env,
            )
            _sync_proc = proc

            for line in proc.stdout:
                _add_sync_log(line)

            proc.wait()

            if proc.returncode == 0:
                _add_sync_log("[FIM] Sincronização concluída com sucesso")
            elif proc.returncode < 0:
                _add_sync_log("[PARADO] Sincronização interrompida.")
            else:
                _add_sync_log(f"[ERRO] Sincronização falhou (exit code {proc.returncode})")
        except Exception as e:
            import traceback
            _add_sync_log(f"[ERRO] {e}")
            _add_sync_log(traceback.format_exc())
        finally:
            _sync_proc = None
            _sync_running = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "mode": mode})


@app.route("/api/sync/logs")
def api_sync_logs():
    since = int(request.args.get("since", 0))
    logs_list = list(_sync_logs)
    lines = logs_list[since:]
    return jsonify({"lines": lines, "total": len(logs_list), "running": _sync_running})


@app.route("/api/sync/status")
def api_sync_status():
    global _sync_running
    if _sync_running and (_sync_proc is None or _sync_proc.poll() is not None):
        _sync_running = False
    return jsonify({"running": _sync_running})


@app.route("/api/sync/stop", methods=["POST"])
def api_sync_stop():
    global _sync_running
    if _sync_proc is not None:
        try:
            _sync_proc.terminate()
        except Exception:
            pass
    _sync_running = False
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Rotas — Update CRM
# ---------------------------------------------------------------------------

@app.route("/api/update/<mode>", methods=["POST"])
def api_update(mode):
    global _update_running

    if mode not in ("dry-run", "test", "execute"):
        return jsonify({"error": "Modo inválido. Use 'dry-run', 'test' ou 'execute'."}), 400

    if _update_running:
        return jsonify({"error": "Atualização já em andamento."}), 409

    body = request.json if request.is_json else {}
    limit = body.get("limit")
    rate = body.get("rate")
    with_address = body.get("withAddress", False)

    _update_running = True
    _update_logs.clear()

    def run():
        global _update_running, _update_proc
        try:
            cmd = [sys.executable, UPDATE_SCRIPT, f"--{mode}"]
            if limit and mode == "execute":
                cmd.extend(["--limit", str(int(limit))])
            if rate is not None:
                cmd.extend(["--rate", str(int(rate))])
            if with_address:
                cmd.append("--with-address")

            _add_update_log(f"[INÍCIO] Update CRM — modo {mode.upper()}")

            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                cwd=str(BASE_DIR),
                env=env,
            )
            _update_proc = proc

            for line in proc.stdout:
                _add_update_log(line)

            proc.wait()

            if proc.returncode == 0:
                _add_update_log("[FIM] Concluído com sucesso (exit code 0)")
            elif proc.returncode < 0:
                _add_update_log("[PARADO] Processo interrompido pelo usuário.")
            else:
                _add_update_log(f"[ERRO] Falhou (exit code {proc.returncode})")
        except Exception as e:
            import traceback
            _add_update_log(f"[ERRO] {e}")
            _add_update_log(traceback.format_exc())
        finally:
            _update_proc = None
            _update_running = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "mode": mode})


@app.route("/api/update/logs")
def api_update_logs():
    since = int(request.args.get("since", 0))
    logs_list = list(_update_logs)
    lines = logs_list[since:]
    return jsonify({"lines": lines, "total": len(logs_list), "running": _update_running})


@app.route("/api/update/status")
def api_update_status():
    global _update_running
    if _update_running and (_update_proc is None or _update_proc.poll() is not None):
        _update_running = False
    return jsonify({"running": _update_running})


@app.route("/api/update/stop", methods=["POST"])
def api_update_stop():
    global _update_running
    if _update_proc is not None:
        try:
            _update_proc.terminate()
        except Exception:
            pass
    _update_running = False
    return jsonify({"ok": True})


@app.route("/api/update/preview")
def api_update_preview():
    preview_path = REPORTS_DIR / "update_preview.csv"
    if not preview_path.exists():
        return jsonify({"rows": [], "error": "Rode dry-run primeiro para gerar o preview."})

    import csv as csv_mod
    rows = []
    with open(preview_path, "r", encoding="utf-8-sig") as f:
        reader = csv_mod.DictReader(f, delimiter=";")
        for i, row in enumerate(reader):
            if i >= 500:
                break
            rows.append(dict(row))
    return jsonify({"rows": rows, "total": len(rows)})


# ---------------------------------------------------------------------------
# Rotas — Saneamento
# ---------------------------------------------------------------------------

@app.route("/api/sanitize/<mode>", methods=["POST"])
def api_sanitize(mode):
    global _sanitize_running

    if mode not in ("dry-run", "test", "execute"):
        return jsonify({"error": "Modo inválido. Use 'dry-run', 'test' ou 'execute'."}), 400

    if _sanitize_running:
        return jsonify({"error": "Saneamento já em andamento."}), 409

    body = request.json if request.is_json else {}
    limit = body.get("limit")
    rate = body.get("rate", 60)

    _sanitize_running = True
    _sanitize_logs.clear()

    def run():
        global _sanitize_running, _sanitize_proc
        try:
            cmd = [sys.executable, SANITIZE_SCRIPT, f"--{mode}",
                   "--rate", str(int(rate))]
            if limit and mode == "execute":
                cmd.extend(["--limit", str(int(limit))])

            _add_sanitize_log(f"[INÍCIO] Saneamento — modo {mode.upper()}")

            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=str(BASE_DIR), env=env,
            )
            _sanitize_proc = proc
            for line in proc.stdout:
                _add_sanitize_log(line)
            proc.wait()

            if proc.returncode == 0:
                _add_sanitize_log("[FIM] Concluído com sucesso (exit code 0)")
            elif proc.returncode < 0:
                _add_sanitize_log("[PARADO] Processo interrompido pelo usuário.")
            else:
                _add_sanitize_log(f"[ERRO] Falhou (exit code {proc.returncode})")
        except Exception as e:
            import traceback
            _add_sanitize_log(f"[ERRO] {e}")
            _add_sanitize_log(traceback.format_exc())
        finally:
            _sanitize_proc = None
            _sanitize_running = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "mode": mode})


@app.route("/api/sanitize/logs")
def api_sanitize_logs():
    since = int(request.args.get("since", 0))
    logs_list = list(_sanitize_logs)
    lines = logs_list[since:]
    return jsonify({"lines": lines, "total": len(logs_list), "running": _sanitize_running})


@app.route("/api/sanitize/status")
def api_sanitize_status():
    global _sanitize_running
    if _sanitize_running and (_sanitize_proc is None or _sanitize_proc.poll() is not None):
        _sanitize_running = False
    return jsonify({"running": _sanitize_running})


@app.route("/api/sanitize/stop", methods=["POST"])
def api_sanitize_stop():
    global _sanitize_running
    if _sanitize_proc is not None:
        try:
            _sanitize_proc.terminate()
        except Exception:
            pass
    _sanitize_running = False
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Rotas — Enriquecimento de duplicatas
# ---------------------------------------------------------------------------

@app.route("/api/enrich/start", methods=["POST"])
def api_enrich_start():
    global _enrich_running
    if _enrich_running:
        return jsonify({"error": "Enriquecimento já em andamento."}), 409

    body = request.json if request.is_json else {}
    rate = body.get("rate", 60)

    _enrich_running = True
    _enrich_logs.clear()

    def run():
        global _enrich_running, _enrich_proc
        try:
            cmd = [sys.executable, ENRICH_SCRIPT, "--rate", str(int(rate))]
            _add_enrich_log("[INÍCIO] Enriquecimento de duplicatas entre leads")
            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            proc = subprocess.Popen(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=str(BASE_DIR), env=env,
            )
            _enrich_proc = proc
            for line in proc.stdout:
                _add_enrich_log(line)
            proc.wait()
            if proc.returncode == 0:
                _add_enrich_log("[FIM] Concluído com sucesso (exit code 0)")
            elif proc.returncode < 0:
                _add_enrich_log("[PARADO] Processo interrompido pelo usuário.")
            else:
                _add_enrich_log(f"[ERRO] Falhou (exit code {proc.returncode})")
        except Exception as e:
            import traceback
            _add_enrich_log(f"[ERRO] {e}")
            _add_enrich_log(traceback.format_exc())
        finally:
            _enrich_proc = None
            _enrich_running = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/enrich/logs")
def api_enrich_logs():
    since = int(request.args.get("since", 0))
    logs_list = list(_enrich_logs)
    lines = logs_list[since:]
    return jsonify({"lines": lines, "total": len(logs_list), "running": _enrich_running})


@app.route("/api/enrich/stop", methods=["POST"])
def api_enrich_stop():
    global _enrich_running
    if _enrich_proc is not None:
        try:
            _enrich_proc.terminate()
        except Exception:
            pass
    _enrich_running = False
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Rotas — Merge
# ---------------------------------------------------------------------------

@app.route("/api/merge/start", methods=["POST"])
def api_merge_start():
    global _merge_running
    if _merge_running:
        return jsonify({"error": "Merge já em andamento."}), 409

    body = request.json if request.is_json else {}
    mode = body.get("mode", "dry-run")
    fase = body.get("fase")
    limit = body.get("limit")
    rate = body.get("rate", 60)

    _merge_running = True
    _merge_logs.clear()

    def run():
        global _merge_running, _merge_proc
        try:
            cmd = [sys.executable, MERGE_SCRIPT, f"--{mode}"]
            if fase:
                cmd += ["--fase", str(int(fase))]
            if limit:
                cmd += ["--limit", str(int(limit))]
            cmd += ["--rate", str(int(rate))]

            _add_merge_log(f"[INÍCIO] Merge de leads — {mode}" +
                           (f" (fase {fase})" if fase else "") +
                           (f" (limit {limit})" if limit else ""))
            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            proc = subprocess.Popen(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=str(BASE_DIR), env=env,
            )
            _merge_proc = proc
            for line in proc.stdout:
                _add_merge_log(line)
            proc.wait()
            if proc.returncode == 0:
                _add_merge_log("[FIM] Concluído com sucesso (exit code 0)")
            elif proc.returncode < 0:
                _add_merge_log("[PARADO] Processo interrompido pelo usuário.")
            else:
                _add_merge_log(f"[ERRO] Falhou (exit code {proc.returncode})")
        except Exception as e:
            import traceback
            _add_merge_log(f"[ERRO] {e}")
            _add_merge_log(traceback.format_exc())
        finally:
            _merge_proc = None
            _merge_running = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/merge/logs")
def api_merge_logs():
    since = int(request.args.get("since", 0))
    logs_list = list(_merge_logs)
    lines = logs_list[since:]
    return jsonify({"lines": lines, "total": len(logs_list), "running": _merge_running})


@app.route("/api/merge/stop", methods=["POST"])
def api_merge_stop():
    global _merge_running
    if _merge_proc is not None:
        try:
            _merge_proc.terminate()
        except Exception:
            pass
    _merge_running = False
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Rotas — Pipeline
# ---------------------------------------------------------------------------

@app.route("/api/pipeline/<mode>", methods=["POST"])
def api_pipeline(mode):
    global _pipeline_running

    if mode not in ("dry-run", "test", "execute"):
        return jsonify({"error": "Modo inválido. Use 'dry-run', 'test' ou 'execute'."}), 400

    if _pipeline_running:
        return jsonify({"error": "Pipeline já em andamento."}), 409

    body = request.json if request.is_json else {}
    limit = body.get("limit")
    rate = body.get("rate")

    _pipeline_running = True
    _pipeline_logs.clear()

    def run():
        global _pipeline_running, _pipeline_proc
        try:
            cmd = [sys.executable, PIPELINE_SCRIPT, f"--{mode}"]
            if limit and mode == "execute":
                cmd.extend(["--limit", str(int(limit))])
            if rate:
                cmd.extend(["--rate", str(int(rate))])

            _add_pipeline_log(f"[INÍCIO] Pipeline — modo {mode.upper()}")

            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=str(BASE_DIR), env=env,
            )
            _pipeline_proc = proc
            for line in proc.stdout:
                _add_pipeline_log(line)
            proc.wait()

            if proc.returncode == 0:
                _add_pipeline_log("[FIM] Concluído com sucesso (exit code 0)")
            elif proc.returncode < 0:
                _add_pipeline_log("[PARADO] Processo interrompido pelo usuário.")
            else:
                _add_pipeline_log(f"[ERRO] Falhou (exit code {proc.returncode})")
        except Exception as e:
            import traceback
            _add_pipeline_log(f"[ERRO] {e}")
            _add_pipeline_log(traceback.format_exc())
        finally:
            _pipeline_proc = None
            _pipeline_running = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "mode": mode})


@app.route("/api/pipeline/logs")
def api_pipeline_logs():
    since = int(request.args.get("since", 0))
    logs_list = list(_pipeline_logs)
    lines = logs_list[since:]
    return jsonify({"lines": lines, "total": len(logs_list), "running": _pipeline_running})


@app.route("/api/pipeline/status")
def api_pipeline_status():
    global _pipeline_running
    if _pipeline_running and (_pipeline_proc is None or _pipeline_proc.poll() is not None):
        _pipeline_running = False
    return jsonify({"running": _pipeline_running})


@app.route("/api/pipeline/stop", methods=["POST"])
def api_pipeline_stop():
    global _pipeline_running
    if _pipeline_proc is not None:
        try:
            _pipeline_proc.terminate()
        except Exception:
            pass
    _pipeline_running = False
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Rotas — Inadimplentes
# ---------------------------------------------------------------------------

@app.route("/api/inadimplentes/<mode>", methods=["POST"])
def api_inadimplentes(mode):
    global _inadimplentes_running

    if mode not in ("dry-run", "execute"):
        return jsonify({"error": "Modo inválido. Use 'dry-run' ou 'execute'."}), 400

    if _inadimplentes_running:
        return jsonify({"error": "Atualização de inadimplentes já em andamento."}), 409

    body = request.json if request.is_json else {}
    rate = body.get("rate")

    _inadimplentes_running = True
    _inadimplentes_logs.clear()

    def run():
        global _inadimplentes_running, _inadimplentes_proc
        try:
            cmd = [sys.executable, INADIMPLENTES_SCRIPT, f"--{mode}"]
            if rate:
                cmd.extend(["--rate", str(int(rate))])

            _add_inadimplentes_log(f"[INÍCIO] Inadimplentes — {mode.upper()}")

            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=str(BASE_DIR), env=env,
            )
            _inadimplentes_proc = proc
            for line in proc.stdout:
                _add_inadimplentes_log(line)
            proc.wait()

            if proc.returncode == 0:
                _add_inadimplentes_log("[FIM] Concluído com sucesso (exit code 0)")
            elif proc.returncode < 0:
                _add_inadimplentes_log("[PARADO] Processo interrompido pelo usuário.")
            else:
                _add_inadimplentes_log(f"[ERRO] Falhou (exit code {proc.returncode})")
        except Exception as e:
            import traceback
            _add_inadimplentes_log(f"[ERRO] {e}")
            _add_inadimplentes_log(traceback.format_exc())
        finally:
            _inadimplentes_proc = None
            _inadimplentes_running = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "mode": mode})


@app.route("/api/inadimplentes/logs")
def api_inadimplentes_logs():
    since = int(request.args.get("since", 0))
    logs_list = list(_inadimplentes_logs)
    lines = logs_list[since:]
    return jsonify({"lines": lines, "total": len(logs_list), "running": _inadimplentes_running})


@app.route("/api/inadimplentes/status")
def api_inadimplentes_status():
    global _inadimplentes_running
    if _inadimplentes_running and (_inadimplentes_proc is None or _inadimplentes_proc.poll() is not None):
        _inadimplentes_running = False
    return jsonify({"running": _inadimplentes_running})


@app.route("/api/inadimplentes/stop", methods=["POST"])
def api_inadimplentes_stop():
    global _inadimplentes_running
    if _inadimplentes_proc is not None:
        try:
            _inadimplentes_proc.terminate()
        except Exception:
            pass
    _inadimplentes_running = False
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Rotas — Concluintes
# ---------------------------------------------------------------------------

@app.route("/api/concluintes/<mode>", methods=["POST"])
def api_concluintes(mode):
    global _concluintes_running

    if mode not in ("dry-run", "execute"):
        return jsonify({"error": "Modo inválido. Use 'dry-run' ou 'execute'."}), 400

    if _concluintes_running:
        return jsonify({"error": "Atualização de concluintes já em andamento."}), 409

    body = request.json if request.is_json else {}
    rate = body.get("rate")

    _concluintes_running = True
    _concluintes_logs.clear()

    def run():
        global _concluintes_running, _concluintes_proc
        try:
            cmd = [sys.executable, CONCLUINTES_SCRIPT, f"--{mode}"]
            if rate:
                cmd.extend(["--rate", str(int(rate))])

            _add_concluintes_log(f"[INÍCIO] Concluintes — {mode.upper()}")

            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=str(BASE_DIR), env=env,
            )
            _concluintes_proc = proc
            for line in proc.stdout:
                _add_concluintes_log(line)
            proc.wait()

            if proc.returncode == 0:
                _add_concluintes_log("[FIM] Concluído com sucesso (exit code 0)")
            elif proc.returncode < 0:
                _add_concluintes_log("[PARADO] Processo interrompido pelo usuário.")
            else:
                _add_concluintes_log(f"[ERRO] Falhou (exit code {proc.returncode})")
        except Exception as e:
            import traceback
            _add_concluintes_log(f"[ERRO] {e}")
            _add_concluintes_log(traceback.format_exc())
        finally:
            _concluintes_proc = None
            _concluintes_running = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "mode": mode})


@app.route("/api/concluintes/logs")
def api_concluintes_logs():
    since = int(request.args.get("since", 0))
    logs_list = list(_concluintes_logs)
    lines = logs_list[since:]
    return jsonify({"lines": lines, "total": len(logs_list), "running": _concluintes_running})


@app.route("/api/concluintes/status")
def api_concluintes_status():
    global _concluintes_running
    if _concluintes_running and (_concluintes_proc is None or _concluintes_proc.poll() is not None):
        _concluintes_running = False
    return jsonify({"running": _concluintes_running})


@app.route("/api/concluintes/stop", methods=["POST"])
def api_concluintes_stop():
    global _concluintes_running
    if _concluintes_proc is not None:
        try:
            _concluintes_proc.terminate()
        except Exception:
            pass
    _concluintes_running = False
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Rotas — Distribuição (proxy para n8n)
# ---------------------------------------------------------------------------

N8N_DIST_GET = "https://n8n-new-n8n.ca31ey.easypanel.host/webhook/api/distribuicao"
N8N_DIST_SAVE = "https://n8n-new-n8n.ca31ey.easypanel.host/webhook/api/atualizar-distribuicao"


@app.route("/api/distribuicao", methods=["GET"])
def api_distribuicao_get():
    try:
        r = _requests.get(N8N_DIST_GET, timeout=15)
        payload = r.json()
        if isinstance(payload, list):
            payload = payload[0] if payload else {}
        return jsonify(payload)
    except Exception as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/distribuicao", methods=["POST"])
def api_distribuicao_save():
    try:
        data = request.json
        r = _requests.post(N8N_DIST_SAVE, json=data, timeout=15,
                           headers={"Content-Type": "application/json"})
        if r.ok:
            return jsonify({"ok": True})
        return jsonify({"ok": False, "error": f"n8n respondeu {r.status_code}"}), 502
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502


# ---------------------------------------------------------------------------
# Motor de Engajamento AVA
# ---------------------------------------------------------------------------

N8N_COMM_WEBHOOK = os.getenv(
    "N8N_COMM_WEBHOOK",
    "https://n8n-new-n8n.ca31ey.easypanel.host/webhook/comm-engagement",
)

NEW_STUDENT_DAYS = 30


def _parse_date_flexible(val):
    """Parse a date string in multiple formats, return datetime or None."""
    if not val:
        return None
    val = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(val[:len(fmt.replace('%', 'X'))], fmt)
        except (ValueError, IndexError):
            continue
    return None


def calculate_engagement_scores():
    """Cross-reference matriculados x acesso_ava and compute engagement scores."""
    conn = get_conn()
    today = datetime.now().date()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id FROM xl_snapshots WHERE tipo='matriculados' ORDER BY id DESC LIMIT 1")
            snap_mat = cur.fetchone()
            cur.execute("SELECT id FROM xl_snapshots WHERE tipo='acesso_ava' ORDER BY id DESC LIMIT 1")
            snap_ava = cur.fetchone()

            if not snap_mat:
                return {"error": "Nenhum snapshot de matriculados encontrado", "processed": 0}

            cur.execute("""
                SELECT data FROM xl_rows WHERE snapshot_id = %s
                AND COALESCE(data->>'rgm_digits', '') != ''
            """, (snap_mat["id"],))
            mat_rows = {r["data"]["rgm_digits"]: r["data"] for r in cur.fetchall()}

            ava_rows = {}
            if snap_ava:
                cur.execute("""
                    SELECT data FROM xl_rows WHERE snapshot_id = %s
                    AND COALESCE(data->>'rgm_digits', '') != ''
                """, (snap_ava["id"],))
                ava_rows = {r["data"]["rgm_digits"]: r["data"] for r in cur.fetchall()}

            all_interactions = []
            all_minutes = []
            for ad in ava_rows.values():
                try:
                    all_interactions.append(int(float(ad.get("interacoes", 0) or 0)))
                except (ValueError, TypeError):
                    pass
                try:
                    all_minutes.append(float(ad.get("minutos", 0) or 0))
                except (ValueError, TypeError):
                    pass
            avg_interactions = sum(all_interactions) / max(len(all_interactions), 1)
            avg_minutes = sum(all_minutes) / max(len(all_minutes), 1)

            processed = 0
            for rgm, mat_data in mat_rows.items():
                sit = (mat_data.get("situacao") or "").strip().lower()
                if sit and sit not in ("em curso", "ativo", "matriculado"):
                    continue

                dt_mat = _parse_date_flexible(mat_data.get("data_mat"))
                days_enrolled = (today - dt_mat.date()).days if dt_mat else None
                is_new = days_enrolled is not None and days_enrolled <= NEW_STUDENT_DAYS

                ava_data = ava_rows.get(rgm)
                if ava_data:
                    dt_access = _parse_date_flexible(ava_data.get("ultimo_acesso"))
                    days_no_access = (today - dt_access.date()).days if dt_access else None
                    try:
                        interactions = int(float(ava_data.get("interacoes", 0) or 0))
                    except (ValueError, TypeError):
                        interactions = 0
                    try:
                        minutes = float(ava_data.get("minutos", 0) or 0)
                    except (ValueError, TypeError):
                        minutes = 0.0
                else:
                    days_no_access = None
                    interactions = 0
                    minutes = 0.0

                recency = 0
                if days_no_access is not None:
                    recency = max(0, 100 - int(days_no_access * 100 / 30))
                elif ava_data is None:
                    recency = 0

                depth = 0
                if avg_interactions > 0 or avg_minutes > 0:
                    int_ratio = min(interactions / max(avg_interactions, 1), 2.0) * 50
                    min_ratio = min(minutes / max(avg_minutes, 1), 2.0) * 50
                    depth = int((int_ratio + min_ratio) / 2)
                elif interactions > 0 or minutes > 0:
                    depth = 50

                frequency = 0
                if ava_data and days_no_access is not None and days_no_access <= 14:
                    frequency = max(0, 100 - days_no_access * 7)
                elif ava_data and days_no_access is not None:
                    frequency = max(0, 30 - (days_no_access - 14) * 2)

                phase_penalty = 0
                if is_new and ava_data is None:
                    phase_penalty = -30
                elif is_new and days_no_access is not None and days_no_access > 3:
                    phase_penalty = -15

                score = int(
                    recency * 0.40
                    + frequency * 0.25
                    + depth * 0.20
                    + max(0, 100 + phase_penalty) * 0.15
                )
                score = max(0, min(100, score))

                if score >= 80:
                    risk = "engajado"
                elif score >= 60:
                    risk = "atencao"
                elif score >= 30:
                    risk = "em_risco"
                else:
                    risk = "critico"

                detail = {
                    "recency": recency, "frequency": frequency,
                    "depth": depth, "phase_penalty": phase_penalty,
                    "is_new": is_new, "nome": mat_data.get("nome", ""),
                    "curso": mat_data.get("curso", ""),
                    "polo": mat_data.get("polo", ""),
                    "email": mat_data.get("email", ""),
                }

                cur.execute("""
                    INSERT INTO ava_engagement
                        (rgm, snapshot_date, score, risk_level, days_since_enrollment,
                         days_since_last_access, access_count, interaction_count,
                         total_minutes, detail)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (rgm, snapshot_date)
                    DO UPDATE SET score=EXCLUDED.score, risk_level=EXCLUDED.risk_level,
                        days_since_enrollment=EXCLUDED.days_since_enrollment,
                        days_since_last_access=EXCLUDED.days_since_last_access,
                        access_count=EXCLUDED.access_count,
                        interaction_count=EXCLUDED.interaction_count,
                        total_minutes=EXCLUDED.total_minutes,
                        detail=EXCLUDED.detail
                """, (
                    rgm, today, score, risk, days_enrolled, days_no_access,
                    1 if ava_data else 0, interactions, minutes,
                    json.dumps(detail, ensure_ascii=False),
                ))
                processed += 1

        conn.commit()
        return {"processed": processed, "date": str(today)}
    except Exception as e:
        conn.rollback()
        app.logger.error("Engagement score error: %s", e)
        return {"error": str(e), "processed": 0}
    finally:
        conn.close()


@app.route("/api/engagement/recalculate", methods=["POST"])
def api_engagement_recalculate():
    result = calculate_engagement_scores()
    return jsonify(result)


@app.route("/api/engagement/scores")
def api_engagement_scores():
    risk = request.args.get("risk", "").strip()
    polo = request.args.get("polo", "").strip()
    curso = request.args.get("curso", "").strip()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    offset = (page - 1) * per_page

    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            conditions = ["snapshot_date = (SELECT MAX(snapshot_date) FROM ava_engagement)"]
            params = []
            if risk:
                conditions.append("risk_level = %s")
                params.append(risk)
            if polo:
                conditions.append("detail->>'polo' ILIKE %s")
                params.append(f"%{polo}%")
            if curso:
                conditions.append("detail->>'curso' ILIKE %s")
                params.append(f"%{curso}%")

            where = " AND ".join(conditions)

            cur.execute(f"SELECT COUNT(*) as cnt FROM ava_engagement WHERE {where}", params)
            total = cur.fetchone()["cnt"]

            cur.execute(f"""
                SELECT rgm, score, risk_level, days_since_enrollment,
                       days_since_last_access, interaction_count, total_minutes,
                       detail, snapshot_date
                FROM ava_engagement WHERE {where}
                ORDER BY score ASC, rgm
                LIMIT %s OFFSET %s
            """, params + [per_page, offset])
            rows = cur.fetchall()

            for r in rows:
                r["snapshot_date"] = str(r["snapshot_date"]) if r["snapshot_date"] else None

            cur.execute("""
                SELECT risk_level, COUNT(*) as cnt
                FROM ava_engagement
                WHERE snapshot_date = (SELECT MAX(snapshot_date) FROM ava_engagement)
                GROUP BY risk_level
            """)
            summary = {r["risk_level"]: r["cnt"] for r in cur.fetchall()}

        return jsonify({
            "scores": rows,
            "total": total,
            "page": page,
            "per_page": per_page,
            "summary": summary,
        })
    finally:
        conn.close()


@app.route("/api/engagement/timeline")
def api_engagement_timeline():
    days = int(request.args.get("days", 90))
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT snapshot_date,
                       ROUND(AVG(score)) as avg_score,
                       COUNT(*) as total,
                       SUM(CASE WHEN risk_level='engajado' THEN 1 ELSE 0 END) as engajados,
                       SUM(CASE WHEN risk_level='atencao' THEN 1 ELSE 0 END) as atencao,
                       SUM(CASE WHEN risk_level='em_risco' THEN 1 ELSE 0 END) as em_risco,
                       SUM(CASE WHEN risk_level='critico' THEN 1 ELSE 0 END) as criticos
                FROM ava_engagement
                WHERE snapshot_date >= CURRENT_DATE - %s
                GROUP BY snapshot_date
                ORDER BY snapshot_date
            """, (days,))
            points = cur.fetchall()
            for p in points:
                p["snapshot_date"] = str(p["snapshot_date"])
        return jsonify({"points": points, "days": days})
    finally:
        conn.close()


@app.route("/api/engagement/student/<rgm>")
def api_engagement_student(rgm):
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM ava_engagement WHERE rgm = %s
                ORDER BY snapshot_date DESC LIMIT 30
            """, (rgm,))
            history = cur.fetchall()
            for h in history:
                h["snapshot_date"] = str(h["snapshot_date"]) if h["snapshot_date"] else None

            cur.execute("""
                SELECT * FROM comm_log WHERE rgm = %s
                ORDER BY sent_at DESC LIMIT 20
            """, (rgm,))
            comms = cur.fetchall()
            for c in comms:
                c["sent_at"] = to_brt(c["sent_at"])

        return jsonify({"history": history, "communications": comms})
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Régua de Comunicação (CRUD)
# ---------------------------------------------------------------------------

@app.route("/api/comm/rules", methods=["GET"])
def api_comm_rules_list():
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM comm_rules ORDER BY priority, id")
            rules = cur.fetchall()
            for r in rules:
                r["created_at"] = to_brt(r["created_at"])
                r["updated_at"] = to_brt(r["updated_at"])
        return jsonify({"rules": rules})
    finally:
        conn.close()


@app.route("/api/comm/rules", methods=["POST"])
def api_comm_rules_create():
    b = request.json or {}
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO comm_rules (name, description, audience, trigger_type, trigger_days,
                    channel, escalation_channel, escalation_after_days, message_template,
                    cooldown_days, max_per_week, priority, enabled)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
            """, (
                b.get("name", "Nova Regra"), b.get("description", ""),
                b.get("audience", "todos"), b.get("trigger_type", "inatividade"),
                int(b.get("trigger_days", 7)), b.get("channel", "email"),
                b.get("escalation_channel"), b.get("escalation_after_days"),
                b.get("message_template", ""), int(b.get("cooldown_days", 3)),
                int(b.get("max_per_week", 2)), int(b.get("priority", 0)),
                b.get("enabled", True),
            ))
            new_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({"ok": True, "id": new_id})
    finally:
        conn.close()


@app.route("/api/comm/rules/<int:rid>", methods=["PUT"])
def api_comm_rules_update(rid):
    b = request.json or {}
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE comm_rules SET
                    name=%s, description=%s, audience=%s, trigger_type=%s, trigger_days=%s,
                    channel=%s, escalation_channel=%s, escalation_after_days=%s,
                    message_template=%s, cooldown_days=%s, max_per_week=%s,
                    priority=%s, enabled=%s, updated_at=NOW()
                WHERE id=%s
            """, (
                b.get("name"), b.get("description"), b.get("audience"),
                b.get("trigger_type"), int(b.get("trigger_days", 7)),
                b.get("channel"), b.get("escalation_channel"),
                b.get("escalation_after_days"), b.get("message_template"),
                int(b.get("cooldown_days", 3)), int(b.get("max_per_week", 2)),
                int(b.get("priority", 0)), b.get("enabled", True), rid,
            ))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route("/api/comm/rules/<int:rid>", methods=["DELETE"])
def api_comm_rules_delete(rid):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM comm_rules WHERE id = %s", (rid,))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Motor de Régua — Avaliação de Gatilhos
# ---------------------------------------------------------------------------

def _render_template(template, data):
    """Replace {{var}} placeholders in a template string."""
    result = template
    for key, val in data.items():
        result = result.replace("{{" + key + "}}", str(val or ""))
    return result


def evaluate_comm_triggers():
    """Daily job: recalculate scores, evaluate rules, enqueue communications."""
    app.logger.info("[ENGAGEMENT] Iniciando avaliação diária de gatilhos")
    calc_result = calculate_engagement_scores()
    app.logger.info("[ENGAGEMENT] Scores recalculados: %s", calc_result)

    conn = get_conn()
    today = datetime.now(BRT)
    today_date = today.date()
    weekday = today.weekday()
    if weekday >= 5:
        app.logger.info("[ENGAGEMENT] Fim de semana — disparos adiados")
        conn.close()
        return

    try:
        enqueued = 0
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM comm_rules WHERE enabled = TRUE ORDER BY priority")
            rules = cur.fetchall()
            if not rules:
                app.logger.info("[ENGAGEMENT] Nenhuma regra ativa")
                conn.close()
                return

            cur.execute("""
                SELECT rgm, score, risk_level, days_since_enrollment,
                       days_since_last_access, detail
                FROM ava_engagement
                WHERE snapshot_date = %s AND risk_level IN ('critico', 'em_risco', 'atencao')
            """, (today_date,))
            students = cur.fetchall()

            for student in students:
                rgm = student["rgm"]
                detail = student["detail"] or {}
                days_enrolled = student["days_since_enrollment"]
                days_no_access = student["days_since_last_access"]
                is_new = days_enrolled is not None and days_enrolled <= NEW_STUDENT_DAYS

                cur.execute("""
                    SELECT COUNT(*) as cnt FROM comm_queue
                    WHERE rgm = %s AND created_at >= %s - INTERVAL '7 days'
                    AND status IN ('pendente', 'enviado')
                """, (rgm, today))
                week_count = cur.fetchone()["cnt"]

                cur.execute("""
                    SELECT MAX(created_at) as last_comm FROM comm_queue
                    WHERE rgm = %s AND status IN ('pendente', 'enviado')
                """, (rgm,))
                last_comm_row = cur.fetchone()
                last_comm = last_comm_row["last_comm"] if last_comm_row else None

                for rule in rules:
                    if rule["audience"] == "novo_aluno" and not is_new:
                        continue
                    if rule["audience"] == "veterano" and is_new:
                        continue

                    if week_count >= rule["max_per_week"]:
                        continue

                    if last_comm and (today - last_comm).days < rule["cooldown_days"]:
                        continue

                    trigger_days = rule["trigger_days"]
                    matches = False
                    if rule["trigger_type"] == "sem_acesso_inicial":
                        if is_new and days_enrolled is not None and days_enrolled >= trigger_days:
                            if days_no_access is None or days_no_access >= trigger_days:
                                matches = True
                    elif rule["trigger_type"] == "inatividade":
                        if days_no_access is not None and days_no_access >= trigger_days:
                            matches = True
                        elif days_no_access is None and days_enrolled and days_enrolled >= trigger_days:
                            matches = True
                    elif rule["trigger_type"] == "score_baixo":
                        if student["score"] <= trigger_days:
                            matches = True

                    if not matches:
                        continue

                    cur.execute("""
                        SELECT id FROM comm_queue
                        WHERE rgm = %s AND rule_id = %s
                        AND created_at >= %s - INTERVAL '30 days'
                        AND status IN ('pendente', 'enviado')
                    """, (rgm, rule["id"], today))
                    if cur.fetchone():
                        continue

                    nome = detail.get("nome", "")
                    primeiro_nome = nome.split()[0] if nome else ""
                    tpl_data = {
                        "nome": nome,
                        "primeiro_nome": primeiro_nome,
                        "curso": detail.get("curso", ""),
                        "polo": detail.get("polo", ""),
                        "email": detail.get("email", ""),
                        "dias_sem_acesso": str(days_no_access or "N/A"),
                        "score": str(student["score"]),
                        "rgm": rgm,
                    }
                    message = _render_template(rule["message_template"], tpl_data)

                    payload = {
                        "event": "ava_engagement_alert",
                        "channel": rule["channel"],
                        "student": {
                            "name": nome,
                            "first_name": primeiro_nome,
                            "email": detail.get("email", ""),
                            "rgm": rgm,
                            "curso": detail.get("curso", ""),
                            "polo": detail.get("polo", ""),
                        },
                        "rule": {
                            "id": rule["id"],
                            "name": rule["name"],
                        },
                        "message": message,
                        "context": {
                            "days_since_enrollment": days_enrolled,
                            "days_since_last_access": days_no_access,
                            "engagement_score": student["score"],
                            "risk_level": student["risk_level"],
                        },
                    }

                    cur.execute("""
                        INSERT INTO comm_queue (rgm, rule_id, channel, status, payload, scheduled_for)
                        VALUES (%s, %s, %s, 'pendente', %s, NOW())
                    """, (rgm, rule["id"], rule["channel"], json.dumps(payload, ensure_ascii=False)))
                    enqueued += 1
                    break

        conn.commit()
        app.logger.info("[ENGAGEMENT] %d comunicações enfileiradas", enqueued)

        _dispatch_pending_comms()

    except Exception as e:
        conn.rollback()
        app.logger.error("[ENGAGEMENT] Erro na avaliação: %s", e)
    finally:
        conn.close()


def _dispatch_pending_comms():
    """Send pending communications to n8n webhook."""
    conn = get_conn()
    sent = 0
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, rgm, rule_id, channel, payload
                FROM comm_queue
                WHERE status = 'pendente'
                ORDER BY created_at
                LIMIT 50
            """)
            pending = cur.fetchall()

            for item in pending:
                try:
                    payload = item["payload"] or {}
                    payload["callback_url"] = request.host_url.rstrip("/") + "/api/comm/callback" if hasattr(request, "host_url") else ""

                    r = _requests.post(N8N_COMM_WEBHOOK, json=payload, timeout=15)
                    n8n_resp = {}
                    try:
                        n8n_resp = r.json() if r.text.strip() else {}
                    except Exception:
                        n8n_resp = {"status_code": r.status_code}

                    if r.ok:
                        cur.execute("""
                            UPDATE comm_queue SET status='enviado', sent_at=NOW(), n8n_response=%s
                            WHERE id=%s
                        """, (json.dumps(n8n_resp), item["id"]))
                        msg_preview = (payload.get("message") or "")[:200]
                        cur.execute("""
                            INSERT INTO comm_log (rgm, rule_id, channel, message_preview, status, metadata)
                            VALUES (%s, %s, %s, %s, 'enviado', %s)
                        """, (item["rgm"], item["rule_id"], item["channel"],
                              msg_preview, json.dumps(n8n_resp)))
                        sent += 1
                    else:
                        cur.execute("""
                            UPDATE comm_queue SET status='falha', n8n_response=%s WHERE id=%s
                        """, (json.dumps(n8n_resp), item["id"]))
                        cur.execute("""
                            INSERT INTO comm_log (rgm, rule_id, channel, message_preview, status, metadata)
                            VALUES (%s, %s, %s, %s, 'falha', %s)
                        """, (item["rgm"], item["rule_id"], item["channel"],
                              "FALHA NO ENVIO", json.dumps(n8n_resp)))

                except Exception as e:
                    app.logger.warning("[COMM] Erro ao enviar para n8n (queue %d): %s", item["id"], e)
                    cur.execute("UPDATE comm_queue SET status='falha' WHERE id=%s", (item["id"],))

        conn.commit()
        app.logger.info("[COMM] %d/%d comunicações enviadas ao n8n", sent, len(pending))
    except Exception as e:
        conn.rollback()
        app.logger.error("[COMM] Erro no dispatch: %s", e)
    finally:
        conn.close()


def _register_engagement_job():
    """Register the daily engagement evaluation job."""
    try:
        scheduler.remove_job("engagement_daily")
    except Exception:
        pass
    trigger = CronTrigger(hour=8, minute=0, timezone="America/Sao_Paulo")
    scheduler.add_job(
        evaluate_comm_triggers,
        trigger=trigger,
        id="engagement_daily",
        replace_existing=True,
        misfire_grace_time=3600,
    )
    app.logger.info("Engagement daily job registered (08:00 BRT)")


@app.route("/api/comm/evaluate", methods=["POST"])
def api_comm_evaluate():
    """Manually trigger the engagement evaluation."""
    def _run():
        with app.app_context():
            evaluate_comm_triggers()
    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify({"ok": True, "message": "Avaliação iniciada em background"})


@app.route("/api/comm/queue")
def api_comm_queue():
    status = request.args.get("status", "").strip()
    limit = int(request.args.get("limit", 50))
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if status:
                cur.execute("""
                    SELECT q.*, r.name as rule_name FROM comm_queue q
                    LEFT JOIN comm_rules r ON r.id = q.rule_id
                    WHERE q.status = %s ORDER BY q.created_at DESC LIMIT %s
                """, (status, limit))
            else:
                cur.execute("""
                    SELECT q.*, r.name as rule_name FROM comm_queue q
                    LEFT JOIN comm_rules r ON r.id = q.rule_id
                    ORDER BY q.created_at DESC LIMIT %s
                """, (limit,))
            items = cur.fetchall()
            for i in items:
                i["created_at"] = to_brt(i["created_at"])
                i["sent_at"] = to_brt(i["sent_at"])
                i["scheduled_for"] = to_brt(i["scheduled_for"])
        return jsonify({"queue": items})
    finally:
        conn.close()


@app.route("/api/comm/log")
def api_comm_log_list():
    limit = int(request.args.get("limit", 50))
    channel = request.args.get("channel", "").strip()
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if channel:
                cur.execute("""
                    SELECT l.*, r.name as rule_name FROM comm_log l
                    LEFT JOIN comm_rules r ON r.id = l.rule_id
                    WHERE l.channel = %s ORDER BY l.sent_at DESC LIMIT %s
                """, (channel, limit))
            else:
                cur.execute("""
                    SELECT l.*, r.name as rule_name FROM comm_log l
                    LEFT JOIN comm_rules r ON r.id = l.rule_id
                    ORDER BY l.sent_at DESC LIMIT %s
                """, (limit,))
            items = cur.fetchall()
            for i in items:
                i["sent_at"] = to_brt(i["sent_at"])
        return jsonify({"log": items})
    finally:
        conn.close()


@app.route("/api/comm/callback", methods=["POST"])
def api_comm_callback():
    """Receive delivery status from n8n."""
    b = request.json or {}
    rgm = b.get("rgm", "")
    status = b.get("status", "")
    queue_id = b.get("queue_id")

    if not rgm and not queue_id:
        return jsonify({"error": "rgm or queue_id required"}), 400

    conn = get_conn()
    try:
        with conn.cursor() as cur:
            if queue_id:
                cur.execute("UPDATE comm_queue SET status=%s WHERE id=%s", (status, queue_id))
            if rgm and status:
                cur.execute("""
                    UPDATE comm_log SET status=%s
                    WHERE rgm=%s AND id = (
                        SELECT id FROM comm_log WHERE rgm=%s ORDER BY sent_at DESC LIMIT 1
                    )
                """, (status, rgm, rgm))

            if status == "respondido" and rgm:
                cur.execute("""
                    UPDATE comm_queue SET status='cancelado'
                    WHERE rgm=%s AND status='pendente'
                """, (rgm,))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route("/api/comm/dispatch", methods=["POST"])
def api_comm_dispatch_manual():
    """Manually dispatch pending communications."""
    try:
        _dispatch_pending_comms()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Rotas — Upload
# ---------------------------------------------------------------------------

UPLOAD_DIR = BASE_DIR


def _find_xlsx():
    for f in UPLOAD_DIR.iterdir():
        if f.suffix.lower() == ".xlsx" and "matriculados" in f.name.lower():
            stat = f.stat()
            return {
                "name": f.name,
                "size": stat.st_size,
                "modified": to_brt(datetime.fromtimestamp(stat.st_mtime, tz=BRT)),
            }
    return None


@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Nenhum arquivo selecionado."}), 400

    fname_lower = f.filename.lower()
    allowed_ext = (".xlsx", ".xlsm", ".zip")
    if not any(fname_lower.endswith(ext) for ext in allowed_ext):
        return jsonify({"error": "Aceitos: .xlsx, .xlsm ou .zip"}), 400

    tipo = request.form.get("tipo", "matriculados").strip().lower()
    if tipo not in XL_TIPOS:
        return jsonify({"error": f"Tipo inválido. Use: {', '.join(XL_TIPOS)}"}), 400

    if tipo == "matriculados":
        for old in UPLOAD_DIR.iterdir():
            if old.suffix.lower() == ".xlsx" and "matriculados" in old.name.lower():
                old.unlink()

    safe_name = f.filename
    if tipo == "matriculados" and "matriculados" not in safe_name.lower():
        safe_name = "Relação de matriculados por polo.xlsx"

    dest = UPLOAD_DIR / safe_name
    f.save(str(dest))

    try:
        if fname_lower.endswith(".zip"):
            snap_count = _handle_zip_upload(str(dest), tipo)
        elif tipo == "inadimplentes" and fname_lower.endswith(".xlsm"):
            tmp_dir = UPLOAD_DIR / f"_tmp_{tipo}"
            tmp_dir.mkdir(exist_ok=True)
            import shutil
            shutil.copy2(str(dest), str(tmp_dir / safe_name))
            entries = _parse_inadimplentes_batch(str(tmp_dir))
            snap_count = _persist_snapshot_entries(entries, tipo, safe_name) if entries else 0
        elif tipo == "sem_rematricula" and fname_lower.endswith((".xlsx", ".xlsm")):
            import shutil
            staging = UPLOAD_DIR / "_staging_sem_rematricula"
            staging.mkdir(exist_ok=True)
            subtipo = request.form.get("subtipo", "").strip().lower()
            if subtipo in ("adimplente", "inadimplente"):
                canonical = f"{subtipo}s.xlsx"
            elif "inadimplente" in safe_name.lower():
                canonical = "inadimplentes.xlsx"
            elif "adimplente" in safe_name.lower():
                canonical = "adimplentes.xlsx"
            else:
                canonical = safe_name
            shutil.copy2(str(dest), str(staging / canonical))
            staged_files = [p.name for p in staging.iterdir() if p.suffix.lower() in (".xlsx", ".xlsm")]
            app.logger.info("sem_rematricula staging: salvou '%s' como '%s'. Arquivos: %s", safe_name, canonical, staged_files)
            entries = _parse_sem_rematricula(str(staging))
            if entries:
                snap_count = _persist_snapshot_entries(entries, tipo, safe_name)
                app.logger.info("sem_rematricula snapshot criado: %d linhas", snap_count)
            else:
                snap_count = 0
                app.logger.info("sem_rematricula: '%s' salvo, aguardando o outro arquivo", canonical)
        else:
            snap_count = _save_xl_snapshot(str(dest), safe_name, tipo)
    except Exception as e:
        import traceback
        app.logger.warning("Erro ao gravar snapshot (%s): %s", tipo, e)
        app.logger.warning("Traceback: %s", traceback.format_exc())
        snap_count = -1

    stat = dest.stat()
    return jsonify({
        "ok": True,
        "tipo": tipo,
        "file": {
            "name": dest.name,
            "size": stat.st_size,
            "modified": to_brt(datetime.fromtimestamp(stat.st_mtime, tz=BRT)),
        },
        "snapshot_rows": snap_count,
    })


def _handle_zip_upload(zip_path, tipo):
    """Extrai um .zip e processa conforme o tipo."""
    import zipfile, shutil, tempfile

    tmp_dir = Path(tempfile.mkdtemp(prefix=f"eduit_{tipo}_"))
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(str(tmp_dir))

        zip_name = Path(zip_path).name

        if tipo == "inadimplentes":
            entries = _parse_inadimplentes_batch(str(tmp_dir))
            return _persist_snapshot_entries(entries, tipo, zip_name) if entries else 0
        elif tipo == "sem_rematricula":
            entries = _parse_sem_rematricula(str(tmp_dir))
            return _persist_snapshot_entries(entries, tipo, zip_name) if entries else 0
        else:
            xlsx_files = list(tmp_dir.glob("*.xlsx"))
            if xlsx_files:
                return _save_xl_snapshot(str(xlsx_files[0]), zip_name, tipo)
            return 0
    finally:
        shutil.rmtree(str(tmp_dir), ignore_errors=True)


@app.route("/api/upload-folder", methods=["POST"])
def api_upload_folder():
    """Processa pastas já presentes no servidor (deploy/scp)."""
    body = request.json or {}
    tipo = body.get("tipo", "").strip().lower()
    folder = body.get("path", "").strip()

    if tipo not in XL_TIPOS:
        return jsonify({"error": f"Tipo inválido. Use: {', '.join(XL_TIPOS)}"}), 400

    if not folder:
        default_folders = {
            "inadimplentes": str(BASE_DIR / "Inadimplentes"),
            "sem_rematricula": str(BASE_DIR / "Sem_Rematricula"),
            "concluintes": str(BASE_DIR / "Concluíntes"),
            "acesso_ava": str(BASE_DIR / "Acesso_AVA"),
        }
        folder = default_folders.get(tipo, "")

    if not folder or not os.path.isdir(folder):
        return jsonify({"error": f"Pasta não encontrada: {folder}"}), 404

    try:
        folder_name = os.path.basename(folder)
        if tipo == "inadimplentes":
            entries = _parse_inadimplentes_batch(folder)
            count = _persist_snapshot_entries(entries, tipo, f"{folder_name} (servidor)") if entries else 0
        elif tipo == "sem_rematricula":
            entries = _parse_sem_rematricula(folder)
            count = _persist_snapshot_entries(entries, tipo, f"{folder_name} (servidor)") if entries else 0
        else:
            xlsx_files = sorted(Path(folder).glob("*.xlsx"))
            if not xlsx_files:
                return jsonify({"error": "Nenhum arquivo .xlsx encontrado na pasta."}), 404
            count = _save_xl_snapshot(str(xlsx_files[0]), xlsx_files[0].name, tipo)

        return jsonify({"ok": True, "tipo": tipo, "snapshot_rows": count, "folder": folder})
    except Exception as e:
        app.logger.exception("Erro ao processar pasta %s", folder)
        return jsonify({"error": str(e)}), 500


_XL_COLUMN_MAP = {
    "nome": ["Nome", "NOME", "Aluno", "Nome Aluno"],
    "cpf": ["CPF"],
    "rgm": ["RGM", "RGM_ALUN"],
    "curso": ["Curso", "DES_CURS"],
    "polo": ["Polo", "NOME_POL"],
    "serie": ["Série", "Serie"],
    "situacao": ["Situação Matrícula", "Situa", "Situação"],
    "tipo_matricula": ["Tipo Matrícula", "Tipo Matr", "Tipo matricula"],
    "data_mat": ["Data Matrícula", "Data Matr"],
    "email": ["Email", "E-mail"],
    "email_acad": ["Email acadêmico", "Email acad", "Email Acadêmico"],
    "fone_cel": ["Fone celular", "Celular", "Telefone"],
    "fone_res": ["Fone Residencial"],
    "fone_com": ["Fone Comercial"],
    "negocio": ["Negócio", "Neg"],
    "empresa": ["Empresa", "NOM_FILI"],
    "bairro": ["Bairro"],
    "cidade": ["Cidade"],
    "sexo": ["Sexo"],
    "data_nasc": ["Data Nascimento"],
    "ciclo": ["Ciclo"],
    "valor": ["Valor", "Valor Devido", "Saldo"],
    "parcela": ["Parcela", "Parcelas"],
    "vencimento": ["Vencimento", "Data Vencimento"],
    "status_financeiro": ["Status", "Status Financeiro", "Situação Financeira"],
    "data_conclusao": ["Data Conclusão", "Data Formatura", "Conclusão"],
    "periodo": ["Período", "Periodo"],
    "modalidade": ["Modalidade"],
    "instituicao": ["Instituição", "Institui"],
    "ultimo_acesso": ["Ultimo Acesso", "Ult Acesso"],
    "interacoes": ["Interações", "Interacoes"],
    "minutos": ["Minutos"],
    "total_registros": ["Total Registros"],
    "id_polo": ["ID_POLO"],
    "cod_inst": ["COD_INST"],
    "tipo_titulo": ["TIPO_TIT"],
    "descricao_titulo": ["DESCRICA"],
    "nr_titulo": ["NR_TITUL"],
    "dt_emissao": ["DTA_EMIS"],
    "dt_vencimento": ["DTA_VCTO"],
    "desconto": ["DESCONTO"],
    "juros": ["JUROS"],
    "valor_titulo": ["VAL_TITU"],
    "dias_atraso": ["ATRASO"],
    "portador": ["PORTADOR"],
    "apto_rematricula": ["Apto Rematricula", "Apta-Rematricula", "Apto-Rematricula"],
}


def _save_xl_snapshot(filepath, filename, tipo="matriculados"):
    """Lê o xlsx e grava um snapshot no banco de dados."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    if not sheet_names:
        wb.close()
        return 0
    ws = wb[sheet_names[0]]
    first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if not first_row:
        wb.close()
        return 0
    header = [cell.value for cell in first_row]
    col_map = {h: i for i, h in enumerate(header) if h}

    def _find(names):
        for n in names:
            if n in col_map:
                return col_map[n]
            for k in col_map:
                if k and n.lower() in k.lower():
                    return col_map[k]
        return None

    idx = {}
    for field, aliases in _XL_COLUMN_MAP.items():
        pos = _find(aliases)
        if pos is not None:
            idx[field] = pos

    unmapped = {}
    for h, i in col_map.items():
        if i not in idx.values() and h:
            safe_key = re.sub(r"\W+", "_", h.strip().lower())[:40]
            if safe_key:
                unmapped[safe_key] = i

    def _get(row, col_idx):
        if col_idx is None or col_idx >= len(row):
            return ""
        v = row[col_idx]
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y")
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v).strip()

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
            continue
        entry = {k: _get(row, v) for k, v in idx.items()}
        for k, v in unmapped.items():
            val = _get(row, v)
            if val:
                entry[k] = val
        entry["cpf_digits"] = _normalize_digits(entry.get("cpf", ""))
        entry["rgm_digits"] = _normalize_digits(entry.get("rgm", ""))
        phones = []
        for pk in ("fone_cel", "fone_res", "fone_com"):
            d = _normalize_digits(entry.get(pk, ""))
            if d:
                phones.append(d)
        entry["phones_digits"] = phones
        entries.append(entry)
    wb.close()

    return _persist_snapshot_entries(entries, tipo, filename)


def _persist_snapshot_entries(entries, tipo, filename):
    """Grava uma lista de dicts no banco como snapshot e retorna row count."""
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO xl_snapshots (tipo, filename, row_count) VALUES (%s, %s, %s) RETURNING id",
                (tipo, filename, len(entries)),
            )
            snap_id = cur.fetchone()[0]

            batch = [(snap_id, json.dumps(e, ensure_ascii=False)) for e in entries]
            psycopg2.extras.execute_batch(
                cur,
                "INSERT INTO xl_rows (snapshot_id, data) VALUES (%s, %s::jsonb)",
                batch,
                page_size=500,
            )
        conn.commit()
    finally:
        conn.close()

    try:
        _compute_snapshot_stats(snap_id, tipo)
    except Exception as e:
        app.logger.warning("Erro ao computar stats para snapshot %s: %s", snap_id, e)

    return len(entries)


def _parse_inadimplentes_batch(folder_path):
    """Consolida todos os .xlsm de inadimplentes de um diretório em uma lista de entries."""
    import openpyxl, glob as _glob

    HEADER_MAP = {
        "ID_POLO": "id_polo", "NOME_POL": "polo", "RGM_ALUN": "rgm",
        "NOME": "nome", "COD_INST": "cod_inst", "NOM_FILI": "empresa",
        "DES_CURS": "curso", "TIPO_TIT": "tipo_titulo", "NR_TITUL": "nr_titulo",
        "DTA_EMIS": "dt_emissao", "DTA_VCTO": "dt_vencimento",
        "DESCONTO": "desconto", "JUROS": "juros", "VAL_TITU": "valor_titulo",
        "ATRASO": "dias_atraso", "PORTADOR": "portador",
    }
    DESCRICA_COLS = ["descricao_titulo", "portador_nome"]

    files = sorted(
        _glob.glob(os.path.join(folder_path, "*.xlsm"))
        + _glob.glob(os.path.join(folder_path, "*.xlsx"))
    )
    if not files:
        return []

    raw_rows = []
    for fpath in files:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            header_row = None
            for i, row in enumerate(ws.iter_rows(max_col=18, values_only=True), 1):
                if row and row[0] and str(row[0]).strip().upper() == "ID_POLO":
                    header_row = i
                    headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(row)]
                    break
            if not header_row:
                wb.close()
                continue
            for row in ws.iter_rows(min_row=header_row + 1, max_col=18, values_only=True):
                if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                    continue
                entry = {}
                descrica_idx = 0
                for j, val in enumerate(row):
                    if j >= len(headers):
                        break
                    h = headers[j].upper().strip()
                    if h == "DESCRICA":
                        key = DESCRICA_COLS[descrica_idx] if descrica_idx < len(DESCRICA_COLS) else f"descrica_{descrica_idx}"
                        descrica_idx += 1
                    else:
                        key = HEADER_MAP.get(h, h.lower()[:40])
                    v = val
                    if v is None:
                        v = ""
                    elif isinstance(v, float) and key in ("rgm", "id_polo", "cod_inst", "tipo_titulo", "portador", "dias_atraso"):
                        v = str(int(v))
                    elif isinstance(v, datetime):
                        v = v.strftime("%d/%m/%Y")
                    else:
                        v = str(v).strip()
                        if key in ("rgm",) and v.endswith(".0"):
                            v = v[:-2]
                    entry[key] = v
                entry["rgm_digits"] = _normalize_digits(entry.get("rgm", ""))
                for fk in ("valor_titulo", "desconto", "juros"):
                    try:
                        entry[fk] = str(round(float(entry.get(fk, "0") or "0"), 2))
                    except (ValueError, TypeError):
                        pass
                raw_rows.append(entry)
            wb.close()
        except Exception as exc:
            app.logger.warning("Erro ao processar %s: %s", fpath, exc)

    aggregated = {}
    for row in raw_rows:
        rgm = row.get("rgm_digits", "")
        if not rgm:
            rgm = row.get("rgm", "unknown")
        if rgm not in aggregated:
            aggregated[rgm] = {
                "nome": row.get("nome", ""),
                "rgm": row.get("rgm", ""),
                "rgm_digits": rgm,
                "polo": row.get("polo", ""),
                "empresa": row.get("empresa", ""),
                "curso": row.get("curso", ""),
                "cpf_digits": "",
                "phones_digits": [],
                "total_titulos": 0,
                "valor_total": 0.0,
                "max_atraso": 0,
                "titulos": [],
            }
        agg = aggregated[rgm]
        agg["total_titulos"] += 1
        try:
            agg["valor_total"] += float(row.get("valor_titulo", "0") or "0")
        except (ValueError, TypeError):
            pass
        try:
            atraso = int(float(row.get("dias_atraso", "0") or "0"))
            if atraso > agg["max_atraso"]:
                agg["max_atraso"] = atraso
        except (ValueError, TypeError):
            pass
        agg["titulos"].append({
            "tipo": row.get("tipo_titulo", ""),
            "descricao": row.get("descricao_titulo", ""),
            "nr": row.get("nr_titulo", ""),
            "dt_emissao": row.get("dt_emissao", ""),
            "dt_vencimento": row.get("dt_vencimento", ""),
            "valor": row.get("valor_titulo", ""),
            "desconto": row.get("desconto", ""),
            "juros": row.get("juros", ""),
            "atraso": row.get("dias_atraso", ""),
            "portador": row.get("portador", ""),
            "portador_nome": row.get("portador_nome", ""),
        })

    entries = []
    for agg in aggregated.values():
        agg["valor_total"] = str(round(agg["valor_total"], 2))
        agg["max_atraso"] = str(agg["max_atraso"])
        agg["total_titulos"] = str(agg["total_titulos"])
        entries.append(agg)
    return entries


def _parse_sem_rematricula(folder_path):
    """Lê adimplentes.xlsx e inadimplentes.xlsx, unifica com flag financeiro."""
    import openpyxl

    HEADER_NORM = {
        "polo": "polo", "aluno": "nome", "telefone": "fone_cel",
        "e-mail": "email", "e_mail": "email", "rgm": "rgm",
        "serie": "serie", "série": "serie",
        "curso": "curso",
        "apto rematricula": "apto_rematricula",
        "apta-rematricula": "apto_rematricula",
        "apto-rematricula": "apto_rematricula",
    }

    def _find_file(folder, keyword, exclude_keyword=None):
        exact = os.path.join(folder, f"{keyword}s.xlsx")
        if os.path.isfile(exact):
            return exact
        exact2 = os.path.join(folder, f"{keyword}.xlsx")
        if os.path.isfile(exact2):
            return exact2
        for fn in os.listdir(folder):
            fl = fn.lower()
            if not fl.endswith((".xlsx", ".xlsm")):
                continue
            if keyword in fl and (exclude_keyword is None or exclude_keyword not in fl):
                return os.path.join(folder, fn)
        return None

    entries = []
    for keyword, flag, excl in [("adimplente", "adimplente", "inadimplente"), ("inadimplente", "inadimplente", None)]:
        fpath = _find_file(folder_path, keyword, exclude_keyword=excl)
        if not fpath:
            continue
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        if not wb.sheetnames:
            wb.close()
            continue
        ws = wb[wb.sheetnames[0]]
        first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not first_row:
            wb.close()
            continue
        raw_header = [c.value for c in first_row]
        col_idx = {}
        for i, h in enumerate(raw_header):
            if h:
                key = HEADER_NORM.get(h.strip().lower(), h.strip().lower()[:40])
                col_idx[key] = i

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                continue
            entry = {}
            for key, idx in col_idx.items():
                v = row[idx] if idx < len(row) else None
                if v is None:
                    v = ""
                elif isinstance(v, datetime):
                    v = v.strftime("%d/%m/%Y")
                elif isinstance(v, float) and v == int(v):
                    v = str(int(v))
                else:
                    v = str(v).strip()
                entry[key] = v
            entry["status_financeiro"] = flag
            entry["rgm_digits"] = _normalize_digits(entry.get("rgm", ""))
            entry["cpf_digits"] = _normalize_digits(entry.get("cpf", ""))
            phones = []
            d = _normalize_digits(entry.get("fone_cel", ""))
            if d:
                phones.append(d)
            entry["phones_digits"] = phones
            entries.append(entry)
        wb.close()
    return entries


# ---------------------------------------------------------------------------
# Rotas — Explorador de Logs
# ---------------------------------------------------------------------------

SAFE_LOG_DIRS = [LOG_DIR, REPORTS_DIR]


def _list_log_files():
    files = []
    for d in SAFE_LOG_DIRS:
        if not d.exists():
            continue
        for f in d.iterdir():
            if f.is_file() and f.suffix.lower() in (".csv", ".log", ".txt"):
                stat = f.stat()
                files.append({
                    "name": f.name,
                    "dir": d.name,
                    "path": f"{d.name}/{f.name}",
                    "size": stat.st_size,
                    "modified": to_brt(datetime.fromtimestamp(stat.st_mtime, tz=BRT)),
                })
    files.sort(key=lambda x: x["modified"], reverse=True)
    return files


def _resolve_log_path(filepath):
    """Resolve and validate a log file path, preventing directory traversal."""
    filepath = filepath.replace("\\", "/")
    if ".." in filepath:
        return None
    for d in SAFE_LOG_DIRS:
        candidate = d.parent / filepath
        try:
            candidate = candidate.resolve()
            if candidate.is_file() and any(str(candidate).startswith(str(sd.resolve())) for sd in SAFE_LOG_DIRS):
                return candidate
        except Exception:
            pass
    return None


@app.route("/api/logs")
def api_logs_list():
    return jsonify({"files": _list_log_files()})


@app.route("/api/logs/view/<path:filepath>")
def api_logs_view(filepath):
    fpath = _resolve_log_path(filepath)
    if not fpath:
        return jsonify({"error": "Arquivo não encontrado."}), 404

    tail = int(request.args.get("tail", 200))

    try:
        with open(fpath, "r", encoding="utf-8-sig", errors="replace") as f:
            lines = f.readlines()

        total = len(lines)
        if tail and tail < total:
            lines = lines[-tail:]

        return jsonify({
            "name": fpath.name,
            "total_lines": total,
            "showing": len(lines),
            "lines": [l.rstrip() for l in lines],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/logs/download/<path:filepath>")
def api_logs_download(filepath):
    fpath = _resolve_log_path(filepath)
    if not fpath:
        return jsonify({"error": "Arquivo não encontrado."}), 404
    return send_file(str(fpath), as_attachment=True)


# ---------------------------------------------------------------------------
# Rotas — Agendamento (Schedules)
# ---------------------------------------------------------------------------

@app.route("/api/schedules")
def api_schedules_list():
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM schedules ORDER BY created_at")
            rows = []
            for r in cur.fetchall():
                row = dict(r)
                for k, v in row.items():
                    if isinstance(v, datetime):
                        row[k] = to_brt(v)
                rows.append(row)

        # Add next run info from scheduler
        for row in rows:
            job = scheduler.get_job(row["id"])
            if job and job.next_run_time:
                row["next_run"] = to_brt(job.next_run_time)
            else:
                row["next_run"] = None

        return jsonify({"schedules": rows})
    except Exception as e:
        return jsonify({"schedules": [], "error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/schedules", methods=["POST"])
def api_schedules_save():
    data = request.json
    if not data:
        return jsonify({"error": "Dados inválidos."}), 400

    job_type = data.get("job_type", "")
    if job_type not in ("sync_delta", "sync_full"):
        return jsonify({"error": "Tipo inválido. Use 'sync_delta' ou 'sync_full'."}), 400

    cron_days = data.get("cron_days", "*")
    cron_hour = int(data.get("cron_hour", 2))
    cron_minute = int(data.get("cron_minute", 0))
    enabled = bool(data.get("enabled", True))
    schedule_id = data.get("id") or f"{job_type}_{cron_hour:02d}{cron_minute:02d}"

    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO schedules (id, job_type, cron_days, cron_hour, cron_minute, enabled)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (id) DO UPDATE SET
                    job_type = EXCLUDED.job_type,
                    cron_days = EXCLUDED.cron_days,
                    cron_hour = EXCLUDED.cron_hour,
                    cron_minute = EXCLUDED.cron_minute,
                    enabled = EXCLUDED.enabled
            """, (schedule_id, job_type, cron_days, cron_hour, cron_minute, enabled))
        conn.commit()

        _register_schedule_job(schedule_id, job_type, cron_days, cron_hour, cron_minute, enabled)

        return jsonify({"ok": True, "id": schedule_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/schedules/<schedule_id>", methods=["DELETE"])
def api_schedules_delete(schedule_id):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM schedules WHERE id = %s", (schedule_id,))
        conn.commit()

        try:
            scheduler.remove_job(schedule_id)
        except Exception:
            pass

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/schedules/<schedule_id>/toggle", methods=["POST"])
def api_schedules_toggle(schedule_id):
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("UPDATE schedules SET enabled = NOT enabled WHERE id = %s RETURNING *", (schedule_id,))
            row = cur.fetchone()
            if not row:
                return jsonify({"error": "Agendamento não encontrado."}), 404
        conn.commit()

        _register_schedule_job(
            row["id"], row["job_type"], row["cron_days"],
            row["cron_hour"], row["cron_minute"], row["enabled"],
        )

        return jsonify({"ok": True, "enabled": row["enabled"]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Rotas — Debug
# ---------------------------------------------------------------------------

@app.route("/api/debug")
def api_debug():
    return jsonify({
        "sync_running": _sync_running,
        "sync_proc_alive": _sync_proc is not None and _sync_proc.poll() is None if _sync_proc else False,
        "sync_log_count": len(_sync_logs),
        "sync_logs_last5": list(_sync_logs)[-5:] if _sync_logs else [],
        "update_running": _update_running,
        "update_log_count": len(_update_logs),
        "python": sys.executable,
        "sync_script": SYNC_SCRIPT,
        "sync_script_exists": Path(SYNC_SCRIPT).exists(),
        "cwd": str(BASE_DIR),
    })


# ---------------------------------------------------------------------------
# APScheduler
# ---------------------------------------------------------------------------

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

scheduler = BackgroundScheduler(timezone="America/Sao_Paulo")

DAY_MAP = {"0": "mon", "1": "tue", "2": "wed", "3": "thu", "4": "fri", "5": "sat", "6": "sun"}


def _run_scheduled_sync(job_type):
    """Executa sync agendado (roda no thread do scheduler)."""
    global _sync_running, _sync_proc

    if _sync_running:
        app.logger.info("Scheduled %s skipped — sync already running", job_type)
        return

    mode = "full" if job_type == "sync_full" else "delta"
    _sync_running = True
    _sync_logs.clear()

    try:
        cmd = [sys.executable, SYNC_SCRIPT]
        if mode == "full":
            cmd.append("--full")

        _add_sync_log(f"[AGENDADO] Sincronização {mode.upper()} iniciada automaticamente")

        env = {**os.environ, "PYTHONUNBUFFERED": "1"}
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, cwd=str(BASE_DIR), env=env,
        )
        _sync_proc = proc

        for line in proc.stdout:
            _add_sync_log(line)

        proc.wait()

        if proc.returncode == 0:
            _add_sync_log("[FIM] Sincronização agendada concluída com sucesso")
        else:
            _add_sync_log(f"[ERRO] Sincronização agendada falhou (exit code {proc.returncode})")

        # Update last_run_at
        try:
            conn = get_conn()
            with conn.cursor() as cur:
                cur.execute("UPDATE schedules SET last_run_at = NOW() WHERE job_type = %s", (job_type,))
            conn.commit()
            conn.close()
        except Exception:
            pass

    except Exception as e:
        _add_sync_log(f"[ERRO] {e}")
    finally:
        _sync_proc = None
        _sync_running = False


def _register_schedule_job(schedule_id, job_type, cron_days, cron_hour, cron_minute, enabled):
    """Register or update a scheduler job."""
    try:
        scheduler.remove_job(schedule_id)
    except Exception:
        pass

    if not enabled:
        return

    if cron_days == "*":
        day_of_week = "*"
    else:
        parts = [d.strip() for d in cron_days.split(",")]
        day_of_week = ",".join(DAY_MAP.get(p, p) for p in parts)

    trigger = CronTrigger(
        day_of_week=day_of_week,
        hour=cron_hour,
        minute=cron_minute,
        timezone="America/Sao_Paulo",
    )

    scheduler.add_job(
        _run_scheduled_sync,
        trigger=trigger,
        args=[job_type],
        id=schedule_id,
        replace_existing=True,
        misfire_grace_time=300,
    )


def _load_schedules_from_db():
    """Load all schedules from DB and register them in APScheduler."""
    try:
        conn = get_conn()
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM schedules")
            for row in cur.fetchall():
                _register_schedule_job(
                    row["id"], row["job_type"], row["cron_days"],
                    row["cron_hour"], row["cron_minute"], row["enabled"],
                )
        conn.close()
        app.logger.info("Schedules loaded from DB")
    except Exception as e:
        app.logger.warning("Could not load schedules: %s", e)


def _ensure_schedules_table():
    """Create the schedules table if it doesn't exist yet."""
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS schedules (
                    id TEXT PRIMARY KEY,
                    job_type TEXT NOT NULL,
                    cron_days TEXT NOT NULL DEFAULT '*',
                    cron_hour INTEGER NOT NULL DEFAULT 2,
                    cron_minute INTEGER NOT NULL DEFAULT 0,
                    enabled BOOLEAN DEFAULT TRUE,
                    last_run_at TIMESTAMPTZ,
                    created_at TIMESTAMPTZ DEFAULT NOW()
                )
            """)
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.warning("Could not ensure schedules table: %s", e)


def _ensure_turmas_table():
    """Create the turmas table if it doesn't exist yet."""
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS turmas (
                    id         SERIAL PRIMARY KEY,
                    nivel      TEXT NOT NULL,
                    nome       TEXT NOT NULL,
                    dt_inicio  DATE NOT NULL,
                    dt_fim     DATE NOT NULL,
                    ano        INTEGER NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE(nivel, nome)
                )
            """)
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.warning("Could not ensure turmas table: %s", e)


def _ensure_ciclos_table():
    """Create the ciclos table if it doesn't exist yet."""
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ciclos (
                    id         SERIAL PRIMARY KEY,
                    nivel      TEXT NOT NULL,
                    nome       TEXT NOT NULL,
                    dt_inicio  DATE NOT NULL,
                    dt_fim     DATE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE(nivel, nome)
                )
            """)
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.warning("Could not ensure ciclos table: %s", e)


XL_TIPOS = ["matriculados", "inadimplentes", "concluintes", "acesso_ava", "sem_rematricula"]

def _ensure_xl_snapshots_table():
    """Create xl_snapshots + xl_rows tables for spreadsheet history."""
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS xl_snapshots (
                    id          SERIAL PRIMARY KEY,
                    tipo        TEXT NOT NULL DEFAULT 'matriculados',
                    filename    TEXT NOT NULL,
                    row_count   INTEGER NOT NULL DEFAULT 0,
                    uploaded_at TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS xl_rows (
                    id          SERIAL PRIMARY KEY,
                    snapshot_id INTEGER NOT NULL REFERENCES xl_snapshots(id) ON DELETE CASCADE,
                    data        JSONB NOT NULL
                )
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_xl_rows_snapshot
                ON xl_rows(snapshot_id)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_xl_rows_cpf
                ON xl_rows ((data->>'cpf_digits'))
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_xl_rows_rgm
                ON xl_rows ((data->>'rgm'))
            """)
            cur.execute("""
                ALTER TABLE xl_snapshots ADD COLUMN IF NOT EXISTS tipo TEXT NOT NULL DEFAULT 'matriculados'
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS xl_snapshot_stats (
                    id          SERIAL PRIMARY KEY,
                    snapshot_id INTEGER NOT NULL REFERENCES xl_snapshots(id) ON DELETE CASCADE,
                    metric      TEXT NOT NULL,
                    value       NUMERIC,
                    detail      JSONB,
                    UNIQUE(snapshot_id, metric)
                )
            """)
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.warning("Could not ensure xl_snapshots table: %s", e)


def _ensure_users_table():
    """Create app_users + user_permissions tables and seed admin from env."""
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS app_users (
                    id         SERIAL PRIMARY KEY,
                    username   TEXT NOT NULL UNIQUE,
                    pw_hash    TEXT NOT NULL,
                    role       TEXT NOT NULL DEFAULT 'viewer',
                    created_at TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS user_permissions (
                    user_id    INTEGER NOT NULL REFERENCES app_users(id) ON DELETE CASCADE,
                    page       TEXT NOT NULL,
                    PRIMARY KEY (user_id, page)
                )
            """)
            cur.execute("SELECT COUNT(*) FROM app_users")
            if cur.fetchone()[0] == 0 and APP_PASS_FALLBACK:
                cur.execute(
                    "INSERT INTO app_users (username, pw_hash, role) VALUES (%s, %s, 'admin')",
                    (APP_USER_FALLBACK, _hash_pw(APP_PASS_FALLBACK)),
                )
                uid = cur.lastrowid
                cur.execute("SELECT id FROM app_users WHERE username = %s", (APP_USER_FALLBACK,))
                uid = cur.fetchone()[0]
                for page in ALL_PAGES:
                    cur.execute("INSERT INTO user_permissions (user_id, page) VALUES (%s, %s)",
                                (uid, page))
                app.logger.info("Admin user seeded from env vars: %s", APP_USER_FALLBACK)
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.warning("Could not ensure users table: %s", e)


def _ensure_engagement_tables():
    """Create ava_engagement, comm_rules, comm_queue, comm_log tables."""
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ava_engagement (
                    id                   SERIAL PRIMARY KEY,
                    rgm                  TEXT NOT NULL,
                    snapshot_date        DATE NOT NULL DEFAULT CURRENT_DATE,
                    score                INTEGER NOT NULL DEFAULT 0,
                    risk_level           TEXT NOT NULL DEFAULT 'critico',
                    days_since_enrollment INTEGER,
                    days_since_last_access INTEGER,
                    access_count         INTEGER DEFAULT 0,
                    interaction_count    INTEGER DEFAULT 0,
                    total_minutes        NUMERIC DEFAULT 0,
                    detail               JSONB,
                    UNIQUE(rgm, snapshot_date)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS comm_rules (
                    id                   SERIAL PRIMARY KEY,
                    name                 TEXT NOT NULL,
                    description          TEXT DEFAULT '',
                    audience             TEXT NOT NULL DEFAULT 'todos',
                    trigger_type         TEXT NOT NULL DEFAULT 'inatividade',
                    trigger_days         INTEGER NOT NULL DEFAULT 7,
                    channel              TEXT NOT NULL DEFAULT 'email',
                    escalation_channel   TEXT,
                    escalation_after_days INTEGER,
                    message_template     TEXT NOT NULL DEFAULT '',
                    cooldown_days        INTEGER NOT NULL DEFAULT 3,
                    max_per_week         INTEGER NOT NULL DEFAULT 2,
                    priority             INTEGER NOT NULL DEFAULT 0,
                    enabled              BOOLEAN NOT NULL DEFAULT TRUE,
                    created_at           TIMESTAMPTZ DEFAULT NOW(),
                    updated_at           TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS comm_queue (
                    id                   SERIAL PRIMARY KEY,
                    rgm                  TEXT NOT NULL,
                    rule_id              INTEGER REFERENCES comm_rules(id) ON DELETE SET NULL,
                    channel              TEXT NOT NULL,
                    status               TEXT NOT NULL DEFAULT 'pendente',
                    payload              JSONB,
                    scheduled_for        TIMESTAMPTZ DEFAULT NOW(),
                    sent_at              TIMESTAMPTZ,
                    n8n_response         JSONB,
                    created_at           TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS comm_log (
                    id                   SERIAL PRIMARY KEY,
                    rgm                  TEXT NOT NULL,
                    rule_id              INTEGER REFERENCES comm_rules(id) ON DELETE SET NULL,
                    channel              TEXT NOT NULL,
                    sent_at              TIMESTAMPTZ DEFAULT NOW(),
                    message_preview      TEXT,
                    status               TEXT NOT NULL DEFAULT 'enviado',
                    metadata             JSONB
                )
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_ava_eng_rgm ON ava_engagement(rgm)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_ava_eng_date ON ava_engagement(snapshot_date)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_ava_eng_risk ON ava_engagement(risk_level)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_comm_queue_status ON comm_queue(status)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_comm_queue_rgm ON comm_queue(rgm)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_comm_log_rgm ON comm_log(rgm)")

            cur.execute("SELECT COUNT(*) FROM comm_rules")
            if cur.fetchone()[0] == 0:
                _seed_default_comm_rules(cur)

        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.warning("Could not ensure engagement tables: %s", e)


def _seed_default_comm_rules(cur):
    """Insert default communication rules (best practices for student retention)."""
    rules = [
        ("Boas-vindas", "Email de boas-vindas com link do AVA", "novo_aluno",
         "sem_acesso_inicial", 0, "email", None, None,
         "Olá {{primeiro_nome}}! Bem-vindo(a) à {{curso}}! Seu ambiente virtual de aprendizagem já está disponível. Acesse agora e comece sua jornada acadêmica.", 3, 2, 10),
        ("Primeiro lembrete - 3 dias", "Lembrete gentil para novos alunos sem acesso", "novo_aluno",
         "sem_acesso_inicial", 3, "email", "whatsapp", 2,
         "Oi {{primeiro_nome}}, notamos que você ainda não acessou o ambiente virtual. Seu espaço de estudos está pronto e esperando por você! Precisa de ajuda?", 3, 2, 20),
        ("Segundo lembrete - 5 dias", "WhatsApp para novos alunos sem acesso", "novo_aluno",
         "sem_acesso_inicial", 5, "whatsapp", None, None,
         "Oi {{primeiro_nome}}! Já se passaram alguns dias desde sua matrícula em {{curso}} e ainda não identificamos seu acesso ao AVA. Precisa de ajuda para entrar? Estamos aqui!", 3, 2, 30),
        ("Alerta - 7 dias sem acesso", "Alerta para novos alunos", "novo_aluno",
         "sem_acesso_inicial", 7, "ambos", None, None,
         "{{primeiro_nome}}, já faz uma semana desde sua matrícula e não identificamos nenhum acesso ao ambiente virtual. É importante iniciar seus estudos o quanto antes. Entre em contato se precisar de suporte.", 3, 2, 40),
        ("Alerta crítico - 14 dias", "Urgência para novos alunos inativos", "novo_aluno",
         "sem_acesso_inicial", 14, "ambos", None, None,
         "{{primeiro_nome}}, notamos que ainda não houve acesso ao AVA desde sua matrícula há 14 dias. Gostaríamos de ajudá-lo(a) a iniciar seus estudos. Por favor, entre em contato conosco.", 5, 1, 50),
        ("Re-engajamento veterano", "Check-in para veteranos inativos há 7 dias", "veterano",
         "inatividade", 7, "email", "whatsapp", 5,
         "Oi {{primeiro_nome}}, sentimos sua falta! Faz alguns dias que você não acessa o AVA. Tem alguma dificuldade? Estamos à disposição.", 5, 2, 60),
        ("Escalação veterano", "WhatsApp para veteranos inativos há 14 dias", "veterano",
         "inatividade", 14, "whatsapp", None, None,
         "{{primeiro_nome}}, tudo bem? Notamos que faz 14 dias sem acessar o ambiente virtual. Podemos ajudar de alguma forma?", 5, 1, 70),
        ("Alerta veterano", "Alerta para veteranos inativos há 21 dias", "veterano",
         "inatividade", 21, "ambos", None, None,
         "{{primeiro_nome}}, já faz 21 dias sem acesso ao AVA. Isso pode impactar seu desempenho acadêmico. Entre em contato para que possamos te ajudar.", 7, 1, 80),
    ]
    for r in rules:
        cur.execute("""
            INSERT INTO comm_rules (name, description, audience, trigger_type, trigger_days,
                channel, escalation_channel, escalation_after_days, message_template,
                cooldown_days, max_per_week, priority)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, r)


# ---------------------------------------------------------------------------
# Rotas — Gestão de usuários
# ---------------------------------------------------------------------------

def _is_admin_or_bootstrap():
    """Allow access if admin role or no users exist yet (first-time setup)."""
    if session.get("role") == "admin":
        return True
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM app_users")
            count = cur.fetchone()[0]
        conn.close()
        return count == 0
    except Exception:
        return False


@app.route("/api/users", methods=["GET"])
def api_users_list():
    if not _is_admin_or_bootstrap():
        return jsonify({"error": "Sem permissão"}), 403
    conn = get_conn()
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT u.id, u.username, u.role, u.created_at,
                   ARRAY(SELECT p.page FROM user_permissions p WHERE p.user_id = u.id ORDER BY p.page) AS pages
            FROM app_users u ORDER BY u.id
        """)
        users = cur.fetchall()
    conn.close()
    for u in users:
        u["created_at"] = to_brt(u["created_at"])
    return jsonify({"users": users, "all_pages": ALL_PAGES})


@app.route("/api/users", methods=["POST"])
def api_users_create():
    if not _is_admin_or_bootstrap():
        return jsonify({"error": "Sem permissão"}), 403
    body = request.json or {}
    username = (body.get("username") or "").strip()
    password = body.get("password", "")
    role = body.get("role", "viewer")
    pages = body.get("pages", [])
    if not username or not password:
        return jsonify({"error": "Usuário e senha são obrigatórios"}), 400
    if role not in ("admin", "viewer"):
        role = "viewer"
    is_bootstrap = session.get("role") != "admin"
    if is_bootstrap:
        role = "admin"
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO app_users (username, pw_hash, role) VALUES (%s, %s, %s) RETURNING id",
                (username, _hash_pw(password), role),
            )
            uid = cur.fetchone()[0]
            if role == "admin":
                pages = list(ALL_PAGES)
            for pg in pages:
                if pg in ALL_PAGES:
                    cur.execute("INSERT INTO user_permissions (user_id, page) VALUES (%s, %s)",
                                (uid, pg))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        conn.close()
        return jsonify({"error": "Usuário já existe"}), 409
    conn.close()
    if is_bootstrap:
        session["user_id"] = uid
        session["username"] = username
        session["role"] = "admin"
    return jsonify({"ok": True, "id": uid})


@app.route("/api/users/<int:uid>", methods=["PUT"])
def api_users_update(uid):
    if session.get("role") != "admin":
        return jsonify({"error": "Sem permissão"}), 403
    body = request.json or {}
    role = body.get("role")
    pages = body.get("pages")
    password = body.get("password")
    conn = get_conn()
    with conn.cursor() as cur:
        if password:
            cur.execute("UPDATE app_users SET pw_hash = %s WHERE id = %s",
                        (_hash_pw(password), uid))
        if role and role in ("admin", "viewer"):
            cur.execute("UPDATE app_users SET role = %s WHERE id = %s", (role, uid))
        if pages is not None:
            if role == "admin":
                pages = list(ALL_PAGES)
            cur.execute("DELETE FROM user_permissions WHERE user_id = %s", (uid,))
            for pg in pages:
                if pg in ALL_PAGES:
                    cur.execute("INSERT INTO user_permissions (user_id, page) VALUES (%s, %s)",
                                (uid, pg))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/users/<int:uid>", methods=["DELETE"])
def api_users_delete(uid):
    if session.get("role") != "admin":
        return jsonify({"error": "Sem permissão"}), 403
    if uid == session.get("user_id"):
        return jsonify({"error": "Não é possível deletar o próprio usuário"}), 400
    conn = get_conn()
    with conn.cursor() as cur:
        cur.execute("DELETE FROM app_users WHERE id = %s", (uid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# Start scheduler
_ensure_schedules_table()
_ensure_turmas_table()
_ensure_ciclos_table()
_ensure_users_table()
_ensure_xl_snapshots_table()
_ensure_engagement_tables()
scheduler.start()
_load_schedules_from_db()
_register_engagement_job()


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5001, threaded=True, use_reloader=False)
